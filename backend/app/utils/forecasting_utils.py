
from flask import jsonify
from sqlalchemy import create_engine, MetaData, Table, text, inspect
from sqlalchemy.orm import sessionmaker
from config import Config
from calendar import month_name
from app.models.user_models import CountryProfile
from dotenv import load_dotenv
from dateutil.relativedelta import relativedelta
from concurrent.futures import ProcessPoolExecutor, as_completed
import calendar
from datetime import datetime, timedelta, date
import pandas as pd
import numpy as np
from io import BytesIO
import os, base64, warnings
from app.models.user_models import Inventory
import json
from statsmodels.tsa.holtwinters import ExponentialSmoothing
import pmdarima as pm
import re  # <-- added
from multiprocessing import cpu_count





warnings.filterwarnings("ignore")



# ============================== ENV / CONFIG ==============================
load_dotenv()
db_url = os.getenv('DATABASE_URL')
db_url2 = os.getenv('DATABASE_AMAZON_URL')
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")  # ChatGPT adjudicator key
ROLLING_HISTORY_MONTHS = 4  # 👈 compare last 4 months of actuals in ChatGPT/local adjudicator
ASP_ELASTICITY = -0.5
ASP_ADJUSTMENT_MIN = 0.80
ASP_ADJUSTMENT_MAX = 1.20
FLAT_FORECAST_REL_TOL = 0.01
FLAT_FORECAST_ABS_TOL = 2.0
FLAT_GUARD_MIN_TREND = 0.02
FLAT_GUARD_MAX_TREND = 0.08
_GPT_ADJUDICATOR_RATE_LIMITED = False

# ============================== SESSIONS & MAPS ==============================
def create_user_session_phormula(db_url):
    user_engine = create_engine(db_url)
    UserSession = sessionmaker(bind=user_engine)
    return UserSession()

def create_user_session(db_url2):
    user_engine1 = create_engine(db_url2)
    UserSession1 = sessionmaker(bind=user_engine1)
    return UserSession1()

MONTHS_REVERSE_MAP = {
    1: "january", 2: "february", 3: "march", 4: "april", 5: "may", 6: "june",
    7: "july", 8: "august", 9: "september", 10: "october", 11: "november", 12: "december"
}

MONTHS_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}

INVENTORY_MARKETPLACE_BY_COUNTRY = {
    "uk": "A1F83G8C2ARO7P",
    "gb": "A1F83G8C2ARO7P",
    "united kingdom": "A1F83G8C2ARO7P",
    "us": "ATVPDKIKX0DER",
    "usa": "ATVPDKIKX0DER",
    "united states": "ATVPDKIKX0DER",
}

INVENTORY_LOCATION_BY_COUNTRY = {
    "uk": "GB",
    "gb": "GB",
    "united kingdom": "GB",
    "us": "US",
    "usa": "US",
    "united states": "US",
}

# ============================== DATE HELPERS ==============================
def month_label(dt: datetime) -> str:
    return dt.strftime("%b'%y")

def add_months(dt: datetime, k: int) -> datetime:
    return dt + relativedelta(months=k)

def target_forecast_labels(req_month_str: str, req_year: int, n: int = 3) -> list[str]:
    m = MONTHS_MAP[req_month_str.lower()]
    start = datetime(req_year, m, 1)
    return [month_label(add_months(start, i)) for i in range(1, n + 1)]

# ============================== DATETIME PARSER (robust) ==============================
# Handles ISO-8601 (Z/+HH:MM), UK/EU day-first, textual months (incl. "Sept"),
# timezone abbreviations, 2-digit years, and fallback via _source_month.
_TZ_ABBREV_TO_OFFSET = {
    "UTC": "+00:00", "GMT": "+00:00",
    "BST": "+01:00",
    "CET": "+01:00", "CEST": "+02:00",
    "EET": "+02:00", "EEST": "+03:00",
    "IST": "+05:30",
    "MSK": "+03:00",
    "AST": "-04:00", "ADT": "-03:00",
    "EST": "-05:00", "EDT": "-04:00",
    "CST": "-06:00", "CDT": "-05:00",
    "MST": "-07:00", "MDT": "-06:00",
    "PST": "-08:00", "PDT": "-07:00",
}
_TZ_ABBREV_REGEX = re.compile(r'\b(' + '|'.join(map(re.escape, _TZ_ABBREV_TO_OFFSET.keys())) + r')\b', re.IGNORECASE)

def _replace_tz_abbrev_with_offset(s: pd.Series) -> pd.Series:
    def repl(m):
        tz = m.group(1).upper()
        return _TZ_ABBREV_TO_OFFSET.get(tz, tz)
    return s.str.replace(_TZ_ABBREV_REGEX, repl, regex=True)

def _expand_two_digit_years(s: pd.Series) -> pd.Series:
    # textual: '1 Aug 25' -> '1 Aug 2025'
    s = s.str.replace(
        r'\b([0-3]?\d)\s+([A-Za-z]{3,9})\s+(\d{2})(\b|[^0-9])',
        lambda m: f"{m.group(1)} {m.group(2)} 20{m.group(3)}{m.group(4)}",
        regex=True
    )
    # numeric day-first: '15/09/25' -> '15/09/2025' (also -, .)
    s = s.str.replace(
        r'\b([0-3]?\d)[/.\-]([01]?\d)[/.\-](\d{2})\b',
        lambda m: f"{m.group(1)}/{m.group(2)}/20{m.group(3)}",
        regex=True
    )
    return s

def parse_order_datetime_series(
    raw: pd.Series,
    source_month_hint: pd.Series | None = None
) -> pd.Series:
    """
    Returns a Series of naive UTC pandas Timestamps.

    GUARANTEES:
    - Output length == input length
    - No rows are dropped
    - Index alignment is preserved
    """

    
    # ------------------
    # Normalize raw strings
    # ------------------
    s = raw.astype(str).str.strip()
    s = s.str.replace(r'\bSept\.?\b', 'Sep', regex=True)
    s = _replace_tz_abbrev_with_offset(s)
    s = s.str.replace(r'\s{2,}', ' ', regex=True)

    # ------------------
    # Pass 1
    # ------------------
    dt1 = pd.to_datetime(
    s,
    errors='coerce',
    utc=True
    )

   
    # ------------------
    # Pass 2 (two-digit years)
    # ------------------
    needs2 = dt1.isna()
   

    if needs2.any():
        s2 = _expand_two_digit_years(s[needs2])
        s2 = _replace_tz_abbrev_with_offset(s2)

       

        dt2 = pd.to_datetime(
            s2,
            errors='coerce',
            utc=True
        )

       

        # 🔥 CRITICAL FIX: force index alignment
        dt2.index = dt1.index[needs2]
        dt1.loc[needs2] = dt2

    

    # ------------------
    # Pass 3 (fallback using source_month_hint)
    # ------------------
    needs3 = dt1.isna()
    

    if needs3.any() and source_month_hint is not None:
        src = source_month_hint.astype(str)

        day_guess = (
            s[needs3]
            .str.extract(r'\b([0-3]?\d)\b', expand=False)
            .astype('float')
            .fillna(1)
            .clip(lower=1, upper=28)
            .astype('int')
        )

        srcm = src[needs3].str.extract(r'([A-Za-z]+)', expand=False)
        srcy = src[needs3].str.extract(r'(\d{4})', expand=False)

        rebuilt = day_guess.astype(str) + " " + srcm + " " + srcy
        rebuilt = _replace_tz_abbrev_with_offset(rebuilt)

        dt3 = pd.to_datetime(
            rebuilt,
            errors='coerce',
            utc=True
        )

      
        # 🔥 CRITICAL FIX: force index alignment
        dt3.index = dt1.index[needs3]
        dt1.loc[needs3] = dt3

   
    # ------------------
    # Convert tz-aware UTC → naive UTC
    # ------------------
    if dt1.dt.tz is not None:
        dt1 = dt1.dt.tz_localize(None)

    

    # ------------------
    # HARD SAFETY GUARDRAIL
    # ------------------
    assert len(dt1) == len(raw), (
        f"[PARSE][FATAL] Row count changed "
        f"{len(raw)} → {len(dt1)}"
    )

    
    return dt1


# ============================== GROWTH HELPER ==============================
def _compute_growth_from_history(recent_hist):
    vals = [float(x) for x in recent_hist if x is not None]
    if len(vals) < 2:
        return 0.0

    eps = 1e-6
    def mom(a, b):  # MoM = a/b - 1
        return (a / max(b, eps)) - 1.0

    g1 = mom(vals[-1], vals[-2])
    g2 = mom(vals[-2], vals[-3]) if len(vals) >= 3 else np.nan
    g3 = mom(vals[-3], vals[-4]) if len(vals) >= 4 else np.nan

    if (g1 < 0) and (not np.isnan(g2) and g2 < 0):
        return 0.0
    if g1 < 0:
        return 0.10

    if g1 > 0.40:
        recent_gs = [x for x in [g1, g2, g3] if not np.isnan(x)]
        if len(recent_gs) >= 2:
            g = np.prod([1.0 + x for x in recent_gs]) ** (1.0 / len(recent_gs)) - 1.0
        else:
            g = g1
        return max(float(g), 0.0)

    return max(float(g1), 0.0)


def _finite_float_values(values):
    out = []
    for x in values:
        try:
            val = float(x)
        except Exception:
            continue
        if np.isnan(val) or np.isinf(val):
            continue
        out.append(val)
    return out


def _values_are_nearly_flat(values) -> bool:
    vals = np.asarray(_finite_float_values(values), dtype=float)
    if len(vals) < 2:
        return False

    spread = float(vals.max() - vals.min())
    scale = max(abs(float(vals.mean())), 1.0)
    return spread <= max(FLAT_FORECAST_ABS_TOL, scale * FLAT_FORECAST_REL_TOL)


def _history_has_movement(values) -> bool:
    vals = np.asarray(_finite_float_values(values), dtype=float)
    if len(vals) < 2:
        return False

    spread = float(vals.max() - vals.min())
    scale = max(abs(float(vals.mean())), 1.0)
    return spread > max(FLAT_FORECAST_ABS_TOL, scale * FLAT_FORECAST_REL_TOL)


def _compute_signed_trend_from_history(recent_hist) -> float:
    vals = np.asarray(_finite_float_values(recent_hist), dtype=float)
    vals = vals[-ROLLING_HISTORY_MONTHS:]

    if len(vals) < 2 or float(vals.max()) <= 0.0:
        return 0.0

    scale = max(float(np.mean(np.abs(vals))), 1.0)
    slope_pct = _slope(vals) / scale if len(vals) >= 3 else (vals[-1] - vals[-2]) / scale

    mom = []
    for prev, cur in zip(vals[:-1], vals[1:]):
        if prev > 0:
            mom.append((cur / prev) - 1.0)

    mom_typical = float(np.median(mom)) if mom else 0.0
    mom_typical = float(np.clip(mom_typical, -FLAT_GUARD_MAX_TREND, FLAT_GUARD_MAX_TREND))
    trend = (0.70 * float(slope_pct)) + (0.30 * mom_typical)

    if _history_has_movement(vals) and abs(trend) < FLAT_GUARD_MIN_TREND:
        direction_seed = slope_pct if abs(slope_pct) > 1e-6 else vals[-1] - vals[0]
        if abs(direction_seed) > 1e-6:
            trend = FLAT_GUARD_MIN_TREND if direction_seed > 0 else -FLAT_GUARD_MIN_TREND

    if np.isnan(trend) or np.isinf(trend):
        return 0.0

    return float(np.clip(trend, -FLAT_GUARD_MAX_TREND, FLAT_GUARD_MAX_TREND))


def _round_to_target_total(values, target_total):
    arr = np.asarray(values, dtype=float)
    arr = np.clip(arr, 0, None)

    if len(arr) == 0:
        return np.asarray([], dtype=int)

    target = int(max(0, np.rint(target_total)))
    floored = np.floor(arr).astype(int)
    remainder = target - int(floored.sum())

    if remainder > 0:
        fractions = arr - np.floor(arr)
        order = np.argsort(-fractions)
        for i in range(remainder):
            floored[order[i % len(order)]] += 1
    elif remainder < 0:
        for _ in range(abs(remainder)):
            candidates = np.where(floored > 0)[0]
            if len(candidates) == 0:
                break
            idx = candidates[np.argmax(floored[candidates])]
            floored[idx] -= 1

    return floored.astype(int)


def _apply_flat_forecast_guard(chosen_df: pd.DataFrame, lastN_daily: pd.Series, sku: str) -> pd.DataFrame:
    out = chosen_df.copy()
    forecasts = pd.to_numeric(out.get("Forecast"), errors="coerce").fillna(0).astype(float).to_numpy()

    if len(forecasts) < 2 or float(forecasts.sum()) <= 0.0:
        return out

    if not _values_are_nearly_flat(forecasts):
        return out

    if lastN_daily is None or lastN_daily.empty:
        return out

    history_months = _mk_monthly(lastN_daily).tail(ROLLING_HISTORY_MONTHS)
    trend = _compute_signed_trend_from_history(history_months.tolist())

    if abs(trend) < 1e-9:
        return out

    weights = np.power(1.0 + trend, np.arange(len(forecasts), dtype=float))
    if not np.isfinite(weights).all() or float(weights.sum()) <= 0.0:
        return out

    total = float(forecasts.sum())
    reshaped = total * weights / float(weights.sum())
    rounded = _round_to_target_total(reshaped, total)

    if _values_are_nearly_flat(rounded) and total >= len(forecasts):
        direction = 1.0 if trend > 0 else -1.0
        centered_offsets = (np.arange(len(forecasts), dtype=float) - ((len(forecasts) - 1) / 2.0)) * direction
        reshaped = (total / len(forecasts)) + centered_offsets
        rounded = _round_to_target_total(reshaped, total)

    old_ints = np.rint(forecasts).astype(int)
    if not np.array_equal(old_ints, rounded):
        out["Forecast"] = rounded
        print(
            f"[FORECAST][FLAT_GUARD] SKU={sku} trend={trend:.3f} "
            f"old={old_ints.tolist()} new={rounded.tolist()}"
        )

    return out


def _prefer_nonflat_candidate(winner: str, lastN_daily: pd.Series, arima_monthly: pd.Series, hybrid_monthly: pd.Series) -> str:
    if winner != "ARIMA":
        return winner

    if not _values_are_nearly_flat(arima_monthly.values):
        return winner

    if _values_are_nearly_flat(hybrid_monthly.values):
        return winner

    history_months = _mk_monthly(lastN_daily).tail(ROLLING_HISTORY_MONTHS)
    if not _history_has_movement(history_months.values):
        return winner

    score_a = _expert_score(history_months, arima_monthly)
    score_h = _expert_score(history_months, hybrid_monthly)
    if not np.isfinite(score_h):
        return winner
    if np.isfinite(score_a) and score_h > max(score_a * 1.25, score_a + 0.05):
        return winner

    return "HYBRID"


def _is_rate_limit_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("429" in msg) or ("too many requests" in msg) or ("rate limit" in msg)

# ============================== REMAINING MONTHS (per-SKU base) ==============================
def calculate_remaining_months_v2(
    user_id,
    country,
    inventory_forecast,
    forecast_horizon_months,
    recent_hist_map,
    base_months_map,
    anchor_months_all,
    last_month_col=None,  # ✅ NEW: pass e.g. "Nov'25 Sold"
):
    """
    Per-SKU extra months:
      extra[sku] = forecast_horizon_months - base
      base = 3 for ARIMA winner, 4 for HYBRID winner
    anchor_months_all: sorted list of the forecast month labels (first 3 are ARIMA anchors).
    Creates numeric future-month columns and fills forecasts using growth rule.

    last_month_col:
      - If provided, uses this column as the base-sales reference instead of
        legacy "Last Month Sales(Units)".
      - Example: last_month_col = "Nov'25 Sold"
    """

    def extra_for(sku):
        base = int(base_months_map.get(sku, 3))
        return max(int(forecast_horizon_months) - base, 0)

    # figure out max number of "extra" months needed across SKUs
    max_extra = 0
    for _, row in inventory_forecast.iterrows():
        s = row.get('sku', '')
        if s == 'Total':
            continue
        max_extra = max(max_extra, extra_for(s))

    if max_extra <= 0:
        return inventory_forecast

    # build future month columns starting after the 3rd forecast month
    third_label = anchor_months_all[2]
    third_dt = datetime.strptime(third_label, "%b'%y")
    start_dt = add_months(third_dt, 1)

    month_names = [month_label(add_months(start_dt, i)) for i in range(max_extra)]
    for m in month_names:
        if m not in inventory_forecast.columns:
            inventory_forecast[m] = 0.0

    # choose which "last month sales" column to use as base
    base_sales_col = last_month_col or 'Last Month Sales(Units)'

    for idx, row in inventory_forecast.iterrows():
        sku_key = row.get('sku', '')
        if sku_key == 'Total':
            continue

        base = int(base_months_map.get(sku_key, 3))
        extra = max(int(forecast_horizon_months) - base, 0)

        # ✅ base sales now comes from dynamic sold column (e.g., "Nov'25 Sold") if provided
        if row.get(base_sales_col, 0) > 0:
            base_sales = float(row[base_sales_col])
        elif row.get('Projected Sales Total', 0) > 0:
            # keep same fallback semantics
            base_sales = float(row['Projected Sales Total']) / 3.0
        else:
            base_sales = 0.0

        g = _compute_growth_from_history(recent_hist_map.get(sku_key, []))

        val = base_sales
        start_index = 0 if base == 3 else 1  # shift one month for HYBRID (its 4th is already in anchors)
        for m in month_names[start_index:start_index + extra]:
            inventory_forecast.at[idx, m] = int(np.rint(val))
            val *= (1.0 + g)

    return inventory_forecast


# ============================== PIPELINE UTILS ==============================
def debug_month_integrity(global_df, months_to_fetch):

    # Normalize type column (defensive)
    if 'type' in global_df.columns:
        global_df = global_df.copy()
        global_df['type'] = global_df['type'].astype(str).str.strip().str.title()
    else:
        global_df = global_df.copy()
        global_df['type'] = 'Unknown'

    valid_types = {'Order', 'Shipment'}

    # 2) Split by 'type' first (Orders + Shipments)
    parsed_dt = pd.to_datetime(global_df['date_time'], errors='coerce')
    by_type = (
        global_df.assign(_m=parsed_dt.dt.to_period('M'))
                 .groupby(['_m', 'type'])['sku']
                 .size()
                 .unstack(fill_value=0)
    )


    # 3) Which rows failed parsing?
    failed_parse = global_df[parsed_dt.isna()]
    if not failed_parse.empty:
        fp_by_month = (
            failed_parse.assign(_m=failed_parse['date_time'].astype(str).str.slice(0, 7))  # rough
                        .groupby('_m')['sku']
                        .size()
        )
    else:
        print("\n[PF DEBUG] No rows failed date parsing.")

    # 4) Orders + Shipments after regex-clean + parsing (matches pipeline)
    cleaned_dt = (
        global_df['date_time'].astype(str)
        .str.replace(r'\s(?:AM|PM)?\s?[A-Z]{3}$', '', regex=True)
        .pipe(pd.to_datetime, errors='coerce', infer_datetime_format=True)
    )

    demand_df = global_df[
        global_df['type'].isin(valid_types) & cleaned_dt.notna()
    ].copy()

    if demand_df.empty:
        return

    demand_df['_m'] = cleaned_dt[demand_df.index].dt.to_period('M')

    demand_range = pd.period_range(
        demand_df['_m'].min(),
        demand_df['_m'].max(),
        freq='M'
    )

    demand_count = (
        demand_df.groupby('_m')['sku']
                 .size()
                 .reindex(demand_range, fill_value=0)
    )

    

    # 5) Which expected months have zero demand after parsing?
    exp_periods = []
    for m in months_to_fetch:
        month = ''.join(filter(str.isalpha, m))
        year = ''.join(filter(str.isdigit, m))
        try:
            dt = pd.to_datetime(f"{month} {year}")
            exp_periods.append(dt.to_period('M'))
        except Exception:
            pass

    exp_periods = pd.PeriodIndex(exp_periods, freq='M')

    zero_months = [p for p in exp_periods if demand_count.get(p, 0) == 0]

    if zero_months:
        print(
            "\n[PF DEBUG] ⚠️ Months with ZERO orders/shipments after parsing:",
            [str(p) for p in zero_months]
        )
    else:
        print("\n[PF DEBUG] ✅ All expected months have some orders/shipments after parsing.")

# ============================== FORECAST PIPELINE ==============================


def process_forecasting(user_id, country, mv, year, engine, table_name_prefix="user"):
    """
    Build an orders+shipments dataframe for the last 12 full months (soft, best-effort):
      - Try per-month tables for each month in the 12M window.
      - If some months are missing, DO NOT raise; just log them.
      - ❌ DO NOT use merged/all-months table (Option 1).
      - Deduplicate and pass to generate_forecast.
    Then apply gating:
      - <5 latest contiguous months  -> ERROR 400
      - 5..11 latest contiguous months -> ARIMA ONLY
      - >=12 latest contiguous months  -> ARIMA + HYBRID
    """
    

    # --- 12-month window ending at last full month ---
    today = datetime.now()
    first_day_of_current_month = today.replace(day=1)
    last_full_month_start = (first_day_of_current_month - timedelta(days=1)).replace(day=1)
    # Normalize boundaries to avoid time offsets
    start_date = pd.Timestamp((last_full_month_start - relativedelta(months=11)).date())
    end_date = pd.Timestamp((last_full_month_start + relativedelta(months=1) - timedelta(days=1)).date()) \
            + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)


    months_to_fetch = [
        f"{month_name[dt.month]}{dt.year}"
        for dt in pd.date_range(start=start_date, end=end_date, freq="MS")
    ]

    # --- Introspect tables once ---
    meta = MetaData()
    meta.reflect(bind=engine)
    normalized_tables = {name.lower(): name for name in meta.tables.keys()}

    fetched_data = []
    missing_months = []

    # --- Fetch each monthly table if present ---
    with engine.connect() as conn:
        for month in months_to_fetch:
            table_name = f"{table_name_prefix}_{user_id}_{country}_{month}_data"
            tkey = table_name.lower()

            if tkey in normalized_tables:
                t_actual = normalized_tables[tkey]
                try:
                    df_m = pd.read_sql(
                        Table(t_actual, meta, autoload_with=engine).select(),
                        conn
                    )
                    df_m['_source_month'] = month
                    fetched_data.append(df_m)
                except Exception as e:
                    print(f"[PF][WARN] Error fetching {t_actual}: {e}")
            else:
                print(f"[PF][MISS] Monthly table not found: {table_name}")
                missing_months.append(month)

    if not fetched_data:
        return {
            "success": False,
            "error": "No data available for the selected window."
        }

    # --- Combine (raw) ---
    global_df = pd.concat(fetched_data, ignore_index=True)

    # Month integrity debug
    debug_month_integrity(global_df, months_to_fetch)


    # ============================================================
    # ✅ ORDER vs SHIPMENT — PRIORITY LOGIC (CRITICAL FIX)
    # ============================================================
    if 'type' in global_df.columns:
        type_norm = (
            global_df['type']
            .astype(str)
            .str.strip()
            .str.lower()
        )

        if (type_norm == 'shipment').any():
            filtered_df = global_df[type_norm == 'shipment'].copy()
            
        elif (type_norm == 'order').any():
            filtered_df = global_df[type_norm == 'order'].copy()

        else:

            filtered_df = global_df.copy()
    else:
   
        filtered_df = global_df.copy()

    

    if filtered_df.empty:
        return {
            "success": False,
            "error": "No usable demand rows available for forecasting."
        }

    if 'date_time' not in filtered_df.columns:
        return {
            "success": False,
            "error": "[PF][FATAL] 'date_time' column missing after fetch."
        }

    # =======================
    # Robust date parsing
    # =======================
    source_hint = filtered_df['_source_month'] if '_source_month' in filtered_df.columns else None

    filtered_df['date_time'] = parse_order_datetime_series(
        filtered_df['date_time'],
        source_hint
    )
 

    # Normalize valid timestamps (FIXED)
    filtered_df.loc[filtered_df['date_time'].notna(), 'date_time'] = (
        filtered_df.loc[filtered_df['date_time'].notna(), 'date_time']
        .dt.normalize()
    )

    


    # 🔥 CRITICAL FIX: impute missing dates using source month
    mask_missing_dt = filtered_df['date_time'].isna() & filtered_df['_source_month'].notna()

    if mask_missing_dt.any():
        filtered_df.loc[mask_missing_dt, 'date_time'] = (
            pd.to_datetime(
                filtered_df.loc[mask_missing_dt, '_source_month'],
                format='%B%Y',
                errors='coerce'
            )
            .dt.to_period('M')
            .dt.to_timestamp()
        )

    # Final safety drop (should be near-zero now)
    # 🔥 IMPUTE missing date_time using source month
    mask_missing_dt = filtered_df['date_time'].isna()

    if mask_missing_dt.any():
        filtered_df.loc[mask_missing_dt, 'date_time'] = (
            pd.to_datetime(
                filtered_df.loc[mask_missing_dt, '_source_month'],
                format='%B%Y',
                errors='coerce'
            )
            .dt.to_period('M')
            .dt.to_timestamp()
        )


    # Ensure required columns exist
    keep_cols = ['sku', 'date_time', 'quantity', 'price_in_gbp', 'type', 'product_name']
    for c in keep_cols:
        if c not in filtered_df.columns:
            filtered_df[c] = np.nan

    new_df = filtered_df[keep_cols].copy()
    new_df['quantity'] = (
        pd.to_numeric(new_df['quantity'], errors='coerce')
        .fillna(0)
        .astype(int)
    )

    new_df = new_df.sort_values(by='date_time').set_index('date_time')


    # 🔴 Clip to 12M window
    new_df = new_df.loc[
        (new_df.index >= start_date) &
        (new_df.index <= end_date)
    ]

    

    if new_df.empty:
        return {
            "success": False,
            "error": "No usable data inside the 12-month window."
        }
    


   
    # ---- Compute contiguous-month streak ----
    def _contiguous_streak_ending_at(end_period: pd.Period, month_index: pd.Index) -> int:
        months_present = set(pd.PeriodIndex(month_index.to_period('M')))
        streak = 0
        p = end_period
        while p in months_present:
            streak += 1
            p = p - 1
        return streak

    last_full_period = pd.Period(
        year=last_full_month_start.year,
        month=last_full_month_start.month,
        freq='M'
    )

    streak = _contiguous_streak_ending_at(last_full_period, new_df.index)
    distinct_months = int(new_df.index.to_period('M').nunique())

    if streak < 5:
        msg = (
            f"Insufficient recent data: only {streak} contiguous month(s) with demand up to {last_full_period}. "
            f"Need at least 5."
        )
        return {
            "success": False,
            "error": msg
        }

    hybrid_allowed = (streak >= 12)

    return generate_forecast(
        user_id,
        new_df,
        country,
        mv,
        year,
        hybrid_allowed=hybrid_allowed
    )


# ============================== ARIMA (3-MONTH FORECAST with DEBUG LOGS) ==============================
def forecast_next_two_months_with_append(sku_id, data, global_last_training_month=None):
    """
    ACCURACY VERSION B2 (model.fit on updated data, NO auto_arima search)
    """
    try:
        from pmdarima import auto_arima

        sku_data = data[data['sku'] == sku_id].copy()
        if sku_data.empty:
            
            return None

        sku_data = sku_data.drop(columns=['sku'])
        sku_data.index = pd.to_datetime(sku_data.index)
        sku_data = sku_data.resample('D').sum()
        sku_data['quantity'] = sku_data['quantity'].interpolate(method='linear').fillna(0)

        # Fit auto_arima ONCE to discover params

        auto_model = auto_arima(
            sku_data['quantity'],
            seasonal=True,
            trace=False,
            suppress_warnings=True,
            stepwise=True
        )

        # Extract discovered order params
        order = auto_model.order
        seasonal_order = auto_model.seasonal_order

        # Fit a model using discovered params
        model = auto_model

        current_data = sku_data['quantity'].copy()
        last_training_date = current_data.index[-1]
        last_obs_period = (
            global_last_training_month
            if global_last_training_month is not None
            else last_training_date.to_period('M')
        )

        # Month labels
        req_anchor = datetime(last_obs_period.year, last_obs_period.month, 1)
        intended_labels = [month_label(add_months(req_anchor, i)) for i in range(1, 4)]

        # Forecast horizon
        third_month_start = add_months(req_anchor, 3)
        end_of_third_month = add_months(third_month_start, 1) - pd.Timedelta(days=1)
        days_needed = (end_of_third_month - last_training_date).days + 1

        full_forecast = []

        for i in range(days_needed):
            next_val = float(model.predict(n_periods=1)[0])
            next_val = max(next_val, 0.0)

            next_date = current_data.index[-1] + pd.Timedelta(days=1)
            current_data.loc[next_date] = next_val
            full_forecast.append((next_date, next_val))

            # Re-fit using updated data — but WITHOUT auto_arima search
            model = auto_arima(
                current_data,
                seasonal=True,
                trace=False,
                suppress_warnings=True,
                stepwise=True,
                start_p=order[0], max_p=order[0],
                start_q=order[2], max_q=order[2],
                start_P=seasonal_order[0], max_P=seasonal_order[0],
                start_Q=seasonal_order[2], max_Q=seasonal_order[2],
                d=order[1], D=seasonal_order[1],
            )

        # Build output
        forecast_df = pd.DataFrame(full_forecast, columns=['Date', 'Forecast']).set_index('Date')
        forecast_df_filtered = forecast_df[forecast_df.index.to_period('M') > last_obs_period]

        monthly_summary = (
            forecast_df_filtered.resample('M').sum()
            .rename(columns={'Forecast': 'Forecast'})
            .reset_index()
        )
        monthly_summary['label'] = monthly_summary['Date'].dt.strftime("%b'%y")
        monthly_summary = monthly_summary.set_index('label').reindex(intended_labels, fill_value=0).reset_index()
        monthly_summary['Month'] = pd.to_datetime(monthly_summary['label'].str.replace("'", ""), format="%b%y")
        monthly_summary['sku'] = sku_id
        monthly_summary['Forecast'] = np.rint(monthly_summary['Forecast']).astype(int)

        return sku_id, monthly_summary[['Month', 'Forecast', 'sku']], sku_data

    except Exception as e:
        print(f"[ARIMA][ERROR] SKU {sku_id}: {e}")
        return None





def _hybrid_forecast_for_sku(
    sku_id,
    data,
    ship_time_weeks: int,
    stock_unit_weeks: int,
    global_last_training_month
):
    """
    Hybrid model (updated):
      - Chooses Croston / SARIMA / ETS automatically.
      - Handles intermittent & sparse series.
      - Forecasts exactly 4 months (no growth extension).
      - Never returns None unless absolutely no data.
      - 100% compatible with generate_forecast().
    """

    try:

        # === CONFIG ===
        SPARSE_ZERO_THRESHOLD   = 0.60
        MIN_SERIES_LENGTH       = 30
        SEASONAL_PERIOD_DAILY   = 7
        SEASONAL_PERIOD_WEEKLY  = 52
        MIN_ARIMA_POINTS_DAILY  = 60
        MIN_ARIMA_POINTS_WEEKLY = 30
        REQUIRE_VAR_POSITIVE    = True
        FORECAST_MONTHS         = 4
        APPLY_GROWTH_EXTENSION  = False

        # === HELPERS ===
        def _zero_ratio(s: pd.Series) -> float:
            return float((s == 0).sum()) / len(s) if len(s) else 1.0

        def _sufficient_for_arima(s: pd.Series, agg: str) -> bool:
            n = len(s)

            if agg == "daily":
                if n < max(MIN_SERIES_LENGTH, MIN_ARIMA_POINTS_DAILY):
                    return False
            else:
                if n < max(MIN_SERIES_LENGTH, MIN_ARIMA_POINTS_WEEKLY):
                    return False

            if REQUIRE_VAR_POSITIVE and np.isclose(s.var(ddof=1), 0.0):
                return False

            return True

        def _croston_sba(y: np.ndarray, h: int, alpha: float = 0.1) -> np.ndarray:
            y = np.asarray(y, dtype=float)

            last_y = None
            last_tau = None
            t = 0
            z = 0

            for v in y:
                t += 1

                if v > 0:
                    if last_y is None:
                        last_y = v
                        last_tau = t if last_tau is None else last_tau
                    else:
                        last_y = last_y + alpha * (v - last_y)
                        last_tau = last_tau + alpha * ((t - z) - last_tau)

                    z = t

            f = (
                0.0
                if (last_y is None or not last_tau)
                else (last_y / last_tau) * (1 - alpha / 2.0)
            )

            return np.full(h, f, dtype=float)

        def _ets(y: pd.Series, seasonal: int, horizon: int) -> np.ndarray:
            model = ExponentialSmoothing(
                y.astype(float),
                trend="add",
                seasonal="add",
                seasonal_periods=seasonal,
                initialization_method="estimated",
            )

            fit = model.fit(optimized=True)

            return fit.forecast(horizon).values.astype(float)

        def _auto_arima(y: pd.Series, seasonal: int, horizon: int) -> np.ndarray:
            model = pm.auto_arima(
                y.astype(float),
                start_p=0,
                start_q=0,
                max_p=3,
                max_q=3,
                start_P=0,
                start_Q=0,
                max_P=2,
                max_Q=2,
                seasonal=True,
                m=seasonal,
                d=None,
                D=None,
                stepwise=True,
                trace=False,
                error_action="ignore",
                suppress_warnings=True,
                information_criterion="aicc",
            )

            fc = model.predict(n_periods=horizon)

            return np.asarray(fc, dtype=float)

        # === DATA PREP ===
        sku_df = data[data['sku'] == sku_id].copy()

        if sku_df.empty:
            return None

        s_raw = sku_df['quantity'].copy()
        s_raw.index = pd.to_datetime(sku_df.index)
        s_raw = s_raw.sort_index()

        if s_raw.empty:
            return None

        last_date = s_raw.index.max()

     

        start_cut = last_date - pd.Timedelta(days=365)

        s_daily = (
            s_raw[s_raw.index >= start_cut]
            .resample("D")
            .sum()
            .astype(float)
            .fillna(0.0)
        )

        zr = _zero_ratio(s_daily)

        if (zr > SPARSE_ZERO_THRESHOLD) or (len(s_daily) < MIN_SERIES_LENGTH):

            s = s_daily.resample("W-MON").sum()

            agg = "weekly"
            seasonal = SEASONAL_PERIOD_WEEKLY
            horizon = 16

            last_dt = s.index.max()

           

            fc_idx = pd.date_range(
                last_dt + pd.offsets.Week(weekday=0),
                periods=horizon,
                freq="W-MON"
            )

        else:

            s = s_daily.asfreq("D", fill_value=0.0)

            agg = "daily"
            seasonal = SEASONAL_PERIOD_DAILY
            horizon = 120

            last_dt = s.index.max()

           

            fc_idx = pd.date_range(
                last_dt + pd.Timedelta(days=1),
                periods=horizon,
                freq="D"
            )

        if len(s) < MIN_SERIES_LENGTH:
         
            return None

        # =========================================================
        # GLOBAL FORECAST ANCHOR (FIX)
        # =========================================================

        anchor_dt = global_last_training_month.to_timestamp()

        future_months = pd.date_range(
            add_months(anchor_dt, 1),
            periods=FORECAST_MONTHS,
            freq="MS"
        )

      

        # === MODEL CHOICE ===
        try:

            if zr > SPARSE_ZERO_THRESHOLD:

                yhat = _croston_sba(s.values, horizon)
                model_name = "Croston-SBA"

            else:

                if _sufficient_for_arima(s, agg):

                    try:
                        yhat = _auto_arima(s, seasonal, horizon)
                        model_name = f"AUTO-SARIMA(m={seasonal})"

                    except Exception:

                        yhat = _ets(s, seasonal, horizon)
                        model_name = "ETS(A,A)"

                else:

                    yhat = _ets(s, seasonal, horizon)
                    model_name = "ETS(A,A)"

        except Exception as e:



            avg_val = (
                s_daily
                .resample("M")
                .sum()
                .tail(3)
                .mean()
            )

       

            monthly_out = pd.DataFrame({
                "sku": sku_id,
                "Month": future_months,
                "Forecast": [float(avg_val)] * len(future_months),
            })

            monthly_out["Forecast"] = (
                np.rint(monthly_out["Forecast"]).astype(int)
            )

    

            return (
                sku_id,
                monthly_out[["Month", "Forecast", "sku"]],
                s_daily
            )

        # === FORECAST BUILD ===

        yhat = np.clip(yhat, 0, None)

        fc = pd.DataFrame({
            "ds": fc_idx,
            "yhat": yhat
        })

        fc["Month"] = (
            fc["ds"]
            .dt.to_period("M")
            .dt.to_timestamp()
        )

        monthly = (
            fc.groupby("Month", as_index=False)["yhat"]
            .sum()
        )



        monthly = monthly[
            monthly["Month"].isin(future_months)
        ]

        monthly["Forecast"] = (
            np.rint(monthly["yhat"]).astype(int)
        )

        monthly_out = pd.DataFrame({
            "sku": sku_id,
            "Month": monthly["Month"],
            "Forecast": monthly["Forecast"]
        })



        return (
            sku_id,
            monthly_out[["Month", "Forecast", "sku"]],
            s_daily
        )

    except Exception as e:


        return None

# ============================== EXPERT ADJUDICATOR (local) ==============================
def _months_with_positive_history(daily_series: pd.Series) -> int:
    monthly = daily_series.resample('M').sum()
    return int((monthly > 0).sum())

def _mk_monthly(series_daily: pd.Series) -> pd.Series:
    return series_daily.resample('M').sum().astype(float)

def _slope(y: np.ndarray) -> float:
    if len(y) < 2:
        return 0.0
    x = np.arange(len(y), dtype=float)
    xm, ym = x.mean(), y.mean()
    denom = ((x - xm) ** 2).sum()
    if denom == 0:
        return 0.0
    return float(((x - xm) * (y - ym)).sum() / denom)

def _safe_ratio(a: float, b: float, eps: float = 1e-6) -> float:
    return float(a / max(b, eps))

def _expert_score(history_m: pd.Series, forecast_m: pd.Series) -> float:
    h = history_m.dropna().astype(float)
    if len(h) == 0 or len(forecast_m) == 0:
        return float('inf')

    h_vals = h.values
    f_vals = forecast_m.values.astype(float)

    h_std  = float(h_vals.std(ddof=1)) if len(h_vals) > 1 else 0.0
    h_slope = _slope(h_vals)
    hg = []
    for i in range(1, len(h_vals)):
        hg.append(_safe_ratio(h_vals[i], h_vals[i-1]) - 1.0)
    h_g_typical = float(np.median(hg)) if hg else 0.0

    f_std = float(f_vals.std(ddof=1)) if len(f_vals) > 1 else 0.0
    f_slope = _slope(f_vals)
    f1_jump = abs((_safe_ratio(f_vals[0], h_vals[-1]) - 1.0)) if len(h_vals) >= 1 else 0.0
    fg = []
    for i in range(1, len(f_vals)):
        fg.append(_safe_ratio(f_vals[i], f_vals[i-1]) - 1.0)
    f_g_typical = float(np.median(fg)) if fg else 0.0
    f_g_maxjump = float(max([abs(x) for x in fg])) if fg else 0.0

    score = (
        0.30 * abs(f_slope - h_slope) / (abs(h_slope) + 1e-3) +
        0.25 * abs(f_g_typical - h_g_typical) +
        0.20 * f1_jump +
        0.10 * (max(0.0, (f_std / (h_std + 1e-6)) - 2.0) if h_std > 0.0 else 0.0) +
        0.10 * abs(f_vals.mean() - h_vals[-1]) / (abs(h_vals[-1]) + 1e-3) +
        0.05 * f_g_maxjump
    )
    return float(score)

def _adjudicate_by_history_trend(lastN_daily: pd.Series, arima_monthly: pd.Series, hybrid_monthly: pd.Series) -> str:
    h_m_all = _mk_monthly(lastN_daily)
    h_lastN = h_m_all.tail(ROLLING_HISTORY_MONTHS) if len(h_m_all) >= 1 else h_m_all
    score_a = _expert_score(h_lastN, arima_monthly)
    score_h = _expert_score(h_lastN, hybrid_monthly)
    return 'HYBRID' if score_h < score_a else 'ARIMA'

# ============================== CHATGPT ADJUDICATOR (primary) ==============================
def call_chatgpt_adjudicator(
    lastN_months: list,
    arima_months: list,
    hybrid_months: list,
    ship_time_weeks: int,
    air_time_weeks: int,
    stock_unit_weeks: int,
    forecast_horizon_months: int,
    sku: str,
    country: str,
) -> str:
    """
    Uses ChatGPT to pick ARIMA vs HYBRID. Returns 'ARIMA' or 'HYBRID'.
    Falls back to local expert if key/lib not available or any error occurs.
    """
    global _GPT_ADJUDICATOR_RATE_LIMITED

    if not OPENAI_API_KEY:
        return None  # no key => signal caller to fallback

    if _GPT_ADJUDICATOR_RATE_LIMITED:
        return None

    try:
        # prefer official SDK if available
        try:
            from openai import OpenAI
            client = OpenAI(api_key=OPENAI_API_KEY)
            messages = [
                {
                    "role": "system",
                    "content": (
                        "You are a forecasting adjudicator. Given the last few months of actual demand and two candidate "
                        "forecasts (ARIMA=3 months, HYBRID=4 months), choose the single model that best continues the "
                        "recent trend. Prefer continuity of trend/growth and reasonable volatility. "
                        "Answer STRICTLY with 'ARIMA' or 'HYBRID'."
                    )
                },
                {
                    "role": "user",
                    "content": json.dumps({
                        "sku": sku,
                        "country": country,
                        "ship_time_weeks": ship_time_weeks,
                        "air_time_weeks": air_time_weeks,
                        "stock_unit_weeks": stock_unit_weeks,
                        "forecast_horizon_months": forecast_horizon_months,
                        "last_N_months_actual": lastN_months,  # 👈 now 4 values
                        "arima_forecast_3m": arima_months,
                        "hybrid_forecast_4m": hybrid_months
                    })
                }
            ]
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=messages,
                temperature=0.0,
                max_tokens=5,
            )
            text = resp.choices[0].message.content.strip().upper()
        except Exception as sdk_error:
            if _is_rate_limit_error(sdk_error):
                _GPT_ADJUDICATOR_RATE_LIMITED = True
                print("[GPT ADJUDICATOR] rate limited; using local adjudicator for remaining SKUs.")
                return None

            # HTTP fallback if SDK isn't available
            import requests
            url = "https://api.openai.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": "gpt-4o-mini",
                "temperature": 0.0,
                "max_tokens": 5,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a forecasting adjudicator. Given the last few months of actual demand and two candidate "
                            "forecasts (ARIMA=3 months, HYBRID=4 months), choose the single model that best continues the "
                            "recent trend. Prefer continuity of trend/growth and reasonable volatility. "
                            "Answer STRICTLY with 'ARIMA' or 'HYBRID'."
                        )
                    },
                    {
                        "role": "user",
                        "content": json.dumps({
                            "sku": sku,
                            "country": country,
                            "ship_time_weeks": ship_time_weeks,
                            "air_time_weeks": air_time_weeks,
                            "stock_unit_weeks": stock_unit_weeks,
                            "forecast_horizon_months": forecast_horizon_months,
                            "last_N_months_actual": lastN_months,  # 👈 now 4 values
                            "arima_forecast_3m": arima_months,
                            "hybrid_forecast_4m": hybrid_months
                        })
                    }
                ]
            }
            r = requests.post(url, headers=headers, data=json.dumps(payload), timeout=15)
            try:
                r.raise_for_status()
            except Exception as http_error:
                if _is_rate_limit_error(http_error):
                    _GPT_ADJUDICATOR_RATE_LIMITED = True
                    print("[GPT ADJUDICATOR] rate limited; using local adjudicator for remaining SKUs.")
                    return None
                raise
            text = r.json()["choices"][0]["message"]["content"].strip().upper()

        if "HYBRID" in text:
            winner = "HYBRID"
        elif "ARIMA" in text:
            winner = "ARIMA"
        else:
            winner = None

       

        return winner

    except Exception as e:
        print(f"[GPT ADJUDICATOR] error: {e}")
        return None



def _norm_sku(x: str) -> str:
    if x is None:
        return ""
    # remove all whitespace and upper-case so joins are robust
    return re.sub(r"\s+", "", str(x)).upper()


def _norm_country_key(country: str | None) -> str:
    return re.sub(r"\s+", " ", str(country or "").strip().lower())


def _inventory_table_country_key(country: str | None) -> str:
    key = _norm_country_key(country)
    if key in {"usa", "united states"}:
        return "us"
    if key in {"gb", "united kingdom"}:
        return "uk"
    return re.sub(r"[^a-z0-9_]", "_", key).strip("_")


def _safe_inventory_table_name(user_id: int, country: str | None, month_num: int, year: int) -> str | None:
    country_key = _inventory_table_country_key(country)
    if not country_key:
        return None

    table_name = f"inventorymonthly_{int(user_id)}_{country_key}_{int(month_num):02d}_{int(year)}"
    if not re.fullmatch(r"[a-z0-9_]+", table_name):
        return None

    return table_name


def _inventory_marketplace_id_for_country(country: str | None) -> str | None:
    return INVENTORY_MARKETPLACE_BY_COUNTRY.get(_norm_country_key(country))


def _inventory_location_for_country(country: str | None) -> str | None:
    return INVENTORY_LOCATION_BY_COUNTRY.get(_norm_country_key(country))


def get_inventory_snapshot_date(selected_month: str, selected_year: int) -> str:
    """
    Rule:
    - If selected month/year is current month/year -> use previous month's last date
    - Else -> use selected month's last date

    Returns: YYYY-MM-DD
    """
    month_num = datetime.strptime(selected_month.strip(), "%B").month
    today = date.today()

    # current month incomplete -> previous month end
    if selected_year == today.year and month_num == today.month:
        if today.month == 1:
            prev_month = 12
            prev_year = today.year - 1
        else:
            prev_month = today.month - 1
            prev_year = today.year

        last_day = calendar.monthrange(prev_year, prev_month)[1]
        return f"{prev_year:04d}-{prev_month:02d}-{last_day:02d}"

    # historical/completed month -> selected month end
    last_day = calendar.monthrange(selected_year, month_num)[1]
    return f"{selected_year:04d}-{month_num:02d}-{last_day:02d}"


def fetch_inventorymonthly_onhand_quantity(
    forecast_totals: pd.DataFrame,
    engine1,
    *,
    user_id: int,
    country: str | None,
    inventory_date: str,
    forecast_sku_col: str = "sku",
    debug: bool = False,
) -> pd.DataFrame:
    """
    Adds total_onhand_quantity from public.inventorymonthly_{user}_{country}_{MM}_{YYYY}.

    The month/year is derived from the same inventory snapshot date used for
    Inventory at Month End, so an August forecast uses the July snapshot table
    when the current-month rule selects July month end.
    """
    out = forecast_totals.copy()

    if forecast_sku_col not in out.columns:
        raise KeyError(f"forecast_totals missing SKU column '{forecast_sku_col}'")

    out["total_onhand_quantity"] = 0

    try:
        snapshot_dt = datetime.strptime(str(inventory_date), "%Y-%m-%d")
    except Exception:
        return out

    table_name = _safe_inventory_table_name(
        user_id=user_id,
        country=country,
        month_num=snapshot_dt.month,
        year=snapshot_dt.year,
    )
    if not table_name:
        return out

    try:
        inspector = inspect(engine1)
        if table_name not in set(inspector.get_table_names(schema="public")):
            if debug:
                print(f"[AWD] Inventory monthly table not found: {table_name}")
            return out

        columns = {col["name"] for col in inspector.get_columns(table_name, schema="public")}
        sku_col = next((c for c in ["msku", "sku", "seller_sku", "SKU"] if c in columns), None)
        if not sku_col or "total_onhand_quantity" not in columns:
            if debug:
                print(f"[AWD] Missing SKU/total_onhand_quantity in {table_name}")
            return out

        awd_df = pd.read_sql(
            text(f"""
                SELECT
                    "{sku_col}" AS "sku",
                    COALESCE(total_onhand_quantity, 0) AS total_onhand_quantity
                FROM public."{table_name}"
            """),
            con=engine1,
        )
    except Exception as e:
        if debug:
            print(f"[AWD] Error reading {table_name}: {e}")
        return out

    if awd_df.empty:
        return out

    awd_df["sku_norm"] = awd_df["sku"].map(_norm_sku)
    awd_df["total_onhand_quantity"] = (
        pd.to_numeric(awd_df["total_onhand_quantity"], errors="coerce")
        .fillna(0)
        .astype(int)
    )
    awd_totals = (
        awd_df[awd_df["sku_norm"] != ""]
        .groupby("sku_norm", as_index=False)["total_onhand_quantity"]
        .sum()
    )

    out["sku_norm"] = out[forecast_sku_col].map(_norm_sku)
    out = out.drop(columns=["total_onhand_quantity"]).merge(
        awd_totals,
        on="sku_norm",
        how="left",
    )
    out.drop(columns=["sku_norm"], inplace=True, errors="ignore")
    out["total_onhand_quantity"] = (
        pd.to_numeric(out["total_onhand_quantity"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    return out


def fetch_currentinventory_dispatch_quantities(
    forecast_totals: pd.DataFrame,
    engine1,
    *,
    user_id: int,
    country: str | None,
    inventory_date: str,
    forecast_sku_col: str = "sku",
    debug: bool = False,
) -> pd.DataFrame:
    """
    Adds the current-inventory dispatch columns from
    public.currentinventory_{user}_{country}_{month}{year}_table.
    """
    out = forecast_totals.copy()
    value_columns = {
        "total_onhand_quantity": "total_onhand_quantity",
        "inbound-shipped": "in_transit_fba",
        "total_inbound_quantity": "in_transit_awd",
        "total_stock": "total_sellable_in_stock",
        "total_transit": "total_sellable_in_transit",
    }

    for target_col in value_columns.values():
        if target_col not in out.columns:
            out[target_col] = 0

    if forecast_sku_col not in out.columns:
        return out

    try:
        snapshot_dt = datetime.strptime(str(inventory_date), "%Y-%m-%d")
    except Exception:
        return out

    country_key = (country or "").strip().lower().replace(" ", "_")
    if not country_key:
        return out

    table_name = (
        f"currentinventory_{int(user_id)}_{country_key}_"
        f"{snapshot_dt.strftime('%B').lower()}{snapshot_dt.year}_table"
    )
    if not re.fullmatch(r"[a-z0-9_]+", table_name):
        return out

    try:
        inspector = inspect(engine1)
        if table_name not in set(inspector.get_table_names(schema="public")):
            if debug:
                print(f"[CURRENT INVENTORY] Table not found: {table_name}")
            return out

        columns = {col["name"] for col in inspector.get_columns(table_name, schema="public")}
        sku_col = next((c for c in ["SKU", "sku", "msku", "seller_sku"] if c in columns), None)
        selected_value_cols = {
            source_col: target_col
            for source_col, target_col in value_columns.items()
            if source_col in columns
        }

        if not sku_col or not selected_value_cols:
            if debug:
                print(f"[CURRENT INVENTORY] Missing SKU/value columns in {table_name}")
            return out

        select_parts = [f'"{sku_col}" AS "sku"']
        for source_col, target_col in selected_value_cols.items():
            select_parts.append(f'COALESCE("{source_col}", 0) AS "{target_col}"')

        inventory_df = pd.read_sql(
            text(f"""
                SELECT {", ".join(select_parts)}
                FROM public."{table_name}"
            """),
            con=engine1,
        )
    except Exception as e:
        if debug:
            print(f"[CURRENT INVENTORY] Error reading {table_name}: {e}")
        return out

    if inventory_df.empty:
        return out

    inventory_df["sku_norm"] = inventory_df["sku"].map(_norm_sku)
    for target_col in selected_value_cols.values():
        inventory_df[target_col] = (
            pd.to_numeric(inventory_df[target_col], errors="coerce")
            .fillna(0)
            .round()
            .astype(int)
        )

    current_totals = (
        inventory_df[inventory_df["sku_norm"] != ""]
        .groupby("sku_norm", as_index=False)[list(selected_value_cols.values())]
        .sum()
    )

    out["sku_norm"] = out[forecast_sku_col].map(_norm_sku)
    out = out.drop(columns=list(value_columns.values()), errors="ignore").merge(
        current_totals,
        on="sku_norm",
        how="left",
    )
    out.drop(columns=["sku_norm"], inplace=True, errors="ignore")

    for target_col in value_columns.values():
        source_values = out[target_col] if target_col in out.columns else pd.Series(0, index=out.index)
        out[target_col] = (
            pd.to_numeric(source_values, errors="coerce")
            .fillna(0)
            .round()
            .astype(int)
        )

    return out


def fetch_fba_inbound_in_transit_quantities(
    forecast_totals: pd.DataFrame,
    engine1,
    *,
    user_id: int,
    country: str | None,
    forecast_sku_col: str = "sku",
    debug: bool = False,
) -> pd.DataFrame:
    """
    Overrides In Transit Inventory > FBA from public.inventory_fba_inbound_shipments.

    It uses rows whose status/shipment_status is IN_TRANSIT for US and SHIPPED
    for UK, grouped by SKU.
    New v2024 rows use msku + quantity; older v0 rows can use seller_sku +
    quantity_shipped as a fallback.
    """
    out = forecast_totals.copy()
    out["in_transit_fba"] = 0

    if forecast_sku_col not in out.columns:
        return out

    marketplace_id = INVENTORY_MARKETPLACE_BY_COUNTRY.get((country or "").strip().lower())
    if not marketplace_id:
        return out
    fba_status = "SHIPPED" if (country or "").strip().lower() in {"uk", "gb", "united kingdom"} else "IN_TRANSIT"

    try:
        inspector = inspect(engine1)
        if "inventory_fba_inbound_shipments" not in set(inspector.get_table_names(schema="public")):
            if debug:
                print("[FBA INBOUND] inventory_fba_inbound_shipments table not found")
            return out

        columns = {col["name"] for col in inspector.get_columns("inventory_fba_inbound_shipments", schema="public")}
        if "user_id" not in columns or "marketplace_id" not in columns:
            return out
        sku_expr = (
            "COALESCE(NULLIF(TRIM(msku), ''), NULLIF(TRIM(seller_sku), ''))"
            if {"msku", "seller_sku"}.issubset(columns)
            else "NULLIF(TRIM(msku), '')"
            if "msku" in columns
            else "NULLIF(TRIM(seller_sku), '')"
            if "seller_sku" in columns
            else None
        )
        if not sku_expr:
            return out

        quantity_expr = (
            "COALESCE(quantity, quantity_shipped, 0)"
            if {"quantity", "quantity_shipped"}.issubset(columns)
            else "COALESCE(quantity, 0)"
            if "quantity" in columns
            else "COALESCE(quantity_shipped, 0)"
            if "quantity_shipped" in columns
            else None
        )
        if not quantity_expr:
            return out

        status_expr = (
            "COALESCE(status, shipment_status, '')"
            if {"status", "shipment_status"}.issubset(columns)
            else "COALESCE(status, '')"
            if "status" in columns
            else "COALESCE(shipment_status, '')"
            if "shipment_status" in columns
            else "''"
        )

        fba_df = pd.read_sql(
            text(f"""
                SELECT
                    {sku_expr} AS sku,
                    SUM({quantity_expr}) AS in_transit_fba
                FROM public.inventory_fba_inbound_shipments
                WHERE user_id = :user_id
                  AND marketplace_id = :marketplace_id
                  AND UPPER(REPLACE({status_expr}, '-', '_')) = :fba_status
                GROUP BY {sku_expr}
            """),
            con=engine1,
            params={
                "user_id": int(user_id),
                "marketplace_id": marketplace_id,
                "fba_status": fba_status,
            },
        )
    except Exception as e:
        if debug:
            print(f"[FBA INBOUND] Error reading inbound shipments: {e}")
        return out

    if fba_df.empty:
        return out

    fba_df["sku_norm"] = fba_df["sku"].map(_norm_sku)
    fba_df["in_transit_fba"] = (
        pd.to_numeric(fba_df["in_transit_fba"], errors="coerce")
        .fillna(0)
        .round()
        .astype(int)
    )
    fba_totals = (
        fba_df[fba_df["sku_norm"] != ""]
        .groupby("sku_norm", as_index=False)["in_transit_fba"]
        .sum()
    )

    out["sku_norm"] = out[forecast_sku_col].map(_norm_sku)
    out = out.drop(columns=["in_transit_fba"], errors="ignore").merge(
        fba_totals,
        on="sku_norm",
        how="left",
    )
    out.drop(columns=["sku_norm"], inplace=True, errors="ignore")
    out["in_transit_fba"] = (
        pd.to_numeric(out["in_transit_fba"], errors="coerce")
        .fillna(0)
        .round()
        .astype(int)
    )

    return out


def fetch_awd_inbound_in_transit_quantities(
    forecast_totals: pd.DataFrame,
    engine1,
    *,
    user_id: int,
    country: str | None,
    inventory_date: str,
    forecast_sku_col: str = "sku",
    debug: bool = False,
) -> pd.DataFrame:
    """
    Adds AWD in-transit units from public.inventory_awd_inbound_shipments.

    Dispatch shows this as In Transit Inventory > AWD. It uses
    expected_unit_quantity for open shipment statuses, grouped by SKU.
    """
    out = forecast_totals.copy()
    out["in_transit_awd"] = 0
    for detail_col in [
        "awd_shipment_id",
        "awd_created_at",
        "awd_updated_at",
        "awd_ship_by",
        "awd_shipment_type",
        "awd_expected_reach_date",
    ]:
        out[detail_col] = ""

    if forecast_sku_col not in out.columns:
        return out

    marketplace_id = INVENTORY_MARKETPLACE_BY_COUNTRY.get((country or "").strip().lower())
    if not marketplace_id:
        return out

    try:
        datetime.strptime(str(inventory_date), "%Y-%m-%d")
    except Exception:
        inventory_date = date.today().strftime("%Y-%m-%d")

    try:
        inspector = inspect(engine1)
        if "inventory_awd_inbound_shipments" not in set(inspector.get_table_names(schema="public")):
            if debug:
                print("[AWD INBOUND] inventory_awd_inbound_shipments table not found")
            return out

        columns = {col["name"] for col in inspector.get_columns("inventory_awd_inbound_shipments", schema="public")}
        required = {
            "user_id",
            "marketplace_id",
            "shipment_status",
            "sku",
            "expected_unit_quantity",
            "shipment_type",
            "expected_reach_date",
        }
        if not required.issubset(columns):
            if debug:
                print(f"[AWD INBOUND] Missing columns: {sorted(required - columns)}")
            return out

        awd_df = pd.read_sql(
            text("""
                SELECT
                    sku,
                    COALESCE(asin, '') AS asin,
                    SUM(COALESCE(expected_unit_quantity, 0)) AS in_transit_awd,
                    STRING_AGG(DISTINCT shipment_id, ', ') AS awd_shipment_id,
                    STRING_AGG(DISTINCT TO_CHAR(created_at AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'), ', ')
                        FILTER (WHERE created_at IS NOT NULL) AS awd_created_at,
                    STRING_AGG(DISTINCT TO_CHAR(updated_at AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'), ', ')
                        FILTER (WHERE updated_at IS NOT NULL) AS awd_updated_at,
                    STRING_AGG(DISTINCT TO_CHAR(ship_by AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'), ', ')
                        FILTER (WHERE ship_by IS NOT NULL) AS awd_ship_by,
                    STRING_AGG(DISTINCT shipment_type, ', ')
                        FILTER (WHERE COALESCE(shipment_type, '') <> '') AS awd_shipment_type,
                    STRING_AGG(DISTINCT TO_CHAR(expected_reach_date, 'YYYY-MM-DD'), ', ')
                        FILTER (WHERE expected_reach_date IS NOT NULL) AS awd_expected_reach_date
                FROM public.inventory_awd_inbound_shipments
                WHERE user_id = :user_id
                  AND marketplace_id = :marketplace_id
                  AND COALESCE(expected_unit_quantity, 0) > 0
                  AND UPPER(REPLACE(COALESCE(shipment_status, ''), '-', '_')) NOT IN ('CANCELLED', 'CLOSED', 'DELIVERED')
                  AND (ship_by IS NULL OR ship_by::date >= CAST(:inventory_date AS date))
                  AND (expected_reach_date IS NULL OR expected_reach_date >= CAST(:inventory_date AS date))
                GROUP BY sku, COALESCE(asin, '')
            """),
            con=engine1,
            params={
                "user_id": int(user_id),
                "marketplace_id": marketplace_id,
                "inventory_date": inventory_date,
            },
        )
    except Exception as e:
        if debug:
            print(f"[AWD INBOUND] Error reading inbound shipments: {e}")
        return out

    if awd_df.empty:
        return out

    awd_df["sku_norm"] = awd_df["sku"].map(_norm_sku)
    awd_df["in_transit_awd"] = (
        pd.to_numeric(awd_df["in_transit_awd"], errors="coerce")
        .fillna(0)
        .round()
        .astype(int)
    )
    awd_totals = (
        awd_df[awd_df["sku_norm"] != ""]
        .groupby("sku_norm", as_index=False)
        .agg({
            "in_transit_awd": "sum",
            "awd_shipment_id": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
            "awd_created_at": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
            "awd_updated_at": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
            "awd_ship_by": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
            "awd_shipment_type": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
            "awd_expected_reach_date": lambda s: ", ".join(dict.fromkeys(v for x in s.dropna() for v in str(x).split(", ") if v)),
        })
    )

    out["sku_norm"] = out[forecast_sku_col].map(_norm_sku)
    out = out.drop(
        columns=[
            "in_transit_awd",
            "awd_shipment_id",
            "awd_created_at",
            "awd_updated_at",
            "awd_ship_by",
            "awd_shipment_type",
            "awd_expected_reach_date",
        ],
        errors="ignore",
    ).merge(
        awd_totals,
        on="sku_norm",
        how="left",
    )
    out.drop(columns=["sku_norm"], inplace=True, errors="ignore")
    out["in_transit_awd"] = (
        pd.to_numeric(out["in_transit_awd"], errors="coerce")
        .fillna(0)
        .round()
        .astype(int)
    )
    for detail_col in [
        "awd_shipment_id",
        "awd_created_at",
        "awd_updated_at",
        "awd_ship_by",
        "awd_shipment_type",
        "awd_expected_reach_date",
    ]:
        out[detail_col] = out[detail_col].fillna("").astype(str)

    return out


def add_air_sea_dispatch_split(
    inventory_forecast: pd.DataFrame,
    forecast_cols: list[str],
    *,
    air_time_weeks: int | float,
    stock_unit_weeks: int | float,
    total_units_col: str = "Shortfall Unit",
) -> pd.DataFrame:
    """
    Split positive shortfall units into AIR and SEA.

    AIR is the urgent quantity needed to cover the air lead-time window:
    air_time_weeks + stock_unit_weeks, converted to whole forecast months.
    SEA receives the remaining positive shortfall.
    """
    out = inventory_forecast.copy()
    usable_forecast_cols = [c for c in forecast_cols if c in out.columns]

    total_units_source = out[total_units_col] if total_units_col in out.columns else pd.Series(0, index=out.index)
    total_units = (
        pd.to_numeric(total_units_source, errors="coerce")
        .fillna(0)
        .clip(lower=0)
    )

    try:
        air_required_weeks = max(float(air_time_weeks or 0), 0) + max(float(stock_unit_weeks or 0), 0)
    except Exception:
        air_required_weeks = 0

    air_month_count = int(np.ceil(air_required_weeks / 4.345)) if air_required_weeks > 0 else 0
    air_month_count = min(max(air_month_count, 0), len(usable_forecast_cols))

    if air_month_count > 0:
        air_demand = (
            out[usable_forecast_cols[:air_month_count]]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .sum(axis=1)
        )
    else:
        air_demand = pd.Series(0, index=out.index)

    if "total_sellable_in_stock" in out.columns:
        stock_source = out["total_sellable_in_stock"]
    elif "Inventory at Month End" in out.columns:
        stock_source = out["Inventory at Month End"]
    else:
        stock_source = pd.Series(0, index=out.index)
    available_stock = (
        pd.to_numeric(stock_source, errors="coerce")
        .fillna(0)
    )
    urgent_air_units = (air_demand - available_stock).clip(lower=0)

    out["AIR"] = np.minimum(total_units, urgent_air_units).round().astype(int)
    out["SEA"] = (total_units - out["AIR"]).clip(lower=0).round().astype(int)

    return out


def fetch_and_merge_inventory_monthwise_sellable(
    forecast_totals: pd.DataFrame,
    engine1,
    *,
    user_id: int | None = None,
    country: str | None = None,
    forecast_sku_col: str = "sku",
    forecast_marketplace_col: str | None = None,
    inventory_date: str | None = None,
    debug: bool = False,
) -> pd.DataFrame:
    """
    Adds 'Inventory at Month End' to forecast_totals by pulling inventory from
    public.monthwise_inventory using ONLY disposition='SELLABLE'.

    Logic:
    - If inventory_date is provided:
        -> pick latest snapshot with date <= inventory_date
    - Else:
        -> pick latest snapshot per key
    - If country is provided:
        -> filter by marketplace_id and ledger location
           UK/GB = A1F83G8C2ARO7P + GB
           US = ATVPDKIKX0DER + US

    Merge:
    - always normalize SKU
    - optionally merge with marketplace_id also
    """

    if forecast_sku_col not in forecast_totals.columns:
        raise KeyError(f"forecast_totals missing SKU column '{forecast_sku_col}'")

    if forecast_marketplace_col is not None and forecast_marketplace_col not in forecast_totals.columns:
        raise KeyError(
            f"forecast_totals missing marketplace column '{forecast_marketplace_col}'"
        )

    use_marketplace = forecast_marketplace_col is not None

    inventory_marketplace_id = _inventory_marketplace_id_for_country(country)
    inventory_location = _inventory_location_for_country(country)

    if _norm_country_key(country) and (not inventory_marketplace_id or not inventory_location):
        raise ValueError(
            f"Unsupported inventory country '{country}'. Expected UK/GB or US."
        )

    params = {}
    where_clauses = ["LOWER(TRIM(disposition)) = 'sellable'"]

    if inventory_date:
        where_clauses.append("date <= :inv_date")
        params["inv_date"] = inventory_date

    if user_id is not None:
        where_clauses.append("user_id = :user_id")
        params["user_id"] = int(user_id)

    if inventory_marketplace_id:
        where_clauses.append("marketplace_id = :inventory_marketplace_id")
        params["inventory_marketplace_id"] = inventory_marketplace_id

    if inventory_location:
        where_clauses.append("UPPER(TRIM(COALESCE(location, ''))) = :inventory_location")
        params["inventory_location"] = inventory_location

    distinct_cols = "msku, marketplace_id" if use_marketplace else "msku"
    marketplace_select = (
        'marketplace_id AS "marketplace_id",\n                    '
        if use_marketplace else ""
    )
    where_sql = "\n                  AND ".join(where_clauses)

    sql = f"""
                SELECT DISTINCT ON ({distinct_cols})
                    msku AS "SKU",
                    {marketplace_select}ending_warehouse_balance AS "Ending Warehouse Balance",
                    date AS snapshot_date
                FROM public.monthwise_inventory
                WHERE {where_sql}
                ORDER BY {distinct_cols}, date DESC
            """

    inv_df = pd.read_sql(text(sql), con=engine1, params=params)

    if debug:

        if not inv_df.empty:
            print(inv_df.head(20).to_string(index=False))

    if inv_df.empty:
        out = forecast_totals.copy()
        out["Inventory at Month End"] = 0
        return out

    keep_cols = ["SKU", "Ending Warehouse Balance"] + (["marketplace_id"] if use_marketplace else [])
    inventory_totals = inv_df[keep_cols].copy()

    inventory_totals.rename(columns={"SKU": "sku"}, inplace=True)
    inventory_totals["Ending Warehouse Balance"] = (
        pd.to_numeric(inventory_totals["Ending Warehouse Balance"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    inventory_totals["sku_norm"] = inventory_totals["sku"].map(_norm_sku)

    # group duplicates safely
    if use_marketplace:
        inventory_totals = (
            inventory_totals
            .groupby(["sku_norm", "marketplace_id"], as_index=False)["Ending Warehouse Balance"]
            .sum()
        )
    else:
        inventory_totals = (
            inventory_totals
            .groupby(["sku_norm"], as_index=False)["Ending Warehouse Balance"]
            .sum()
        )

    out = forecast_totals.copy()
    out["sku_norm"] = out[forecast_sku_col].map(_norm_sku)

    if debug:
        forecast_keys = set(out["sku_norm"])
        inventory_keys = set(inventory_totals["sku_norm"])
        common = forecast_keys & inventory_keys



        missing = forecast_keys - inventory_keys


    if use_marketplace:
        inv_merge = inventory_totals[["sku_norm", "marketplace_id", "Ending Warehouse Balance"]].copy()

        out = out.merge(
            inv_merge,
            left_on=["sku_norm", forecast_marketplace_col],
            right_on=["sku_norm", "marketplace_id"],
            how="left"
        )
        out.drop(columns=["sku_norm", "marketplace_id"], inplace=True, errors="ignore")
    else:
        out = out.merge(
            inventory_totals[["sku_norm", "Ending Warehouse Balance"]],
            on="sku_norm",
            how="left"
        )
        out.drop(columns=["sku_norm"], inplace=True, errors="ignore")

    out.rename(columns={"Ending Warehouse Balance": "Inventory at Month End"}, inplace=True)
    out["Inventory at Month End"] = (
        pd.to_numeric(out["Inventory at Month End"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    return out

# def fetch_skuwise_monthly_sales(engine, meta, user_id, country, dt):
#     """
#     Fetch sku-wise monthly quantity and ADD label in-memory
#     """
#     table_name = f"skuwisemonthly_{user_id}_{country}_{dt.strftime('%B').lower()}{dt.year}"
#     table_key = table_name.lower()

#     if table_key not in {t.lower(): t for t in meta.tables}:
#         print(f"[SOLD][WARN] Table not found: {table_name}")
#         return pd.DataFrame(columns=['sku', 'total_quantity', 'Label'])

#     try:
#         tbl = Table(table_key, meta, autoload_with=engine)
#         with engine.connect() as conn:
#             df = pd.read_sql(tbl.select(), conn)

#         df['total_quantity'] = pd.to_numeric(df['total_quantity'], errors='coerce').fillna(0)

#         # ✅ CREATE LABEL HERE
#         df['Label'] = month_label(dt)   # e.g. "Sep'25"

#         return df[['sku', 'total_quantity', 'Label']]

#     except Exception as e:
#         print(f"[SOLD][ERROR] {table_name}: {e}")
#         return pd.DataFrame(columns=['sku', 'total_quantity', 'Label'])

def fetch_skuwise_monthly_sales(engine, meta, user_id, country, dt):
    """
    Fetch sku-wise monthly quantity + ASP and ADD label in-memory.

    ASP will be used internally for forecast adjustment.
    ASP will NOT be added to Excel output.
    """

    table_name = f"skuwisemonthly_{user_id}_{country}_{dt.strftime('%B').lower()}{dt.year}"
    table_key = table_name.lower()

    normalized_tables = {t.lower(): t for t in meta.tables.keys()}

    if table_key not in normalized_tables:
        print(f"[SOLD][WARN] Table not found: {table_name}")
        return pd.DataFrame(columns=["sku", "total_quantity", "asp", "Label"])

    try:
        actual_table_name = normalized_tables[table_key]
        tbl = Table(actual_table_name, meta, autoload_with=engine)

        with engine.connect() as conn:
            df = pd.read_sql(tbl.select(), conn)

        if "sku" not in df.columns:
            print(f"[SOLD][WARN] sku column missing in table: {table_name}")
            return pd.DataFrame(columns=["sku", "total_quantity", "asp", "Label"])

        if "total_quantity" not in df.columns:
            df["total_quantity"] = 0

        if "asp" not in df.columns:
            print(f"[ASP][WARN] asp column missing in table: {table_name}")
            df["asp"] = np.nan

        df["total_quantity"] = (
            pd.to_numeric(df["total_quantity"], errors="coerce")
            .fillna(0)
        )

        df["asp"] = pd.to_numeric(df["asp"], errors="coerce")

        # ASP is valid only where units sold > 0
        df.loc[df["total_quantity"] <= 0, "asp"] = np.nan

        df["Label"] = month_label(dt)

        return df[["sku", "total_quantity", "asp", "Label"]]

    except Exception as e:
        print(f"[SOLD][ERROR] {table_name}: {e}")
        return pd.DataFrame(columns=["sku", "total_quantity", "asp", "Label"])

def _norm_sku(x: str) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", "", str(x)).upper()


def _unique_cols(cols):
    return list(dict.fromkeys(cols))


def _to_monthly_series(df_m: pd.DataFrame) -> pd.Series:
    if df_m is None or df_m.empty:
        return pd.Series(dtype=float)

    month_idx = (
        pd.to_datetime(df_m["Month"], errors="coerce")
        .dt.to_period("M")
        .dt.to_timestamp("M")
    )
    good = ~month_idx.isna()
    if not good.any():
        return pd.Series(dtype=float)

    s = pd.Series(
        pd.to_numeric(df_m.loc[good, "Forecast"], errors="coerce"),
        index=month_idx[good],
    )
    s = s.dropna()
    s = s[~s.index.duplicated(keep="first")].sort_index()
    return s


def arima_forecast_3m_monthly(sku_id: str, sku_daily_qty: pd.Series, global_last_training_month) -> tuple | None:
    """
    Fast ARIMA: fit ONCE on monthly data and forecast 3 months.
    Returns: (sku, monthly_df, daily_series_used)
    monthly_df columns: Month, Forecast, sku
    """
    try:
        if sku_daily_qty is None or sku_daily_qty.empty:
            return None

        # last 12 months only (keeps model light)
        last_date = sku_daily_qty.index.max()
        cut = last_date - pd.Timedelta(days=365)
        s_daily = sku_daily_qty[sku_daily_qty.index >= cut].asfreq("D", fill_value=0.0)

        # monthly series
        s_m = s_daily.resample("M").sum().astype(float)
        s_m = s_m[s_m.index.notna()]

        if len(s_m) < 5:
            return None

        # Fit ARIMA once
        model = pm.auto_arima(
            s_m.values,
            seasonal=False,
            stepwise=True,
            suppress_warnings=True,
            error_action="ignore",
            trace=False,
        )

        fc = model.predict(n_periods=3)
        fc = np.clip(fc, 0, None)

        # anchor labels
        anchor_dt = global_last_training_month.to_timestamp()
        months = [add_months(anchor_dt, i) for i in range(1, 4)]

        monthly_df = pd.DataFrame(
            {
                "Month": months,
                "Forecast": np.rint(fc).astype(int),
                "sku": sku_id,
            }
        )

        return (sku_id, monthly_df[["Month", "Forecast", "sku"]], s_daily)

    except Exception as e:
        print(f"[ARIMA][ERROR] SKU {sku_id}: {e}")
        return None


def _build_sku_daily_map(new_df: pd.DataFrame) -> dict[str, pd.Series]:
    """
    Build minimal per-SKU daily series so we don't pickle/copy whole dataframes to worker processes.
    """
    out = {}
    # new_df index is date_time already
    for sku in new_df["sku"].unique():
        s = new_df[new_df["sku"] == sku]["quantity"]
        s = s.resample("D").sum().astype(float).fillna(0.0)
        out[sku] = s
    return out

def _safe_float(x, default=np.nan):
    try:
        if x is None:
            return default

        val = float(x)

        if np.isnan(val) or np.isinf(val):
            return default

        return val

    except Exception:
        return default


def build_latest_asp_change_map(sold_df: pd.DataFrame) -> dict:
    """
    Builds SKU-wise ASP movement using actual monthly ASP.

    Uses latest two valid ASP months where total_quantity > 0.

    Returns:
    {
        "SKU123": {
            "previous_asp": 10.0,
            "latest_asp": 12.0,
            "asp_change": 0.20
        }
    }
    """

    asp_map = {}

    if sold_df is None or sold_df.empty:
        return asp_map

    required_cols = {"sku", "asp", "Label"}

    if not required_cols.issubset(set(sold_df.columns)):
        return asp_map

    temp = sold_df.copy()

    temp["sku_norm"] = temp["sku"].map(_norm_sku)
    temp["asp"] = pd.to_numeric(temp["asp"], errors="coerce")

    if "total_quantity" in temp.columns:
        temp["total_quantity"] = (
            pd.to_numeric(temp["total_quantity"], errors="coerce")
            .fillna(0)
        )

        temp = temp[temp["total_quantity"] > 0]

    temp = temp.dropna(subset=["sku_norm", "asp", "Label"])

    if temp.empty:
        return asp_map

    temp["Label_dt"] = pd.to_datetime(
        temp["Label"].astype(str).str.replace("'", ""),
        format="%b%y",
        errors="coerce"
    )

    temp = temp.dropna(subset=["Label_dt"])
    temp = temp.sort_values(["sku_norm", "Label_dt"])

    for sku_norm, g in temp.groupby("sku_norm"):
        g = g.dropna(subset=["asp"]).sort_values("Label_dt")

        if len(g) < 2:
            continue

        previous_asp = _safe_float(g.iloc[-2]["asp"])
        latest_asp = _safe_float(g.iloc[-1]["asp"])

        if (
            np.isnan(previous_asp)
            or np.isnan(latest_asp)
            or previous_asp <= 0
            or latest_asp <= 0
        ):
            continue

        asp_change = (latest_asp / previous_asp) - 1.0

        asp_map[sku_norm] = {
            "previous_asp": previous_asp,
            "latest_asp": latest_asp,
            "asp_change": float(asp_change),
        }

    return asp_map


def apply_asp_adjustment_to_forecast(
    chosen_df: pd.DataFrame,
    sku: str,
    asp_change_map: dict,
    elasticity: float = ASP_ELASTICITY,
    min_factor: float = ASP_ADJUSTMENT_MIN,
    max_factor: float = ASP_ADJUSTMENT_MAX,
) -> pd.DataFrame:
    """
    Adjust forecasted units using actual ASP movement.

    Important:
    - Does NOT multiply forecast by ASP.
    - Does NOT add ASP columns to final Excel.
    - Only modifies Forecast internally.
    """

    out = chosen_df.copy()

    sku_norm = _norm_sku(sku)
    asp_info = asp_change_map.get(sku_norm)

    if not asp_info:
        return out

    asp_change = _safe_float(asp_info.get("asp_change"))

    if np.isnan(asp_change):
        return out

    adjustment_factor = 1.0 + (elasticity * asp_change)
    adjustment_factor = float(np.clip(adjustment_factor, min_factor, max_factor))

    out["Forecast"] = (
        pd.to_numeric(out["Forecast"], errors="coerce")
        .fillna(0)
        .astype(float)
        * adjustment_factor
    )

    out["Forecast"] = (
        np.rint(out["Forecast"])
        .clip(lower=0)
        .astype(int)
    )

    return out


def generate_forecast(user_id, new_df, country, mv, year, hybrid_allowed: bool = True):
    import time
    start_time = time.time()


    engine = create_engine(db_url)
    engine1 = create_engine(db_url2)  # Amazon DB
    meta = MetaData()
    meta.reflect(bind=engine)
    meta.reflect(bind=engine1)

    req_year = int(year)
    req_month_num = MONTHS_MAP[mv.lower()]

    # Anchor based on last available training month
    last_training_ts = new_df.index.max()
    global_last_training_month = last_training_ts.to_period("M")

    unique_skus = new_df["sku"].unique()
    all_forecasts = pd.DataFrame()
    model_winner = {}
    asp_change_map = {}

    profile = CountryProfile.query.filter_by(user_id=user_id, country=country).first()
    if not profile:
        raise ValueError(f"Country profile not found for user {user_id} and country {country}")
    ship_time_weeks = int(profile.ship_time_weeks or 0)
    air_time_weeks = int(profile.air_time_weeks or 0)
    stock_unit_weeks = int(profile.stock_unit_weeks or 0)

    # Forecast files are month-based. Use sea transit as the primary
    # replenishment route and convert the complete weekly requirement
    # into the required number of forecast months.
    total_requirement_weeks = ship_time_weeks + stock_unit_weeks
    forecast_horizon_months = max(
        1,
        int(np.ceil(total_requirement_weeks / 4.345)),
    )

    # HARD CAP workers to prevent EC2 OOM
    max_workers = min(2, max(1, cpu_count() - 1))


    # Build small per-SKU series map (cheap + avoids pickling huge DF)
    sku_daily_map = _build_sku_daily_map(new_df)

    # ---------------- ARIMA PARALLEL ----------------

    arima_results = {}
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futs = {
            executor.submit(arima_forecast_3m_monthly, sku, sku_daily_map[sku], global_last_training_month): sku
            for sku in unique_skus
            if sku in sku_daily_map
        }
        for fut in as_completed(futs):
            sku = futs[fut]
            try:
                res = fut.result()
                if res is not None:
                    arima_results[sku] = res
            except Exception as e:
                print(f"[ARIMA][ERROR] SKU={sku}: {e}")


    
    # ---------------- HYBRID PARALLEL ----------------
    hybrid_results = {}
    if hybrid_allowed:

        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futs = {
                executor.submit(
                    _hybrid_forecast_for_sku,
                    sku,
                    new_df,
                    ship_time_weeks,
                    stock_unit_weeks,
                    global_last_training_month
                ): sku
                for sku in unique_skus
            }
            for fut in as_completed(futs):
                sku = futs[fut]
                try:

                    res = fut.result()

                    if res is not None:
                        hybrid_results[sku] = res
                except Exception as e:
                    print(
                        f"[HYBRID][ERROR] SKU={sku}: {e} "
                        f"(fallback to ARIMA)"
                    )


    else:

        print(
            "[HYBRID] Disabled — ARIMA-only path based on streak gate."
        )


    # ---------------- SOLD + ASP HISTORY FOR ADJUSTMENT ----------------
    sold_anchor_dt = add_months(add_months(global_last_training_month.to_timestamp(), 1), -1)

    sold_m1 = month_label(sold_anchor_dt)
    sold_m2 = month_label(add_months(sold_anchor_dt, -1))
    sold_m3 = month_label(add_months(sold_anchor_dt, -2))

    sold_labels = [sold_m3, sold_m2, sold_m1]

    sold_month_dts = [
        add_months(sold_anchor_dt, -2),
        add_months(sold_anchor_dt, -1),
        sold_anchor_dt,
    ]

    sold_frames = []

    for dt in sold_month_dts:
        df_m = fetch_skuwise_monthly_sales(
            engine=engine,
            meta=meta,
            user_id=user_id,
            country=country,
            dt=dt,
        )

        if not df_m.empty:
            sold_frames.append(df_m)

    sold_df = (
        pd.concat(sold_frames, ignore_index=True)
        if sold_frames
        else pd.DataFrame(columns=["sku", "total_quantity", "asp", "Label"])
    )

    asp_change_map = build_latest_asp_change_map(sold_df)



    # ---------------- Adjudicate ----------------
    for sku in unique_skus:
        arima_res = arima_results.get(sku)
        hybrid_res = hybrid_results.get(sku)

        if arima_res is None and hybrid_res is None:
            continue

        # history for adjudication
        s_daily = sku_daily_map.get(sku, pd.Series(dtype=float))
        h_start = s_daily.index.max() - pd.DateOffset(months=ROLLING_HISTORY_MONTHS) if not s_daily.empty else None
        lastN_daily = s_daily[s_daily.index > h_start] if h_start is not None else s_daily

        if arima_res is not None and hybrid_res is not None:
            _, a_monthly_df, _ = arima_res
            _, h_monthly_df, _ = hybrid_res

            arima_series = _to_monthly_series(a_monthly_df)
            hybrid_series = _to_monthly_series(h_monthly_df)

            lastN_m = _mk_monthly(lastN_daily).tail(ROLLING_HISTORY_MONTHS).astype(float).tolist()
            a_list = a_monthly_df.sort_values("Month")["Forecast"].astype(float).tolist()[:3]
            h_list = h_monthly_df.sort_values("Month")["Forecast"].astype(float).tolist()[:4]

            gpt_choice = call_chatgpt_adjudicator(
                lastN_months=lastN_m,
                arima_months=a_list,
                hybrid_months=h_list,
                ship_time_weeks=ship_time_weeks,
                air_time_weeks=air_time_weeks,
                stock_unit_weeks=stock_unit_weeks,
                forecast_horizon_months=forecast_horizon_months,
                sku=sku,
                country=country,
            )

            if gpt_choice in ("ARIMA", "HYBRID"):
                winner = gpt_choice
            else:
                winner = _adjudicate_by_history_trend(lastN_daily, arima_series, hybrid_series)

            guarded_winner = _prefer_nonflat_candidate(winner, lastN_daily, arima_series, hybrid_series)
            if guarded_winner != winner:
                print(f"[FORECAST][FLAT_GUARD] SKU={sku} switched {winner} -> {guarded_winner} to avoid flat ARIMA.")
                winner = guarded_winner

            model_winner[sku] = winner
            chosen_df = h_monthly_df if winner == "HYBRID" else a_monthly_df

        elif arima_res is not None:
            model_winner[sku] = "ARIMA"
            _, chosen_df, _ = arima_res
        else:
            model_winner[sku] = "HYBRID"
            _, chosen_df, _ = hybrid_res

        chosen_df = chosen_df.sort_values("Month").copy()

        if model_winner[sku] == "ARIMA":
            chosen_df = chosen_df.iloc[:3]

        chosen_df = apply_asp_adjustment_to_forecast(
            chosen_df=chosen_df,
            sku=sku,
            asp_change_map=asp_change_map,
        )
        chosen_df = _apply_flat_forecast_guard(chosen_df, lastN_daily, sku)

        # attach price
        try:
            price_gbp_value = new_df[new_df["sku"] == sku]["price_in_gbp"].iloc[0]
        except Exception:
            price_gbp_value = None

        chosen_df["price_in_gbp"] = price_gbp_value
        chosen_df["sku"] = sku
        chosen_df["Forecast"] = pd.to_numeric(chosen_df["Forecast"], errors="coerce").fillna(0).pipe(np.rint).astype(int)

        all_forecasts = pd.concat(
            [all_forecasts, chosen_df[["Month", "Forecast", "sku", "price_in_gbp"]]],
            ignore_index=True,
        )

    # ---- forecasts_for file bytes ----
    forecast_filename = f"forecasts_for_{user_id}_{country}.xlsx"
    _af = all_forecasts.rename(columns={"Month": "month", "Forecast": "forecast"}).copy()
    forecast_buf = BytesIO()
    _af[["sku", "month", "forecast", "price_in_gbp"]].to_excel(forecast_buf, index=False, engine="openpyxl")
    forecast_buf.seek(0)
    forecast_bytes = forecast_buf.getvalue()

    # ---- pivot to month labels ----
    all_forecasts["Month"] = pd.to_datetime(all_forecasts["Month"], errors="coerce")
    all_forecasts = all_forecasts.dropna(subset=["Month"])
    all_forecasts["Month"] = all_forecasts["Month"].dt.strftime("%b'%y")

    anchor_dt = global_last_training_month.to_timestamp()
    arima_months = [month_label(add_months(anchor_dt, i)) for i in range(1, 4)]

    forecast_pivot = (
        all_forecasts.pivot_table(index="sku", columns="Month", values="Forecast", aggfunc="sum")
        .reset_index()
        .fillna(0)
        .round()
    )

    forecast_month_count = max(forecast_horizon_months, 3)
    requested_start_dt = datetime(req_year, req_month_num, 1)
    required_forecast_months = [
        month_label(add_months(requested_start_dt, i))
        for i in range(forecast_month_count)
    ]

    for col in _unique_cols(arima_months + required_forecast_months):
        if col not in forecast_pivot.columns:
            forecast_pivot[col] = 0

    monthwise_forecast_cols = [
        c for c in required_forecast_months
        if c in forecast_pivot.columns
    ]
    forecast_totals = forecast_pivot[["sku"] + monthwise_forecast_cols].copy()

    # ---- inventory snapshot fetch ----
    month_start = datetime(req_year, req_month_num, 1)
    month_end = (month_start + relativedelta(months=1)).replace(day=1)
    # ---- inventory snapshot fetch ----
    snapshot_date = get_inventory_snapshot_date(mv.title(), req_year)

    inventory_forecast = fetch_and_merge_inventory_monthwise_sellable(
        forecast_totals,
        engine1,
        user_id=user_id,
        country=country,
        inventory_date=snapshot_date,
    )
    inventory_forecast = fetch_inventorymonthly_onhand_quantity(
        inventory_forecast,
        engine1,
        user_id=user_id,
        country=country,
        inventory_date=snapshot_date,
    )
    inventory_forecast = fetch_currentinventory_dispatch_quantities(
        inventory_forecast,
        engine,
        user_id=user_id,
        country=country,
        inventory_date=snapshot_date,
    )
    inventory_forecast = fetch_fba_inbound_in_transit_quantities(
        inventory_forecast,
        engine1,
        user_id=user_id,
        country=country,
    )
    inventory_forecast = fetch_awd_inbound_in_transit_quantities(
        inventory_forecast,
        engine1,
        user_id=user_id,
        country=country,
        inventory_date=snapshot_date,
    )
    # ✅ PRODUCT NAME from monthly user_* tables (already present in new_df)
    product_names = pd.DataFrame(columns=["sku", "Product Name"])

    if "product_name" in new_df.columns:
        product_names = (
            new_df.reset_index()
            .dropna(subset=["sku"])
            .sort_values("date_time")
            .groupby("sku")["product_name"]
            .first()
            .reset_index()
            .rename(columns={"product_name": "Product Name"})
        )

    inventory_forecast = inventory_forecast.merge(product_names, on="sku", how="left")
    inventory_forecast["Product Name"] = inventory_forecast["Product Name"].fillna("")


    last3_sold_pivot = (
        sold_df.pivot_table(index="sku", columns="Label", values="total_quantity", aggfunc="sum")
        .reset_index()
        .fillna(0)
    )
    for lbl in sold_labels:
        if lbl not in last3_sold_pivot.columns:
            last3_sold_pivot[lbl] = 0

    last3_sold_pivot = last3_sold_pivot.rename(columns={lbl: f"{lbl} Sold" for lbl in sold_labels})
    inventory_forecast = inventory_forecast.merge(
        last3_sold_pivot[["sku", f"{sold_m3} Sold", f"{sold_m2} Sold", f"{sold_m1} Sold"]],
        on="sku",
        how="left",
    ).fillna(0)

    last_month_col = f"{sold_m1} Sold"

    # ---- recent hist map for growth ----
    monthly_actuals = (
        new_df.groupby("sku")["quantity"]
        .resample("M")
        .sum()
        .rename_axis(index=["sku", "Month"])
        .reset_index()
    )

    recent_hist_map = {}
    for sku, g in monthly_actuals.groupby("sku"):
        last4 = g.sort_values("Month").tail(4)["quantity"].tolist()
        if len(last4) >= 2:
            recent_hist_map[sku] = last4

    # ---- extend remaining months ----
    extra_months = max(forecast_horizon_months - 3, 0)
    start_after_third = add_months(anchor_dt, 4)
    future_month_columns = [month_label(add_months(start_after_third, i)) for i in range(extra_months)]
    future_month_columns = [m for m in future_month_columns if m not in monthwise_forecast_cols]
    for m in future_month_columns:
        if m not in inventory_forecast.columns:
            inventory_forecast[m] = 0.0

    base_months_map = {sku: (4 if model_winner.get(sku) == "HYBRID" else 3) for sku in inventory_forecast["sku"].tolist() if sku != "Total"}

    inventory_forecast = calculate_remaining_months_v2(
        user_id,
        country,
        inventory_forecast,
        forecast_horizon_months,
        recent_hist_map,
        base_months_map=base_months_map,
        anchor_months_all=monthwise_forecast_cols,
        last_month_col=last_month_col,
    )

    all_month_cols = _unique_cols(monthwise_forecast_cols + future_month_columns)

    inventory_forecast["Projected Sales Total"] = (
        inventory_forecast[all_month_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1).round().astype(int)
    )

    inventory_forecast.loc[inventory_forecast["sku"] == "Total", "Projected Sales Total"] = 0

    if "total_onhand_quantity" in inventory_forecast:
        inventory_forecast["total_onhand_quantity"] = (
            pd.to_numeric(inventory_forecast["total_onhand_quantity"], errors="coerce")
            .fillna(0)
            .round()
            .astype(int)
        )
    else:
        inventory_forecast["total_onhand_quantity"] = 0

    dispatch_in_stock = (
        inventory_forecast["Inventory at Month End"]
        + inventory_forecast["total_onhand_quantity"]
    ).round().astype(int)
    inventory_forecast["total_sellable_in_stock"] = dispatch_in_stock

    inventory_forecast["Dispatch"] = (
        (inventory_forecast["Projected Sales Total"] - dispatch_in_stock)
        .clip(lower=0)
        .round()
        .astype(int)
    )

    inventory_forecast["Current Inventory + Dispatch"] = (
        inventory_forecast["Dispatch"] + dispatch_in_stock
    ).astype(int)
    in_transit_fba = (
        inventory_forecast["in_transit_fba"]
        if "in_transit_fba" in inventory_forecast.columns
        else pd.Series(0, index=inventory_forecast.index)
    )
    in_transit_awd = (
        inventory_forecast["in_transit_awd"]
        if "in_transit_awd" in inventory_forecast.columns
        else pd.Series(0, index=inventory_forecast.index)
    )
    inventory_forecast["total_sellable_in_transit"] = (
        pd.to_numeric(in_transit_fba, errors="coerce").fillna(0)
        + pd.to_numeric(in_transit_awd, errors="coerce").fillna(0)
    ).round().astype(int)
    inventory_forecast["Shortfall Unit"] = (
        inventory_forecast["Projected Sales Total"]
        - dispatch_in_stock
    ).clip(lower=0).round().astype(int)
    inventory_forecast["To be Dispatch"] = (
        inventory_forecast["Shortfall Unit"]
        - inventory_forecast["total_sellable_in_transit"]
    ).clip(lower=0).round().astype(int)
    inventory_forecast = add_air_sea_dispatch_split(
        inventory_forecast,
        all_month_cols,
        air_time_weeks=air_time_weeks,
        stock_unit_weeks=stock_unit_weeks,
        total_units_col="To be Dispatch",
    )

    divisor = pd.to_numeric(inventory_forecast[last_month_col], errors="coerce").replace(0, np.nan)
    coverage_inventory = pd.to_numeric(
        inventory_forecast["total_sellable_in_stock"],
        errors="coerce",
    ).fillna(0)
    coverage = (coverage_inventory / divisor).round(2)
    inventory_forecast["Inventory Coverage Ratio Before Dispatch"] = coverage.where(coverage.notna(), "-")

    # totals row
    sold_cols = [f"{sold_m3} Sold", f"{sold_m2} Sold", f"{sold_m1} Sold"]
    numeric_columns = _unique_cols(
        [
            "Projected Sales Total",
            "Inventory at Month End",
            "total_onhand_quantity",
            "in_transit_fba",
            "in_transit_awd",
            "total_sellable_in_stock",
            "total_sellable_in_transit",
            "Dispatch",
            "Current Inventory + Dispatch",
            "Shortfall Unit",
            "To be Dispatch",
            "SEA",
            "AIR",
            last_month_col,
        ]
        + sold_cols
        + monthwise_forecast_cols
        + future_month_columns
    )
    numeric_columns = [c for c in numeric_columns if c in inventory_forecast.columns]

    sums = pd.to_numeric(inventory_forecast[numeric_columns].stack(), errors="coerce").unstack().fillna(0).sum()
    total_row = pd.DataFrame([np.rint(sums).astype(int)], columns=numeric_columns)
    total_row.insert(0, "sku", "Total")
    total_row["Product Name"] = "Total"
    total_row["SKU Type"] = "-"
    total_row["Inventory Coverage Ratio Before Dispatch"] = "-"
    inventory_forecast = pd.concat([inventory_forecast, total_row], ignore_index=True)

    # final columns
    final_columns = _unique_cols(
        [
            "sku",
            "Product Name",
            last_month_col,
            f"{sold_m3} Sold",
            f"{sold_m2} Sold",
            "Projected Sales Total",
            "Inventory at Month End",
            "total_onhand_quantity",
            "in_transit_fba",
            "in_transit_awd",
            "total_sellable_in_stock",
            "total_sellable_in_transit",
            "Inventory Coverage Ratio Before Dispatch",
            "Dispatch",
            "Current Inventory + Dispatch",
            "Shortfall Unit",
            "To be Dispatch",
            "SEA",
            "AIR",
        ]
        + monthwise_forecast_cols
        + future_month_columns
    )
    final_columns = [c for c in final_columns if c in inventory_forecast.columns]
    inventory_forecast = inventory_forecast[final_columns]

    # save bytes
    current_month = datetime.now().strftime("%b").lower()
    inventory_filename = f"inventory_forecast_{user_id}_{country}_{current_month}+2.xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ----------------- Build 2-sheet Excel -----------------
    inventory_sheet_cols = [
        "sku",
        "Product Name",
        f"{sold_m3} Sold",
        f"{sold_m2} Sold",
        f"{sold_m1} Sold",
    ] + monthwise_forecast_cols + future_month_columns + [
        "Projected Sales Total",
    ]

    inventory_sheet_cols = [c for c in inventory_sheet_cols if c in inventory_forecast.columns]

    dispatch_sheet_cols = [
        "sku",
        "Product Name",
        "Inventory at Month End",
        "total_onhand_quantity",
        "in_transit_fba",
        "in_transit_awd",
        "total_sellable_in_stock",
        "total_sellable_in_transit",
        "Projected Sales Total",
        "Inventory Coverage Ratio Before Dispatch",
        "Shortfall Unit",
        "To be Dispatch",
        "SEA",
        "AIR",
    ]
    dispatch_sheet_cols = [c for c in dispatch_sheet_cols if c in inventory_forecast.columns]

    inventory_sheet_df = inventory_forecast[inventory_sheet_cols].copy()
    dispatch_sheet_df = inventory_forecast[dispatch_sheet_cols].copy()
    dispatch_sheet_df.rename(
        columns={
            "Inventory at Month End": "FBA",
            "total_onhand_quantity": "AWD",
            "in_transit_fba": "FBA",
            "in_transit_awd": "AWD",
            "total_sellable_in_stock": "In stock",
            "total_sellable_in_transit": "In transit",
        },
        inplace=True,
    )

    current_month = datetime.now().strftime("%b").lower()
    inventory_filename = f"inventory_forecast_{user_id}_{country}_{current_month}+2.xlsx"

    # ---------- workbook ----------
    wb = Workbook()
    ws_inventory = wb.active
    ws_inventory.title = "Inventory"
    ws_dispatch = wb.create_sheet("Dispatch")

    # ---------- styles ----------
    title_fill = PatternFill("solid", fgColor="1F3B2D")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    meta_font = Font(bold=False, size=11)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, title_text, df, grouped_headers=None):
        # Title row
        ws["A1"] = title_text
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        # Merge title across columns
        max_cols = max(len(df.columns), 6)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
        company_name = "Skin"

        # Meta info
        ws["A2"] = f"Company Name : {company_name}"
        ws["A4"] = f"Country : {str(country).upper()}"
        ws["A5"] = "Platform : Phormula"

        for row in range(2, 6):
            ws[f"A{row}"].font = meta_font

        if grouped_headers:
            group_row = 6
            start_row = 7

            for c_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=group_row, column=c_idx, value="")
                cell.border = border
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for group_title, start_col, end_col in grouped_headers:
                ws.cell(row=group_row, column=start_col, value=group_title)
                ws.merge_cells(
                    start_row=group_row,
                    start_column=start_col,
                    end_row=group_row,
                    end_column=end_col,
                )
        else:
            start_row = 7

        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border

                if r_idx == start_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Autofit-ish widths
        from openpyxl.utils import get_column_letter

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

    write_sheet(ws_inventory, "Inventory Report", inventory_sheet_df)
    write_sheet(
        ws_dispatch,
        "Dispatch Report",
        dispatch_sheet_df,
        grouped_headers=[
            ("Inventory at Month End", 3, 4),
            ("In Transit Inventory", 5, 6),
            ("Total Sellable Inventory", 7, 8),
            ("Dispatch by Mode", 13, 14),
        ],
    )

    inv_buf = BytesIO()
    wb.save(inv_buf)
    inv_buf.seek(0)



    return {
        "forecast_filename": forecast_filename,
        "forecast_bytes": forecast_bytes,
        "inventory_filename": inventory_filename,
        "inventory_bytes": inv_buf.getvalue(),
        "model_winner": model_winner,
    }

# ============================== FILE ENCODER ==============================
def encode_file_to_base64(file_path):
    try:
        with open(file_path, "rb") as file:
            return base64.b64encode(file.read()).decode("utf-8")
    except Exception as e:
        print(f"Error encoding file to base64: {e}")
        return None

