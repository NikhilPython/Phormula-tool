from flask import Blueprint, request, jsonify , send_file , send_from_directory
import jwt
import os
from sqlalchemy import create_engine, MetaData, text, inspect, Table
import pandas as pd
from config import Config
SECRET_KEY = Config.SECRET_KEY
from app.utils.data_utils import MONTHS_MAP, MONTHS_REVERSE_MAP 
from app.utils.data_utils import send_forecast_email, send_pnlforecast_email 
from app.utils.token_utils import get_effective_user_id_from_token
from app.utils.forecasting_utils import process_forecasting
from app.models.user_models import UploadHistory , User, CountryProfile, StoredFile, db
from app.utils.manual_forecast_utils import generate_manual_forecast
from app.services.forecast_service import _normalize_forecast_month
from calendar import month_name
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from datetime import datetime
import numpy as np
import re
from decimal import Decimal
from sqlalchemy.orm import sessionmaker
import tempfile
from pathlib import Path
from io import BytesIO
from sqlalchemy.exc import IntegrityError
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)

load_dotenv()
db_url = os.getenv('DATABASE_URL')
db_url1= os.getenv('DATABASE_ADMIN_URL')
forecast_bp = Blueprint('forecast_bp', __name__)


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def send_db_file(stored: StoredFile, download_name: str):
    """Return a StoredFile as flask send_file response."""
    return send_file(
        BytesIO(stored.data),
        mimetype=stored.content_type or XLSX_MIME,
        as_attachment=True,
        download_name=download_name
    )



def ingest_xlsx_path_to_db(*, path: str, user_id, country, filename, kind, month=None, year=None, content_type=XLSX_MIME):
    """Read XLSX from a filesystem path (anywhere), store into DB, then delete it."""
    with open(path, "rb") as f:
        file_bytes = f.read()

    save_file_to_db(
        user_id=user_id,
        country=country,
        filename=filename,
        file_bytes=file_bytes,
        kind=kind,
        month=month,
        year=year,
        content_type=content_type,
    )

    try:
        os.remove(path)
    except Exception as e:
        print(f"⚠ Could not delete temp file {path}: {e}")

    return file_bytes


def find_latest_matching_file(pattern: str, search_dirs: list[str]) -> str | None:
    """
    Search for the newest file matching pattern in provided directories.
    pattern can include wildcards: e.g. 'inventory_forecast_12_us_*.xlsx'
    """
    newest_path = None
    newest_mtime = -1

    for d in search_dirs:
        try:
            for p in Path(d).glob(pattern):
                try:
                    mt = p.stat().st_mtime
                    if mt > newest_mtime:
                        newest_mtime = mt
                        newest_path = str(p)
                except Exception:
                    continue
        except Exception:
            continue

    return newest_path


def default_temp_dirs():
    """
    Places where legacy utils might write.
    Add more if your deployment uses a fixed folder.
    """
    dirs = [tempfile.gettempdir(), os.getcwd()]
    # optional: if you have env-configured temp output folder
    extra = os.getenv("FORECAST_TEMP_DIR")
    if extra:
        dirs.insert(0, extra)
    return dirs


def ingest_disk_xlsx_to_db(*, path: str, user_id, country, filename, kind, month=None, year=None):
    """Read XLSX from disk, store into DB, delete disk file."""
    with open(path, "rb") as f:
        file_bytes = f.read()

    save_file_to_db(
        user_id=user_id,
        country=country,
        filename=filename,
        file_bytes=file_bytes,
        kind=kind,
        month=month,
        year=year,
        content_type=XLSX_MIME,
    )

    try:
        os.remove(path)
    except Exception as e:
        print(f"⚠ Could not delete temp file {path}: {e}")

    return file_bytes

def load_forecast_df(*, user_id, country, filename, disk_fallback_path=None):
    stored = load_file_from_db(user_id=user_id, country=country, filename=filename)

    if stored:
        b = stored.data

        if country == "global":
            return pd.read_excel(BytesIO(b)), b   # ✅ FIX
        else:
            return pd.read_excel(BytesIO(b), header=6), b

    if disk_fallback_path and os.path.exists(disk_fallback_path):
        with open(disk_fallback_path, "rb") as f:
            b = f.read()

        if country == "global":
            return pd.read_excel(BytesIO(b)), b   # ✅ FIX
        else:
            return pd.read_excel(BytesIO(b), header=6), b

    return None, None

def clean_inventory_forecast_excel_bytes(file_bytes, mv, year, country):
    from openpyxl import load_workbook
    from io import BytesIO
    import re
    import pandas as pd

    month_lookup = {
        "january": 1, "jan": 1,
        "february": 2, "feb": 2,
        "march": 3, "mar": 3,
        "april": 4, "apr": 4,
        "may": 5,
        "june": 6, "jun": 6,
        "july": 7, "jul": 7,
        "august": 8, "aug": 8,
        "september": 9, "sep": 9,
        "october": 10, "oct": 10,
        "november": 11, "nov": 11,
        "december": 12, "dec": 12,
    }

    selected_month_num = month_lookup.get(str(mv).strip().lower())
    selected_year = int(year)

    if not selected_month_num:
        raise ValueError(f"Invalid forecast month: {mv}")

    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    header_row = 1 if str(country).lower() == "global" else 7

    cols_to_delete = []

    for cell in ws[header_row]:
        col_name = str(cell.value).strip() if cell.value is not None else ""

        # keep Feb'26 Sold / Mar'26 Sold / Apr'26 Sold
        if col_name.endswith(" Sold"):
            continue

        # only delete forecast month cols like Mar'26 / Apr'26
        if not re.match(r"^[A-Za-z]{3}'\d{2}$", col_name):
            continue

        dt = pd.to_datetime(col_name.replace("'", ""), format="%b%y")

        if dt.year < selected_year or (
            dt.year == selected_year and dt.month < selected_month_num
        ):
            cols_to_delete.append(cell.column)

    for col_idx in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col_idx)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_file_to_db(*, user_id, country, filename, file_bytes, kind, month=None, year=None, content_type=None):
    if content_type is None:
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # UPSERT-like behavior
    existing = StoredFile.query.filter_by(
        user_id=user_id,
        country=country,
        filename=filename
    ).first()

    if existing:
        existing.data = file_bytes
        existing.kind = kind
        existing.month = month
        existing.year = str(year) if year is not None else existing.year
        existing.content_type = content_type
        db.session.commit()
        return existing

    row = StoredFile(
        user_id=user_id,
        country=country,
        filename=filename,
        data=file_bytes,   # ✅ correct field
        kind=kind,
        month=month,
        year=str(year) if year is not None else None,
        content_type=content_type,
    )
    db.session.add(row)

    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()

        row2 = StoredFile.query.filter_by(
            user_id=user_id,
            country=country,
            filename=filename
        ).first()

        if row2:
            row2.data = file_bytes
            row2.kind = kind
            row2.month = month
            row2.year = str(year) if year is not None else row2.year
            row2.content_type = content_type
            db.session.commit()
            return row2

        raise

    return row

def load_file_from_db(*, user_id, country, filename):
    return (
        StoredFile.query
        .filter_by(user_id=user_id, country=country, filename=filename)
        .order_by(StoredFile.created_at.desc())
        .first()
    )

@forecast_bp.route('/api/forecast_allmonths', methods=['GET']) 
def forecast_allmonths():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        if not user_id:
            return jsonify({'error': 'Invalid token payload: user_id missing'}), 401

        country = request.args.get('country')
        mv = request.args.get('month')
        year = request.args.get('year')

        if not all([country, mv, year]):
            return jsonify({'error': 'Missing required parameters: country, month, or year'}), 400

        # NOTE: utils currently writes file using current month suffix (+2).
        current_month = datetime.now().strftime("%b").lower()

        
        output_file = (
            f'inventory_forecast_{user_id}_global_{current_month}+2.xlsx'
            if country == 'global'
            else f'inventory_forecast_{user_id}_{country}_{current_month}+2.xlsx'
        )

        stored = load_file_from_db(user_id=user_id, country=country, filename=output_file)
        if not stored:
            return jsonify({'error': 'Forecast file not found in DB. Please generate it first.'}), 404

        df = pd.read_excel(BytesIO(stored.data))


        # Identify month columns (normalize if needed)
        month_columns = []
        full_month_names = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        month_abbrev = {
            "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr", 
            "May": "May", "June": "Jun", "July": "Jul", "August": "Aug", 
            "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec"
        }

        year_short = str(year)[-2:] if year else datetime.now().strftime("%y")
        for col in df.columns:
            col_str = str(col)
            if re.match(r"^[A-Za-z]{3}'\d{2}\s+Sold$", col_str):
                month_columns.append(col_str)
            elif re.match(r"^[A-Za-z]{3}'\d{2}$", col_str):
                month_columns.append(col_str)
            elif col_str in full_month_names:
                abbrev = month_abbrev[col_str]
                new_col = f"{abbrev}'{year_short}"
                df.rename(columns={col_str: new_col}, inplace=True)
                month_columns.append(new_col)

        def month_key(col):
            try:
                return pd.to_datetime(col.replace(" Sold", "").replace("'", ""), format="%b%y")
            except:
                return pd.Timestamp.max

        month_columns = sorted(set(month_columns), key=month_key)

        # requested forecast start month, example: May 2026
        selected_month_num = MONTHS_MAP.get(str(mv).strip().lower())
        selected_year = int(year)

        def col_month_year(col):
            clean = str(col).replace(" Sold", "").replace("'", "")
            dt = pd.to_datetime(clean, format="%b%y")
            return dt.year, dt.month

        sold_columns = []
        forecast_columns = []

        for col in month_columns:
            col_str = str(col)

            if col_str.endswith(" Sold"):
                sold_columns.append(col_str)
            else:
                y, m = col_month_year(col_str)

                # only keep forecast month from selected month onwards
                if (y > selected_year) or (y == selected_year and m >= selected_month_num):
                    forecast_columns.append(col_str)

        month_columns = sold_columns + forecast_columns

        if country.lower() == 'global':
            df_filtered = df[df['sku'] != 'Total'].copy()
            selected_columns = ["Product Name"] + month_columns
            df_selected = df_filtered[selected_columns].copy()
            df_aggregated = df_selected.groupby('Product Name', as_index=False).sum()

            totals_row = {'Product Name': 'Total'}
            for col in month_columns:
                totals_row[col] = df_aggregated[col].sum()
            df_aggregated = pd.concat([df_aggregated, pd.DataFrame([totals_row])], ignore_index=True)

            df_aggregated.insert(0, 'S.no', range(1, len(df_aggregated) + 1))
            df_aggregated.iloc[-1, 0] = '-'
            columns = list(df_aggregated.columns)
        else:
            selected_columns = ["sku", "Product Name"] + month_columns

            if "Projected Sales Total" in df.columns:
                selected_columns.append("Projected Sales Total")
            df_aggregated = df[selected_columns].copy()
            columns = selected_columns

        filtered_df = df_aggregated.where(pd.notnull(df_aggregated), None)
        data = filtered_df.to_dict(orient='records')

        print("Selected columns passed to frontend:", columns)
        return jsonify({'columns': columns, 'data': data}), 200

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401
    except Exception as e:
        import traceback
        print("Unexpected error during forecast generation:")
        print(traceback.format_exc())
        return jsonify({'error': 'Internal server error', 'message': str(e)}), 500


@forecast_bp.route('/api/forecast_monthrange', methods=['GET'])
def forecast_monthrange():
    import re
    import traceback
    from io import BytesIO
    from datetime import datetime

    import jwt
    import pandas as pd
    from flask import request, jsonify

    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)

        if not user_id:
            return jsonify({'error': 'Invalid token payload: user_id missing'}), 401

        country = request.args.get('country')
        if not country:
            return jsonify({'error': 'Missing required parameter: country'}), 400

        country = str(country).strip().lower()
        current_month = datetime.now().strftime("%b").lower()

        output_file = (
            f'inventory_forecast_{user_id}_global_{current_month}+2.xlsx'
            if country == 'global'
            else f'inventory_forecast_{user_id}_{country}_{current_month}+2.xlsx'
        )

        stored = load_file_from_db(
            user_id=user_id,
            country=country,
            filename=output_file
        )

        if not stored:
            return jsonify({
                'error': 'Forecast file not found',
                'expected_filename': output_file
            }), 404

        # Global forecast file is generated with headers at row 0.
        # UK/US forecast files have report title rows, so headers are at row 7/index 6.
        if country == "global":
            df = pd.read_excel(BytesIO(stored.data), engine="openpyxl")
        else:
            df = pd.read_excel(BytesIO(stored.data), header=6, engine="openpyxl")

        # Normalize columns like "Jan'26 Sold" -> "Jan'26"
        df.columns = [str(c).strip() for c in df.columns]

        month_pattern = re.compile(r"^[A-Za-z]{3}'\d{2}(\s+Sold)?$")

        month_columns = [
            col for col in df.columns
            if month_pattern.match(str(col))
        ]

        if not month_columns:
            return jsonify({
                'error': 'No month columns found',
                'expected_format': "Jan'26, Feb'26, etc.",
                'columns': [str(c) for c in df.columns]
            }), 400

        def parse_month(col):
            return pd.to_datetime(
                str(col).replace(" Sold", "").replace("'", ""),
                format="%b%y"
            )

        month_columns_sorted = sorted(month_columns, key=parse_month)

        first_month = month_columns_sorted[0]
        last_month = month_columns_sorted[-1]

        return jsonify({
            "first_month": first_month,
            "last_month": last_month,
            "month_range": f"{first_month}-{last_month}",
            "month_columns": month_columns_sorted
        }), 200

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token expired'}), 401

    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    except Exception as e:
        print("Unexpected error in /api/forecast_monthrange:")
        print(traceback.format_exc())

        return jsonify({
            'error': 'Internal server error',
            'message': str(e)
        }), 500
    
       

def generate_forecast_core(user_id, country, mv, year, send_email_flag=True):
    country = str(country).strip().lower()
    mv = str(mv).strip().lower()
    year = str(year).strip()

    month_map = {
        "jan": "january",
        "feb": "february",
        "mar": "march",
        "apr": "april",
        "may": "may",
        "jun": "june",
        "jul": "july",
        "aug": "august",
        "sep": "september",
        "oct": "october",
        "nov": "november",
        "dec": "december",
    }
    mv = month_map.get(mv, mv)

from app.services.forecast_service import generate_forecast_for_user

@forecast_bp.route('/api/forecast', methods=['GET'])
def get_forecast():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        if not user_id:
            return jsonify({'error': 'Invalid token payload: user_id missing'}), 401

        country = request.args.get('country')
        year = request.args.get('year')

        mv = request.args.get('month')

        if mv:
            mv = _normalize_forecast_month(mv)
        else:
            return jsonify({"error": "Month parameter is required"}), 400

        if not all([country, mv, year]):
            return jsonify({'error': 'Missing required parameters: country, month, or year'}), 400

        country = str(country).strip().lower()
        mv = _normalize_forecast_month(mv)
        year = str(year).strip()

        result = generate_forecast_for_user(
            user_id=user_id,
            country=country,
            mv=mv,
            year=year,
            send_email=True,
        )

        if not result.get("success"):
            return jsonify({'error': result.get("error", "Forecast generation failed.")}), 500

        stored_inv = load_file_from_db(
            user_id=user_id,
            country=country,
            filename=result["filename"],
        )

        if not stored_inv:
            return jsonify({'error': 'Saved forecast not found in DB after generation.'}), 500

        cleaned_bytes = clean_inventory_forecast_excel_bytes(
            stored_inv.data,
            mv,
            year,
            country
        )

        stored_inv.data = cleaned_bytes
        db.session.commit()

        return send_file(
            BytesIO(cleaned_bytes),
            mimetype=XLSX_MIME,
            as_attachment=True,
            download_name=stored_inv.filename
        )

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': 'Internal server error', 'message': str(e)}), 500
    


@forecast_bp.route('/forecast_global', methods=['GET', 'POST'])
def forecast_global():
    if request.method == 'OPTIONS':
        return jsonify({'message': 'CORS Preflight OK'}), 200

    try:
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Authorization token is missing or invalid'}), 401

        token = auth_header.split(' ')[1]
        payload, user_id, member_id = get_effective_user_id_from_token(token)

        mv = request.args.get('month')
        year = request.args.get('year')

        if not all([mv, year]):
            return jsonify({'error': 'Missing required parameters: month or year'}), 400

        mv = str(mv).strip().lower()
        year = str(year).strip()

        current_month = datetime.now().strftime("%b").lower()

        uk_name = f'inventory_forecast_{user_id}_uk_{current_month}+2.xlsx'
        us_name = f'inventory_forecast_{user_id}_us_{current_month}+2.xlsx'
        global_name = f'inventory_forecast_{user_id}_global_{current_month}+2.xlsx'

        search_dirs = default_temp_dirs()

        uk_disk = None
        us_disk = None

        for d in search_dirs:
            p1 = os.path.join(d, uk_name)
            p2 = os.path.join(d, us_name)

            if not uk_disk and os.path.exists(p1):
                uk_disk = p1

            if not us_disk and os.path.exists(p2):
                us_disk = p2

        df_uk, uk_bytes = load_forecast_df(
            user_id=user_id,
            country='uk',
            filename=uk_name,
            disk_fallback_path=uk_disk
        )

        df_us, us_bytes = load_forecast_df(
            user_id=user_id,
            country='us',
            filename=us_name,
            disk_fallback_path=us_disk
        )

        present = []

        if df_uk is not None:
            present.append('uk')

        if df_us is not None:
            present.append('us')

        if not present:
            return jsonify({
                'error': 'No forecast files found for global processing',
                'expected_files': {
                    'uk': uk_name,
                    'us': us_name
                }
            }), 404

        if len(present) == 1:
            only = present[0]

            if only == 'uk':
                if uk_bytes and not load_file_from_db(
                    user_id=user_id,
                    country='uk',
                    filename=uk_name
                ):
                    save_file_to_db(
                        user_id=user_id,
                        country='uk',
                        filename=uk_name,
                        file_bytes=uk_bytes,
                        kind="inventory_forecast",
                        month=mv,
                        year=year,
                        content_type=XLSX_MIME
                    )

                stored = load_file_from_db(
                    user_id=user_id,
                    country='uk',
                    filename=uk_name
                )

                return send_db_file(stored, download_name=uk_name)

            if only == 'us':
                if us_bytes and not load_file_from_db(
                    user_id=user_id,
                    country='us',
                    filename=us_name
                ):
                    save_file_to_db(
                        user_id=user_id,
                        country='us',
                        filename=us_name,
                        file_bytes=us_bytes,
                        kind="inventory_forecast",
                        month=mv,
                        year=year,
                        content_type=XLSX_MIME
                    )

                stored = load_file_from_db(
                    user_id=user_id,
                    country='us',
                    filename=us_name
                )

                return send_db_file(stored, download_name=us_name)

        def normalize_forecast_df(df):
            df = df.copy()

            df.columns = [str(c).strip() for c in df.columns]

            rename_map = {}

            for col in df.columns:
                clean = str(col).strip().lower()

                if clean in ['product name', 'product_name', 'productname']:
                    rename_map[col] = 'Product Name'

                if clean == 'sku':
                    rename_map[col] = 'sku'

            df.rename(columns=rename_map, inplace=True)

            if 'sku' in df.columns:
                df['sku'] = df['sku'].astype(str).str.strip()

            if 'Product Name' in df.columns:
                df['Product Name'] = df['Product Name'].fillna('').astype(str).str.strip()

            return df

        df_uk = normalize_forecast_df(df_uk)
        df_us = normalize_forecast_df(df_us)

        required_cols = ['Product Name', 'sku']

        missing = {
            'uk_missing': [c for c in required_cols if c not in df_uk.columns],
            'us_missing': [c for c in required_cols if c not in df_us.columns],
        }

        if missing['uk_missing'] or missing['us_missing']:
            return jsonify({
                'error': 'Required merge columns missing',
                'missing': missing,
                'uk_columns': list(df_uk.columns),
                'us_columns': list(df_us.columns)
            }), 400

        month_pattern = re.compile(r"^[A-Za-z]{3}'\d{2}(\s+Sold)?$")

        month_cols_uk = [
            c for c in df_uk.columns
            if month_pattern.match(str(c))
        ]

        month_cols_us = [
            c for c in df_us.columns
            if month_pattern.match(str(c))
        ]

        def month_sort_key(col):
            try:
                return pd.to_datetime(str(col).replace(" Sold", "").replace("'", ""), format="%b%y")
            except Exception:
                return pd.Timestamp.max

        forecast_cols = sorted(
            set(month_cols_uk) | set(month_cols_us),
            key=month_sort_key
        )

        if not forecast_cols:
            return jsonify({
                'error': 'No forecast month columns found',
                'uk_columns': list(df_uk.columns),
                'us_columns': list(df_us.columns)
            }), 400

        keep_uk = ['Product Name', 'sku'] + month_cols_uk
        keep_us = ['Product Name', 'sku'] + month_cols_us

        df_uk = df_uk[keep_uk].copy()
        df_us = df_us[keep_us].copy()

        for col in month_cols_uk:
            df_uk[col] = pd.to_numeric(df_uk[col], errors='coerce').fillna(0)

        for col in month_cols_us:
            df_us[col] = pd.to_numeric(df_us[col], errors='coerce').fillna(0)

        merged = pd.merge(
            df_uk,
            df_us,
            on=['Product Name', 'sku'],
            how='outer',
            suffixes=('_uk', '_us')
        )

        for col in forecast_cols:
            col_uk = f"{col}_uk"
            col_us = f"{col}_us"

            if col_uk in merged.columns and col_us in merged.columns:
                merged[col] = (
                    pd.to_numeric(merged[col_uk], errors='coerce').fillna(0)
                    + pd.to_numeric(merged[col_us], errors='coerce').fillna(0)
                )
            elif col_uk in merged.columns:
                merged[col] = pd.to_numeric(merged[col_uk], errors='coerce').fillna(0)
            elif col_us in merged.columns:
                merged[col] = pd.to_numeric(merged[col_us], errors='coerce').fillna(0)
            elif col in merged.columns:
                merged[col] = pd.to_numeric(merged[col], errors='coerce').fillna(0)
            else:
                merged[col] = 0

        global_df = merged[['Product Name', 'sku'] + forecast_cols].copy()

        sku_df = (
            global_df.groupby('Product Name')['sku']
            .apply(lambda x: ', '.join(sorted(set(str(v).strip() for v in x if str(v).strip() and str(v).strip().lower() != 'nan'))))
            .reset_index()
        )

        global_df = (
            global_df
            .groupby('Product Name', as_index=False)[forecast_cols]
            .sum()
        )

        global_df = pd.merge(global_df, sku_df, on='Product Name', how='left')

        total_row = {
            'Product Name': 'Total',
            'sku': 'Total'
        }

        for col in forecast_cols:
            total_row[col] = global_df[col].sum()

        global_df = pd.concat(
            [global_df, pd.DataFrame([total_row])],
            ignore_index=True
        )

        buf = BytesIO()
        global_df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)

        save_file_to_db(
            user_id=user_id,
            country='global',
            filename=global_name,
            file_bytes=buf.getvalue(),
            kind="inventory_forecast",
            month=mv,
            year=year,
            content_type=XLSX_MIME
        )

        stored = load_file_from_db(
            user_id=user_id,
            country='global',
            filename=global_name
        )

        if not stored:
            return jsonify({
                'error': 'Global forecast generated but not found in DB'
            }), 500

        return send_db_file(
            stored,
            download_name=global_name
        )

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401

    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    except Exception as e:
        import traceback
        print("Unexpected error during global forecast generation:")
        print(traceback.format_exc())

        return jsonify({
            'error': 'Internal server error',
            'message': str(e)
        }), 500
    


@forecast_bp.route('/api/manual_forecast', methods=['POST'])
def manual_forecast():
    """
    Manual forecast route (no ML).

    Flow:
      1) Enforce that the last 4 months of sales tables (mv-1 .. mv-4) exist; if not, 404 with the missing list.
      2) Build orders-only new_df from those months (force each table's rows into its month if date parsing is bad).
      3) If preview=1 -> return JSON rows (no email, no file).
      4) If finalize (no preview) -> write files and return XLSX for the requested month.
    """
    from pandas.tseries.offsets import MonthEnd
    import re

    def _year_month_from_title(mon_title: str) -> tuple[int, int]:
        """mon_title like 'September2025' -> (2025, 9)"""
        m = re.match(r"([A-Za-z]+)(\d{4})$", mon_title)
        if not m:
            raise ValueError(f"Bad month token: {mon_title}")
        mon_name_s, year_s = m.groups()
        try:
            exp_month = list(month_name).index(mon_name_s)  # 1..12
        except ValueError:
            raise ValueError(f"Unknown month name in token: {mon_title}")
        return int(year_s), exp_month

    # --- Auth ---
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    try:
        # --- Inputs ---
        body = request.get_json(silent=True) or {}
        country_raw = (body.get('country') or request.args.get('country') or '').strip()
        mv_raw = (body.get('month') or request.args.get('month') or '').strip()
        year_raw = (body.get('year') or request.args.get('year') or '').strip()
        preview = str(request.args.get('preview', '') or body.get('preview', '')).lower() in ('1', 'true', 'yes')

        if not all([country_raw, mv_raw, year_raw]):
            return jsonify({'error': 'Missing required parameters: country, month, or year'}), 400

        # Normalize/validate inputs
        country = country_raw.lower()
        try:
            mv_lower = mv_raw.strip().lower()
            mv_num = MONTHS_MAP[mv_lower]
        except KeyError:
            return jsonify({'error': f'Invalid month value: {mv_raw}'}), 400

        try:
            req_year = int(str(year_raw))
        except Exception:
            return jsonify({'error': 'Year must be an integer'}), 400

        # Growth map normalization: supports list[{sku,pct}] or dict{sku:pct}
        growth_raw = body.get('growth', {})
        custom_growth_map = {}
        if isinstance(growth_raw, dict):
            for k, v in growth_raw.items():
                try:
                    custom_growth_map[str(k)] = float(v)
                except (TypeError, ValueError):
                    custom_growth_map[str(k)] = 0.0
        elif isinstance(growth_raw, list):
            for item in growth_raw:
                if not item:
                    continue
                sku = str(item.get('sku')) if 'sku' in item else None
                pct = item.get('pct', 0)
                if sku:
                    try:
                        custom_growth_map[sku] = float(pct)
                    except (TypeError, ValueError):
                        custom_growth_map[sku] = 0.0

        # --- Country profile (for horizon) ---
        profile = CountryProfile.query.filter_by(user_id=user_id, country=country).first()
        if not profile:
            return jsonify({'error': f'Country profile not found for user {user_id} and country {country}'}), 404

        transit_time = int(profile.transit_time or 0)
        stock_unit = int(profile.stock_unit or 0)

        # --- 1) Enforce "last 4 months" availability (mv-1 .. mv-4) ---
        ref_month = datetime(req_year, mv_num, 1)
        needed_tokens = []
        cur = ref_month - relativedelta(months=4)
        while cur <= ref_month - relativedelta(months=1):
            needed_tokens.append(f"{month_name[cur.month]}{cur.year}")  # e.g., "September2025"
            cur = cur + relativedelta(months=1)

        engine = create_engine(db_url)
        meta = MetaData()
        meta.reflect(bind=engine)
        all_tables = {t.lower(): t for t in meta.tables.keys()}

        missing = []
        for mon_title in needed_tokens:
            tname = f"user_{user_id}_{country}_{mon_title}_data"
            if tname.lower() not in all_tables:
                missing.append(mon_title)

        if missing:
            return jsonify({
                'error': 'missing_sales_months',
                'detail': 'Required sales tables are missing',
                'months': missing,
                'needed': needed_tokens
            }), 404

        # --- 2) Build orders-only new_df from those months (fix date_time per table) ---
        fetched_frames = []
        with engine.connect() as conn:
            for mon_title in needed_tokens:
                tname = f"user_{user_id}_{country}_{mon_title}_data"
                key = tname.lower()
                table = Table(all_tables[key], MetaData(), autoload_with=engine)

                try:
                    df_month = pd.read_sql(table.select(), conn)
                except Exception as e:
                    print(f"[manual_forecast] Error reading {tname}: {e}")
                    continue

                if df_month.empty:
                    print(f"[manual_forecast] Empty table: {tname}")
                    continue

                # --- DEBUG: before fixing dates
                try:
                    labels_before = sorted(
                        pd.to_datetime(df_month.get('date_time'), errors='coerce')
                        .dt.strftime("%b'%y").dropna().unique().tolist()
                    )
                except Exception:
                    labels_before = []
                print(f"[manual_forecast] {tname} rows={len(df_month)} date_labels_before={labels_before}")

                # Expected month/year from token, force if mismatch or NaT
                exp_year, exp_month = _year_month_from_title(mon_title)

                # 1) Parse to tz-aware UTC then make tz-naive (consistent dtype)
                dt_raw = pd.to_datetime(df_month.get('date_time'), errors='coerce', utc=True)
                try:
                    dt_raw = dt_raw.dt.tz_convert('UTC').dt.tz_localize(None)
                except Exception:
                    # if already tz-naive or all NaT
                    dt_raw = dt_raw.dt.tz_localize(None)

                # 2) Any NaT or mismatched month/year → force to the table's month-end
                mismatch = dt_raw.isna() | (dt_raw.dt.month != exp_month) | (dt_raw.dt.year != exp_year)
                fix_count = int(mismatch.sum())
                df_month['date_time'] = dt_raw
                df_month.loc[mismatch, 'date_time'] = (datetime(exp_year, exp_month, 1) + MonthEnd(0))

                # Normalize numeric types and SKU
                if 'quantity' in df_month.columns:
                    df_month['quantity'] = pd.to_numeric(df_month['quantity'], errors='coerce').fillna(0.0)
                if 'price_in_gbp' in df_month.columns:
                    df_month['price_in_gbp'] = pd.to_numeric(df_month['price_in_gbp'], errors='coerce')
                if 'sku' in df_month.columns:
                    df_month['sku'] = df_month['sku'].astype(str).str.strip()

                # --- DEBUG: after fixing dates
                labels_after = sorted(
                    pd.to_datetime(df_month['date_time'], errors='coerce')
                    .dt.strftime("%b'%y").dropna().unique().tolist()
                )
                print(f"[manual_forecast] {tname} fixed_dates -> forced={fix_count} date_labels_after={labels_after}")

                fetched_frames.append(df_month)

        if not fetched_frames:
            return jsonify({'error': f'No sales tables found for months: {needed_tokens}'}), 404

        global_df = pd.concat(fetched_frames, ignore_index=True)

        orders = global_df[global_df.get('type') == 'Order'].copy()
        if orders.empty:
            return jsonify({'error': 'No Order rows found in selected months'}), 404

        # Build new_df for the util
        new_df = orders[['sku', 'date_time', 'quantity', 'price_in_gbp']].copy()

        # Ensure tz-naive again (handles any leftovers)
        new_dt = pd.to_datetime(new_df['date_time'], errors='coerce', utc=True)
        new_dt = new_dt.dt.tz_convert('UTC').dt.tz_localize(None)  # now tz-naive
        new_df['date_time'] = new_dt

        # quantities numeric
        new_df['quantity'] = pd.to_numeric(new_df['quantity'], errors='coerce').fillna(0.0)

        # Optional debug: what months we are passing forward
        labels_pass = sorted(new_df['date_time'].dt.strftime("%b'%y").dropna().unique().tolist())
        print(f"[manual_forecast] concatenated orders: rows={len(new_df)} date_labels_passed={labels_pass}")

        # Now safe to sort/index
        new_df = new_df.sort_values(by='date_time').set_index('date_time')

        # --- 3) Run manual forecast util ---
        resp, status = generate_manual_forecast(
            user_id=user_id,
            new_df=new_df,
            country=country,
            mv=mv_lower,
            year=req_year,
            custom_growth_map=custom_growth_map,
            transit_time=transit_time,
            stock_unit=stock_unit,
            preview=preview,
        )

        if preview:
            # PREVIEW MODE: JSON rows for the UI; no email, no file
            return resp, status

        # --- 4) FINALIZE: email (optional) and return XLSX for the requested month ---
        try:
            send_forecast_email(user_id, None, mv_lower, req_year)
        except Exception as e:
            print(f"[manual_forecast] Email sending failed: {e}")

        requested_token = datetime(req_year, mv_num, 1).strftime("%b").lower()
        output_file = f'inventory_forecast_{user_id}_{country}_{requested_token}+2.xlsx'
        requested_token = datetime(req_year, mv_num, 1).strftime("%b").lower()
        output_file = f'inventory_forecast_{user_id}_{country}_{requested_token}+2.xlsx'

        # If util already saved into DB elsewhere, just return it
        stored = load_file_from_db(user_id=user_id, country=country, filename=output_file)
        if stored:
            return send_db_file(stored, download_name=output_file)

        
        search_dirs = default_temp_dirs()
        disk_path = None
        for d in search_dirs:
            p = os.path.join(d, output_file)
            if os.path.exists(p):
                disk_path = p
                break

        if not disk_path:
            disk_path = find_latest_matching_file(output_file, search_dirs)

        if not disk_path or not os.path.exists(disk_path):
            return resp, status  # util likely returned JSON already

        # Ingest into DB + delete disk file
        ingest_xlsx_path_to_db(
            path=disk_path,
            user_id=user_id,
            country=country,
            filename=output_file,
            kind="inventory_forecast",
            month=mv_lower,
            year=str(req_year),
        )

        stored = load_file_from_db(user_id=user_id, country=country, filename=output_file)
        return send_db_file(stored, download_name=output_file)


    except Exception as e:
        import traceback
        print("Unexpected error in /api/manual_forecast:")
        print(traceback.format_exc())
        return jsonify({'error': 'Internal server error', 'message': str(e)}), 500

@forecast_bp.route('/api/Pnlforecast', methods=['GET', 'POST'])
def Pnlforecast():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)

        country = request.args.get('country')
        month = request.args.get('month')
        year = request.args.get('year')

        if not country or not month or not year:
            return jsonify({'error': 'Missing required parameters: country, month or year'}), 400

        country = str(country).strip().lower()
        month = str(month).strip().lower()
        year = int(year)

        print(f"User ID: {user_id}, Country: {country}, Year: {year}, month: {month}")

        forecast_filename = f'forecasts_for_{user_id}_{country}.xlsx'
        stored = load_file_from_db(user_id=user_id, country=country, filename=forecast_filename)

        if not stored:
            return jsonify({'error': f'Forecast file for user {user_id} not found in DB'}), 404

        df = pd.read_excel(BytesIO(stored.data), engine='openpyxl')
        df.columns = [str(col).strip().lower() for col in df.columns]

        required_columns = ['sku', 'forecast', 'price_in_gbp', 'month']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            column_mapping = {
                'price_in_gbp': ['price in gbp', 'price_in_gbp', 'price_gbp', 'price'],
                'forecast': ['forecast', 'forecasted', 'forecast_qty', 'qty'],
                'month': ['date', 'forecast_date', 'period']
            }

            for req_col in missing_columns[:]:
                for alternative in column_mapping.get(req_col, []):
                    if alternative in df.columns:
                        df[req_col] = df[alternative]
                        missing_columns.remove(req_col)
                        print(f"Mapped column {alternative} to {req_col}")
                        break

        if missing_columns:
            return jsonify({'error': f'Missing required columns in forecast file: {missing_columns}'}), 500

        # Safe defaults
        if 'product_name' not in df.columns:
            df['product_name'] = ''

        df['sku'] = df['sku'].astype(str).str.strip()
        df['forecast'] = pd.to_numeric(df['forecast'], errors='coerce').fillna(0)
        df['price_in_gbp'] = pd.to_numeric(df['price_in_gbp'], errors='coerce').fillna(0)
        df['profit_percentage'] = 0.0
        df['avg_sales_price'] = 0.0
        df['product_name'] = df['product_name'].fillna('').astype(str)

        engine = create_engine(db_url)
        inspector = inspect(engine)

        month_table = f"skuwisemonthly_{user_id}_{country}_{month}{year}"
        month_table_alt = f"skuwisemonthly_{user_id}_{country}_{month}{year}_table"
        merge_table = f"skuwisemonthly_{user_id}_{country}"

        selected_month_table = None
        if inspector.has_table(month_table):
            selected_month_table = month_table
        elif inspector.has_table(month_table_alt):
            selected_month_table = month_table_alt

        merge_table_exists = inspector.has_table(merge_table)

        print(
            f"Table check => month_table: {month_table}={inspector.has_table(month_table)}, "
            f"month_table_alt: {month_table_alt}={inspector.has_table(month_table_alt)}, "
            f"merge_table: {merge_table}={merge_table_exists}"
        )

        from sqlalchemy import text

        with engine.connect() as connection:
            for sku in df['sku'].unique():
                result = None

                # First try month-specific table if present
                if selected_month_table:
                    try:
                        query = text(f"""
                            SELECT profit_percentage, net_sales, quantity, product_name
                            FROM {selected_month_table}
                            WHERE sku = :sku
                            LIMIT 1
                        """)
                        result = connection.execute(query, {"sku": sku}).fetchone()
                    except Exception as e:
                        print(f"Error querying month table for SKU {sku}: {str(e)}")
                        result = None

                # Fallback to merge table
                if result is None and merge_table_exists:
                    try:
                        query_merge = text(f"""
                            SELECT profit_percentage, net_sales, quantity, product_name
                            FROM {merge_table}
                            WHERE sku = :sku
                            ORDER BY
                                CAST(year AS INTEGER) DESC,
                                CASE LOWER(month)
                                    WHEN 'january' THEN 1
                                    WHEN 'february' THEN 2
                                    WHEN 'march' THEN 3
                                    WHEN 'april' THEN 4
                                    WHEN 'may' THEN 5
                                    WHEN 'june' THEN 6
                                    WHEN 'july' THEN 7
                                    WHEN 'august' THEN 8
                                    WHEN 'september' THEN 9
                                    WHEN 'october' THEN 10
                                    WHEN 'november' THEN 11
                                    WHEN 'december' THEN 12
                                END DESC
                            LIMIT 1
                        """)
                        result = connection.execute(query_merge, {"sku": sku}).fetchone()
                    except Exception as e:
                        print(f"Error querying merge table for SKU {sku}: {str(e)}")
                        result = None

                if result:
                    profit_percentage, net_sales, quantity, product_name = result
                    avg_sales_price = (net_sales / quantity) if quantity not in [0, None] else 0

                    df.loc[df['sku'] == sku, 'profit_percentage'] = float(profit_percentage or 0)
                    df.loc[df['sku'] == sku, 'avg_sales_price'] = float(avg_sales_price or 0)
                    df.loc[df['sku'] == sku, 'product_name'] = str(product_name or '')
                else:
                    print(f"⚠ SKU {sku} not found in month table or merge table. Using defaults.")
                    df.loc[df['sku'] == sku, 'profit_percentage'] = 0
                    df.loc[df['sku'] == sku, 'avg_sales_price'] = 0
                    # keep product_name as blank

        df['profit_percentage'] = pd.to_numeric(df['profit_percentage'], errors='coerce').fillna(0)
        df['avg_sales_price'] = pd.to_numeric(df['avg_sales_price'], errors='coerce').fillna(0)
        df['product_name'] = df['product_name'].fillna('').astype(str)

        df['Total_Sales'] = df['forecast'] * df.apply(
            lambda row: row['avg_sales_price']
            if pd.notnull(row['avg_sales_price']) and row['avg_sales_price'] > 0
            else row['price_in_gbp'],
            axis=1
        )

        df['profit'] = (df['profit_percentage'] / 100) * df['Total_Sales']
        df.loc[df['forecast'] == 0, 'profit_percentage'] = 0

        df['month'] = pd.to_datetime(df['month'], errors='coerce')
        df = df.sort_values(by=['sku', 'month'])
        df['forecast_month'] = df.groupby('sku').cumcount() + 1

        ordinal_map = {1: '1st', 2: '2nd', 3: '3rd'}
        df['forecast_month'] = df['forecast_month'].map(ordinal_map)

        if 'product_name' not in df.columns:
            df['product_name'] = ''

        product_names = df[['sku', 'product_name']].drop_duplicates(subset='sku')

        df_pivot = df.pivot(
            index='sku',
            columns='forecast_month',
            values=['forecast', 'profit_percentage', 'Total_Sales', 'profit']
        )

        df_pivot.columns = [f'{col[0]}_{col[1]}' for col in df_pivot.columns]
        df_pivot.reset_index(inplace=True)
        df_pivot = df_pivot.merge(product_names, on='sku', how='left')

        df_pivot['forecast_sum'] = df_pivot.filter(regex='^forecast_').sum(axis=1, skipna=True)
        df_pivot['profit_percentage_sum'] = df_pivot.filter(regex='^profit_percentage_').mean(axis=1, skipna=True)
        df_pivot['profit_sum'] = df_pivot.filter(regex='^profit_[^p]').sum(axis=1, skipna=True)
        df_pivot['Total_Sales_sum'] = df_pivot.filter(regex='^Total_Sales_').sum(axis=1, skipna=True)

        total_values = df_pivot.select_dtypes(include=['number']).sum()

        profit_percentage_columns = [col for col in df_pivot.columns if col.startswith('profit_percentage_')]
        for col in profit_percentage_columns:
            period = col.split('_')[-1]
            sales_col = f'Total_Sales_{period}'
            profit_col = f'profit_{period}'

            if sales_col in df_pivot.columns and profit_col in df_pivot.columns:
                total_profit = df_pivot[profit_col].sum()
                total_sales = df_pivot[sales_col].sum()
                total_values[col] = (total_profit / total_sales) * 100 if total_sales > 0 else 0

        total_row = pd.DataFrame([total_values])
        total_row.insert(0, 'sku', 'Total')
        total_row['product_name'] = 'Total'

        df_pivot = pd.concat([df_pivot, total_row], ignore_index=True)

        months = MONTHS_MAP.get(month.lower(), '1')
        try:
            months = int(months)
        except (ValueError, TypeError):
            months = 1

        prev_months = [(months - i) % 12 if (months - i) % 12 != 0 else 12 for i in range(3)]
        prev_years = [year if months - i > 0 else year - 1 for i in range(3)]

        acos_values = []
        reimbursement_vs_cm2_values = []

        for m, y in zip(prev_months, prev_years):
            month_name = MONTHS_REVERSE_MAP.get(m, 'january')
            record = UploadHistory.query.filter_by(
                user_id=user_id,
                country=country,
                month=month_name,
                year=str(y)
            ).first()

            if record and record.acos is not None:
                try:
                    acos_values.append(float(record.acos))
                except (ValueError, TypeError):
                    pass

            if record and record.rembursment_vs_cm2_margins is not None:
                try:
                    reimbursement_vs_cm2_values.append(float(record.rembursment_vs_cm2_margins))
                except (ValueError, TypeError):
                    pass

        avg_acos = sum(acos_values) / len(acos_values) if acos_values else 0

        prev_months_5 = [(months - i) % 12 if (months - i) % 12 != 0 else 12 for i in range(5)]
        prev_years_5 = [year if months - i > 0 else year - 1 for i in range(5)]

        platform_fee_values = []
        for m, y in zip(prev_months_5, prev_years_5):
            month_name = MONTHS_REVERSE_MAP.get(m, 'january')
            record = UploadHistory.query.filter_by(
                user_id=user_id,
                country=country,
                month=month_name,
                year=str(y)
            ).first()

            if record and record.platform_fee is not None:
                try:
                    platform_fee_values.append(float(record.platform_fee))
                except (ValueError, TypeError):
                    pass

        avg_platform_fee_percentage = (
            sum(platform_fee_values) / len(platform_fee_values)
            if platform_fee_values else 0
        )

        requested_month = month.lower()
        requested_year = int(year)

        current_month_name = datetime.now().strftime("%B").lower()
        current_year_num = datetime.now().year

        if requested_month == current_month_name and requested_year == current_year_num:
            fallback_month_num = datetime.now().month - 1
            fallback_year = requested_year

            if fallback_month_num == 0:
                fallback_month_num = 12
                fallback_year -= 1

            fallback_month = MONTHS_REVERSE_MAP[fallback_month_num]
            fallback_year = str(fallback_year)
        else:
            fallback_month = requested_month
            fallback_year = str(requested_year)

        print(f"📌 Using upload_history from {fallback_month} {fallback_year}")

        upload_history = UploadHistory.query.filter_by(
            user_id=user_id,
            country=country,
            month=fallback_month,
            year=fallback_year
        ).first()

        if not upload_history:
            return jsonify({'error': 'No upload history found for specified parameters'}), 404

        try:
            total_sales_1st = df_pivot.loc[df_pivot['sku'] == 'Total', 'Total_Sales_1st'].values[0] if 'Total_Sales_1st' in df_pivot.columns else 0
            total_sales_2nd = df_pivot.loc[df_pivot['sku'] == 'Total', 'Total_Sales_2nd'].values[0] if 'Total_Sales_2nd' in df_pivot.columns else 0
            total_sales_3rd = df_pivot.loc[df_pivot['sku'] == 'Total', 'Total_Sales_3rd'].values[0] if 'Total_Sales_3rd' in df_pivot.columns else 0

            cm1_profit_1st = df_pivot.loc[df_pivot['sku'] == 'Total', 'profit_1st'].values[0] if 'profit_1st' in df_pivot.columns else 0
            cm1_profit_2nd = df_pivot.loc[df_pivot['sku'] == 'Total', 'profit_2nd'].values[0] if 'profit_2nd' in df_pivot.columns else 0
            cm1_profit_3rd = df_pivot.loc[df_pivot['sku'] == 'Total', 'profit_3rd'].values[0] if 'profit_3rd' in df_pivot.columns else 0
        except (IndexError, KeyError):
            total_sales_1st = total_sales_2nd = total_sales_3rd = 0
            cm1_profit_1st = cm1_profit_2nd = cm1_profit_3rd = 0

        historical_total_sales = []
        advertising_total_values = []
        reimbursement_fee_values = []

        for m, y in zip(prev_months, prev_years):
            month_name = MONTHS_REVERSE_MAP.get(m, 'january')
            record = UploadHistory.query.filter_by(
                user_id=user_id,
                country=country,
                month=month_name,
                year=str(y)
            ).first()

            if record and record.total_sales is not None:
                try:
                    historical_total_sales.append(float(record.total_sales))
                except (ValueError, TypeError):
                    pass

            if record and record.advertising_total is not None:
                try:
                    advertising_total_values.append(float(record.advertising_total))
                except (ValueError, TypeError):
                    pass

            if record and record.rembursement_fee is not None:
                try:
                    reimbursement_fee_values.append(float(record.rembursement_fee))
                except (ValueError, TypeError):
                    pass

        total_sales_sum_3months = sum(historical_total_sales) if historical_total_sales else 0
        advertising_total_sum_3months = sum(advertising_total_values) if advertising_total_values else 0
        reimbursement_fee_sum_3months = sum(reimbursement_fee_values) if reimbursement_fee_values else 0

        acos1_value = abs((advertising_total_sum_3months / total_sales_sum_3months) * 100) if total_sales_sum_3months > 0 else 0
        acos2_value = acos1_value
        acos3_value = acos1_value

        advertising_total1 = abs((total_sales_1st * acos1_value) / 100)
        advertising_total2 = abs((total_sales_2nd * acos2_value) / 100)
        advertising_total3 = abs((total_sales_3rd * acos3_value) / 100)

        platform_fees1_value = abs(avg_platform_fee_percentage)
        platform_fees2_value = abs(avg_platform_fee_percentage)
        platform_fees3_value = abs(avg_platform_fee_percentage)

        cm2profit1_value = cm1_profit_1st - advertising_total1 - platform_fees1_value
        cm2profit2_value = cm1_profit_2nd - advertising_total2 - platform_fees2_value
        cm2profit3_value = cm1_profit_3rd - advertising_total3 - platform_fees3_value

        cm2margin1_value = (cm2profit1_value / total_sales_1st) * 100 if total_sales_1st > 0 else 0
        cm2margin2_value = (cm2profit2_value / total_sales_2nd) * 100 if total_sales_2nd > 0 else 0
        cm2margin3_value = (cm2profit3_value / total_sales_3rd) * 100 if total_sales_3rd > 0 else 0

        reimbursement_percentage = abs((reimbursement_fee_sum_3months / total_sales_sum_3months) * 100) if total_sales_sum_3months > 0 else 0

        NetReimbursement1_value = abs((total_sales_1st * reimbursement_percentage) / 100)
        NetReimbursement2_value = abs((total_sales_2nd * reimbursement_percentage) / 100)
        NetReimbursement3_value = abs((total_sales_3rd * reimbursement_percentage) / 100)

        ReimbursementvsCM2Margins1_value = (NetReimbursement1_value / abs(cm2profit1_value)) * 100 if abs(cm2profit1_value) > 0 else 0
        ReimbursementvsCM2Margins2_value = (NetReimbursement2_value / abs(cm2profit2_value)) * 100 if abs(cm2profit2_value) > 0 else 0
        ReimbursementvsCM2Margins3_value = (NetReimbursement3_value / abs(cm2profit3_value)) * 100 if abs(cm2profit3_value) > 0 else 0

        Reimbursementvssales1_value = abs(reimbursement_percentage)
        Reimbursementvssales2_value = abs(reimbursement_percentage)
        Reimbursementvssales3_value = abs(reimbursement_percentage)

        platform_fees_rows = pd.DataFrame([
            {'sku': 'Platform_Fees1', 'value': abs(platform_fees1_value)},
            {'sku': 'Platform_Fees2', 'value': abs(platform_fees2_value)},
            {'sku': 'Platform_Fees3', 'value': abs(platform_fees3_value)}
        ])

        advertising_rows = pd.DataFrame([
            {'sku': 'advertising_total1', 'value': abs(advertising_total1)},
            {'sku': 'advertising_total2', 'value': abs(advertising_total2)},
            {'sku': 'advertising_total3', 'value': abs(advertising_total3)}
        ])

        cm2profit_rows = pd.DataFrame([
            {'sku': 'cm2profit1', 'value': cm2profit1_value},
            {'sku': 'cm2profit2', 'value': cm2profit2_value},
            {'sku': 'cm2profit3', 'value': cm2profit3_value}
        ])

        cm2margin_rows = pd.DataFrame([
            {'sku': 'cm2margin1', 'value': cm2margin1_value},
            {'sku': 'cm2margin2', 'value': cm2margin2_value},
            {'sku': 'cm2margin3', 'value': cm2margin3_value}
        ])

        NetReimbursement_rows = pd.DataFrame([
            {'sku': 'NetReimbursement1', 'value': abs(NetReimbursement1_value)},
            {'sku': 'NetReimbursement2', 'value': abs(NetReimbursement2_value)},
            {'sku': 'NetReimbursement3', 'value': abs(NetReimbursement3_value)}
        ])

        ReimbursementvsCM2Margins_rows = pd.DataFrame([
            {'sku': 'ReimbursementvsCM2Margins1', 'value': abs(ReimbursementvsCM2Margins1_value)},
            {'sku': 'ReimbursementvsCM2Margins2', 'value': abs(ReimbursementvsCM2Margins2_value)},
            {'sku': 'ReimbursementvsCM2Margins3', 'value': abs(ReimbursementvsCM2Margins3_value)}
        ])

        Reimbursementvssales_rows = pd.DataFrame([
            {'sku': 'Reimbursementvssales1', 'value': abs(Reimbursementvssales1_value)},
            {'sku': 'Reimbursementvssales2', 'value': abs(Reimbursementvssales2_value)},
            {'sku': 'Reimbursementvssales3', 'value': abs(Reimbursementvssales3_value)}
        ])

        acos_rows = pd.DataFrame([
            {'sku': 'acos1', 'value': abs(acos1_value)},
            {'sku': 'acos2', 'value': abs(acos2_value)},
            {'sku': 'acos3', 'value': abs(acos3_value)}
        ])

        try:
            platform_fees_total = abs(platform_fees1_value) + abs(platform_fees2_value) + abs(platform_fees3_value)
            advertising_total = abs(advertising_total1) + abs(advertising_total2) + abs(advertising_total3)

            total_sales_sum = df_pivot.loc[df_pivot['sku'] == 'Total', 'Total_Sales_sum'].values[0] if 'Total_Sales_sum' in df_pivot.columns else 0
            cm1_profit_sum = df_pivot.loc[df_pivot['sku'] == 'Total', 'profit_sum'].values[0] if 'profit_sum' in df_pivot.columns else 0

            cm2profit_total = cm1_profit_sum - advertising_total - platform_fees_total
            cm2margin_total = (cm2profit_total / total_sales_sum) * 100 if total_sales_sum > 0 else 0
            acos_total = abs((advertising_total / total_sales_sum) * 100) if total_sales_sum > 0 else 0

            NetReimbursement_total = abs(NetReimbursement1_value) + abs(NetReimbursement2_value) + abs(NetReimbursement3_value)
            ReimbursementvsCM2Margins_total = abs((NetReimbursement_total / abs(cm2profit_total)) * 100) if abs(cm2profit_total) > 0 else 0
            Reimbursementvssales_total = abs((NetReimbursement_total / total_sales_sum) * 100) if total_sales_sum > 0 else 0
        except Exception as e:
            print(f"Error calculating quarter-end totals: {str(e)}")
            platform_fees_total = advertising_total = cm2profit_total = cm2margin_total = 0
            acos_total = NetReimbursement_total = ReimbursementvsCM2Margins_total = Reimbursementvssales_total = 0

        quarterend_row = pd.DataFrame([
            {'sku': 'platform_fees_total', 'value': abs(platform_fees_total)},
            {'sku': 'advertising_total', 'value': abs(advertising_total)},
            {'sku': 'cm2profit_total', 'value': cm2profit_total},
            {'sku': 'cm2margin_total', 'value': cm2margin_total},
            {'sku': 'acos_total', 'value': abs(acos_total)},
            {'sku': 'NetReimbursement_total', 'value': abs(NetReimbursement_total)},
            {'sku': 'ReimbursementvsCM2Margins_total', 'value': abs(ReimbursementvsCM2Margins_total)},
            {'sku': 'Reimbursementvssales_total', 'value': abs(Reimbursementvssales_total)}
        ])

        columns_list = df_pivot.columns.tolist()
        additional_dfs = [
            acos_rows, platform_fees_rows, advertising_rows, cm2profit_rows,
            NetReimbursement_rows, ReimbursementvsCM2Margins_rows,
            Reimbursementvssales_rows, cm2margin_rows, quarterend_row
        ]

        for df_add in additional_dfs:
            for col in columns_list:
                if col not in df_add.columns:
                    df_add[col] = None

        final_df = pd.concat([df_pivot] + additional_dfs, ignore_index=True)

        output_filename = f'forecastpnl_{user_id}_{country}_{month}_{year}_table.xlsx'

        buf = BytesIO()
        final_df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)

        save_file_to_db(
            user_id=user_id,
            country=country,
            filename=output_filename,
            file_bytes=buf.getvalue(),
            kind="pnl_forecast",
            month=month.lower() if month else None,
            year=str(year),
        )

        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PnL_forecast_{country}_{month}_{year}.xlsx'
        )

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401
    except Exception as e:
        print(f"Error in Pnlforecast: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
from flask import current_app

@forecast_bp.route('/api/Pnlforecast/global', methods=['GET', 'POST'])
def Pnlforecasts():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)

        month = request.args.get('month')
        year = request.args.get('year')

        if not month or not year:
            return jsonify({'error': 'Missing required parameters: month or year'}), 400

        month = str(month).strip().lower()
        year = int(year)

        global_filename = f'forecastpnl_{user_id}_global_{month}_{year}_table.xlsx'
        uk_filename = f'forecastpnl_{user_id}_uk_{month}_{year}_table.xlsx'
        us_filename = f'forecastpnl_{user_id}_us_{month}_{year}_table.xlsx'

        def get_usd_conversion_rate(month, year):
            try:
                engine_conv = create_engine(db_url1)
                with engine_conv.connect() as conn:
                    query = text("""
                        SELECT conversion_rate
                        FROM currency_conversion
                        WHERE lower(user_currency) = 'gbp'
                        AND lower(country) = 'us'
                        AND lower(selected_currency) = 'usd'
                        AND lower(month) = :month
                        AND year = :year
                        LIMIT 1
                    """)
                    result = conn.execute(query, {
                        "month": month.lower(),
                        "year": int(year)
                    }).fetchone()

                    if result and result[0] is not None:
                        rate = float(result[0])
                        return rate

                    print(f"⚠ No GBP -> USD conversion rate found for {month} {year}. Using 1.0")
                    return 1.0

            except Exception as e:
                print(f"❌ Error fetching conversion rate: {str(e)}")
                return 1.0

        def ensure_country_pnl_exists(country_code: str, filename: str):
            stored = load_file_from_db(
                user_id=user_id,
                country=country_code,
                filename=filename
            )
            if stored:
                print(f"✅ {country_code.upper()} PnL already exists in DB")
                return stored

            print(f"🔄 Missing {country_code.upper()} PnL file in DB, generating now...")

            with current_app.test_request_context(
                f'/api/Pnlforecast?country={country_code}&month={month}&year={year}',
                method='GET',
                headers={'Authorization': f'Bearer {token}'}
            ):
                resp = Pnlforecast()

            if isinstance(resp, tuple):
                response_obj, status_code = resp
                if status_code != 200:
                    print(f"❌ Failed generating {country_code.upper()} PnL: status={status_code}")
                    return None
            else:
                status_code = getattr(resp, 'status_code', 200)
                if status_code != 200:
                    print(f"❌ Failed generating {country_code.upper()} PnL: status={status_code}")
                    return None

            stored = load_file_from_db(
                user_id=user_id,
                country=country_code,
                filename=filename
            )

            if stored:
                print(f"✅ {country_code.upper()} PnL generated and saved in DB")
            else:
                print(f"❌ {country_code.upper()} PnL generation call succeeded but file not found in DB")

            return stored

        uk_stored = ensure_country_pnl_exists('uk', uk_filename)
        us_stored = ensure_country_pnl_exists('us', us_filename)

        uk_exists = uk_stored is not None
        us_exists = us_stored is not None

        print(f"📁 DB File status - UK: {uk_exists}, US: {us_exists}")

        if not uk_exists and not us_exists:
            return jsonify({'error': 'No PnL forecast files found for UK or US in DB'}), 404

        conversion_rate = get_usd_conversion_rate(month, year)
        print(f"🔎 FINAL conversion_rate used for global: {conversion_rate}")

        # IMPORTANT: always regenerate global from latest available country files
        if uk_exists and us_exists:
            print("📊 Creating combined global forecast from UK and US data")

            df_uk_original = pd.read_excel(BytesIO(uk_stored.data), engine='openpyxl')
            df_us_original = pd.read_excel(BytesIO(us_stored.data), engine='openpyxl')

            df_uk = df_uk_original.copy()
            df_us = df_us_original.copy()

            df_uk['country'] = 'UK'
            df_us['country'] = 'US'

            df_uk_converted = convert_uk_to_usd(df_uk, conversion_rate)

            df_global = process_combined_forecast_data(df_us, df_uk_converted)

        elif uk_exists and not us_exists:
            print("🇬🇧 Creating global forecast from UK only and converting GBP -> USD")

            df_uk_original = pd.read_excel(BytesIO(uk_stored.data), engine='openpyxl')
            df_uk = df_uk_original.copy()
            df_uk['country'] = 'UK'

            df_global = convert_uk_to_usd(df_uk, conversion_rate)

        elif us_exists and not uk_exists:
            print("🇺🇸 Creating global forecast from US only")
            df_global = pd.read_excel(BytesIO(us_stored.data), engine='openpyxl')

        else:
            return jsonify({'error': 'Unexpected scenario while generating global PnL'}), 500

        if 'country' in df_global.columns:
            df_global = df_global.drop(columns=['country'])

        buf = BytesIO()
        df_global.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)

        save_file_to_db(
            user_id=user_id,
            country='global',
            filename=global_filename,
            file_bytes=buf.getvalue(),
            kind="pnl_forecast",
            month=month,
            year=str(year),
            content_type=XLSX_MIME
        )

        global_stored = load_file_from_db(
            user_id=user_id,
            country='global',
            filename=global_filename
        )

        return send_db_file(
            global_stored,
            download_name=f'PnL_forecast_global_{month}_{year}.xlsx'
        )

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401
    except Exception as e:
        print(f"❌ Error in Pnlforecasts route: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    

    
def convert_uk_to_usd(df_uk, conversion_rate):
    """
    Convert UK PnL values from GBP to USD.
    Multiplies monetary columns by GBP->USD rate.
    Does NOT convert percentage fields.
    """
    df_converted = df_uk.copy()
    conversion_rate = float(conversion_rate or 1.0)

    financial_columns = [
        'Total_Sales_1st', 'Total_Sales_2nd', 'Total_Sales_3rd', 'Total_Sales_sum',
        'profit_1st', 'profit_2nd', 'profit_3rd', 'profit_sum'
    ]

    for col in financial_columns:
        if col in df_converted.columns:
            df_converted[col] = pd.to_numeric(df_converted[col], errors='coerce').fillna(0)
            df_converted[col] = df_converted[col] * conversion_rate

    financial_metrics = [
        'Platform_Fees1', 'Platform_Fees2', 'Platform_Fees3',
        'advertising_total1', 'advertising_total2', 'advertising_total3',
        'cm2profit1', 'cm2profit2', 'cm2profit3',
        'NetReimbursement1', 'NetReimbursement2', 'NetReimbursement3',
        'platform_fees_total', 'advertising_total', 'cm2profit_total',
        'NetReimbursement_total'
    ]

    if 'sku' in df_converted.columns and 'value' in df_converted.columns:
        df_converted['sku'] = df_converted['sku'].astype(str).str.strip()
        df_converted['value'] = pd.to_numeric(df_converted['value'], errors='coerce').fillna(0)

        for metric in financial_metrics:
            mask = df_converted['sku'] == metric
            if mask.any():
                df_converted.loc[mask, 'value'] = df_converted.loc[mask, 'value'] * conversion_rate

    return df_converted



def process_combined_forecast_data(df_us, df_uk_converted):
    """
    Combine US PnL + converted UK PnL into one global dataframe.
    Product rows are grouped by product_name.
    Financial summary rows are grouped by sku.
    """
    df_combined = pd.concat([df_us, df_uk_converted], ignore_index=True)

    financial_metrics = [
        'acos1', 'acos2', 'acos3',
        'Platform_Fees1', 'Platform_Fees2', 'Platform_Fees3',
        'advertising_total1', 'advertising_total2', 'advertising_total3',
        'cm2profit1', 'cm2profit2', 'cm2profit3',
        'NetReimbursement1', 'NetReimbursement2', 'NetReimbursement3',
        'ReimbursementvsCM2Margins1', 'ReimbursementvsCM2Margins2', 'ReimbursementvsCM2Margins3',
        'Reimbursementvssales1', 'Reimbursementvssales2', 'Reimbursementvssales3',
        'cm2margin1', 'cm2margin2', 'cm2margin3',
        'platform_fees_total', 'advertising_total', 'cm2profit_total', 'cm2margin_total',
        'acos_total', 'NetReimbursement_total', 'ReimbursementvsCM2Margins_total',
        'Reimbursementvssales_total', 'Total'
    ]

    df_combined['sku'] = df_combined['sku'].astype(str).str.strip()

    product_rows = df_combined[~df_combined['sku'].isin(financial_metrics)].copy()
    financial_rows = df_combined[df_combined['sku'].isin(financial_metrics)].copy()

    if not product_rows.empty:
        if 'product_name' not in product_rows.columns:
            product_rows['product_name'] = 'Unknown'
        product_rows['product_name'] = product_rows['product_name'].fillna('Unknown').astype(str)

        numeric_cols_to_sum = [
            'forecast_1st', 'forecast_2nd', 'forecast_3rd', 'forecast_sum',
            'Total_Sales_1st', 'Total_Sales_2nd', 'Total_Sales_3rd', 'Total_Sales_sum',
            'profit_1st', 'profit_2nd', 'profit_3rd', 'profit_sum'
        ]

        agg_dict = {'sku': 'first'}

        for col in numeric_cols_to_sum:
            if col in product_rows.columns:
                product_rows[col] = pd.to_numeric(product_rows[col], errors='coerce').fillna(0)
                agg_dict[col] = 'sum'

        grouped_products = product_rows.groupby('product_name', as_index=False).agg(agg_dict)

        for period in ['1st', '2nd', '3rd', 'sum']:
            profit_col = f'profit_{period}'
            sales_col = f'Total_Sales_{period}'
            percentage_col = f'profit_percentage_{period}'

            if profit_col in grouped_products.columns and sales_col in grouped_products.columns:
                grouped_products[percentage_col] = np.where(
                    grouped_products[sales_col] != 0,
                    (grouped_products[profit_col] / grouped_products[sales_col]) * 100,
                    0
                )
    else:
        grouped_products = pd.DataFrame()

    grouped_financial_list = []

    if not financial_rows.empty:
        numeric_cols = financial_rows.select_dtypes(include=[np.number]).columns.tolist()

        for sku in financial_rows['sku'].unique():
            sku_data = financial_rows[financial_rows['sku'] == sku].copy()
            result_row = sku_data.iloc[0].copy()

            for col in numeric_cols:
                sku_data[col] = pd.to_numeric(sku_data[col], errors='coerce').fillna(0)
                result_row[col] = sku_data[col].sum()

            grouped_financial_list.append(result_row)

        grouped_financial = pd.DataFrame(grouped_financial_list)
    else:
        grouped_financial = pd.DataFrame()

    if not grouped_products.empty and not grouped_financial.empty:
        result_df = pd.concat([grouped_products, grouped_financial], ignore_index=True)
    elif not grouped_products.empty:
        result_df = grouped_products
    elif not grouped_financial.empty:
        result_df = grouped_financial
    else:
        result_df = pd.DataFrame()

    if not result_df.empty:
        result_df = recalculate_dependent_metrics(result_df)

    return result_df

def recalculate_dependent_metrics(df):
    """
    Recalculate ACOS, CM2 margins, reimbursement percentages
    after global aggregation.
    """
    total_row = df[df['sku'] == 'Total']
    if total_row.empty:
        print("⚠️ No 'Total' row found for recalculation")
        return df

    total_sales_1st = float(total_row['Total_Sales_1st'].iloc[0]) if 'Total_Sales_1st' in total_row.columns else 0
    total_sales_2nd = float(total_row['Total_Sales_2nd'].iloc[0]) if 'Total_Sales_2nd' in total_row.columns else 0
    total_sales_3rd = float(total_row['Total_Sales_3rd'].iloc[0]) if 'Total_Sales_3rd' in total_row.columns else 0
    total_sales_sum = float(total_row['Total_Sales_sum'].iloc[0]) if 'Total_Sales_sum' in total_row.columns else 0

    for period, sales in [('1', total_sales_1st), ('2', total_sales_2nd), ('3', total_sales_3rd)]:
        platform_fees_mask = df['sku'] == f'Platform_Fees{period}'
        acos_mask = df['sku'] == f'acos{period}'

        if platform_fees_mask.any() and acos_mask.any() and sales > 0:
            platform_fees_value = float(df.loc[platform_fees_mask, 'value'].iloc[0])
            df.loc[acos_mask, 'value'] = (platform_fees_value / sales) * 100

    total_platform_fees_mask = df['sku'] == 'platform_fees_total'
    acos_total_mask = df['sku'] == 'acos_total'
    if total_platform_fees_mask.any() and acos_total_mask.any() and total_sales_sum > 0:
        total_platform_fees = float(df.loc[total_platform_fees_mask, 'value'].iloc[0])
        df.loc[acos_total_mask, 'value'] = (total_platform_fees / total_sales_sum) * 100

    for period, sales in [('1', total_sales_1st), ('2', total_sales_2nd), ('3', total_sales_3rd)]:
        cm2profit_mask = df['sku'] == f'cm2profit{period}'
        cm2margin_mask = df['sku'] == f'cm2margin{period}'

        if cm2profit_mask.any() and cm2margin_mask.any() and sales > 0:
            cm2profit_value = float(df.loc[cm2profit_mask, 'value'].iloc[0])
            df.loc[cm2margin_mask, 'value'] = (cm2profit_value / sales) * 100

    total_cm2profit_mask = df['sku'] == 'cm2profit_total'
    cm2margin_total_mask = df['sku'] == 'cm2margin_total'
    if total_cm2profit_mask.any() and cm2margin_total_mask.any() and total_sales_sum > 0:
        total_cm2profit = float(df.loc[total_cm2profit_mask, 'value'].iloc[0])
        df.loc[cm2margin_total_mask, 'value'] = (total_cm2profit / total_sales_sum) * 100

    for period, sales in [('1', total_sales_1st), ('2', total_sales_2nd), ('3', total_sales_3rd)]:
        net_reimb_mask = df['sku'] == f'NetReimbursement{period}'
        reimb_vs_sales_mask = df['sku'] == f'Reimbursementvssales{period}'
        reimb_vs_cm2_mask = df['sku'] == f'ReimbursementvsCM2Margins{period}'
        cm2profit_mask = df['sku'] == f'cm2profit{period}'

        if net_reimb_mask.any():
            net_reimb_value = float(df.loc[net_reimb_mask, 'value'].iloc[0])

            if reimb_vs_sales_mask.any() and sales > 0:
                df.loc[reimb_vs_sales_mask, 'value'] = (net_reimb_value / sales) * 100

            if reimb_vs_cm2_mask.any() and cm2profit_mask.any():
                cm2profit_value = float(df.loc[cm2profit_mask, 'value'].iloc[0])
                if cm2profit_value != 0:
                    df.loc[reimb_vs_cm2_mask, 'value'] = (net_reimb_value / cm2profit_value) * 100

    total_net_reimb_mask = df['sku'] == 'NetReimbursement_total'
    total_reimb_vs_sales_mask = df['sku'] == 'Reimbursementvssales_total'
    total_reimb_vs_cm2_mask = df['sku'] == 'ReimbursementvsCM2Margins_total'

    if total_net_reimb_mask.any():
        total_net_reimb = float(df.loc[total_net_reimb_mask, 'value'].iloc[0])

        if total_reimb_vs_sales_mask.any() and total_sales_sum > 0:
            df.loc[total_reimb_vs_sales_mask, 'value'] = (total_net_reimb / total_sales_sum) * 100

        if total_reimb_vs_cm2_mask.any() and total_cm2profit_mask.any():
            total_cm2profit = float(df.loc[total_cm2profit_mask, 'value'].iloc[0])
            if total_cm2profit != 0:
                df.loc[total_reimb_vs_cm2_mask, 'value'] = (total_net_reimb / total_cm2profit) * 100

    return df



@forecast_bp.route('/api/Pnlforecast/previous_months', methods=['GET', 'POST'])
def Pnlforecast_previous_months():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        month = request.args.get('month')
        year = request.args.get('year')
        period_type = request.args.get('period_type', 'monthly')
        record_country = request.args.get('country', 'global')

        if not month or not year:
            return jsonify({'error': 'Missing required parameters: month or year'}), 400

        month = month.lower()
        year = int(year)
        month_name = month

        engine = create_engine(db_url)
        SessionLocal = sessionmaker(bind=engine)
        db_session = SessionLocal()
        inspector = inspect(engine)

        def get_best_key(d, keys, default=0):
            for key in keys:
                val = d.get(key)
                if val not in [None, '']:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        continue
            return float(default)

        try:
            if period_type == 'monthly':
                if record_country.lower() == "global":
                    table_name = f"skuwisemonthly_{user_id}_{record_country.lower()}_{month_name.lower()}{year}_table"
                else:
                    table_name = f"skuwisemonthly_{user_id}_{record_country.lower()}_{month_name.lower()}{year}"

            if not inspector.has_table(table_name):
                return jsonify({'error': f'Table {table_name} does not exist'}), 404

            data_query = f"SELECT * FROM {table_name}"
            result = db_session.execute(text(data_query))
            rows = result.fetchall()

            if not rows:
                return jsonify({
                    'data': [],
                    'totals': {
                        'net_sales_total': 0,
                        'advertising_total': 0,
                        'amazon_fee_total': 0,
                        'cm2_profit_total': 0,
                        'profit_total': 0,
                        'cost_of_unit_sold_total': 0
                    }
                }), 200

            columns = result.keys()
            data = []

            net_sales_total = 0
            advertising_total = 0
            amazon_fee_total = 0
            cm2_profit_total = 0
            profit_total = 0
            cost_of_unit_sold_total = 0

           

            for row in rows:
                row_dict = {}
                for col in columns:
                    value = getattr(row, col)
                    row_dict[col] = float(value) if isinstance(value, (int, float, Decimal)) else value

                # Skip TOTAL row for calculations but include it in data
                if str(row_dict.get('sku', '')).strip().upper() == 'TOTAL':
                    data.append(row_dict)
                    continue

                # Based on your database screenshots, here are the correct column mappings:
                
                # Net Sales - exact match from your DB
                net_sales_total += get_best_key(row_dict, ['net_sales', 'netsales'])

                # Advertising Total - from your screenshots, the column appears to be 'advertising_total'
                # But it might be empty/zero, so let's also check for variations
                advertising_total += get_best_key(row_dict, [
                    'advertising_total', 
                    'advertisingtotal', 
                    'advertising_costs',
                    'advertising',
                    'ad_spend',
                    'ppc_spend'
                ])

                # Amazon Fee - exact match from your DB
                amazon_fee_total += get_best_key(row_dict, ['amazon_fee', 'amazonfee'])

                # CM2 Profit - from your screenshots, this column appears to be 'cm2_profit'
                # But it might be empty/zero, so let's also check for variations
                cm2_profit_total += get_best_key(row_dict, [
                    'cm2_profit', 
                    'cm2profit',
                    'cm2_profit_loss',
                    'contribution_margin_2'
                ])

                # CM1 Profit (profit column) - exact match from your DB  
                profit_total += get_best_key(row_dict, ['profit', 'cm1_profit', 'cm1profit'])

                # Cost of Unit Sold - exact match from your DB
                cost_of_unit_sold_total += get_best_key(row_dict, ['cost_of_unit_sold', 'costofunitsold', 'cogs'])

                data.append(row_dict)

            # Check if we have a TOTAL row and extract values from it if individual rows are empty
            total_row = next((row for row in data if str(row.get('sku', '')).strip().upper() == 'TOTAL'), None)
            
            if total_row and (advertising_total == 0 or cm2_profit_total == 0):
                
                
                if advertising_total == 0:
                    advertising_total = get_best_key(total_row, [
                        'advertising_total', 
                        'advertisingtotal', 
                        'advertising_costs',
                        'advertising'
                    ])
                   
                
                if cm2_profit_total == 0:
                    cm2_profit_total = get_best_key(total_row, [
                        'cm2_profit', 
                        'cm2profit',
                        'cm2_profit_loss'
                    ])
            

            totals = {
                'net_sales_total': round(net_sales_total, 2),
                'advertising_total': round(advertising_total, 2),
                'amazon_fee_total': round(amazon_fee_total, 2),
                'cm2_profit_total': round(cm2_profit_total, 2),
                'profit_total': round(profit_total, 2),
                'cost_of_unit_sold_total': round(cost_of_unit_sold_total, 2)
            }

            return jsonify({
                'data': data,
                'totals': totals,
                'record_count': len(data),
                'table_name': table_name,
                'debug_info': {
                    'available_columns': list(columns),
                    'sample_row': dict(rows[0]._mapping) if rows else None,
                    'column_mapping_status': {
                        'net_sales_found': net_sales_total > 0,
                        'advertising_found': advertising_total > 0,
                        'amazon_fee_found': amazon_fee_total > 0,
                        'cm2_profit_found': cm2_profit_total > 0,
                        'profit_found': profit_total > 0,
                        'cogs_found': cost_of_unit_sold_total > 0
                    },
                    'zero_value_columns': {
                        'advertising_columns_checked': ['advertising_total', 'advertisingtotal', 'advertising_costs', 'advertising'],
                        'cm2_profit_columns_checked': ['cm2_profit', 'cm2profit', 'cm2_profit_loss', 'contribution_margin_2']
                    }
                }
            }), 200

        except Exception as e:
            db_session.rollback()
            print(f"Database error: {str(e)}")
            return jsonify({'error': f'Database query failed: {str(e)}'}), 500

        finally:
            db_session.close()

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401
    except Exception as e:
        print(f"General error: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    


@forecast_bp.route('/api/save_pnl_forecast', methods=['POST'])
def save_pnl_forecast():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    month = request.form.get('month')
    year = request.form.get('year')
    country = request.form.get('country')
    print(f"Received Month: {month}, Year: {year}, Country: {country}")

    if not all([month, year, country]):
        return jsonify({"error": "month, year, and country are required"}), 400

    # ✅ Read bytes directly (no disk)
    file_bytes = file.read()
    if not file_bytes:
        return jsonify({"error": "Uploaded file is empty"}), 400

    # Keep your naming
    new_filename = f"PNLforecast_{user_id}_{month}_{year}.xlsx"
    try:
        save_file_to_db(
            user_id=user_id,
            country=country.lower(),
            filename=new_filename,
            file_bytes=file_bytes,
            kind="pnl_forecast_upload",
            month=month.lower(),
            year=str(year),
            content_type=file.mimetype or XLSX_MIME
        )
        print(f"✅ Saved PNL forecast to DB: {new_filename}")
    except Exception as e:
        print(f"❌ Error saving file to DB: {str(e)}")
        return jsonify({"error": "Failed to save file to DB", "message": str(e)}), 500

    upload_history = UploadHistory.query.filter_by(
        user_id=user_id,
        country=country,
        month=month,
        year=year
    ).first()

    if upload_history:
        if upload_history.pnl_email_sent:
            print("📭 Email already sent for this upload. Skipping...")
        else:
            try:
                from app import db
                # NOTE: your email util currently receives a filename.
                # If it reads from disk, you must update it to read from DB using load_file_from_db.
                send_pnlforecast_email(user_id, new_filename, month, year)
                upload_history.pnl_email_sent = True
                db.session.commit()
                print("✅ Email sent and status updated.")
            except Exception as e:
                print(f"❌ Error sending email: {str(e)}")

    return jsonify({
        "message": "PNL forecast saved successfully",
        "filename": new_filename,
        "stored_in": "db"
    }), 200


