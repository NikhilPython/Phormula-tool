from flask import Blueprint, request, jsonify
import jwt
from io import BytesIO
import calendar
from werkzeug.utils import secure_filename
import pandas as pd
from sqlalchemy import create_engine, Table, MetaData, Column, Integer, String, Float, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy import inspect
import os
import io, re
import base64
from io import BytesIO
from config import Config
SECRET_KEY = Config.SECRET_KEY
from app import db
from app.models.user_models import UploadHistory , CountryProfile , User
from app.utils.data_utils import generate_pnl_report 
from app.utils.token_utils import get_effective_user_id_from_token
from app.utils.us_process_utils  import process_skuwise_us_data , process_us_yearly_skuwise_data, process_us_quarterly_skuwise_data
from app.utils.uk_process_utils import process_skuwise_data , process_quarterly_skuwise_data, process_yearly_skuwise_data 
from app.utils.plotting_utils import (
    get_referral_fees , apply_modifications 
)
from app.utils.currency_utils import (  process_global_yearly_skuwise_data ,
    process_global_quarterly_skuwise_data , 
    process_global_monthly_skuwise_data
)
from app.utils.history_graph_utils import get_performance_trend
from dotenv import load_dotenv
from sqlalchemy import MetaData
from sqlalchemy import text



load_dotenv()
db_url = os.getenv('DATABASE_URL')
db_url1 = os.getenv('DATABASE_ADMIN_URL')

user_engine = create_engine(
    db_url,
    pool_pre_ping=True,
    pool_size=5,
    max_overflow=10,
    pool_recycle=1800,
)

admin_engine = create_engine(
    db_url1,
    pool_pre_ping=True,
    pool_size=5,
    max_overflow=10,
    pool_recycle=1800,
)

upload_bp = Blueprint('upload_bp', __name__)

def encode_file_to_base64(file_path):
    with open(file_path, 'rb') as file:
        return base64.b64encode(file.read()).decode('utf-8')




COLUMN_MAPPING = {
    'date/time': 'date_time',
    'settlement id': 'settlement_id',
    'type': 'type',
    'order id': 'order_id',
    'sku': 'sku',
    'description': 'description',
    'quantity': 'quantity',
    'marketplace': 'marketplace',
    'fulfilment': 'fulfilment',
    'fulfillment': 'fulfillment',
    'order city': 'order_city',
    'order state': 'order_state',
    'order postal': 'order_postal',
    'tax collection model': 'tax_collection_model',
    'product sales': 'product_sales',
    'product sales tax': 'product_sales_tax',
    'postage credits': 'postage_credits',
    'shipping credits tax': 'shipping_credits_tax',
    'gift wrap credits': 'gift_wrap_credits',
    'giftwrap credits tax': 'giftwrap_credits_tax',
    'promotional rebates': 'promotional_rebates',
    'promotional rebates tax': 'promotional_rebates_tax',
    'sales tax collected': 'sales_tax_collected',
    'marketplace withheld tax': 'marketplace_facilitator_tax',
    'selling fees': 'selling_fees',
    'fba fees': 'fba_fees',
    'other transaction fees': 'other_transaction_fees',
    'other': 'other',
    'total': 'total',
    'account type': 'account_type',
    'Regulatory Fee': 'regulatory_fee',
    'Tax On Regulatory Fee': 'tax_on_regulatory_fee',
    'Bucket': 'bucket',
    'shipping credits': 'shipping_credits',
    'regulatory fee': 'regulatory_fee',
    'tax on regulatory fee': 'tax_on_regulatory_fee',

}

MONTHS_REVERSE_MAP = {
    1: "january", 2: "february", 3: "march", 4: "april", 5: "may", 6: "june",
    7: "july", 8: "august", 9: "september", 10: "october", 11: "november", 12: "december"
}


MONTHS_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}

def get_next_month_year(month, year):
    year = int(year)
    month_num = MONTHS_MAP[month]
    if month_num == 12:
        month_num = 1
        month_next = MONTHS_REVERSE_MAP[month_num]
       

        return month_next, year + 1
    month_next = month_num + 1
    month_next1 = MONTHS_REVERSE_MAP[month_next]
    

    return month_next1, year

def table_exists(engine, table_name, schema='public'):
    inspector = inspect(engine)
    return inspector.has_table(table_name, schema=schema)


@upload_bp.route('/upload', methods=['GET', 'POST'])
def upload():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # Retrieve user using SQLAlchemy ORM
    user = User.query.get(user_id)
    if user is None:
        return jsonify({'error': 'User not found'}), 404

    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'success': False, 'message': 'Both files are required'}), 400
    
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    country = request.form['country'].lower()
    month = request.form['month'].lower()
    if month not in ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']:
        return jsonify({'error': 'Invalid month provided'}), 400
    profile = request.form['profile_id']
    year = request.form['year']

    profile = CountryProfile.query.filter(
    db.func.lower(CountryProfile.country) == country,
    CountryProfile.user_id == user_id).first()

    transit_time = int(profile.transit_time)  # Transit time in months
    stock_unit = int(profile.stock_unit)

    table_name = f"user_{user_id}_{country}_{month}{year}_data".lower()
    country_table_name = f"sku_{user_id}_data_table"
    consolidated_table_name = f"user_{user_id}_{country}_merge_data_of_all_months"
    inventory_file_name = f"user_{user_id}_{country}_{month}{year}_inventory_file.xlsx"
    currency_table = f"currency_conversion"
    countris_table_name = f"user_{user_id}_{country}_table"


    # Create SQLAlchemy engine with PostgreSQL
    engine = user_engine
    engine1 = admin_engine
    meta = MetaData()

    with engine.connect() as connection:
        connection.execute(text(f"DROP TABLE IF EXISTS {table_name}"))
        connection.commit()

    # Define the table - keep the original column case
    user_monthly_data = Table(
        table_name, meta,
        Column('id', Integer, primary_key=True),
        Column('date_time', String),
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('account_type', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('regulatory_fee', Float),
        Column('tax_on_regulatory_fee', Float),
        Column('bucket', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('product_name', String),
        Column('currency', String),
        Column('advertising_cost', Float),
        Column('net_reimbursement', Float),
        Column('platform_fees', Float),
        Column('product_group', String),
    )

    user_consolidated_data = Table(
        consolidated_table_name, meta,
        Column('id', Integer, primary_key=True),
        Column('date_time', String),
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('account_type', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('regulatory_fee', Float),
        Column('tax_on_regulatory_fee', Float),
        Column('bucket', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('month', String),
        Column('year', String),
        Column('product_name', String),
        Column('currency', String),
        Column('advertising_cost', Float),
        Column('net_reimbursement', Float),
        Column('platform_fees', Float),
        Column('product_group', String),
    )
    meta.create_all(engine)

    with engine.connect() as connection:
        connection.execute(user_monthly_data.delete())  # Delete previous records
        connection.execute(text(f"DELETE FROM {user_consolidated_data} WHERE month = '{month}' AND year = '{year}'"))
        connection.commit()

    # Handle the file upload
    if file1.filename == '' or file2.filename == '':
        return jsonify({'error': 'Both files must be uploaded'}), 400
    
    # Save file1
    file1_path = os.path.join( f'user_{user_id}_{country}_{month}_{year}_mtd_file.xlsx')
    file1.save(file1_path)
    
    # Save file2 (inventory file)
    if file2.filename.endswith('.csv'):
        # Read CSV file and save as Excel
        df_file2 = pd.read_csv(file2)
        file2_path = os.path.join( f'user_{user_id}_{country}_{month}{year}_inventory_file.xlsx')
        df_file2.to_excel(file2_path, index=False, engine='openpyxl')  # Save as Excel file
    else:
        # Save the file directly if it's already an Excel file
        file2_path = os.path.join( secure_filename(inventory_file_name))
        file2.save(file2_path)


    if file1.filename.endswith('.csv'):
        with open(file1_path, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
            header_line = None
            # Look for header row
            for i, line in enumerate(lines):
                if any(keyword in line.lower() for keyword in ['date/time', 'date / time', 'date time']):
                    header_line = i
                    break
        if header_line is None:
            return jsonify({'error': 'Could not find header row with "date/time" column'}), 400

        df = pd.read_csv(file1_path, skiprows=header_line, encoding="utf-8-sig", dayfirst=True, skip_blank_lines=True)  # Use utf-8-sig to remove BOM
        df.columns = df.columns.str.strip()
        # Replace problematic characters

    elif file1.filename.endswith(('.xls', '.xlsx')):
        import openpyxl
    
        wb = openpyxl.load_workbook(file1_path, read_only=True)
        ws = wb.active
        header_line = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
            if row is not None and any(
                cell is not None and any(keyword in str(cell).lower() for keyword in ['date/time', 'date / time', 'date time'])
                for cell in row
            ):
                header_line = i
                break
        if header_line is None:
            return jsonify({'error': 'Could not find header row with "date/time" column in Excel'}), 400

        # Now read excel from the detected header line
        df = pd.read_excel(file1_path, skiprows=header_line)
        df.columns = df.columns.str.strip()
    else:
        return jsonify({'error': 'Invalid file format. Only .csv and .xlsx files are allowed'}), 400
    
    df.columns = [c.lower() for c in df.columns]  # Lowercase all column names for consistency

    # Function to clean numeric columns
    def clean_numeric_value(val):
        if isinstance(val, str):
            # Remove commas and other non-numeric characters
            if ',' in val:
                val = val.replace(',', '')
            try:
                return float(val)
            except ValueError:
                return None
        return val

    # Apply cleaning function to all numeric columns
    numeric_columns = ['quantity', 'price_in_gbp', 'cost_of_unit_sold', 'product_sales', 
                       'product_sales_tax', 'postage_credits', 'shipping_credits_tax',
                       'gift_wrap_credits', 'giftwrap_credits_tax', 'promotional_rebates',
                       'promotional_rebates_tax', 'sales_tax_collected', 'marketplace_withheld_tax',
                       'marketplace_facilitator_tax', 'selling_fees', 'percentage1', 'fba_fees',
                       'percentage2', 'other_transaction_fees', 'other', 'total','advertising_cost',
                        'net_reimbursement',
                        'platform_fees']
    
    # Make sure all numeric columns exist in the dataframe
    numeric_columns = [col for col in numeric_columns if col in df.columns]
    
    for col in numeric_columns:
        df[col] = df[col].apply(clean_numeric_value)

    df.rename(columns=COLUMN_MAPPING, inplace=True)
    for col in COLUMN_MAPPING.values():
        if col not in df.columns:
            if col in numeric_columns:
                df[col] = 0
            else:
                df[col] = None  # or use np.nan if preferred
    
    
    if country.upper() == 'UK':
        sku_column = 'sku_uk'
    elif country.upper() == 'US':
        sku_column = 'sku_us'
    else:
        raise ValueError("Unsupported country")

            
    # with engine.connect() as conn:
    #     country_df = pd.read_sql(f"SELECT {sku_column} AS sku, price,currency, product_name FROM {country_table_name}", conn)

    # df = df.merge(country_df, on='sku', how='left')

    

    MONTH_NUM = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12
    }

    target_month_num = MONTH_NUM.get(month.lower())
    if not target_month_num:
        raise ValueError(f"Invalid month: {month}. Expected january..december")

    target_year = int(year)  # form year string -> int

    month_case_sql = """
    CASE lower(month)
        WHEN 'january' THEN 1
        WHEN 'february' THEN 2
        WHEN 'march' THEN 3
        WHEN 'april' THEN 4
        WHEN 'may' THEN 5
        WHEN 'june' THEN 6
        WHEN 'july' THEN 7
        WHEN 'august' THEN 8
        WHEN 'september' THEN 9
        WHEN 'october' THEN 10
        WHEN 'november' THEN 11
        WHEN 'december' THEN 12
        ELSE NULL
    END
    """

    # year is string in DB, so cast it for comparisons
    price_asof_query = text(f"""
        SELECT sku, price, currency, product_name
        FROM (
            SELECT
                {sku_column} AS sku,
                price,
                currency,
                product_name,
                CAST(year AS INTEGER) AS year_int,
                {month_case_sql} AS month_num,
                ROW_NUMBER() OVER (
                    PARTITION BY {sku_column}
                    ORDER BY CAST(year AS INTEGER) DESC, {month_case_sql} DESC
                ) AS rn
            FROM {country_table_name}
            WHERE
                year IS NOT NULL
                AND trim(year) <> ''
                AND {month_case_sql} IS NOT NULL
                AND (
                    CAST(year AS INTEGER) < :target_year
                    OR (CAST(year AS INTEGER) = :target_year AND {month_case_sql} <= :target_month_num)
                )
        ) x
        WHERE x.rn = 1
    """)

    with engine.connect() as conn:
        country_df = pd.read_sql(
            price_asof_query,
            conn,
            params={"target_year": target_year, "target_month_num": target_month_num}
        )

    df = df.merge(country_df, on="sku", how="left")


    

    with engine.connect() as conn:
        countries_df = pd.read_sql(f"SELECT sku, product_group FROM {countris_table_name}", conn)

    df = df.merge(countries_df, on='sku', how='left')

    with engine1.connect() as conn:
        currency_query = text("""
            SELECT conversion_rate
            FROM currency_conversion 
            WHERE lower(user_currency) = :currency 
            AND lower(country) = :country 
            AND lower(month) = :month 
            AND year = :year
            LIMIT 1
        """)

        result = conn.execute(
            currency_query,
            {
                "currency": country_df['currency'].dropna().iloc[0].lower(),  # Get any available currency from df
                "country": country.lower(),
                "month": month.lower(),
                "year": year
            }
        ).fetchone()

# Step 3: Apply conversion rate to calculate price_in_gbp
    conversion_rate = result[0] if result else None

    if conversion_rate:
        df['price_in_gbp'] = df['price'] * conversion_rate
    else:
        df['price_in_gbp'] = None  # Or handle fallback logic if conversion rate not found

    
    # Calculate cost_of_unit_sold where both values are not null
    df['cost_of_unit_sold'] = df.apply(
        lambda row: row['price_in_gbp'] * row['quantity']
        if str(row.get('type', '')).strip().lower() == 'shipment'
        and pd.notnull(row['price_in_gbp'])
        and pd.notnull(row['quantity'])
        else 0.0,
        axis=1
    )

    df.drop(columns=['price'], inplace=True)

    # Clean the 'total' column specifically since it appears in the error message
    if 'total' in df.columns:
        df['total'] = df['total'].apply(clean_numeric_value)

    # Handle the case where any numeric column might still have commas or other formatting issues
    for col in [c for c in df.columns if c in numeric_columns]:
        if col in df.columns:
            # Convert to float with NaN for any unconvertible values
            df[col] = pd.to_numeric(df[col], errors='coerce')

    for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
    for col in df.columns:
                if df[col].dtype == 'object':  # likely to have commas or bad data
                    # Remove commas and convert to numeric
                    df[col] = df[col].str.replace(',', '', regex=True)

            # Now convert all possible numeric columns to float
    df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

    df.to_sql(table_name, con=engine, if_exists='append', index=False)
    
    df['month'] = month  # Assigning month from form data
    df['year'] = year 

    for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
    for col in df.columns:
                if df[col].dtype == 'object':  # likely to have commas or bad data
                    # Remove commas and convert to numeric
                    df[col] = df[col].str.replace(',', '', regex=True)

            # Now convert all possible numeric columns to float
    df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

    
    

    df.to_sql(consolidated_table_name, con=engine, if_exists='append', index=False)


    # ✅ Step 1: Currency Rates Dictionary
    if country.lower() == 'uk':
        currency1 = 'gbp'
    elif country.lower() == 'us':
        currency1 = 'usd'
    elif country.lower() == 'canada':
        currency1 = 'cad'
    else:
        currency1 = 'usd'  # fallback/default if unknown

    # Step 2: Fetch conversion rate from currency_conversion table
    with engine1.connect() as conn:
        currency_query = text("""
            SELECT conversion_rate
            FROM currency_conversion 
            WHERE lower(user_currency) = :currency1
            AND lower(country) = 'us'
            AND lower(month) = :month 
            AND year = :year
            LIMIT 1
        """)
        
        result = conn.execute(currency_query, {
            "currency1": currency1,
            "country": "us",
            "month": month.lower(),
            "year": year
        }).fetchone()

    # Step 3: Use the conversion rate
    currency_rate  = result[0] if result else None

    # Step 4: Create a USD-converted copy if conversion rate exists
    df_usd = df.copy()

    # ✅ Step 3: Convert monetary columns to USD if country found
    if currency_rate :
        monetary_columns = [
            'product_sales', 'product_sales_tax', 'postage_credits', 
            'shipping_credits_tax', 'gift_wrap_credits', 'giftwrap_credits_tax', 
            'promotional_rebates', 'promotional_rebates_tax', 'sales_tax_collected',
            'marketplace_facilitator_tax', 'selling_fees', 'fba_fees', 
            'other_transaction_fees', 'other', 'total', 'price_in_gbp', 
            'cost_of_unit_sold'
        ]

        for col in monetary_columns:
            if col in df_usd.columns:
                df_usd[col] = pd.to_numeric(df_usd[col], errors='coerce') * currency_rate 
    else:
        print("⚠️ No conversion rate found for:", currency1, country, month, year)

    # ✅ Step 4: Insert the USD converted data into global table
    global_table_name = f"user_{user_id}_total_country_global_data"

    user_global_table = Table(
        global_table_name, meta,
        Column('id', Integer, primary_key=True),  # optional, just for reflection
        Column('date_time', String),  # Changed from DateTime to String
        Column('settlement_id', String),
        Column('type', String),
        Column('order_id', String),
        Column('sku', String),
        Column('description', String),
        Column('quantity', Integer),
        Column('price_in_gbp', Float),
        Column('cost_of_unit_sold', Float),
        Column('marketplace', String),
        Column('fulfilment', String),
        Column('fulfillment', String),
        Column('order_city', String),
        Column('order_state', String),
        Column('order_postal', String),
        Column('tax_collection_model', String),
        Column('product_sales', Float),
        Column('product_sales_tax', Float),
        Column('postage_credits', Float),
        Column('shipping_credits', Float),
        Column('shipping_credits_tax', Float),
        Column('gift_wrap_credits', Float),
        Column('giftwrap_credits_tax', Float),
        Column('promotional_rebates', Float),
        Column('promotional_rebates_tax', Float),
        Column('sales_tax_collected', Float),
        Column('marketplace_withheld_tax', Float),
        Column('marketplace_facilitator_tax', Float),
        Column('selling_fees', Float),
        Column('percentage1', Float),
        Column('fba_fees', Float),
        Column('percentage2', Float),
        Column('other_transaction_fees', Float),
        Column('other', Float),
        Column('total', Float),
        Column('month', String),  # Added column
        Column('year', String),
        Column('product_name', String),
        Column('country', String),
        
    )
    meta.create_all(engine)

    with engine.connect() as connection:
        # Delete previous records from user_total_country_global_data
        connection.execute(
            text(f"DELETE FROM {global_table_name} WHERE month = :month AND year = :year AND country = :country"),
            {"month": month, "year": year, "country": country}
        )
        connection.commit()

    # Add month, year, and country columns to the dataframe
    df_usd['month'] = month
    df_usd['year'] = year
    df_usd['country'] = country

    # Ensure all data types are correct before insertion
    # Convert any problematic columns to appropriate types
    for col in df_usd.columns:
        if df_usd[col].dtype == 'object' and col not in ['date_time', 'settlement_id', 'type', 'order_id', 'sku', 
                                                        'description', 'marketplace', 'fulfilment', 'order_city', 
                                                        'order_state', 'order_postal', 'tax_collection_model',
                                                        'month', 'year', 'country', 'product_name']:
            # Try to convert string columns that should be numeric to float
            df_usd[col] = pd.to_numeric(df_usd[col], errors='coerce')

    # Handle NaN values explicitly
    df_usd = df_usd.fillna({col: 0.0 for col in df_usd.select_dtypes(include=['float64']).columns})
    df_usd = df_usd.fillna({col: 0 for col in df_usd.select_dtypes(include=['int64']).columns})
    df_usd = df_usd.fillna('')  # Fill string columns with empty string

    try:
        # Convert DataFrame to dict of records for proper SQL insertion
        records = df_usd.to_dict(orient='records')
        
        with engine.begin() as connection:
            for chunk in [records[i:i+1000] for i in range(0, len(records), 1000)]:
                connection.execute(user_global_table.insert(), chunk)
        
    except Exception as e:
        return jsonify({'error': f'Error inserting into global table: {str(e)}'}), 500

    # Update required columns to match the actual column names in the database
    REQUIRED_COLUMNS = [
            "order_id", "sku", "description", "product_sales", "product_sales_tax", 
            "postage_credits", "shipping_credits_tax", "gift_wrap_credits", "giftwrap_credits_tax", 
            "promotional_rebates", "promotional_rebates_tax", "marketplace_facilitator_tax", 
            "selling_fees", "fba_fees", "other_transaction_fees", "errorstatus", "answer", 
            "difference", "fbaerrorstatus", "fbaanswer"  # Use original capitalization
    ]
        
    # Convert both files to base64
    def encode_file_to_base64(file_path):
        with open(file_path, "rb") as file:
            return base64.b64encode(file.read()).decode()
             
    def replace_nan_with_null(data):
        if isinstance(data, dict):
            return {key: replace_nan_with_null(value) for key, value in data.items()}
        elif isinstance(data, list):
            return [replace_nan_with_null(item) for item in data]
        elif isinstance(data, float) and (data != data):  # NaN check
            return None
        return data
        
    if file1 and file1.filename != '':
        try:
            if file1.filename.endswith('.csv'):
                with open(file1_path, 'r', encoding='utf-8-sig') as f:
                    lines = f.readlines()
                    header_line = None
                    # Look for header row
                    for i, line in enumerate(lines):
                        if any(keyword in line.lower() for keyword in ['date/time', 'date / time', 'date time']):
                            header_line = i
                            break
                if header_line is None:
                    return jsonify({'error': 'Could not find header row with "date/time" column'}), 400

                df = pd.read_csv(file1_path, skiprows=header_line, encoding="utf-8-sig", dayfirst=True, skip_blank_lines=True)
                df.columns = df.columns.str.strip()

            elif file1.filename.endswith(('.xls', '.xlsx')):
                import openpyxl
    
                wb = openpyxl.load_workbook(file1_path, read_only=True)
                ws = wb.active
                header_line = None
                for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
                    if row is not None and any(
                        cell is not None and any(keyword in str(cell).lower() for keyword in ['date/time', 'date / time', 'date time'])
                        for cell in row
                    ):
                        header_line = i
                        break
                if header_line is None:
                    return jsonify({'error': 'Could not find header row with "date/time" column in Excel'}), 400

                # Now read excel from the detected header line
                df = pd.read_excel(file1_path, skiprows=header_line)
                df.columns = df.columns.str.strip()
            else:
                return jsonify({'error': 'Invalid file format. Only .csv and .xlsx files are allowed'}), 400

            df.columns = [c.lower() for c in df.columns]  # Lowercase all column names for consistency
            
            # Clean numeric columns again for this new dataframe
            for col in [c for c in df.columns if c in numeric_columns]:
                df[col] = df[col].apply(clean_numeric_value)
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
            df.rename(columns=COLUMN_MAPPING, inplace=True)
            for col in COLUMN_MAPPING.values():
                if col not in df.columns:
                    if col in numeric_columns:
                        df[col] = 0
                    else:
                        df[col] = None  # or use np.nan if preferred
            
            if country.upper() == 'UK':
                sku_column = 'sku_uk'
            elif country.upper() == 'US':
                sku_column = 'sku_us'
            else:
                raise ValueError("Unsupported country")

            
            # with engine.connect() as conn:
            #     country_df = pd.read_sql(f"SELECT {sku_column} AS sku, price,currency, product_name FROM {country_table_name}", conn)

            # df = df.merge(country_df, on='sku', how='left')

            
            MONTH_NUM = {
                "january": 1, "february": 2, "march": 3, "april": 4,
                "may": 5, "june": 6, "july": 7, "august": 8,
                "september": 9, "october": 10, "november": 11, "december": 12
            }

            target_month_num = MONTH_NUM.get(month.lower())
            if not target_month_num:
                raise ValueError(f"Invalid month: {month}. Expected january..december")

            target_year = int(year)  # form year string -> int

            month_case_sql = """
            CASE lower(month)
                WHEN 'january' THEN 1
                WHEN 'february' THEN 2
                WHEN 'march' THEN 3
                WHEN 'april' THEN 4
                WHEN 'may' THEN 5
                WHEN 'june' THEN 6
                WHEN 'july' THEN 7
                WHEN 'august' THEN 8
                WHEN 'september' THEN 9
                WHEN 'october' THEN 10
                WHEN 'november' THEN 11
                WHEN 'december' THEN 12
                ELSE NULL
            END
            """

            # year is string in DB, so cast it for comparisons
            price_asof_query = text(f"""
                SELECT sku, price, currency, product_name
                FROM (
                    SELECT
                        {sku_column} AS sku,
                        price,
                        currency,
                        product_name,
                        CAST(year AS INTEGER) AS year_int,
                        {month_case_sql} AS month_num,
                        ROW_NUMBER() OVER (
                            PARTITION BY {sku_column}
                            ORDER BY CAST(year AS INTEGER) DESC, {month_case_sql} DESC
                        ) AS rn
                    FROM {country_table_name}
                    WHERE
                        year IS NOT NULL
                        AND trim(year) <> ''
                        AND {month_case_sql} IS NOT NULL
                        AND (
                            CAST(year AS INTEGER) < :target_year
                            OR (CAST(year AS INTEGER) = :target_year AND {month_case_sql} <= :target_month_num)
                        )
                ) x
                WHERE x.rn = 1
            """)

            with engine.connect() as conn:
                country_df = pd.read_sql(
                    price_asof_query,
                    conn,
                    params={"target_year": target_year, "target_month_num": target_month_num}
                )

            df = df.merge(country_df, on="sku", how="left")


            with engine.connect() as conn:
                countries_df = pd.read_sql(f"SELECT sku, product_group FROM {countris_table_name}", conn)

            df = df.merge(countries_df, on='sku', how='left')

            with engine1.connect() as conn:
                currency_query = text("""
                    SELECT conversion_rate
                    FROM currency_conversion 
                    WHERE lower(user_currency) = :currency 
                    AND lower(country) = :country 
                    AND lower(month) = :month 
                    AND year = :year
                    LIMIT 1
            """)

                result = conn.execute(
                    currency_query,
                    {
                        "currency": country_df['currency'].dropna().iloc[0].lower(),  # Get any available currency from df
                        "country": country.lower(),
                        "month": month.lower(),
                        "year": year
                    }
                ).fetchone()

        # Step 3: Apply conversion rate to calculate price_in_gbp
            conversion_rate = result[0] if result else None

            if conversion_rate:
                df['price_in_gbp'] = df['price'] * conversion_rate
            else:
                df['price_in_gbp'] = None  # Or handle fallback logic if conversion rate not found

            # Calculate cost_of_unit_sold where both values are not null
            df['cost_of_unit_sold'] = df.apply(
                lambda row: row['price_in_gbp'] * row['quantity']
                if str(row.get('type', '')).strip().lower() == 'shipment'
                and pd.notnull(row['price_in_gbp'])
                and pd.notnull(row['quantity'])
                else 0.0,
                axis=1
            )

            df.drop(columns=['price'], inplace=True)

            # Convert any remaining string numeric values to float
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            for col in df.columns:
                if df[col].dtype == 'object':  # likely a string column
                    if df[col].str.contains(',').any():
                        print(f"Column {col} contains comma-formatted numbers!")
            for col in df.columns:
                        if df[col].dtype == 'object':  # likely to have commas or bad data
                            # Remove commas and convert to numeric
                            df[col] = df[col].str.replace(',', '', regex=True)

                    # Now convert all possible numeric columns to float
            df = df.apply(lambda x: pd.to_numeric(x, errors='ignore') if x.dtype == 'object' else x)

            # df.to_sql(table_name, con=engine, if_exists='append', index=False)
            df.to_sql(table_name, con=engine, if_exists='append', index=False)

            file1_base64 = encode_file_to_base64(file1_path)
            file2_base64 = encode_file_to_base64(file2_path)

            sku_list = df['sku'].dropna().tolist()  # Filter out NaN values

            referral_fees = get_referral_fees(user_id, country, sku_list)

            if referral_fees is not None:
                df_modified = apply_modifications(df, country)

                # IMPORTANT: apply_modifications may recalculate COGS for every row.
                # Re-apply shipment-only COGS so Refund/FBA reimbursement/other rows stay zero.
                type_normalized = (
                    df_modified['type']
                    .fillna('')
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )
                shipment_mask = type_normalized.eq('shipment')

                df_modified['price_in_gbp'] = pd.to_numeric(
                    df_modified['price_in_gbp'], errors='coerce'
                )
                df_modified['quantity'] = pd.to_numeric(
                    df_modified['quantity'], errors='coerce'
                )

                df_modified['cost_of_unit_sold'] = 0.0
                df_modified.loc[shipment_mask, 'cost_of_unit_sold'] = (
                    df_modified.loc[shipment_mask, 'price_in_gbp'].fillna(0.0)
                    * df_modified.loc[shipment_mask, 'quantity'].fillna(0.0)
                )
                
                # Ensure numeric columns are properly converted before saving to database
                for col in numeric_columns:
                    if col in df_modified.columns:
                        df_modified[col] = pd.to_numeric(df_modified[col], errors='coerce')
                
                df_modified.to_sql(table_name, con=engine, if_exists='replace', index=False, method='multi')

                excel_output = io.BytesIO()
                df_modified.to_excel(excel_output, index=False)
                excel_output.seek(0)

                pnl_report = generate_pnl_report(year, month)

                table_name = f"user_{user_id}_{country}_{month}{year}_data"
    
                with engine.connect() as conn:
                    query = f"""
                    SELECT * FROM {table_name} 
                    WHERE ErrorStatus IN ('cases to be inquired', 'NoReferralFee')
                    AND sku <> '0'
                    AND sku IS NOT NULL
                    AND TRIM(sku) <> ''
                    """
                    error_df = pd.read_sql(query, conn)
                
                # Filter error_df to keep only the required columns
                error_df = error_df[[col for col in REQUIRED_COLUMNS if col in error_df.columns]]

                # Save the error file if there are errors
                error_file_path = None
                error_file_base64 = None
                if not error_df.empty:
                    error_filename = f"error_file_{user_id}{country}{month}_{year}.xlsx"
                    error_file_path = os.path.join( error_filename)                    
                    error_df.to_excel(error_file_path, index=False)
                    
                    # Encode error file to base64
                    with open(error_file_path, "rb") as error_file:
                        error_file_base64 = base64.b64encode(error_file.read()).decode()

                quarter_mapping = {
                    'january': 'Q1', 'february': 'Q1', 'march': 'Q1',
                    'april': 'Q2', 'may': 'Q2', 'june': 'Q2',
                    'july': 'Q3', 'august': 'Q3', 'september': 'Q3',
                    'october': 'Q4', 'november': 'Q4', 'december': 'Q4'
                }

                if month.lower() not in quarter_mapping:
                    return jsonify({'error': 'Invalid month provided'}), 400

                quarter = quarter_mapping[month.lower()]
                quarterly_table = f"quater{quarter[-1]}{country}{year}_table"
                

                # Generate sales pie chart
                if country.lower() == 'uk':
                    total_cous, total_amazon_fee, cm2_profit, rembursement_fee, platform_fee, total_expense, total_profit, total_fba_fees, advertising_total, taxncredit, reimbursement_vs_sales, cm2_margins, acos, rembursment_vs_cm2_margins, total_sales, unit_sold, total_product_sales = process_skuwise_data(user_id, country, month, year)
                    ytd_pie_chart = process_yearly_skuwise_data(user_id, country, year)
                    qtd_pie_chart = process_quarterly_skuwise_data(user_id, country, month, year, quarter, db_url)
                    # sales_pie_chart, platform_fee, rembursement_fee = create_sales_pie_chart(df_modified)
                    # expense_pie_chart,   otherwplatform = create_expense_pie_chart(df_modified, country, month, year)
                    
                elif country.lower() == 'us':
                    platform_fee, rembursement_fee, total_cous, total_amazon_fee,  total_profit, total_expense, total_fba_fees, cm2_profit, cm2_margins, acos, rembursment_vs_cm2_margins, advertising_total, reimbursement_vs_sales, unit_sold, total_sales, otherwplatform, taxncredit = process_skuwise_us_data(user_id, country, month, year)
                    ytd_pie_chart = process_us_yearly_skuwise_data(user_id, country, year)
                    qtd_pie_chart = process_us_quarterly_skuwise_data(user_id, country, month, year, quarter, db_url)
                

                process_global_monthly_skuwise_data(user_id, country, year, month)
                process_global_quarterly_skuwise_data(user_id, country, month, year, quarter , db_url)
                process_global_yearly_skuwise_data(user_id, country, year)


                
                
                # Check if entry already exists using string month
                existing_entry = UploadHistory.query.filter_by(user_id=user_id, country=country, month=month, year=year).first()
                if existing_entry:
                    db.session.delete(existing_entry)
                    db.session.commit()

                new_upload = UploadHistory(
                    user_id=user_id,
                    year=year,
                    month=month,
                    country=country,
                    file_name=secure_filename(file1.filename), 
                    sales_chart_img=None,
                    expense_chart_img=None,
                    total_sales=float(total_sales),
                    total_product_sales=float(total_product_sales),
                    total_profit=float(total_profit),
                    otherwplatform=float(platform_fee),
                    taxncredit = float(taxncredit) if taxncredit is not None else 0.0,
                    total_expense=float(total_expense),
                    qtd_pie_chart=qtd_pie_chart,
                    ytd_pie_chart=ytd_pie_chart,
                    total_cous=float(total_cous),
                    total_amazon_fee=float(total_amazon_fee),
                    total_fba_fees=float(total_fba_fees),
                    platform_fee=float(platform_fee),  # Include platform_fee here
                    rembursement_fee=float(rembursement_fee),
                    cm2_profit= float(cm2_profit),
                    cm2_margins= float(cm2_margins),
                    acos= float(acos),
                    rembursment_vs_cm2_margins= float(rembursment_vs_cm2_margins),
                    advertising_total= float(advertising_total),
                    reimbursement_vs_sales= float(reimbursement_vs_sales),
                    unit_sold = int(unit_sold)
                )
                db.session.add(new_upload)
                db.session.commit()       

                next_month, next_year = get_next_month_year(month, year)
                next_year = str(next_year)
                next_table = f"skuwisemonthly_{user_id}_{country}_{next_month}{next_year}"

                if table_exists(engine, next_table):
                    print("Table exists! Running next month logic...")
    # Run for next month too
                    if country.lower() == 'uk':
                        total_cous, total_amazon_fee, cm2_profit, rembursement_fee, platform_fee, total_expense, total_profit, total_fba_fees, advertising_total, taxncredit, reimbursement_vs_sales, cm2_margins, acos, rembursment_vs_cm2_margins, total_sales, unit_sold, total_product_sales = process_skuwise_data(user_id, country, next_month, next_year)
                    elif country.lower() == 'us':
                        platform_fee, rembursement_fee, total_cous, total_amazon_fee,  total_profit, total_expense, total_fba_fees, cm2_profit, cm2_margins, acos, rembursment_vs_cm2_margins, advertising_total, reimbursement_vs_sales, unit_sold, total_sales, otherwplatform,taxncredit = process_skuwise_us_data(user_id, country, next_month, next_year)

                else:
                    print("Table does NOT exist for next month.")

                      
            
                response_data = {
                    'success': True,
                    # 'sales_chart_img': sales_pie_chart,
                    # 'expense_chart_img': expense_pie_chart,
                    'total_sales': total_sales,
                    'total_product_sales': total_product_sales,
                    'total_profit': total_profit,
                    'otherwplatform': platform_fee,
                    'taxncredit': taxncredit,
                    'total_expense': total_expense,
                    'total_fba_fees': total_fba_fees,  
                    'excel_file': base64.b64encode(excel_output.getvalue()).decode(),
                    'file1': file1_base64,  # Original file 1 in base64
                    'file2': file2_base64,  # Original file 2 in base64
                    'error_file': error_file_base64 if 'error_file_base64' in locals() and error_file_base64 else None,
                    'platform_fee': platform_fee,
                }
                # Clean up NaN values before returning the response
                response_data_cleaned = replace_nan_with_null(response_data)
                return jsonify(response_data_cleaned)
            else:
                return jsonify({'success': False, 'message': 'Referral fee not found for the SKUs in the uploaded file.'}), 400
        except Exception as e:
            print("Error:", e)
            return jsonify({'success': False, 'message': 'An error occurred while processing your request.'}), 500

    else:
        return jsonify({'success': False, 'message': 'No file selected. Please select a file.'}), 400





# ---------- Helpers ----------

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Trim, lowercase, replace spaces/dashes with underscores."""
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'[\s\-]+', '_', regex=True)
    )
    return df

def _promote_first_row_to_header_if_needed(df: pd.DataFrame) -> pd.DataFrame:
    """Detect sheets where pandas used 'Unnamed' headers and real headers are on row 0."""
    total = len(df.columns)
    if total == 0:
        return df
    unnamed_count = sum(str(c).lower().startswith('unnamed') for c in df.columns)
    if unnamed_count / total >= 0.5 and len(df) > 0:
        first_row = df.iloc[0]
        candidate_headers = [str(x).strip() if x is not None else "" for x in first_row.tolist()]
        known = {
            's. no.', 's_no', 'product name', 'product_name',
            'product barcode', 'product_barcode', 'asin', 'sku',
            'sku_uk', 'sku_us', 'sku_canada', 'landing cost',
            'landing_cost', 'your price', 'your_price', 'currency',
            'amazon-store', 'amazon_store', 'marketplace'
        }
        matches = sum(1 for x in candidate_headers if x.lower() in known)
        if matches >= 2:
            df = df.copy()
            df.columns = candidate_headers
            df = df.iloc[1:].reset_index(drop=True)
    return df

def _pick(row: dict, keys):
    for k in keys:
        if k in row and row[k] not in (None, '', '—', '--'):
            return row[k]
    return None

def _marketplace_to_country(store):
    if not store:
        return None
    s = str(store).strip().upper()
    if s in {'GB', 'UK', 'GBR', 'UNITED KINGDOM'}:
        return 'UK'
    if s in {'US', 'USA', 'UNITED STATES'}:
        return 'US'
    return None


@upload_bp.route('/ConfirmationFeepreview', methods=['GET'])
def ConfirmationFeepreview():
    return jsonify({'message': 'ConfirmationFeepreview successful!'}), 200

def _quote_ident(engine, name):
    return engine.dialect.identifier_preparer.quote(name)


def _norm_key_sql(col_sql):
    return f"lower(trim(({col_sql})::text))"


def _find_column(columns, candidates):
    lookup = {c["name"].lower(): c["name"] for c in columns}

    for candidate in candidates:
        if candidate.lower() in lookup:
            return lookup[candidate.lower()]

    normalized_lookup = {
        re.sub(r"[^a-z0-9]", "", c["name"].lower()): c["name"]
        for c in columns
    }

    for candidate in candidates:
        key = re.sub(r"[^a-z0-9]", "", candidate.lower())
        if key in normalized_lookup:
            return normalized_lookup[key]

    return None


def _infer_country_from_table(table_name):
    t = table_name.lower()

    if "_uk_" in t or t.endswith("_uk") or "_uk" in t:
        return "uk"

    if "_us_" in t or t.endswith("_us") or "_us" in t:
        return "us"

    if "_canada_" in t or t.endswith("_canada") or "_canada" in t:
        return "canada"

    if "_global_" in t or "global" in t:
        return None

    return None


def _target_table_prefixes(user_id):
    return [
        f"currentinventory_{user_id}_",
        f"nse_{user_id}_",
        f"quarter1_{user_id}_",
        f"quarter2_{user_id}_",
        f"quarter3_{user_id}_",
        f"quarter4_{user_id}_",
        f"sku_{user_id}_",
        f"skuwisemonthly_{user_id}_",
        f"skuwiseyearly_{user_id}_",
    ]



def sync_uploaded_sku_to_all_user_tables(engine, user_id, sku_rows):
    """
    Updates ONLY sku and product_name columns in all matching user tables.

    Table patterns updated:
      currentinventory_{user_id}_
      nse_{user_id}_
      quarter1_{user_id}_
      quarter2_{user_id}_
      quarter3_{user_id}_
      quarter4_{user_id}_
      sku_{user_id}_
      skuwisemonthly_{user_id}_
      skuwiseyearly_{user_id}_

    Matching logic:
      1. If table sku matches uploaded sku, update product_name.
      2. If table product_name matches uploaded product_name, update sku.
      3. UK tables use sku_uk.
      4. US tables use sku_us.
      5. Canada tables use sku_canada.
      6. Global tables use all available sku mappings.
    """

    inspector = inspect(engine)

    all_tables = inspector.get_table_names(schema="public")

    prefixes = _target_table_prefixes(user_id)

    target_tables = [
        table_name
        for table_name in all_tables
        if any(table_name.lower().startswith(prefix.lower()) for prefix in prefixes)
    ]

    mappings = []

    for row in sku_rows:
        product_name = row.get("product_name")

        if not product_name:
            continue

        product_name = str(product_name).strip()

        if product_name.lower() in ("", "nan", "none", "null", "total"):
            continue

        sku_uk = row.get("sku_uk")
        sku_us = row.get("sku_us")
        sku_canada = row.get("sku_canada")

        if sku_uk and str(sku_uk).strip():
            mappings.append({
                "country": "uk",
                "sku": str(sku_uk).strip(),
                "product_name": product_name,
            })

        if sku_us and str(sku_us).strip():
            mappings.append({
                "country": "us",
                "sku": str(sku_us).strip(),
                "product_name": product_name,
            })

        if sku_canada and str(sku_canada).strip():
            mappings.append({
                "country": "canada",
                "sku": str(sku_canada).strip(),
                "product_name": product_name,
            })

    # remove duplicate mappings
    unique_mappings = {}

    for mapping in mappings:
        key = (
            mapping["country"].lower().strip(),
            mapping["sku"].lower().strip(),
            mapping["product_name"].lower().strip(),
        )
        unique_mappings[key] = mapping

    mappings = list(unique_mappings.values())

    if not mappings:
        return {
            "tables_checked": len(target_tables),
            "tables_updated": [],
            "message": "No valid SKU mappings found",
        }

    updated_tables = []

    with engine.begin() as conn:
        # Important with SQLAlchemy pool:
        # temp tables can remain on reused PostgreSQL connections.
        conn.execute(text("DROP TABLE IF EXISTS tmp_unique_product_mapping"))
        conn.execute(text("DROP TABLE IF EXISTS tmp_uploaded_sku_mapping"))

        conn.execute(text("""
            CREATE TEMP TABLE tmp_uploaded_sku_mapping (
                country TEXT,
                sku TEXT,
                product_name TEXT,
                sku_key TEXT,
                product_key TEXT
            ) ON COMMIT DROP
        """))

        conn.execute(
            text("""
                INSERT INTO tmp_uploaded_sku_mapping
                    (country, sku, product_name, sku_key, product_key)
                VALUES
                    (
                        :country,
                        :sku,
                        :product_name,
                        lower(trim(:sku)),
                        lower(trim(:product_name))
                    )
            """),
            mappings
        )

        conn.execute(text("DROP TABLE IF EXISTS tmp_unique_product_mapping"))

        conn.execute(text("""
            CREATE TEMP TABLE tmp_unique_product_mapping ON COMMIT DROP AS
            SELECT
                country,
                product_key,
                MAX(sku) AS sku,
                COUNT(DISTINCT sku_key) AS sku_count
            FROM tmp_uploaded_sku_mapping
            GROUP BY country, product_key
            HAVING COUNT(DISTINCT sku_key) = 1
        """))

        for table_name in target_tables:
            # Skip the master SKU table because it was just uploaded/recreated.
            if table_name.lower() == f"sku_{user_id}_data_table".lower():
                continue

            try:
                columns = inspector.get_columns(table_name, schema="public")
            except Exception:
                continue

            sku_col = _find_column(columns, [
                "sku",
                "SKU",
                "seller_sku",
                "seller sku",
                "merchant_sku",
                "merchant sku",
            ])

            product_col = _find_column(columns, [
                "product_name",
                "Product Name",
                "product name",
                "product",
                "title",
            ])

            # Update only tables that have both sku and product_name columns.
            if not sku_col or not product_col:
                continue

            q_table = _quote_ident(engine, table_name)
            q_sku_col = _quote_ident(engine, sku_col)
            q_product_col = _quote_ident(engine, product_col)

            table_country = _infer_country_from_table(table_name)

            params = {}

            if table_country:
                country_filter_m = "AND m.country = :country"
                country_filter_u = "AND u.country = :country"
                params["country"] = table_country
            else:
                country_filter_m = ""
                country_filter_u = ""

            # 1. Same SKU found => update product_name.
            update_product_sql = f"""
                UPDATE {q_table} AS t
                SET {q_product_col} = m.product_name
                FROM tmp_uploaded_sku_mapping AS m
                WHERE {_norm_key_sql(f't.{q_sku_col}')} = m.sku_key
                  {country_filter_m}
                  AND m.product_name IS NOT NULL
                  AND trim(m.product_name) <> ''
                  AND (
                        t.{q_product_col} IS NULL
                        OR trim(t.{q_product_col}::text) IS DISTINCT FROM trim(m.product_name)
                  )
            """

            result1 = conn.execute(text(update_product_sql), params)

            # 2. Same product_name found => update SKU.
            # This only updates if that product_name maps to exactly one SKU.
            update_sku_sql = f"""
                UPDATE {q_table} AS t
                SET {q_sku_col} = u.sku
                FROM tmp_unique_product_mapping AS u
                WHERE {_norm_key_sql(f't.{q_product_col}')} = u.product_key
                  {country_filter_u}
                  AND u.sku IS NOT NULL
                  AND trim(u.sku) <> ''
                  AND (
                        t.{q_sku_col} IS NULL
                        OR trim(t.{q_sku_col}::text) IS DISTINCT FROM trim(u.sku)
                  )
            """

            result2 = conn.execute(text(update_sku_sql), params)

            affected_rows = (result1.rowcount or 0) + (result2.rowcount or 0)

            if affected_rows > 0:
                updated_tables.append({
                    "table": table_name,
                    "rows_changed": affected_rows,
                })

        # clean manually also, not only ON COMMIT DROP
        conn.execute(text("DROP TABLE IF EXISTS tmp_unique_product_mapping"))
        conn.execute(text("DROP TABLE IF EXISTS tmp_uploaded_sku_mapping"))

    return {
        "tables_checked": len(target_tables),
        "tables_updated": updated_tables,
    }



@upload_bp.route('/multiCountry', methods=['POST'])
def multiCountry():
    # ---------- Auth ----------
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ---------- File ----------
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400
    if not (file.filename.lower().endswith('.csv') or file.filename.lower().endswith('.xlsx')):
        return jsonify({'error': 'Invalid file type. Only CSV or XLSX files are allowed.'}), 400

    # ---------- DB ----------
    engine = user_engine
    inspector = inspect(engine)
    metadata = MetaData()
    table_name = f"sku_{user_id}_data_table"

    # ✅ Updated schema includes local_stock + in_transit_units
    user_specific_table = Table(
        table_name, metadata,
        Column('id', Integer, primary_key=True),
        Column('user_id', Integer, nullable=False),
        Column('s_no', Integer, nullable=True),
        Column('product_name', String(255), nullable=True),
        Column('product_barcode', String(255), nullable=True),
        Column('sku_uk', String(255), nullable=True),
        Column('sku_us', String(255), nullable=True),
        Column('sku_canada', String(255), nullable=True),
        Column('asin', String(255), nullable=True),
        Column('price', Float, nullable=True),
        Column('currency', String(255), nullable=True),

        # ✅ NEW
        Column('local_stock', Integer, nullable=True, server_default=text("0")),
        Column('in_transit_units', Integer, nullable=True, server_default=text("0")),

        Column('month', String(20), nullable=True),
        Column('year', String(20), nullable=True),
    )

    # ---------- helpers ----------
    def _month_name_from_num(mi: int):
        if 1 <= mi <= 12:
            return calendar.month_name[mi]  # "January"
        return None

    def _parse_mm_yyyy(val):
        """
        Accepts '01/2024', '1/2024', '01-2024', '2024-01'
        Returns (month_name, year_str) or (None, None)
        """
        if val in (None, ''):
            return None, None

        s = str(val).strip()
        if not s:
            return None, None

        s = s.replace('.', '/').replace('-', '/')
        parts = s.split('/')

        if len(parts) == 2 and parts[0].strip().isdigit() and parts[1].strip().isdigit():
            a, b = parts[0].strip(), parts[1].strip()

            # MM/YYYY
            if len(b) == 4:
                mi = int(a)
                month_name = _month_name_from_num(mi)
                if month_name:
                    return month_name, b

            # YYYY/MM
            if len(a) == 4:
                yi = a
                mi = int(b)
                month_name = _month_name_from_num(mi)
                if month_name:
                    return month_name, yi

        return None, None

    def _extract_month_year(row_dict):
        """
        Priority:
          1) date column (normalized key: 'date') -> pandas Timestamp/datetime
          2) date string MM/YYYY (csv)
          3) fallback separate month/year columns
        """
        date_val = _pick(row_dict, ['date', 'Date'])

        month = None
        year = None

        # 1) datetime-like
        if date_val not in (None, '') and hasattr(date_val, 'month') and hasattr(date_val, 'year'):
            try:
                mi = int(date_val.month)
                yi = int(date_val.year)
                month = _month_name_from_num(mi)
                year = str(yi)
                if month and year:
                    return month, year
            except Exception:
                pass

        # 2) string MM/YYYY
        m2, y2 = _parse_mm_yyyy(date_val)
        if m2 and y2:
            return m2, y2

        # 3) fallback month/year columns
        month_raw = _pick(row_dict, ['month', 'Month', 'mon', 'mm'])
        year_raw = _pick(row_dict, ['year', 'Year', 'yyyy', 'yy'])

        # normalize month
        if month_raw not in (None, ''):
            s = str(month_raw).strip()
            if s.isdigit():
                month = _month_name_from_num(int(s))
            else:
                s_lower = s.lower()
                for i in range(1, 13):
                    if s_lower in (calendar.month_name[i].lower(), calendar.month_abbr[i].lower()):
                        month = calendar.month_name[i]
                        break

        # normalize year
        if year_raw not in (None, ''):
            y = str(year_raw).strip()
            if y.isdigit() and len(y) == 2:
                y = "20" + y
            year = y

        return month, year

    def _safe_int(val, default=0):
        try:
            if val in (None, ''):
                return default
            return int(float(str(val).strip()))
        except Exception:
            return default

    def _s(x):
        if x is None:
            return None
        x = str(x).strip()
        return x if x else None

    session = None
    try:
        # ---------- Read into DataFrame ----------
        if file.filename.lower().endswith('.csv'):
            try:
                df = pd.read_csv(BytesIO(file.read()))
            except UnicodeDecodeError:
                file.stream.seek(0)
                df = pd.read_csv(BytesIO(file.read()), encoding='latin-1')
        else:
            df = pd.read_excel(BytesIO(file.read()))

        # Fix “Unnamed” header case, then normalize
        df = _promote_first_row_to_header_if_needed(df)
        df = _normalize_columns(df)
        df = df.where(pd.notnull(df), None)

        # (Re)create fresh table per upload
        if inspector.has_table(table_name):
            Table(table_name, MetaData(), autoload_with=engine).drop(engine, checkfirst=True)
        metadata.create_all(engine)

        Session = sessionmaker(bind=engine)
        session = Session()

        inserts = []
        for _, r in df.iterrows():
            row = r.to_dict()

            s_no = _pick(row, ['s_no', 's_no.', 's._no.', 'no', '#'])
            try:
                s_no = int(s_no) if s_no not in (None, '') else None
            except Exception:
                s_no = None

            product_name = _pick(row, ['product_name', 'product-name', 'title', 'item_name'])
            product_barcode = _pick(row, ['product_barcode', 'barcode', 'ean', 'upc'])
            asin = _pick(row, ['asin'])

            amazon_store = _pick(row, ['amazon_store', 'amazon-store', 'marketplace'])
            country = _marketplace_to_country(amazon_store)

            raw_sku = _pick(row, ['sku', 'seller_sku', 'merchant_sku', 'sku_uk', 'sku_us'])
            sku_uk = _pick(row, ['sku_uk', 'uk_sku'])
            sku_us = _pick(row, ['sku_us', 'us_sku'])
            sku_canada = _pick(row, ['sku_canada', 'canada_sku', 'sku_ca', 'ca_sku'])
            if not sku_uk and not sku_us and not sku_canada and raw_sku:
                if country == 'UK':
                    sku_uk = str(raw_sku).strip()
                elif country == 'US':
                    sku_us = str(raw_sku).strip()
                elif country == 'Canada':
                    sku_canada = str(raw_sku).strip()

            price_value = _pick(row, ['landing_cost', 'your_price', 'sales_price', 'price'])
            try:
                price_value = float(price_value) if price_value not in (None, '') else None
            except Exception:
                price_value = None

            currency = _pick(row, ['currency'])

            # ✅ month/year from template "Date"
            month, year = _extract_month_year(row)

            # ✅ NEW: read stock columns (support different headers)
            local_stock = _safe_int(_pick(row, ['local_stock', 'local stock', 'Local Stock']), default=0)
            in_transit_units = _safe_int(_pick(row, ['in_transit_units', 'in transit units', 'In Transit Units']), default=0)

            # Skip fully empty lines (include stock too)
            if not any([
                s_no, product_name, product_barcode, asin, sku_uk, sku_us, sku_canada,
                price_value, currency, month, year, local_stock, in_transit_units
            ]):
                continue

            inserts.append({
                'user_id': user_id,
                's_no': s_no,
                'product_name': _s(product_name),
                'product_barcode': _s(product_barcode),
                'sku_uk': _s(sku_uk),
                'sku_us': _s(sku_us),
                'sku_canada': _s(sku_canada),
                'asin': _s(asin),
                'price': price_value,
                'currency': _s(currency),

                # ✅ NEW
                'local_stock': local_stock,
                'in_transit_units': in_transit_units,

                'month': _s(month),
                'year': _s(year),
            })

        sync_result = None

        if inserts:
            session.execute(user_specific_table.insert(), inserts)
            session.commit()

            # ✅ Update ONLY sku and product_name in all user tables
            sync_result = sync_uploaded_sku_to_all_user_tables(
                engine=engine,
                user_id=user_id,
                sku_rows=inserts
            )

            msg = 'SKU file uploaded and sku/product_name synced successfully'
        else:
            msg = 'File processed, but no valid rows found to insert.'

        session.close()

        return jsonify({
            'success': True,
            'message': msg,
            'sync_result': sync_result
        }), 200

    except Exception as e:
        try:
            if session is not None:
                session.rollback()
                session.close()
        except Exception:
            pass
        print(f"Error processing file: {str(e)}")
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500



@upload_bp.route('/file-upload-status', methods=['GET'])
def check_file_upload_status():
    # ---------- Auth ----------
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ---------- DB ----------
    engine = user_engine
    inspector = inspect(engine)
    table_name = f"sku_{user_id}_data_table"

    try:
        if not inspector.has_table(table_name):
            return jsonify({'file_uploaded': False}), 200

        with engine.connect() as conn:
            result = conn.execute(
                text(f"""
                    SELECT COUNT(*) FROM {table_name}
                    WHERE user_id = :user_id
                      AND (sku_us IS NOT NULL OR sku_uk IS NOT NULL)
                """),
                {"user_id": user_id}
            ).scalar() or 0

        return jsonify({'file_uploaded': result > 0}), 200

    except Exception as e:
        print(f"Error checking file upload status: {str(e)}")
        return jsonify({'error': 'Server error'}), 500



################################################################################################################
# Revised upload_history with performance trend logic integrated donot delete this comment

@upload_bp.route('/upload_history', methods=['GET'])
def upload_history():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # ✅ read optional params (existing)
    country_param = (request.args.get('country') or "").strip().lower()
    home_currency = (request.args.get('homeCurrency') or "").strip().lower()

    # (Optional) If you want "latest" logic stable
    uploads = (
        UploadHistory.query
        .filter_by(user_id=user_id)
        .order_by(UploadHistory.year.desc(), UploadHistory.month.desc())
        .all()
    )

    month_names = {
        1: 'january', 2: 'february', 3: 'march', 4: 'april',
        5: 'may', 6: 'june', 7: 'july', 8: 'august',
        9: 'september', 10: 'october', 11: 'november', 12: 'december'
    }

    response = []

    for upload in uploads:
        upload_country = (upload.country or "").strip().lower()

        # ✅ IMPORTANT FIX:
        # When requesting GLOBAL history:
        # - if homeCurrency is provided => only return global_<currency>
        # - else => only return base global
        if country_param == "global":
            if upload_country not in ["uk", "us"]:
                continue

        # ✅ Optional: if FE passes specific country, filter by it
        elif country_param:
            if upload_country != country_param:
                continue

        # Convert numeric month to month name for display
        month_name = month_names.get(upload.month, str(upload.month))
        table_name = f"user_{upload_country}_{month_name}{upload.year}_data"

        response.append({
            'month': month_name,
            'month_num': upload.month,
            'year': upload.year,
            'country': upload_country,
            'file_name': table_name,

            'total_sales': upload.total_sales,
            'total_product_sales': upload.total_product_sales,
            'total_profit': upload.total_profit,
            'total_expense': upload.total_expense,
            'total_fba_fees': upload.total_fba_fees,

            'platform_fee': upload.platform_fee,
            'rembursement_fee': upload.rembursement_fee,

            'expense_chart_img': upload.expense_chart_img,
            'sales_chart_img': upload.sales_chart_img,
            'qtd_pie_chart': upload.qtd_pie_chart,
            'ytd_pie_chart': upload.ytd_pie_chart,

            'total_cous': upload.total_cous,
            'total_amazon_fee': upload.total_amazon_fee,
            'profit_chart_img': upload.profit_chart_img,

            'cm2_profit': upload.cm2_profit,
            'cm2_margins': upload.cm2_margins,
            'acos': upload.acos,

            'rembursment_vs_cm2_margins': upload.rembursment_vs_cm2_margins,
            'advertising_total': upload.advertising_total,
            'reimbursement_vs_sales': upload.reimbursement_vs_sales,
            'taxncredit': upload.taxncredit,
            'unit_sold': upload.unit_sold,

            'otherwplatform': upload.platform_fee,
        })

    # ---------------- PERFORMANCE TREND (moved here, down below) ----------------
    # Make sure you have:
    # from utils.performance_trend import get_performance_trend

    period = (request.args.get("period") or "monthly").strip().lower()
    timeline = (request.args.get("timeline") or "ALL").strip().upper()
    year = request.args.get("year", type=int)

    metric = (request.args.get("metric", "net_sales") or "net_sales").strip().lower()
    if metric not in ("net_sales", "units", "asp"):
        metric = "net_sales"

    # Use requested country if provided, else infer from first upload, else default "uk"
    trend_country = country_param if country_param else (response[0]["country"] if response else "uk")
    trend_country = trend_country.strip().lower()

    # ✅ If FE doesn't send year, infer from filtered history (latest year available)
    if year is None:
        year = max((u["year"] for u in response), default=None)

    # ✅ Your original summary logic required monthly timeline 1..12.
    # If FE sends ALL for monthly, normalize it to "12" (last 12 months style)
    if period == "monthly" and timeline == "ALL":
        timeline = "12"

    performance_trend = None
    try:
        if year is not None:
            performance_trend = get_performance_trend(
                user_id=user_id,
                country=trend_country,
                period=period,
                timeline=timeline,
                year=year,
            )
    except Exception as e:
        # Keep upload_history usable even if trend fails
        import traceback
        print("Unexpected error computing performance_trend in /upload_history:", e)
        print(traceback.format_exc())

    return jsonify({
        'uploads': response,
        'performance_trend': performance_trend,
        'performance_trend_metric': metric,
        'trend_country': trend_country,
        'trend_period': period,
        'trend_timeline': timeline,
        'trend_year': year,
    }), 200


####################################################################################################################


def resolve_country(country, currency):
    country = (country or "").lower()
    currency = (currency or "").lower()

    # 1. If country = global
    if country == "global":
        if currency == "usd":
            return "global"
        elif currency == "inr":
            return "global_inr"
        elif currency == "gbp":
            return "global_gbp"
        elif currency == "cad":
            return "global_cad"
        else:
            return "global"  # default fallback

    # 2. If country = uk
    if country == "uk":
        if currency == "usd":
            return "uk_usd"
        else:
            return "uk"  # default for all other currencies

    # 3. Default (no special logic)
    return country


@upload_bp.route('/upload_history2', methods=['GET'])
def upload_history2():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    range_type = request.args.get('range')
    month = request.args.get('month')
    year = request.args.get('year')
    quarter = request.args.get('quarter')

    country_param = (request.args.get('country', '') or '').lower()

    # ✅ only use homeCurrency for GLOBAL
    is_global = country_param == "global"

    if is_global:
        country_list = ["uk", "us"]
    else:
        country_list = [country_param]

    def base_upload_query(y=None, m=None, months=None):
        q = UploadHistory.query.filter(
            UploadHistory.user_id == user_id,
            UploadHistory.country.in_(country_list)
        )

        if y is not None:
            q = q.filter(UploadHistory.year == y)

        if m is not None:
            q = q.filter(UploadHistory.month == m.lower())

        if months is not None:
            q = q.filter(UploadHistory.month.in_(months))

        return q

    def _get_gbp_to_usd_rate(month, year):
        with admin_engine.connect() as admin_conn:
            rate = admin_conn.execute(
                text("""
                    SELECT conversion_rate
                    FROM currency_conversion
                    WHERE LOWER(user_currency) = 'gbp'
                    AND LOWER(selected_currency) = 'usd'
                    AND LOWER(month) = :month
                    AND year = :year
                    ORDER BY id DESC
                    LIMIT 1
                """),
                {
                    "month": str(month).lower().strip(),
                    "year": int(year),
                }
            ).scalar()

        return float(rate or 1)
    
    # Try to infer range_type if not provided
    if not range_type:
        if month and year:
            range_type = 'monthly'
        elif quarter and year:
            range_type = 'quarterly'
        elif year:
            range_type = 'yearly'
        else:
            return jsonify({
                'error': 'Invalid range parameters. Must specify range type or provide appropriate parameters to infer range.'
            }), 400

    try:
        year_num = int(year)
    except (TypeError, ValueError):
        return jsonify({'error': 'Year must be a valid number.'}), 400

    def summarize_uploads(uploads):
        summary = {
            "total_sales": 0,
            "total_product_sales": 0,
            "total_profit": 0,
            "total_expense": 0,
            "advertising_total": 0,
            "cm2_profit": 0,
            "total_amazon_fee": 0,
            "total_cous": 0,
            "otherwplatform": 0,
            "taxncredit": 0,
            "unit_sold": 0,
        }

        for upload in uploads:
            upload_country = (upload.country or "").lower()
            upload_month = (upload.month or "").lower()
            upload_year = upload.year

            multiplier = 1

            if is_global and upload_country == "uk":
                multiplier = _get_gbp_to_usd_rate(upload_month, upload_year)

            summary["total_sales"] += (upload.total_sales or 0) * multiplier
            summary["total_product_sales"] += (upload.total_product_sales or 0) * multiplier
            summary["total_profit"] += (upload.total_profit or 0) * multiplier
            summary["total_expense"] += (upload.total_expense or 0) * multiplier
            summary["advertising_total"] += (upload.advertising_total or 0) * multiplier
            summary["cm2_profit"] += (upload.cm2_profit or 0) * multiplier
            summary["total_amazon_fee"] += (upload.total_amazon_fee or 0) * multiplier
            summary["total_cous"] += (upload.total_cous or 0) * multiplier
            summary["otherwplatform"] += (upload.platform_fee or 0) * multiplier
            summary["taxncredit"] += (upload.taxncredit or 0) * multiplier
            summary["unit_sold"] += upload.unit_sold or 0

        return summary

    # ---------------- comparison helpers ----------------

    month_order = [
        'january', 'february', 'march',
        'april', 'may', 'june',
        'july', 'august', 'september',
        'october', 'november', 'december'
    ]

    quarter_months = {
        'Q1': ['january', 'february', 'march'],
        'Q2': ['april', 'may', 'june'],
        'Q3': ['july', 'august', 'september'],
        'Q4': ['october', 'november', 'december']
    }

    def get_previous_month(m: str, y: int):
        m = (m or '').lower()
        if m not in month_order:
            return None, None
        idx = month_order.index(m)
        if idx == 0:
            return month_order[-1], y - 1
        return month_order[idx - 1], y

    def get_quarter_from_month(m: str):
        m = (m or '').lower()
        for q, months in quarter_months.items():
            if m in months:
                return q
        return None

    def get_previous_quarter(q: str, y: int):
        order = ['Q1', 'Q2', 'Q3', 'Q4']
        if q not in order:
            return None, None
        idx = order.index(q)
        if idx == 0:
            return 'Q4', y - 1
        return order[idx - 1], y

    def fetch_monthly_summary(m: str, y: int):
        if not m or y is None:
            return None
        ups = base_upload_query(y=y, m=m).all()
        return summarize_uploads(ups) if ups else None

    def fetch_quarterly_summary(q: str, y: int):
        if not q or y is None or q not in quarter_months:
            return None
        ups = base_upload_query(y=y, months=quarter_months[q]).all()
        return summarize_uploads(ups) if ups else None

    def fetch_yearly_summary(y: int):
        if y is None:
            return None
        ups = base_upload_query(y=y).all()
        return summarize_uploads(ups) if ups else None

    # ---------------- main logic ----------------

    if range_type == 'monthly' and month and year:
        month_l = month.lower()

        uploads = base_upload_query(y=year_num, m=month_l).all()

        current_summary = summarize_uploads(uploads)
        # last month (prev month)
        prev_m, prev_y = get_previous_month(month_l, year_num)
        last_month_summary = fetch_monthly_summary(prev_m, prev_y) if prev_m else None

        # last quarter (previous quarter from the month you are viewing)
        current_q = get_quarter_from_month(month_l)
        prev_q, prev_q_y = get_previous_quarter(current_q, year_num) if current_q else (None, None)
        last_quarter_summary = fetch_quarterly_summary(prev_q, prev_q_y) if prev_q else None

        # last year (same month previous year)
        last_year_summary = fetch_monthly_summary(month_l, year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': last_month_summary,
                'lastQuarter': last_quarter_summary,
                'lastYear': last_year_summary
            }
        })

    elif range_type == 'quarterly' and quarter and year:
        if quarter not in quarter_months:
            return jsonify({'error': 'Quarter must be one of: Q1, Q2, Q3, Q4'}), 400

        uploads = base_upload_query(y=year_num, months=quarter_months[quarter]).all()

        current_summary = summarize_uploads(uploads)

        # last quarter (previous quarter)
        prev_q, prev_q_y = get_previous_quarter(quarter, year_num)
        last_quarter_summary = fetch_quarterly_summary(prev_q, prev_q_y) if prev_q else None

        # last year (same quarter last year)
        last_year_summary = fetch_quarterly_summary(quarter, year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': None,
                'lastQuarter': last_quarter_summary,
                'lastYear': last_year_summary
            }
        })

    elif range_type == 'yearly' and year:
        uploads = base_upload_query(y=year_num).all()

        current_summary = summarize_uploads(uploads)

        # last year
        last_year_summary = fetch_yearly_summary(year_num - 1)

        return jsonify({
            'uploads': [u.id for u in uploads],
            'summary': current_summary,
            'summaryComparisons': {
                'lastMonth': None,
                'lastQuarter': None,
                'lastYear': last_year_summary
            }
        })

    else:
        return jsonify({'error': 'Invalid range parameters'}), 400



@upload_bp.route('/upload_historyforacos', methods=['GET'])
def upload_historyforacos():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Authorization token is missing or invalid'}), 401

    token = auth_header.split(' ')[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token has expired'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Invalid token'}), 401

    # Fetching query params
    country = request.args.get('country')
    month = request.args.get('month')
    year = request.args.get('year')

    # Month name to number mapping
    month_map = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'may': 5, 'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }

    # Convert month name to number if it's a string
    if month:
        try:
            month_num = int(month)
        except ValueError:
            month_num = month_map.get(month.lower())
            if month_num is None:
                return jsonify({'error': f'Invalid month name: {month}'}), 400
            month = month_num  # Store as integer for filtering

    # Convert year to integer if present
    if year:
        try:
            year = int(year)
        except ValueError:
            return jsonify({'error': 'Invalid year format'}), 400

    # Query with optional filters
    uploads = UploadHistory.query.filter_by(user_id=user_id)
    
    if country:
        uploads = uploads.filter_by(country=country)
    if month:
        uploads = uploads.filter_by(month=month)
    if year:
        uploads = uploads.filter_by(year=year)
    
    uploads = uploads.all()

    # Month number to name mapping
    month_names = {
        1: 'january', 2: 'february', 3: 'march', 4: 'april',
        5: 'may', 6: 'june', 7: 'july', 8: 'august',
        9: 'september', 10: 'october', 11: 'november', 12: 'december'
    }

    response = []
    for upload in uploads:
        # Convert numeric month to month name for display
        month_name = month_names.get(upload.month, str(upload.month))
        
        table_name = f"user_{upload.country}_{month_name}{upload.year}_data"
        response.append({
            'month': month_name,  # Use month name for display
            'month_num': upload.month,  # Keep numeric month for filtering if needed
            'year': upload.year,
            'country': upload.country,
            'file_name': table_name,
            'total_sales': upload.total_sales,
            'total_product_sales': upload.total_product_sales,
            'total_profit': upload.total_profit,
            'total_expense': upload.total_expense,
            'total_fba_fees': upload.total_fba_fees,
            'platform_fee': upload.platform_fee,
            'rembursement_fee': upload.rembursement_fee,
            'expense_chart_img': upload.expense_chart_img,
            'sales_chart_img': upload.sales_chart_img,
            'profit_chart_img': upload.profit_chart_img,
            'cm2_profit': upload.cm2_profit,
            'cm2_margins': upload.cm2_margins,
            'acos': upload.acos,
            'rembursment_vs_cm2_margins': upload.rembursment_vs_cm2_margins,
            'advertising_total': upload.advertising_total,
            'reimbursement_vs_sales': upload.reimbursement_vs_sales,
            'taxncredit': upload.taxncredit,
        })

    return jsonify({'uploads': response})


