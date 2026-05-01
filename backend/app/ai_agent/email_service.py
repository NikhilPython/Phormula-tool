from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, Optional, List
import logging
from flask import current_app
from flask_mail import Message
import csv
import io
from app import mail
from app.models.user_models import User
logger = logging.getLogger(__name__)
from openpyxl import Workbook
from openpyxl.styles import Font

def send_agent_email(
    *,
    user_id: int,
    subject: str,
    html_body: str,
    recipient: Optional[str] = None,
    attachments: Optional[List[Dict[str, Any]]] = None,  # 👈 NEW
) -> Dict[str, Any]:

    user = User.query.filter_by(id=user_id).first()
    to_email = recipient or (user.email if user else None)

    if not to_email:
        raise ValueError("Recipient email not available")

    msg = Message(
        subject=subject,
        sender=current_app.config.get("MAIL_DEFAULT_SENDER"),
        recipients=[to_email],
    )

    msg.html = html_body

    # -------- ✅ NEW: Attachments handling --------
    if attachments:
        for file in attachments:
            msg.attach(
                filename=file["filename"],
                content_type=file["content_type"],
                data=file["data"],
            )

    mail.send(msg)

    return {
        "status": "sent",
        "recipient": to_email,
        "subject": subject,
        "attachments": [f["filename"] for f in attachments or []],  # 👈 optional but useful
    }

def build_email_html(state: Dict[str, Any]) -> str:
    metric = state.get("current_metrics") or {}
    comparison = state.get("comparison") or {}
    analysis = state.get("analysis_result") or {}
    advice = state.get("advice") or []
    event_plan = state.get("event_plan_result") or {}
    sku_intel = state.get("sku_intelligence_result") or {}

    # -------- FORMATTER --------
    def fmt(v):
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return "0.00"

    parts = ["<html><body style='font-family:Arial,sans-serif;color:#1f2937'>"]
    parts.append("<h2>Phormula AI Report</h2>")

    # -------- EVENT PLAN --------
    if event_plan:
        parts.append("<p><strong>Event plan generated</strong></p>")
        summary = event_plan.get("summary") or []
        if summary:
            items = "".join(f"<li>{x}</li>" for x in summary[:12])
            parts.append(f"<ul>{items}</ul>")
        actions = event_plan.get("actions") or []
        if actions:
            items = "".join(f"<li>{x}</li>" for x in actions[:12])
            parts.append(f"<p><strong>Actions</strong></p><ul>{items}</ul>")
        parts.append(f"<p>Generated at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC.</p></body></html>")
        return "".join(parts)

    # -------- SKU INTELLIGENCE --------
    if sku_intel:
        product = sku_intel.get("product_match") or "selected product"
        parts.append(f"<p><strong>SKU intelligence for:</strong> {product}</p>")

        current = sku_intel.get("current", {})
        previous = sku_intel.get("previous", {})

        parts.append("<h3>Current Performance</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Metric</th><th>Value</th></tr>")
        parts.append(f"<tr><td>Sales</td><td>{fmt(current.get('net_sales'))}</td></tr>")
        parts.append(f"<tr><td>Profit</td><td>{fmt(current.get('profit'))}</td></tr>")
        parts.append(f"<tr><td>Units</td><td>{fmt(current.get('total_quantity'))}</td></tr>")
        parts.append(f"<tr><td>ASP</td><td>{fmt(current.get('asp'))}</td></tr>")
        parts.append("</table>")

        if previous:
            parts.append("<h3>Previous Period</h3>")
            parts.append("<table border='1' cellpadding='6'>")
            parts.append("<tr><th>Metric</th><th>Value</th></tr>")
            parts.append(f"<tr><td>Sales</td><td>{fmt(previous.get('net_sales'))}</td></tr>")
            parts.append(f"<tr><td>Profit</td><td>{fmt(previous.get('profit'))}</td></tr>")
            parts.append(f"<tr><td>Units</td><td>{fmt(previous.get('total_quantity'))}</td></tr>")
            parts.append(f"<tr><td>ASP</td><td>{fmt(previous.get('asp'))}</td></tr>")
            parts.append("</table>")

    # -------- HEADER --------
    period_label = metric.get("period_label") or "selected period"
    metric_name = metric.get("metric") or state.get("metric_name") or "metric"
    total = metric.get("total")

    parts.append(f"<p><strong>Metric:</strong> {metric_name.replace('_', ' ').title()}</p>")
    parts.append(f"<p><strong>Period:</strong> {period_label}</p>")

    if total is not None:
        parts.append(f"<p><strong>Total:</strong> {fmt(total)}</p>")

    # -------- COMPARISON --------
    if comparison:
        left = comparison.get("left", {})
        right = comparison.get("right", {})
        pct = comparison.get("pct_change")

        parts.append("<h3>Comparison</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Metric</th><th>Value</th></tr>")
        parts.append(f"<tr><td>Current</td><td>{fmt(left.get('total', 0))}</td></tr>")
        parts.append(f"<tr><td>Previous</td><td>{fmt(right.get('total', 0))}</td></tr>")
        parts.append(f"<tr><td>Change %</td><td>{f'{pct:.2f}%' if pct else 'N/A'}</td></tr>")
        parts.append("</table>")

    # -------- TREND / GROWTH --------
    if analysis.get("type") in ["trend", "growth"]:
        rows = analysis.get("series_display") or analysis.get("series", [])

        parts.append("<h3>Trend</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Period</th><th>Value</th><th>% Change</th></tr>")

        prev_val = None

        for r in rows[:12]:
            curr_val = float(r.get('__metric__') or 0)

            if prev_val is None or prev_val == 0:
                pct_change = "—"
            else:
                change = ((curr_val - prev_val) / prev_val) * 100
                pct_change = f"{change:.2f}%"

            parts.append(
                f"<tr>"
                f"<td>{r.get('period_label')}</td>"
                f"<td>{fmt(curr_val)}</td>"
                f"<td>{pct_change}</td>"
                f"</tr>"
            )

            prev_val = curr_val

        parts.append("</table>")

    # -------- BREAKDOWN --------
    elif analysis.get("type") == "breakdown":
        rows = analysis.get("per_sku", [])[:10]

        parts.append("<h3>Top Products</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Product</th><th>Value</th></tr>")

        for r in rows:
            parts.append(
                f"<tr><td>{r.get('product_name') or r.get('sku') or 'Unknown'}</td>"
                f"<td>{fmt(r.get('__metric__'))}</td></tr>"
            )

        parts.append("</table>")

    # -------- SUMMARY --------
    elif analysis.get("type") == "summary":
        metrics = analysis.get("metrics", {})

        parts.append("<h3>Business Summary</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Metric</th><th>Value</th></tr>")

        for k, v in metrics.items():
            parts.append(f"<tr><td>{k}</td><td>{fmt(v)}</td></tr>")

        parts.append("</table>")

                # -------- 🔥 ADD AI INSIGHTS TO EMAIL --------
        try:
            insights = state.get("insights")

            # fallback if not already computed
            if not insights:
                insights = state.get("insights")

            if insights:
                parts.append("<h3>Insights</h3>")
                parts.append(
                    f"<div style='white-space: pre-line; font-size: 14px;'>"
                    f"{insights}"
                    f"</div>"
                )

        except Exception:
            logger.exception("[EMAIL_INSIGHTS_ERROR]")

    # -------- MULTI-DIM --------
    elif analysis.get("type") == "multi_dimensional":
        rows = analysis.get("data", [])[:30]

        parts.append("<h3>Detailed Breakdown</h3>")
        parts.append("<table border='1' cellpadding='6'>")
        parts.append("<tr><th>Month</th><th>Product</th><th>Metric</th><th>Value</th></tr>")

        for r in rows:
            parts.append(
                f"<tr><td>{r.get('month')}</td>"
                f"<td>{r.get('product')}</td>"
                f"<td>{r.get('metric')}</td>"
                f"<td>{fmt(r.get('value'))}</td></tr>"
            )

        parts.append("</table>")

    # -------- ADVICE --------
    if advice:
        parts.append("<h3>Recommendations</h3>")
        parts.append("<ul>")
        for a in advice:
            parts.append(f"<li>{a}</li>")
        parts.append("</ul>")

    parts.append(f"<p>Generated at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC.</p>")
    parts.append("</body></html>")

    return "".join(parts)


def build_excel_attachment(state):
    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)

    # -------- HEADER --------
    ws.cell(row=1, column=1, value="P&L Productwise Breakdown").font = bold
    ws.cell(row=2, column=1, value="Company Name: Skin Elements")
    ws.cell(row=3, column=1, value=f"Currency: {'£' if state.get('country') == 'uk' else '$'}")
    ws.cell(row=4, column=1, value=f"Country: {state.get('country')}")
    ws.cell(row=5, column=1, value="Platform: Amazon")

    # -------- TABLE HEADER --------
    headers = [
        "S. no", "Product Name", "SKU",
        "Units Sold", "Return", "Net Units Sold",
        "ASP", "Gross Sales", "Sales - Refund",
        "Taxes and Credits", "Net Sales",
        "Promotions", "Promotions %",
        "COGS", "Selling Fees", "FBA Fees",
        "Net Taxes", "Net Credits", "Misc. Transactions",
        "CM1 Profit", "CM1 Profit Per Unit", "CM1 Profit %"
    ]

    start_row = 7

    for col, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=h).font = bold

    # -------- DATA --------
    rows = state.get("analysis_result", {}).get("per_sku", [])

    for i, r in enumerate(rows, start=1):
        row_idx = start_row + i

        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=r.get("product_name"))
        ws.cell(row=row_idx, column=3, value=r.get("sku"))

        ws.cell(row=row_idx, column=4, value=r.get("units_sold"))
        ws.cell(row=row_idx, column=5, value=r.get("returns"))
        ws.cell(row=row_idx, column=6, value=r.get("net_units"))

        ws.cell(row=row_idx, column=7, value=r.get("asp"))
        ws.cell(row=row_idx, column=8, value=r.get("gross_sales"))
        ws.cell(row=row_idx, column=9, value=r.get("refunds"))

        ws.cell(row=row_idx, column=10, value=r.get("taxes"))
        ws.cell(row=row_idx, column=11, value=r.get("net_sales"))

        ws.cell(row=row_idx, column=12, value=r.get("promotions"))
        ws.cell(row=row_idx, column=13, value=r.get("promotions_pct"))

        ws.cell(row=row_idx, column=14, value=r.get("cogs"))
        ws.cell(row=row_idx, column=15, value=r.get("selling_fees"))
        ws.cell(row=row_idx, column=16, value=r.get("fba_fees"))

        ws.cell(row=row_idx, column=17, value=r.get("net_taxes"))
        ws.cell(row=row_idx, column=18, value=r.get("net_credits"))
        ws.cell(row=row_idx, column=19, value=r.get("misc"))

        ws.cell(row=row_idx, column=20, value=r.get("cm1_profit"))
        ws.cell(row=row_idx, column=21, value=r.get("cm1_per_unit"))
        ws.cell(row=row_idx, column=22, value=r.get("cm1_pct"))

    # -------- SAVE --------
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return {
        "filename": "amazon_pnl.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "data": file_stream.read(),
    }

def send_agent_report(state):

    subject = f"Phormula AI {state.get('metric_name')} - {state.get('current_metrics', {}).get('period_label')}"

    html = build_email_html(state)

    attachments = []

    if state.get("include_csv"):
        attachments.append(build_excel_attachment(state))

    return send_agent_email(
        user_id=state["user_id"],
        subject=subject,
        html_body=html,
        attachments=attachments,
    )