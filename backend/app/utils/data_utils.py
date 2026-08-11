from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from config import Config
SECRET_KEY = Config.SECRET_KEY
import os
import pandas as pd
import numpy as np 
from app.models.user_models import User
from flask_mail import Message
from flask import current_app
from app import mail
from datetime import datetime, date
from dotenv import load_dotenv

import warnings
warnings.filterwarnings("ignore") 


load_dotenv()
db_url = os.getenv('DATABASE_URL')


MONTHS_REVERSE_MAP = {
    1: "january", 2: "february", 3: "march", 4: "april", 5: "may", 6: "june",
    7: "july", 8: "august", 9: "september", 10: "october", 11: "november", 12: "december"
}


MONTHS_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12
}


def create_user_session(db_url):
    user_engine = create_engine(db_url)
    UserSession = sessionmaker(bind=user_engine)
    return UserSession()



from app.models.user_models import StoredFile  

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _inventory_sheet_only_bytes(file_bytes):
    try:
        from io import BytesIO
        from openpyxl import load_workbook

        raw_bytes = file_bytes.tobytes() if isinstance(file_bytes, memoryview) else bytes(file_bytes)
        workbook = load_workbook(BytesIO(raw_bytes))
        inventory_sheet = next(
            (
                sheet_name
                for sheet_name in workbook.sheetnames
                if sheet_name.strip().lower() == "inventory"
            ),
            None,
        )
        if not inventory_sheet:
            return raw_bytes

        for sheet_name in list(workbook.sheetnames):
            if sheet_name != inventory_sheet:
                del workbook[sheet_name]

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as exc:
        print(f"[EMAIL][WARN] Could not build inventory-only attachment: {exc}")
        return file_bytes.tobytes() if isinstance(file_bytes, memoryview) else file_bytes

def _guess_kind_from_filename(file_name: str) -> str:
    name = (file_name or "").lower()
    if name.startswith("inventory_forecast_"):
        return "inventory_forecast"
    if name.startswith("forecasts_for_"):
        return "forecasts_for"
    if name.startswith("forecastpnl_"):
        return "pnl_forecast"
    if name.startswith("pnlforecast_"):
        return "pnl_forecast_upload"
    return "file"

def send_forecast_email(user_id, file_name, month, year, *, country=None):
    try:
        user = User.query.filter_by(id=user_id).first()
        if not user:
            raise ValueError(f"No user found with ID {user_id}")

        user_email = user.email
        user_name = (user.name or "there").strip()

        msg = Message(
            "Your Forecast Report",
            sender=current_app.config.get("MAIL_DEFAULT_SENDER"),
            recipients=[user_email],
        )

        month_title = str(month).capitalize()
        year_text = str(year)

        msg.html = f"""
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">

  <style>
    @media only screen and (max-width: 600px) {{
      .email-container {{
        width: 100% !important;
        max-width: 100% !important;
      }}

      .top-report-title {{
        font-size: 14px !important;
        line-height: 18px !important;
      }}

      .content-cell {{
        padding: 22px 24px 26px 24px !important;
      }}

      .cta-wrap {{
        text-align: center !important;
      }}

      .cta-button {{
        display: inline-block !important;
        margin: 0 auto !important;
        text-align: center !important;
      }}
    }}
  </style>
</head>

<body style="margin:0; padding:0; font-family:Arial, Helvetica, sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="padding:16px 0;">
    <tr>
      <td align="center">

        <table class="email-container" width="600" cellpadding="0" cellspacing="0" border="0" style="
          background:#ffffff;
          width:600px;
          max-width:600px;
          border-collapse:collapse;
        ">

          <!-- top green bar -->
          <tr>
            <td style="background:#5ea68e; padding:18px 24px; color:#ffffff;">
              <table width="100%" cellpadding="0" cellspacing="0" border="0" style="table-layout:fixed; border-collapse:collapse;">
                <tr>
                  <td width="110" style="text-align:left; vertical-align:middle; white-space:nowrap;">
                    <img
                      src="https://res.cloudinary.com/du58s6gdz/image/upload/f_auto,q_auto/output-onlinepngtools_ypplvv"
                      alt="Phormula"
                      width="40"
                      style="display:block; width:40px; max-width:40px; height:auto; border:0;"
                    />
                  </td>

                  <td width="382" align="right" class="top-report-title" style="
                    font-size:16px;
                    line-height:18px;
                    color:#f8edce;
                    text-align:right;
                    vertical-align:middle;
                    white-space:nowrap;
                  ">
                    Inventory Report
                  </td>
                </tr>
              </table>
            </td>
          </tr>

          <!-- logo/title -->
          <tr>
            <td align="center" style="
              padding:28px 30px 18px 30px;
              background:#ffffff;
              border-left:1px solid #e4e7ec;
              border-right:1px solid #e4e7ec;
            ">
              <img
                src="https://res.cloudinary.com/du58s6gdz/image/upload/f_auto,q_auto/Logo_Phormula_pmbp8q"
                alt="Phormula Logo"
                width="220"
                style="display:block; width:220px; max-width:220px; height:auto; margin:0 auto 14px auto; border:0;"
              />

              <div style="font-size:18px; color:#4a4a4a; line-height:1.4;">
                Your {month_title} {year_text} Forecast Report is ready
              </div>
            </td>
          </tr>

          <!-- divider -->
          <tr>
            <td style="border-top:1px solid #dddddd; font-size:1px; line-height:1px;">&nbsp;</td>
          </tr>

          <!-- body -->
          <tr>
            <td class="content-cell" style="
              padding:22px 32px 26px 32px;
              color:#444444;
              font-size:14px;
              line-height:1.7;
              text-align:justify;
              text-justify:inter-word;
              border-left:1px solid #e4e7ec;
              border-right:1px solid #e4e7ec;
            ">
              <p style="margin:0 0 18px 0; text-align:left;">
                Hey {user_name},
              </p>

              <p style="margin:0 0 14px 0; text-align:justify; text-justify:inter-word;">
                Please find attached your inventory forecast report for
                {month_title} {year_text}. It covers your current stock positions
                and demand projections for the period.
              </p>

              <p style="margin:0 0 14px 0; text-align:justify; text-justify:inter-word;">
                Kindly review at your earliest convenience and flag anything that needs attention.
              </p>

              <p style="margin:0 0 14px 0; text-align:justify; text-justify:inter-word;">
                Thank you for your continued partnership. We look forward to hearing from you.
              </p>

              <p style="margin:18px 0 0 0; text-align:left;">
                Warm regards,
              </p>

              <p style="margin:0; text-align:left;">
                <strong>The Phormula Team</strong>
              </p>

              <p style="margin:0; text-align:left;">
                <a href="mailto:care@phormula.io" style="color:#37455f; text-decoration:none;">
                  care@phormula.io
                </a>
              </p>
            </td>
          </tr>

          <!-- full-width note section -->
          <tr>
            <td style="
              border-top:1px solid #dddddd;
              padding:14px 32px 16px 32px;
              background:#ffffff;
              font-size:12px;
              color:#999999;
              line-height:1.6;
              text-align:left;
              border-left:1px solid #e4e7ec;
              border-right:1px solid #e4e7ec;
            ">
              This email was generated automatically by Phormula.
            </td>
          </tr>

          <!-- footer -->
          <tr>
            <td align="center" style="
              background:#5ea68e;
              padding:12px 18px;
              color:#f8edce;
              font-size:12px;
              line-height:1.5;
              text-align:center;
            ">
              © 2026 Phormula. All rights reserved.
            </td>
          </tr>

        </table>
      </td>
    </tr>
  </table>
</body>
</html>
"""

        q = StoredFile.query.filter_by(user_id=user_id, filename=file_name)
        if country:
            q = q.filter_by(country=country.lower())

        stored = q.first()
        if not stored:
            raise FileNotFoundError(
                f"File not found in DB for user_id={user_id}, filename={file_name}"
            )

        content_type = stored.content_type or XLSX_MIME
        attachment_data = stored.data
        if str(file_name or "").lower().endswith(".xlsx"):
            attachment_data = _inventory_sheet_only_bytes(attachment_data)
        msg.attach(file_name, content_type, attachment_data)

        mail.send(msg)
        print(f"✅ Forecast email sent to {user_email} (attached from DB: {file_name})")

    except Exception as e:
        print(f"Failed to send forecast email: {e}")
        raise
    
def generate_pnl_report(year: int, month: str) -> dict:
    """
    Generate a simple P&L (profit and Loss) report for a given month and year.

    Parameters:
        year (int): The year for the report.
        month (str): The month for the report.

    Returns:
        dict: A dictionary containing sales, expenses, FBA fees, and calculated profit.
    """
    
    # Placeholder data - replace this with actual database queries in a real scenario
    sales_data = {
        'sales': 10000,
        'expenses': 5000,
        'fba_fees': 1000,
    }
    
    # Calculate profit
    profit = sales_data['sales'] - sales_data['expenses'] - sales_data['fba_fees']
    
    return {
        'year': year,
        'month': month,
        'sales': sales_data['sales'],
        'expenses': sales_data['expenses'],
        'fba_fees': sales_data['fba_fees'],
        'profit': profit,
    }



def send_pnlforecast_email(user_id, file_name, month, year, *, country=None):
    try:
        user = User.query.filter_by(id=user_id).first()
        if not user:
            raise ValueError(f"No user found with ID {user_id}")

        user_email = user.email

        msg = Message(
            'Your PnL Forecast Report',
            sender=current_app.config.get('MAIL_DEFAULT_SENDER'),
            recipients=[user_email]
        )

        msg.body = f"""
Dear {user.email},

Please find attached the PNL forecast report for next 3 months of {month} {year} that you requested.

Best regards,
The Phormula Team
"""

        # ✅ Load bytes from DB
        q = StoredFile.query.filter_by(user_id=user_id, filename=file_name)
        if country:
            q = q.filter_by(country=country.lower())
        stored = q.first()

        if not stored:
            raise FileNotFoundError(f"PnL file not found in DB for user_id={user_id}, filename={file_name}")

        content_type = stored.content_type or XLSX_MIME
        msg.attach(file_name, content_type, stored.data)

        mail.send(msg)
        print(f"✅ PnL email sent to {user_email} (attached from DB: {file_name})")

    except Exception as e:
        print(f"Failed to send pnl forecast email: {e}")
        raise



def get_previous_month_year(month, year):
    """Calculate the previous month and year."""

    year = int(year)
    prev_month_num = MONTHS_MAP[month] - 1
    if prev_month_num == 0:
        prev_month_num = 12
        year -= 1
    prev_month = MONTHS_REVERSE_MAP[prev_month_num]

    return prev_month, year


