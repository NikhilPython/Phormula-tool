import io
from datetime import datetime , date
import calendar, json, hashlib
import re
from sqlalchemy import func, text
from sqlalchemy.dialects.postgresql import insert
import jwt
import pandas as pd
from flask import Blueprint, jsonify, request, send_file, Response
from app import db
from config import Config
from app.utils.token_utils import get_effective_user_id_from_token
from app.models.user_models import amazon_user, amazon_sponsored_products , amazon_sponsored_display_advertised_products
from app.utils.amazon_ads_utils_reporting import (
    build_ads_lwa_auth_url,
    exchange_code_for_tokens,
    get_ads_access_token_from_refresh,
    list_top_level_profiles_all_regions,
    find_manager_profile_id,
    list_child_profiles_all_regions,
    pick_profile_id,
    tokeninfo,
    fetch_report_rows_for_profiles,
    flatten_ads_profiles,
    normalize_ads_country_filter,
)
from app.models.user_models import amazon_sponsored_brands_keywords
from openpyxl.utils import get_column_letter
from app.ads_report_services import (
    run_sp_advertised_product_report_service
)

from app.utils.ads_helpers import (
    _get_user_row,
    _to_date,
    _to_float,
    _to_int,
    _safe_div,
    _norm_key,
    _pick,
)

SECRET_KEY = Config.SECRET_KEY
advertisement_api_routes_bp = Blueprint("advertisement_api_routes", __name__)


def _require_jwt_user_id() -> int:
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise PermissionError("Authorization token is missing or invalid")

    token = auth_header.split(" ")[1]
    payload, user_id, member_id = get_effective_user_id_from_token(token)
    return int(payload["user_id"])




@advertisement_api_routes_bp.route("/api/ads/connect_url", methods=["GET"])
def ads_connect_url():
    """
    Returns URL to start Ads LWA OAuth.
    """
    try:
        user_id = _require_jwt_user_id()
        state = f"user:{user_id}"
        url = build_ads_lwa_auth_url(state=state)
        return jsonify({"url": url})
    except Exception as e:
        return jsonify({"error": str(e)}), 400





@advertisement_api_routes_bp.route("/api/ads/callback", methods=["GET"])
def ads_callback():
    """
    - exchange code -> refresh_token
    - store refresh token
    - refresh -> access_token
    - list top profiles (no scope) -> find manager profile id (if any)
    - if manager exists: list child advertiser profiles via scope
      else: child profiles == top profiles
    - save best-effort UK/US/CA advertiser profile IDs
    """
    try:
        code = request.args.get("code")
        state = request.args.get("state") or ""

        if not code:
            return jsonify({"error": "Missing code"}), 400
        if not state.startswith("user:"):
            return jsonify({"error": "Invalid state"}), 400

        user_id = int(state.split("user:")[1])

        tokens = exchange_code_for_tokens(code)
        refresh_token = tokens.get("refresh_token")

        if not refresh_token:
            payload = {
                "error": "Missing refresh_token from Amazon. Reconnect with prompt=consent.",
                "token_response": tokens
            }
            if request.args.get("format") == "json":
                return jsonify(payload), 400

            html = f"""
            <!doctype html>
            <html>
              <head><meta charset="utf-8" /></head>
              <body>
                <script>
                  (function () {{
                    try {{
                      if (window.opener && !window.opener.closed) {{
                        window.opener.postMessage(
                          {{ type: "amazon_ads_connected", ok: false, error: {json.dumps(payload["error"])} }},
                          "*"
                        );
                      }}
                    }} catch (e) {{}}
                    try {{ window.close(); }} catch (e) {{}}
                    document.body.innerHTML = "Connection failed. You can close this window.";
                  }})();
                </script>
              </body>
            </html>
            """
            return Response(html, mimetype="text/html"), 400

        u = _get_user_row(user_id)
        u.amazon_ads_refresh_token = refresh_token
        u.amazon_ads_refresh_token_updated_at = datetime.utcnow()

        access_token = get_ads_access_token_from_refresh(refresh_token)

        # 1) Top-level profiles across regions (no scope)
        top_profiles_by_region = list_top_level_profiles_all_regions(access_token)
        manager_profile_id = find_manager_profile_id(top_profiles_by_region)
        u.amazon_ads_manager_profile_id = manager_profile_id

        # 2) Child profiles (advertisers) if manager; otherwise top-level is your advertisers
        if manager_profile_id:
            child_profiles_by_region = list_child_profiles_all_regions(access_token, manager_profile_id)
        else:
            child_profiles_by_region = top_profiles_by_region

        # Pick a single advertiser profile id and store it in one column
        eu_child = child_profiles_by_region.get("EU", []) or []
        na_child = child_profiles_by_region.get("NA", []) or []

        selected_profile_id = (
            pick_profile_id(eu_child, {"GB", "UK"}) or
            pick_profile_id(na_child, {"US"}) or
            pick_profile_id(na_child, {"CA"})
        )

        u.amazon_ads_profile_id = str(selected_profile_id) if selected_profile_id else None

        db.session.commit()

        payload = {
            "message": "Amazon Ads connected successfully",
            "saved": {
                "amazon_ads_refresh_token_updated_at": (
                    u.amazon_ads_refresh_token_updated_at.isoformat()
                    if u.amazon_ads_refresh_token_updated_at else None
                ),
                "amazon_ads_manager_profile_id": u.amazon_ads_manager_profile_id,
                "amazon_ads_profile_id": u.amazon_ads_profile_id,
            },
            "counts": {
                "top_level": {k: len(v or []) for k, v in top_profiles_by_region.items()},
                "child": {k: len(v or []) for k, v in child_profiles_by_region.items()},
            },
        }

        # Debug mode if you want JSON:
        # /api/ads/callback?code=...&state=...&format=json
        if request.args.get("format") == "json":
            return jsonify(payload)

        # Normal popup flow: notify opener + close popup
        html = f"""
        <!doctype html>
        <html>
          <head><meta charset="utf-8" /></head>
          <body>
            <script>
              (function () {{
                try {{
                  if (window.opener && !window.opener.closed) {{
                    window.opener.postMessage(
                      {{ type: "amazon_ads_connected", ok: true, payload: {json.dumps(payload)} }},
                      "*"
                    );
                  }}
                }} catch (e) {{}}

                try {{ window.close(); }} catch (e) {{}}

                document.body.innerHTML = "Connected. You can close this window.";
              }})();
            </script>
          </body>
        </html>
        """
        return Response(html, mimetype="text/html")

    except Exception as e:
        err_payload = {"error": str(e)}

        if request.args.get("format") == "json":
            return jsonify(err_payload), 400

        html = f"""
        <!doctype html>
        <html>
          <head><meta charset="utf-8" /></head>
          <body>
            <script>
              (function () {{
                try {{
                  if (window.opener && !window.opener.closed) {{
                    window.opener.postMessage(
                      {{ type: "amazon_ads_connected", ok: false, error: {json.dumps(str(e))} }},
                      "*"
                    );
                  }}
                }} catch (e) {{}}

                try {{ window.close(); }} catch (e) {{}}

                document.body.innerHTML = "Connection failed. You can close this window.";
              }})();
            </script>
          </body>
        </html>
        """
        return Response(html, mimetype="text/html"), 400



@advertisement_api_routes_bp.route("/api/ads/status", methods=["GET"])
def ads_status():
    # -------- auth (same as amazon_status) --------
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({"success": False, "error": "Authorization token is missing or invalid"}), 401

    token = auth_header.split(" ")[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = payload["user_id"]
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401

    # -------- load from DB for this user --------
    # If your ads tokens are stored on the main User row, this is correct.
    # If you store ads tokens in a separate model, swap this query accordingly.
    u = _get_user_row(user_id)  # or: User.query.filter_by(id=user_id).first()

    # No row at all (very unlikely if users always exist)
    if not u:
        return jsonify({
            "success": False,
            "status": "no_record",
            "has_refresh_token": False,
        }), 200

    # Refresh token missing => OAuth not completed
    if not getattr(u, "amazon_ads_refresh_token", None):
        return jsonify({
            "success": False,
            "status": "pending",
            "has_refresh_token": False,
            "saved": {
                "amazon_ads_refresh_token_updated_at": None,
                "amazon_ads_manager_profile_id": getattr(u, "amazon_ads_manager_profile_id", None),
                "amazon_ads_profile_id": getattr(u, "amazon_ads_profile_id", None),
            }
        }), 200

    # We DO have a refresh token in DB -> optionally validate by getting access token + calling profiles endpoint
    refresh_token = u.amazon_ads_refresh_token

    try:
        access_token = get_ads_access_token_from_refresh(refresh_token)

        # light validation call: if this succeeds, you're basically connected
        # (you can also skip this call and just return connected if you only care about token existence)
        profiles_by_region = list_top_level_profiles_all_regions(access_token)

        return jsonify({
            "success": True,
            "status": "connected",
            "has_refresh_token": True,
            "saved": {
                "amazon_ads_refresh_token_updated_at": u.amazon_ads_refresh_token_updated_at.isoformat()
                if getattr(u, "amazon_ads_refresh_token_updated_at", None) else None,

                "amazon_ads_manager_profile_id": getattr(u, "amazon_ads_manager_profile_id", None),

                "amazon_ads_profile_id": getattr(u, "amazon_ads_profile_id", None),
            },
            "counts": {
                "top_level": {k: len(v or []) for k, v in (profiles_by_region or {}).items()}
            }
        }), 200

    except Exception as e:
        # token exists but API call failed
        return jsonify({
            "success": False,
            "status": "ads_api_error",
            "has_refresh_token": True,
            "saved": {
                "amazon_ads_refresh_token_updated_at": u.amazon_ads_refresh_token_updated_at.isoformat()
                if getattr(u, "amazon_ads_refresh_token_updated_at", None) else None,

                "amazon_ads_manager_profile_id": getattr(u, "amazon_ads_manager_profile_id", None),

                "amazon_ads_profile_id": getattr(u, "amazon_ads_profile_id", None),
            },
            "error": str(e)
        }), 502




# --- helpers (keep once in your file; don't duplicate) ---
def _safe_ident(name: str) -> str:
    """Allow only safe Postgres identifier chars (letters/numbers/_)."""
    name = str(name or "")
    return re.sub(r"[^a-zA-Z0-9_]+", "_", name)


def _public_table_exists(table_name: str) -> bool:
    """Return True only when the requested table already exists in public schema."""
    safe_table_name = _safe_ident(table_name).lower()
    return bool(db.session.execute(text("""
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema = 'public'
              AND table_name = :table_name
        )
    """), {"table_name": safe_table_name}).scalar())


def _missing_ads_table_response(table_name: str, month: int, year: int, country: str, table_type: str):
    month_name = calendar.month_name[month]
    return jsonify({
        "success": False,
        "error": f"{month_name} {year} {table_type} ads table is not present.",
        "message": f"Requested {table_type} ads table public.{table_name.lower()} does not exist.",
        "table_name": f"public.{table_name.lower()}",
        "country": country,
        "month": month,
        "month_name": month_name,
        "year": year,
    }), 404


def _latest_end_date_for_month(model, user_id, country, first_day, last_day):
    """
    Picks latest end_date for rows whose start_date falls within [first_day, last_day].
    This prevents double-counting when you have multiple runs like:
      2026-02-01 -> 2026-02-11
      2026-02-01 -> 2026-02-12   (latest)
    """
    return (
        db.session.query(func.max(model.end_date))
        .filter(
            model.user_id == user_id,
            model.country == country,
            model.start_date >= first_day,
            model.start_date <= last_day,
        )
        .scalar()
    )

# ------------------------------------------ Sponsored Products Advertised Product Report Route ------------------------------------------

@advertisement_api_routes_bp.route("/api/ads/manager/sp_advertised_product_report", methods=["POST"])
def manager_sp_advertised_product_report():
    """
    Creates SP advertised product report across all accessible advertiser profiles,
    merges into one Excel AND saves rows to DB (idempotent UPSERT).

    Body:
      {
        "start_date": "YYYY-MM-DD",
        "end_date": "YYYY-MM-DD",
        "time_unit": "SUMMARY" | "DAILY",
        "countries": ["UK","US"],        # optional
        "return_excel": true            # optional (default true)
      }
    """
    try:
        user_id = _require_jwt_user_id()
        u = _get_user_row(user_id)

        data = request.get_json(force=True) or {}
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        time_unit = (data.get("time_unit") or "SUMMARY").upper()
        return_excel = bool(data.get("return_excel", True))

        wanted_countries = normalize_ads_country_filter(data.get("countries"))

        if time_unit not in {"DAILY", "SUMMARY"}:
            return jsonify({"error": "time_unit must be DAILY or SUMMARY"}), 400
        if not start_date or not end_date:
            return jsonify({"error": "start_date and end_date required (YYYY-MM-DD)"}), 400

        # ✅ HARD VALIDATE REQUEST DATES ONCE
        try:
            sd = _to_date(start_date, strict=True, field_name="start_date")
            ed = _to_date(end_date, strict=True, field_name="end_date")
        except ValueError as ve:
            return jsonify({"error": str(ve)}), 400

        if not u.amazon_ads_refresh_token:
            return jsonify({"error": "Amazon Ads not connected for this user."}), 400

        access_token = get_ads_access_token_from_refresh(u.amazon_ads_refresh_token)

        top_profiles = list_top_level_profiles_all_regions(access_token)
        manager_profile_id = find_manager_profile_id(top_profiles)

        if manager_profile_id:
            child_by_region = list_child_profiles_all_regions(access_token, manager_profile_id)
        else:
            child_by_region = top_profiles

        all_profiles = flatten_ads_profiles(child_by_region, wanted_countries)

        if not all_profiles:
            return jsonify({
                "error": "No advertiser profiles found (or your country filter removed all). "
                         "This usually means the Amazon login that consented has no API-linked Ads profiles."
            }), 400

        max_workers = int(data.get("max_workers") or 4)
        merged_rows, download_errors = fetch_report_rows_for_profiles(
            access_token=access_token,
            profiles=all_profiles,
            create_method_name="create_sp_advertised_product_report",
            start_date=start_date,
            end_date=end_date,
            time_unit=time_unit,
            max_wait_seconds=int(data.get("max_wait_seconds") or 1800),
            poll_every_seconds=int(data.get("poll_every_seconds") or 10),
            max_workers=max_workers,
            strict_row_dicts=True,
        )

        if not merged_rows:
            return jsonify({
                "error": "Reports returned no rows",
                "download_errors": download_errors[:50],
            }), 400

        df = pd.DataFrame(merged_rows)

        # numeric safety (v3 columns)
        numeric_cols = [
            "impressions", "clicks", "cost", "costPerClick", "clickThroughRate",
            "sales7d", "purchases7d", "unitsSoldClicks7d",
            "acosClicks7d", "roasClicks7d",
            "attributedSalesSameSku7d", "salesOtherSku7d",
            "purchasesSameSku7d", "unitsSoldSameSku7d", "unitsSoldOtherSku7d",
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        def sdiv(a, b):
            # convert scalars to series
            a = a if isinstance(a, pd.Series) else pd.Series([a] * len(df))
            b = b if isinstance(b, pd.Series) else pd.Series([b] * len(df))
            b = b.replace({0: pd.NA})
            return (a / b).fillna(0.0)

        # =========================================================
        # ✅ BUILD OUTPUT (DATES ALWAYS FROM REQUEST)
        # =========================================================
        out = pd.DataFrame()

        if time_unit == "DAILY":
            out["Start Date"] = df.get("date", df.get("startDate", start_date))
            out["End Date"] = df.get("date", df.get("endDate", df.get("startDate", end_date)))
        else:
            out["Start Date"] = df.get("startDate", start_date)
            out["End Date"] = df.get("endDate", df.get("startDate", end_date))

        out["Country"] = df.get("_country", "")
        out["Profile ID"] = df.get("_profileId", "")

        out["Portfolio ID"] = df.get("portfolioId", "")
        out["Currency"] = df.get("campaignBudgetCurrencyCode", "")

        out["Campaign ID"] = df.get("campaignId", "")
        out["Campaign Name"] = df.get("campaignName", "")

        out["Ad Group ID"] = df.get("adGroupId", "")
        out["Ad Group Name"] = df.get("adGroupName", "")

        out["Advertised SKU"] = df.get("advertisedSku", "")
        out["Advertised ASIN"] = df.get("advertisedAsin", "")

        out["Impressions"] = df.get("impressions", 0)
        out["Clicks"] = df.get("clicks", 0)

        if "clickThroughRate" in df.columns:
            out["Click-Thru Rate (CTR)"] = df.get("clickThroughRate", 0.0)
        else:
            out["Click-Thru Rate (CTR)"] = sdiv(df.get("clicks", 0.0), df.get("impressions", 0.0))

        if "costPerClick" in df.columns:
            out["Cost Per Click (CPC)"] = df.get("costPerClick", 0.0)
        else:
            out["Cost Per Click (CPC)"] = sdiv(df.get("cost", 0.0), df.get("clicks", 0.0))

        out["Spend"] = df.get("cost", 0.0)

        out["7 Day Total Sales"] = df.get("sales7d", 0.0)
        out["7 Day Total Orders (#)"] = df.get("purchases7d", 0.0)
        out["7 Day Total Units (#)"] = df.get("unitsSoldClicks7d", 0.0)

        out["Total Advertising Cost of Sales (ACOS)"] = df.get("acosClicks7d", 0.0)
        out["Total Return on Advertising Spend (ROAS)"] = df.get("roasClicks7d", 0.0)

        out["7 Day Conversion Rate"] = sdiv(df.get("purchases7d", 0.0), df.get("clicks", 0.0))

        out["7 Day Advertised SKU Sales"] = df.get("attributedSalesSameSku7d", 0.0)
        out["7 Day Other SKU Sales"] = df.get("salesOtherSku7d", 0.0)

        out["7 Day Advertised SKU Orders (#)"] = df.get("purchasesSameSku7d", 0.0)
        out["7 Day Advertised SKU Units (#)"] = df.get("unitsSoldSameSku7d", 0.0)
        out["7 Day Other SKU Units (#)"] = df.get("unitsSoldOtherSku7d", 0.0)

        # =========================================================
        # ✅ DEDUPE IN-BATCH ON UNIQUE KEY
        # =========================================================
        key_cols = [
            "Start Date", "End Date", "Country", "Profile ID",
            "Campaign ID", "Ad Group ID", "Advertised SKU", "Advertised ASIN"
        ]

        for c in ["Country", "Profile ID", "Campaign ID", "Ad Group ID", "Advertised SKU", "Advertised ASIN"]:
            if c in out.columns:
                out[c] = out[c].fillna("").astype(str)

        out = out.drop_duplicates(subset=key_cols, keep="last").reset_index(drop=True)

        # =========================================================
        # ✅ SAVE TO DATABASE (UPSERT) — PERMANENT NOT-NULL FIX
        # =========================================================
        now = datetime.utcnow()

        rows_to_insert = []
        for rec in out.to_dict(orient="records"):
            rows_to_insert.append({
                "user_id": user_id,
                "created_at": now,
                "updated_at": now,
                "time_unit": time_unit,

                # ✅ DAILY fix:
                # Save each Amazon row's own date instead of the full request range.
                # For SUMMARY reports, this will still safely fall back to request dates.
                "start_date": _to_date(
                    rec.get("Start Date") or start_date,
                    strict=True,
                    field_name="Start Date"
                ),
                "end_date": _to_date(
                    rec.get("End Date") or rec.get("Start Date") or end_date,
                    strict=True,
                    field_name="End Date"
                ),

                "country": (rec.get("Country") or None),
                "profile_id": str(rec.get("Profile ID") or "") or None,
                "portfolio_id": str(rec.get("Portfolio ID") or "") or None,
                "currency": rec.get("Currency") or None,

                "campaign_id": str(rec.get("Campaign ID") or "") or None,
                "campaign_name": rec.get("Campaign Name") or None,

                "ad_group_id": str(rec.get("Ad Group ID") or "") or None,
                "ad_group_name": rec.get("Ad Group Name") or None,

                "advertised_sku": rec.get("Advertised SKU") or None,
                "advertised_asin": rec.get("Advertised ASIN") or None,

                "impressions": _to_int(rec.get("Impressions")),
                "clicks": _to_int(rec.get("Clicks")),

                "ctr": _to_float(rec.get("Click-Thru Rate (CTR)")),
                "cpc": _to_float(rec.get("Cost Per Click (CPC)")),
                "spend": _to_float(rec.get("Spend")),

                "sales_7d": _to_float(rec.get("7 Day Total Sales")),
                "orders_7d": _to_float(rec.get("7 Day Total Orders (#)")),
                "units_7d": _to_float(rec.get("7 Day Total Units (#)")),

                "acos": _to_float(rec.get("Total Advertising Cost of Sales (ACOS)")),
                "roas": _to_float(rec.get("Total Return on Advertising Spend (ROAS)")),
                "conv_rate_7d": _to_float(rec.get("7 Day Conversion Rate")),

                "adv_sku_sales_7d": _to_float(rec.get("7 Day Advertised SKU Sales")),
                "other_sku_sales_7d": _to_float(rec.get("7 Day Other SKU Sales")),

                "adv_sku_orders_7d": _to_float(rec.get("7 Day Advertised SKU Orders (#)")),
                "adv_sku_units_7d": _to_float(rec.get("7 Day Advertised SKU Units (#)")),
                "other_sku_units_7d": _to_float(rec.get("7 Day Other SKU Units (#)")),
            })

        # ✅ extra safety: never let null dates reach SQL
        if any(r["start_date"] is None or r["end_date"] is None for r in rows_to_insert):
            return jsonify({"error": "Sanity check failed: start_date/end_date became null before insert"}), 500

        if rows_to_insert:
            table = amazon_sponsored_products.__table__
            stmt = insert(table).values(rows_to_insert)

            conflict_cols = [
                "user_id",
                "start_date",
                "end_date",
                "country",
                "profile_id",
                "campaign_id",
                "ad_group_id",
                "advertised_sku",
                "advertised_asin",
            ]

            update_cols = {
                "updated_at": stmt.excluded.updated_at,
                "time_unit": stmt.excluded.time_unit,
                "portfolio_id": stmt.excluded.portfolio_id,
                "currency": stmt.excluded.currency,

                "campaign_name": stmt.excluded.campaign_name,
                "ad_group_name": stmt.excluded.ad_group_name,

                "impressions": stmt.excluded.impressions,
                "clicks": stmt.excluded.clicks,
                "ctr": stmt.excluded.ctr,
                "cpc": stmt.excluded.cpc,
                "spend": stmt.excluded.spend,

                "sales_7d": stmt.excluded.sales_7d,
                "orders_7d": stmt.excluded.orders_7d,
                "units_7d": stmt.excluded.units_7d,

                "acos": stmt.excluded.acos,
                "roas": stmt.excluded.roas,
                "conv_rate_7d": stmt.excluded.conv_rate_7d,

                "adv_sku_sales_7d": stmt.excluded.adv_sku_sales_7d,
                "other_sku_sales_7d": stmt.excluded.other_sku_sales_7d,
                "adv_sku_orders_7d": stmt.excluded.adv_sku_orders_7d,
                "adv_sku_units_7d": stmt.excluded.adv_sku_units_7d,
                "other_sku_units_7d": stmt.excluded.other_sku_units_7d,
            }

            stmt = stmt.on_conflict_do_update(
                index_elements=conflict_cols,
                set_=update_cols
            )

            db.session.execute(stmt)
            db.session.commit()

        # =========================================================
        # ✅ API JSON RESPONSE (no excel)
        # =========================================================
        if not return_excel:
            return jsonify({
                "message": "Saved Sponsored Products report rows (UPSERT)",
                "rows_saved": len(rows_to_insert),
                "start_date": start_date,
                "end_date": end_date,
                "time_unit": time_unit,
                "countries": list(wanted_countries) if wanted_countries else None,
            }), 200

        # =========================================================
        # ✅ RETURN EXCEL
        # =========================================================
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="AdvertisedProduct")
        output.seek(0)

        filename = f"SP_Advertised_Product_{start_date}_to_{end_date}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401
    except PermissionError as e:
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ------------------------------------------- Combine Sponsored product and display routes ------------------------------------------------
@advertisement_api_routes_bp.route("/api/ads/monthly_sp_sd_to_db", methods=["POST"])
def monthly_sp_sd_to_db():
    """
    Save monthly aggregated Sponsored Products + Sponsored Display + Sponsored Brands (keywords) into:
      public.adsmonthly_{user_id}_{country}_{month}_{year}

    ✅ Adds 3 columns:
      - product_spend (SP spend)
      - display_spend (SD spend)
      - brand_spend   (SB spend)

    IMPORTANT:
    - Does NOT delete previous raw report rows in amazon_sponsored_* tables.
    - To avoid double-counting, it aggregates ONLY the latest run for that month
      (latest end_date for the month/user/country) *when end_date exists*.
    - SB keywords table often cannot map to SKU/ASIN. In that case, SB spend is still included
      as a grouped row with empty sku/asin and included in Grand Total.
    """
    try:
        user_id = _require_jwt_user_id()
        payload = request.get_json(force=True) or {}

        # ---- inputs ----
        month = int(payload.get("month") or 0)
        year = int(payload.get("year") or 0)
        country = str(payload.get("country") or "").upper().strip()

        country_aliases = [country]
        if country in ("UK", "GB"):
            country_aliases = ["UK", "GB"]

        include = payload.get("include") or ["SP", "SD", "SB"]
        include = {str(x).upper().strip() for x in include}

        # Force SB for UK/US monthly ads, even if frontend sends only SP + SD
        if country in ("UK", "GB", "US"):
            include.add("SB")

        if not (1 <= month <= 12):
            return jsonify({"error": "month must be 1..12"}), 400
        if not (2000 <= year <= 2100):
            return jsonify({"error": "year looks invalid"}), 400
        if not country:
            return jsonify({"error": "country is required (e.g. UK/US/CA)"}), 400
        if not include.intersection({"SP", "SD", "SB"}):
            return jsonify({"error": "include must contain SP and/or SD and/or SB"}), 400

        # Create or refresh the requested monthly ads table.
        table_name = _safe_ident(f"adsmonthly_{user_id}_{country}_{month}_{year}")

        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])

        frames = []

        # ---------------- helpers ----------------
        def _safe_div(a, b):
            try:
                a = float(a or 0.0)
                b = float(b or 0.0)
                return (a / b) if b else 0.0
            except Exception:
                return 0.0

        def _pick_attr(obj, names):
            """Return first non-empty attribute value among candidates."""
            for n in names:
                if hasattr(obj, n):
                    v = getattr(obj, n)
                    if v is not None and str(v).strip() != "":
                        return str(v).strip()
            return ""

        def _apply_filters_if_exist(q, model):
            """Apply user/country/date/time_unit filters only if those columns exist on the model."""
            if hasattr(model, "user_id"):
                q = q.filter(model.user_id == user_id)
            if hasattr(model, "country"):
                q = q.filter(model.country == country)
            if hasattr(model, "start_date"):
                q = q.filter(model.start_date >= first_day, model.start_date <= last_day)

            # ✅ Monthly is now built from DAILY rows.
            # This allows one fetch only: fetch DAILY once, then create both daily and monthly tables.
            if hasattr(model, "time_unit"):
                q = q.filter(func.upper(func.coalesce(model.time_unit, "")) == "DAILY")

            return q

        # =========================
        # 1) Sponsored Products (SP)
        # =========================
        if "SP" in include:
            # ✅ Monthly is built by summing all DAILY SP rows in the month.
            # Do not filter by latest end_date, otherwise only the last day can be picked.
            q = amazon_sponsored_products.query
            q = _apply_filters_if_exist(q, amazon_sponsored_products)

            sp_rows = q.all()

            if sp_rows:
                sp_df = pd.DataFrame([{
                    "source": "SP",
                    "advertised_sku": (getattr(r, "advertised_sku", None) or ""),
                    "advertised_asin": (getattr(r, "advertised_asin", None) or ""),
                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    # SP uses 7d attribution in your schema (keep safe)
                    "sales": float(getattr(r, "sales_7d", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders_7d", 0.0) or 0.0),
                    "units": float(getattr(r, "units_7d", 0.0) or 0.0),

                    "sp_ads_sales": float(getattr(r, "sales_7d", 0.0) or 0.0),
                    "sd_ads_sales": 0.0,
                    "sb_ads_sales": 0.0,

                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),
                    "advertised_unit_sale": float(getattr(r, "adv_sku_units_7d", 0.0) or 0.0),
                    "other_unit_sale": float(getattr(r, "other_sku_units_7d", 0.0) or 0.0),

                    # ✅ new spend columns
                    "product_spend": float(getattr(r, "spend", 0.0) or 0.0),
                    "display_spend": 0.0,
                    "brand_spend": 0.0,
                } for r in sp_rows])
                frames.append(sp_df)

        # =========================
        # 2) Sponsored Display (SD)
        # =========================
        if "SD" in include:
            # ✅ Monthly is built by summing all DAILY SD rows in the month.
            # Do not filter by latest end_date, otherwise only the last day can be picked.
            q = amazon_sponsored_display_advertised_products.query
            q = _apply_filters_if_exist(q, amazon_sponsored_display_advertised_products)

            sd_rows = q.all()

            if sd_rows:
                sd_df = pd.DataFrame([{
                    "source": "SD",
                    "advertised_sku": (getattr(r, "advertised_sku", None) or ""),
                    "advertised_asin": (getattr(r, "advertised_asin", None) or ""),
                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    # SD uses 14d attribution in your schema (keep safe)
                    "sales": float(getattr(r, "sales_14d", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders_14d", 0.0) or 0.0),
                    "units": float(getattr(r, "units_14d", 0.0) or 0.0),

                    "sp_ads_sales": 0.0,
                    "sd_ads_sales": float(getattr(r, "sales_14d", 0.0) or 0.0),
                    "sb_ads_sales": 0.0,

                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),

                    "advertised_unit_sale": 0.0,
                    "other_unit_sale": 0.0,

                    # ✅ new spend columns
                    "product_spend": 0.0,
                    "display_spend": float(getattr(r, "spend", 0.0) or 0.0),
                    "brand_spend": 0.0,
                } for r in sd_rows])
                frames.append(sd_df)

        # =========================
        # 3) Sponsored Brands (SB) - keywords
        # =========================
        if "SB" in include:
            # ✅ Monthly is built by summing all DAILY SB rows in the month.
            # Do not use latest end_date here.
            q = amazon_sponsored_brands_keywords.query.filter(
                amazon_sponsored_brands_keywords.user_id == user_id,
                func.upper(func.trim(amazon_sponsored_brands_keywords.country)).in_(country_aliases),
                amazon_sponsored_brands_keywords.start_date >= first_day,
                amazon_sponsored_brands_keywords.start_date <= last_day,

                # ✅ Monthly is now built from DAILY rows
                func.upper(func.coalesce(amazon_sponsored_brands_keywords.time_unit, "")) == "DAILY",
            )

            sb_rows = q.all()

            if sb_rows:
                sb_df = pd.DataFrame([{
                    "source": "SB",

                    # SB keywords often cannot map to SKU/ASIN; try common field names if present
                    "advertised_sku": _pick_attr(r, ["advertised_sku", "sku", "product_sku"]) or "",
                    "advertised_asin": _pick_attr(r, ["advertised_asin", "asin", "product_asin"]) or "",

                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    # keyword report may not have these -> keep 0
                    "sales": float(getattr(r, "sales", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders", 0.0) or 0.0),
                    "units": float(getattr(r, "units", 0.0) or 0.0),

                    "sp_ads_sales": 0.0,
                    "sd_ads_sales": 0.0,
                    "sb_ads_sales": float(getattr(r, "sales", 0.0) or 0.0),

                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),
                    "advertised_unit_sale": 0.0,
                    "other_unit_sale": 0.0,

                    # ✅ new spend columns
                    "product_spend": 0.0,
                    "display_spend": 0.0,
                    "brand_spend": float(getattr(r, "spend", 0.0) or 0.0),
                } for r in sb_rows])
                frames.append(sb_df)

        if not frames:
            return jsonify({
                "error": "No rows found for this user/country/month (or no latest end_date found)."
            }), 404

        df = pd.concat(frames, ignore_index=True)

        # Ensure missing columns exist
        for col in [
            "impressions", "clicks", "spend", "sales", "orders", "units",
            "advertised_unit_sale", "other_unit_sale", "new_to_brand_sales",
            "product_spend", "display_spend", "brand_spend",
            "sp_ads_sales", "sd_ads_sales", "sb_ads_sales",
        ]:
            if col not in df.columns:
                df[col] = 0.0

        df["advertised_sku"] = df.get("advertised_sku", "").fillna("").astype(str).str.strip()
        df["advertised_asin"] = df.get("advertised_asin", "").fillna("").astype(str).str.strip()

        # ---- group by SKU + ASIN ----
        g = df.groupby(["advertised_sku", "advertised_asin"], as_index=False).agg({
            "impressions": "sum",
            "clicks": "sum",
            "spend": "sum",
            "sales": "sum",
            "orders": "sum",
            "units": "sum",
            "advertised_unit_sale": "sum",
            "other_unit_sale": "sum",
            "new_to_brand_sales": "sum",

            # ✅ new spend columns
            "product_spend": "sum",
            "display_spend": "sum",
            "brand_spend": "sum",

            "sp_ads_sales": "sum",
            "sd_ads_sales": "sum",
            "sb_ads_sales": "sum",

            # ✅ keep which sources contributed
            "source": lambda s: ",".join(sorted(set([str(x).upper() for x in s if x])))
        })

        # ---- output in your monthly table shape ----
        out = pd.DataFrame()
        out["sno"] = range(1, len(g) + 1)
        out["products"] = g["advertised_sku"]
        out["asin"] = g["advertised_asin"]

        def _ad_type_from_source(src: str) -> str:
            parts = set((src or "").split(","))
            labels = []
            if "SD" in parts:
                labels.append("sponsored_display")
            if "SP" in parts:
                labels.append("sponsored_product")
            if "SB" in parts:
                labels.append("sponsored_brands")
            return ", ".join(labels) if labels else None

        out["ad_type"] = g["source"].apply(_ad_type_from_source)
        out["match_type"] = None

        out["impressions"] = g["impressions"].astype(int)
        out["clicks"] = g["clicks"].astype(int)

        # percentages 0..100
        out["ctr"] = [
            _safe_div(c, i) * 100.0
            for c, i in zip(g["clicks"].tolist(), g["impressions"].tolist())
        ]
        out["cpc"] = [
            _safe_div(sp, c)
            for sp, c in zip(g["spend"].tolist(), g["clicks"].tolist())
        ]

        out["spend"] = g["spend"].astype(float)
        out["sale_units"] = g["units"].astype(float)
        out["sale_amount"] = g["sales"].astype(float)

        out["sp_ads_sales"] = g["sp_ads_sales"].astype(float)
        out["sd_ads_sales"] = g["sd_ads_sales"].astype(float)
        out["sb_ads_sales"] = g["sb_ads_sales"].astype(float)

        out["advertised_unit_sale"] = g["advertised_unit_sale"].astype(float)
        out["other_unit_sale"] = g["other_unit_sale"].astype(float)
        out["new_to_brand_sales"] = g["new_to_brand_sales"].astype(float)

        # ✅ NEW columns
        out["product_spend"] = g["product_spend"].astype(float)
        out["display_spend"] = g["display_spend"].astype(float)
        out["brand_spend"] = g["brand_spend"].astype(float)

        out["conversion_rate"] = [
            _safe_div(o, c) * 100.0
            for o, c in zip(g["orders"].tolist(), g["clicks"].tolist())
        ]
        out["roas"] = [
            _safe_div(sa, sp)
            for sa, sp in zip(g["sales"].tolist(), g["spend"].tolist())
        ]
        out["acos"] = [
            _safe_div(sp, sa) * 100.0
            for sp, sa in zip(g["spend"].tolist(), g["sales"].tolist())
        ]

        # ---- Grand Total row ----

        # 1) Get raw SB spend safely from DB
        raw_sb_brand_spend = 0.0

        if "SB" in include:
            try:
                # ✅ Monthly is now built from DAILY rows.
                # Sum all daily SB spend in the month. No latest_end_date filter.
                sb_q = db.session.query(
                    func.coalesce(func.sum(amazon_sponsored_brands_keywords.spend), 0.0)
                ).filter(
                    amazon_sponsored_brands_keywords.user_id == user_id,
                    func.upper(func.trim(amazon_sponsored_brands_keywords.country)).in_(country_aliases),
                    amazon_sponsored_brands_keywords.start_date >= first_day,
                    amazon_sponsored_brands_keywords.start_date <= last_day,
                    func.upper(func.coalesce(amazon_sponsored_brands_keywords.time_unit, "")) == "DAILY",
                )

                raw_sb_brand_spend = round(float(sb_q.scalar() or 0.0), 2)

            except Exception as e:
                print("[WARN] SB brand_spend total fetch failed:", e)
                raw_sb_brand_spend = 0.0


        # 2) Check if SB spend already exists in out rows
        existing_out_brand_spend = (
            float(pd.to_numeric(out["brand_spend"], errors="coerce").fillna(0.0).sum())
            if "brand_spend" in out.columns
            else 0.0
        )

        # 3) If some SB spend is missing from out, add the missing amount as a row
        if raw_sb_brand_spend > 0 and abs(existing_out_brand_spend - raw_sb_brand_spend) > 0.01:
            missing_brand_spend = raw_sb_brand_spend - existing_out_brand_spend

            sb_manual_row = {
                "sno": len(out) + 1,
                "products": "Sponsored Brands",
                "asin": None,
                "ad_type": "sponsored_brands",
                "match_type": None,

                "impressions": 0,
                "clicks": 0,
                "ctr": 0.0,
                "cpc": 0.0,

                "spend": missing_brand_spend,
                "product_spend": 0.0,
                "display_spend": 0.0,
                "brand_spend": missing_brand_spend,

                "sale_units": 0.0,
                "sale_amount": 0.0,
                "advertised_unit_sale": 0.0,
                "other_unit_sale": 0.0,
                "new_to_brand_sales": 0.0,
                "sp_ads_sales": 0.0,
                "sd_ads_sales": 0.0,
                "sb_ads_sales": 0.0,

                "conversion_rate": 0.0,
                "roas": 0.0,
                "acos": 0.0,
            }

            out = pd.concat([out, pd.DataFrame([sb_manual_row])], ignore_index=True)


        # 4) Recalculate totals after possible SB manual row
        total_impr = int(pd.to_numeric(out["impressions"], errors="coerce").fillna(0).sum())
        total_clicks = int(pd.to_numeric(out["clicks"], errors="coerce").fillna(0).sum())

        total_product_spend = float(
            pd.to_numeric(out["product_spend"], errors="coerce").fillna(0.0).sum()
        )

        total_display_spend = float(
            pd.to_numeric(out["display_spend"], errors="coerce").fillna(0.0).sum()
        )

        total_brand_spend = float(
            pd.to_numeric(out["brand_spend"], errors="coerce").fillna(0.0).sum()
        )

        total_spend = round(
            total_product_spend + total_display_spend + total_brand_spend,
            2,
        )

        total_sales_amt = float(
            pd.to_numeric(out["sale_amount"], errors="coerce").fillna(0.0).sum()
        )

        total_orders = float(g["orders"].sum()) if "orders" in g.columns else 0.0

        total_units = float(
            pd.to_numeric(out["sale_units"], errors="coerce").fillna(0.0).sum()
        )
        total_sp_ads_sales = float(
            pd.to_numeric(out["sp_ads_sales"], errors="coerce").fillna(0.0).sum()
        )

        total_sd_ads_sales = float(
            pd.to_numeric(out["sd_ads_sales"], errors="coerce").fillna(0.0).sum()
        )

        total_sb_ads_sales = float(
            pd.to_numeric(out["sb_ads_sales"], errors="coerce").fillna(0.0).sum()
        )

        total_row = {
            "sno": None,
            "products": "Grand Total",
            "asin": None,
            "ad_type": None,
            "match_type": None,
            "impressions": total_impr,
            "clicks": total_clicks,
            "ctr": _safe_div(total_clicks, total_impr) * 100.0,
            "cpc": _safe_div(total_spend, total_clicks),
            "spend": total_spend,

            # ✅ product/display unchanged, only brand fixed
            "product_spend": total_product_spend,
            "display_spend": total_display_spend,
            "brand_spend": total_brand_spend,

            "sale_units": total_units,
            "sale_amount": total_sales_amt,
            "sp_ads_sales": total_sp_ads_sales,
            "sd_ads_sales": total_sd_ads_sales,
            "sb_ads_sales": total_sb_ads_sales,
            "advertised_unit_sale": float(out["advertised_unit_sale"].sum()),
            "other_unit_sale": float(out["other_unit_sale"].sum()),
            "new_to_brand_sales": float(out["new_to_brand_sales"].sum()),
            "conversion_rate": _safe_div(total_orders, total_clicks) * 100.0,
            "roas": _safe_div(total_sales_amt, total_spend),
            "acos": _safe_div(total_spend, total_sales_amt) * 100.0,
        }

        out = pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)
        items = out.where(pd.notnull(out), None).to_dict(orient="records")

        # ---- target table is created automatically before insert ----

        # ---- SQL ----
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS public.{table_name} (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            country VARCHAR(10) NOT NULL,
            month INT NOT NULL,
            year INT NOT NULL,

            sno INT,
            products TEXT,
            asin VARCHAR(32),
            ad_type TEXT,
            match_type TEXT,

            impressions BIGINT,
            clicks BIGINT,
            ctr DOUBLE PRECISION,
            cpc DOUBLE PRECISION,
            spend DOUBLE PRECISION,

            -- ✅ NEW
            product_spend DOUBLE PRECISION,
            display_spend DOUBLE PRECISION,
            brand_spend DOUBLE PRECISION,

            sale_units DOUBLE PRECISION,
            sale_amount DOUBLE PRECISION,
            sp_ads_sales DOUBLE PRECISION,
            sd_ads_sales DOUBLE PRECISION,
            sb_ads_sales DOUBLE PRECISION,

            advertised_unit_sale DOUBLE PRECISION,
            other_unit_sale DOUBLE PRECISION,
            new_to_brand_sales DOUBLE PRECISION,

            conversion_rate DOUBLE PRECISION,
            roas DOUBLE PRECISION,
            acos DOUBLE PRECISION,

            created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW()
        );
        """

        insert_sql = f"""
        INSERT INTO public.{table_name} (
            user_id, country, month, year,
            sno, products, asin, ad_type, match_type,
            impressions, clicks, ctr, cpc, spend,

            product_spend, display_spend, brand_spend,

            sale_units, sale_amount,
            sp_ads_sales, sd_ads_sales, sb_ads_sales,
            advertised_unit_sale, other_unit_sale, new_to_brand_sales,
            conversion_rate, roas, acos
        ) VALUES (
            :user_id, :country, :month, :year,
            :sno, :products, :asin, :ad_type, :match_type,
            :impressions, :clicks, :ctr, :cpc, :spend,

            :product_spend, :display_spend, :brand_spend,

            :sale_units, :sale_amount,
            :sp_ads_sales, :sd_ads_sales, :sb_ads_sales,
            :advertised_unit_sale, :other_unit_sale, :new_to_brand_sales,
            :conversion_rate, :roas, :acos
        );
        """

        try:
            # Create the monthly table automatically when it is missing.
            db.session.execute(text(create_sql))

            # ✅ if table existed previously, add missing columns safely
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS product_spend DOUBLE PRECISION;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS display_spend DOUBLE PRECISION;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS brand_spend DOUBLE PRECISION;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sp_ads_sales DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sd_ads_sales DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sb_ads_sales DOUBLE PRECISION DEFAULT 0;'))

            # ✅ wipe monthly output table before inserting
            db.session.execute(text(f"TRUNCATE TABLE public.{table_name};"))

            params = []
            for r in items:
                params.append({
                    "user_id": user_id,
                    "country": country,
                    "month": month,
                    "year": year,

                    "sno": r.get("sno"),
                    "products": r.get("products"),
                    "asin": r.get("asin"),
                    "ad_type": r.get("ad_type"),
                    "match_type": r.get("match_type"),

                    "impressions": int(r.get("impressions") or 0),
                    "clicks": int(r.get("clicks") or 0),
                    "ctr": float(r.get("ctr") or 0.0),
                    "cpc": float(r.get("cpc") or 0.0),
                    "spend": float(r.get("spend") or 0.0),

                    # ✅ NEW
                    "product_spend": float(r.get("product_spend") or 0.0),
                    "display_spend": float(r.get("display_spend") or 0.0),
                    "brand_spend": float(r.get("brand_spend") or 0.0),

                    "sale_units": float(r.get("sale_units") or 0.0),
                    "sale_amount": float(r.get("sale_amount") or 0.0),
                    "sp_ads_sales": float(r.get("sp_ads_sales") or 0.0),
                    "sd_ads_sales": float(r.get("sd_ads_sales") or 0.0),
                    "sb_ads_sales": float(r.get("sb_ads_sales") or 0.0),

                    "advertised_unit_sale": float(r.get("advertised_unit_sale") or 0.0),
                    "other_unit_sale": float(r.get("other_unit_sale") or 0.0),
                    "new_to_brand_sales": float(r.get("new_to_brand_sales") or 0.0),

                    "conversion_rate": float(r.get("conversion_rate") or 0.0),
                    "roas": float(r.get("roas") or 0.0),
                    "acos": float(r.get("acos") or 0.0),
                })

            if params:
                db.session.execute(text(insert_sql), params)
                month_name = calendar.month_name[month].lower()
                skuwise_table_name = _safe_ident(f"skuwisemonthly_{user_id}_{country.lower()}_{month_name}{year}")

                # ---------------------------------------------------------
                # Sync monthly ads values into skuwisemonthly table
                # Columns filled on every route hit:
                # product_spend, display_spend, brand_spend, ad_type,
                # sp_ads_sales, sd_ads_sales, sb_ads_sales
                # ---------------------------------------------------------
                skuwise_table_exists = db.session.execute(text("""
                    SELECT EXISTS (
                        SELECT 1
                        FROM information_schema.tables
                        WHERE table_schema = 'public'
                          AND table_name = :table_name
                    )
                """), {"table_name": skuwise_table_name}).scalar()

                if not skuwise_table_exists:
                    raise RuntimeError(
                        f"SKU-wise monthly table public.{skuwise_table_name} does not exist"
                    )

                # Add all required columns safely when an older SKU-wise table is used.
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS product_spend DOUBLE PRECISION DEFAULT 0;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS display_spend DOUBLE PRECISION DEFAULT 0;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS brand_spend DOUBLE PRECISION DEFAULT 0;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS ad_type TEXT;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS sp_ads_sales DOUBLE PRECISION DEFAULT 0;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS sd_ads_sales DOUBLE PRECISION DEFAULT 0;'
                ))
                db.session.execute(text(
                    f'ALTER TABLE public.{skuwise_table_name} '
                    'ADD COLUMN IF NOT EXISTS sb_ads_sales DOUBLE PRECISION DEFAULT 0;'
                ))

                # Refresh actual SKU-wise columns after ALTER statements.
                skuwise_columns = set(db.session.execute(text("""
                    SELECT LOWER(column_name)
                    FROM information_schema.columns
                    WHERE table_schema = 'public'
                      AND table_name = :table_name
                """), {"table_name": skuwise_table_name}).scalars().all())

                total_conditions = []
                if "sku" in skuwise_columns:
                    total_conditions.append("""
                        UPPER(TRIM(COALESCE(sku::text, ''))) IN
                        ('TOTAL', 'GRAND TOTAL', 'GRAND_TOTAL')
                    """)
                if "product_name" in skuwise_columns:
                    total_conditions.append("""
                        UPPER(TRIM(COALESCE(product_name::text, ''))) IN
                        ('TOTAL', 'GRAND TOTAL', 'GRAND_TOTAL')
                    """)

                total_where_sql = " OR ".join(total_conditions) if total_conditions else "FALSE"

                # Reset all non-total rows first so stale ads values cannot remain.
                db.session.execute(text(f"""
                    UPDATE public.{skuwise_table_name}
                    SET
                        product_spend = 0,
                        display_spend = 0,
                        brand_spend = 0,
                        ad_type = NULL,
                        sp_ads_sales = 0,
                        sd_ads_sales = 0,
                        sb_ads_sales = 0
                    WHERE NOT ({total_where_sql})
                """))

                # Build one deterministic ads source row per normalized SKU.
                # Blank SKUs (normally unmapped Sponsored Brands rows) are excluded
                # from product rows and remain represented in the Grand Total only.
                if "sku" in skuwise_columns:
                    db.session.execute(text(f"""
                        WITH ads_by_sku AS (
                            SELECT
                                UPPER(TRIM(products::text)) AS join_sku,
                                SUM(COALESCE(product_spend, 0)) AS product_spend,
                                SUM(COALESCE(display_spend, 0)) AS display_spend,
                                SUM(COALESCE(brand_spend, 0)) AS brand_spend,
                                SUM(COALESCE(sp_ads_sales, 0)) AS sp_ads_sales,
                                SUM(COALESCE(sd_ads_sales, 0)) AS sd_ads_sales,
                                SUM(COALESCE(sb_ads_sales, 0)) AS sb_ads_sales,
                                STRING_AGG(DISTINCT NULLIF(TRIM(ad_type), ''), ', ')
                                    FILTER (WHERE NULLIF(TRIM(ad_type), '') IS NOT NULL) AS ad_type
                            FROM public.{table_name}
                            WHERE products IS NOT NULL
                              AND TRIM(products::text) <> ''
                              AND UPPER(TRIM(products::text)) NOT IN
                                  ('TOTAL', 'GRAND TOTAL', 'GRAND_TOTAL')
                            GROUP BY UPPER(TRIM(products::text))
                        )
                        UPDATE public.{skuwise_table_name} AS s
                        SET
                            product_spend = a.product_spend,
                            display_spend = a.display_spend,
                            brand_spend = a.brand_spend,
                            ad_type = a.ad_type,
                            sp_ads_sales = a.sp_ads_sales,
                            sd_ads_sales = a.sd_ads_sales,
                            sb_ads_sales = a.sb_ads_sales
                        FROM ads_by_sku AS a
                        WHERE UPPER(TRIM(COALESCE(s.sku::text, ''))) = a.join_sku
                          AND NOT ({total_where_sql})
                    """))

                # ASIN fallback is used only for still-unmatched rows and only when
                # the SKU-wise table actually contains an ASIN column.
                if "asin" in skuwise_columns:
                    db.session.execute(text(f"""
                        WITH ads_by_asin AS (
                            SELECT
                                UPPER(TRIM(asin::text)) AS join_asin,
                                SUM(COALESCE(product_spend, 0)) AS product_spend,
                                SUM(COALESCE(display_spend, 0)) AS display_spend,
                                SUM(COALESCE(brand_spend, 0)) AS brand_spend,
                                SUM(COALESCE(sp_ads_sales, 0)) AS sp_ads_sales,
                                SUM(COALESCE(sd_ads_sales, 0)) AS sd_ads_sales,
                                SUM(COALESCE(sb_ads_sales, 0)) AS sb_ads_sales,
                                STRING_AGG(DISTINCT NULLIF(TRIM(ad_type), ''), ', ')
                                    FILTER (WHERE NULLIF(TRIM(ad_type), '') IS NOT NULL) AS ad_type
                            FROM public.{table_name}
                            WHERE asin IS NOT NULL
                              AND TRIM(asin::text) <> ''
                              AND UPPER(TRIM(COALESCE(products::text, ''))) NOT IN
                                  ('TOTAL', 'GRAND TOTAL', 'GRAND_TOTAL')
                            GROUP BY UPPER(TRIM(asin::text))
                        )
                        UPDATE public.{skuwise_table_name} AS s
                        SET
                            product_spend = a.product_spend,
                            display_spend = a.display_spend,
                            brand_spend = a.brand_spend,
                            ad_type = a.ad_type,
                            sp_ads_sales = a.sp_ads_sales,
                            sd_ads_sales = a.sd_ads_sales,
                            sb_ads_sales = a.sb_ads_sales
                        FROM ads_by_asin AS a
                        WHERE UPPER(TRIM(COALESCE(s.asin::text, ''))) = a.join_asin
                          AND NULLIF(TRIM(COALESCE(s.ad_type, '')), '') IS NULL
                          AND NOT ({total_where_sql})
                    """))

                # Fill the SKU-wise Grand Total from the visible SKU rows for
                # Sponsored Product / Display spend. Sponsored Brand totals can
                # still come from the ads table because those rows may be
                # account-level and not SKU-attributed.
                if total_conditions:
                    db.session.execute(text(f"""
                        WITH sku_ad_totals AS (
                            SELECT
                                COALESCE(SUM(product_spend), 0) AS product_spend,
                                COALESCE(SUM(display_spend), 0) AS display_spend,
                                COALESCE(SUM(sp_ads_sales), 0) AS sp_ads_sales,
                                COALESCE(SUM(sd_ads_sales), 0) AS sd_ads_sales
                            FROM public.{skuwise_table_name}
                            WHERE NOT ({total_where_sql})
                        ),
                        ads_total AS (
                            SELECT
                                COALESCE(SUM(brand_spend), 0) AS brand_spend,
                                COALESCE(SUM(sb_ads_sales), 0) AS sb_ads_sales
                            FROM public.{table_name}
                            WHERE UPPER(TRIM(COALESCE(products::text, ''))) NOT IN
                                  ('TOTAL', 'GRAND TOTAL', 'GRAND_TOTAL')
                        )
                        UPDATE public.{skuwise_table_name}
                        SET
                            product_spend = s.product_spend,
                            display_spend = s.display_spend,
                            brand_spend = t.brand_spend,
                            ad_type = CASE
                                WHEN s.product_spend <> 0
                                 AND s.display_spend <> 0
                                 AND t.brand_spend <> 0
                                    THEN 'sponsored_product, sponsored_display, sponsored_brands'
                                WHEN s.product_spend <> 0 AND s.display_spend <> 0
                                    THEN 'sponsored_product, sponsored_display'
                                WHEN s.product_spend <> 0 AND t.brand_spend <> 0
                                    THEN 'sponsored_product, sponsored_brands'
                                WHEN s.display_spend <> 0 AND t.brand_spend <> 0
                                    THEN 'sponsored_display, sponsored_brands'
                                WHEN s.product_spend <> 0 THEN 'sponsored_product'
                                WHEN s.display_spend <> 0 THEN 'sponsored_display'
                                WHEN t.brand_spend <> 0 THEN 'sponsored_brands'
                                ELSE NULL
                            END,
                            sp_ads_sales = s.sp_ads_sales,
                            sd_ads_sales = s.sd_ads_sales,
                            sb_ads_sales = t.sb_ads_sales
                        FROM sku_ad_totals AS s
                        CROSS JOIN ads_total AS t
                        WHERE {total_where_sql}
                    """))


                # ---------------------------------------------------------
                # Recalculate CM2 and total advertising profitability fields
                # using the same formulas as amazon_api_routes.
                #
                # Product-level formulas:
                #   ads_spend = product_spend + display_spend
                #   cm2_profit = profit - ads_spend
                #   cm2 margin = cm2_profit / net_sales * 100
                #
                # Total profitability formulas:
                #   total_ads = product_spend + display_spend
                #               + brand_spend + dealsvouchar_ads
                #   total_cm2_profit = cm2_profit - brand_spend
                #                      - dealsvouchar_ads
                #                      - abs(platform_fee)
                #                      - abs(shipment_fees)
                # ---------------------------------------------------------
                for target_column in [
                    "cm2_profit",
                    "cm2_profit_percentage",
                    "cm2_margins",
                    "cm2_profit_per_unit",
                    "total_ads",
                    "total_cm2_profit",
                    "total_cm2_margins",
                    "tacos_total_advertising_cost_of_sale",
                ]:
                    db.session.execute(text(
                        f'ALTER TABLE public.{skuwise_table_name} '
                        f'ADD COLUMN IF NOT EXISTS {target_column} DOUBLE PRECISION DEFAULT 0;'
                    ))

                # Keep ads_spend synchronized too when that existing Amazon
                # calculation column is available in the SKU-wise table.
                skuwise_columns = set(db.session.execute(text("""
                    SELECT LOWER(column_name)
                    FROM information_schema.columns
                    WHERE table_schema = 'public'
                      AND table_name = :table_name
                """), {"table_name": skuwise_table_name}).scalars().all())

                def _numeric_sql(column_name, *, absolute=False):
                    if column_name not in skuwise_columns:
                        return "0.0"
                    expression = f"COALESCE({column_name}, 0.0)"
                    return f"ABS({expression})" if absolute else expression

                profit_expr = _numeric_sql("profit")
                net_sales_expr = _numeric_sql("net_sales")
                product_spend_expr = _numeric_sql("product_spend", absolute=True)
                display_spend_expr = _numeric_sql("display_spend", absolute=True)

                # Amazon route primarily uses total_quantity. Fall back to
                # quantity - return_quantity for older monthly tables.
                if "total_quantity" in skuwise_columns:
                    units_expr = "ABS(COALESCE(total_quantity, 0.0))"
                elif "quantity" in skuwise_columns and "return_quantity" in skuwise_columns:
                    units_expr = (
                        "GREATEST(COALESCE(quantity, 0.0) "
                        "- COALESCE(return_quantity, 0.0), 0.0)"
                    )
                elif "quantity" in skuwise_columns:
                    units_expr = "ABS(COALESCE(quantity, 0.0))"
                else:
                    units_expr = "0.0"

                product_ads_expr = f"({product_spend_expr} + {display_spend_expr})"
                cm2_profit_expr = f"({profit_expr} - {product_ads_expr})"

                set_clauses = []
                if "ads_spend" in skuwise_columns:
                    set_clauses.append(f"ads_spend = ROUND(({product_ads_expr})::numeric, 2)")

                set_clauses.extend([
                    f"cm2_profit = ROUND(({cm2_profit_expr})::numeric, 2)",
                    (
                        "cm2_profit_percentage = ROUND((CASE "
                        f"WHEN {net_sales_expr} <> 0 THEN "
                        f"({cm2_profit_expr} / {net_sales_expr}) * 100.0 "
                        "ELSE 0.0 END)::numeric, 2)"
                    ),
                    (
                        "cm2_margins = ROUND((CASE "
                        f"WHEN {net_sales_expr} <> 0 THEN "
                        f"({cm2_profit_expr} / {net_sales_expr}) * 100.0 "
                        "ELSE 0.0 END)::numeric, 2)"
                    ),
                    (
                        "cm2_profit_per_unit = ROUND((CASE "
                        f"WHEN {units_expr} <> 0 THEN "
                        f"({cm2_profit_expr} / {units_expr}) "
                        "ELSE 0.0 END)::numeric, 2)"
                    ),
                ])

                db.session.execute(text(f"""
                    UPDATE public.{skuwise_table_name}
                    SET {", ".join(set_clauses)}
                """))

                if total_conditions:
                    db.session.execute(text(f"""
                        UPDATE public.{skuwise_table_name}
                        SET
                            total_ads = 0,
                            total_cm2_profit = 0,
                            total_cm2_margins = 0,
                            tacos_total_advertising_cost_of_sale = 0
                        WHERE NOT ({total_where_sql})
                    """))

                # Productwise CM2 total must be the sum of SKU rows. The
                # account-level value stays in total_cm2_profit and follows
                # the lower P&L section: cost ads, shipping, storage,
                # inventory reimbursement, and other fees.
                if total_conditions:
                    total_net_sales_expr = (
                        "COALESCE(s.net_sales, 0.0)"
                        if "net_sales" in skuwise_columns
                        else "0.0"
                    )
                    total_units_expr = (
                        "ABS(COALESCE(s.total_quantity, 0.0))"
                        if "total_quantity" in skuwise_columns
                        else (
                            "GREATEST(COALESCE(s.quantity, 0.0) "
                            "- COALESCE(s.return_quantity, 0.0), 0.0)"
                            if "quantity" in skuwise_columns and "return_quantity" in skuwise_columns
                            else (
                                "ABS(COALESCE(s.quantity, 0.0))"
                                if "quantity" in skuwise_columns
                                else "0.0"
                            )
                        )
                    )
                    total_product_expr = (
                        "ABS(COALESCE(s.product_spend, 0.0))"
                        if "product_spend" in skuwise_columns
                        else "0.0"
                    )
                    total_display_expr = (
                        "ABS(COALESCE(s.display_spend, 0.0))"
                        if "display_spend" in skuwise_columns
                        else "0.0"
                    )
                    total_brand_expr = (
                        "ABS(COALESCE(s.brand_spend, 0.0))"
                        if "brand_spend" in skuwise_columns
                        else "0.0"
                    )
                    total_deals_expr = (
                        "ABS(COALESCE(s.dealsvouchar_ads, 0.0))"
                        if "dealsvouchar_ads" in skuwise_columns
                        else "0.0"
                    )
                    total_shipping_expr = (
                        "ABS(COALESCE(s.shipping_charges, 0.0))"
                        if "shipping_charges" in skuwise_columns
                        else "0.0"
                    )
                    total_storage_expr = (
                        "ABS(COALESCE(s.storage_fee, 0.0))"
                        if "storage_fee" in skuwise_columns
                        else "0.0"
                    )
                    total_lost_expr = (
                        "ABS(COALESCE(s.lost_total, 0.0))"
                        if "lost_total" in skuwise_columns
                        else "0.0"
                    )
                    total_disposal_expr = (
                        "ABS(COALESCE(s.fba_disposal, 0.0))"
                        if "fba_disposal" in skuwise_columns
                        else "0.0"
                    )
                    total_platform_management_expr = (
                        "ABS(COALESCE(s.platform_management_fees, 0.0))"
                        if "platform_management_fees" in skuwise_columns
                        else "0.0"
                    )
                    total_other_adjustment_expr = (
                        "ABS(COALESCE(s.other_adjustment, 0.0))"
                        if "other_adjustment" in skuwise_columns
                        else "0.0"
                    )
                    total_ads_expr = (
                        f"({total_product_expr} + {total_display_expr} "
                        f"+ {total_brand_expr} + {total_deals_expr})"
                    )
                    inventory_reimbursement_expr = (
                        f"({total_lost_expr} - {total_disposal_expr})"
                    )
                    other_fees_expr = (
                        f"({total_platform_management_expr} - {total_other_adjustment_expr})"
                    )
                    total_cm2_expr = (
                        f"(pt.cm2_profit - {total_brand_expr} - {total_deals_expr} "
                        f"- {total_shipping_expr} - {total_storage_expr} "
                        f"+ {inventory_reimbursement_expr} - {other_fees_expr})"
                    )

                    db.session.execute(text(f"""
                        WITH product_totals AS (
                            SELECT
                                COALESCE(SUM(COALESCE(cm2_profit, 0.0)), 0.0) AS cm2_profit
                            FROM public.{skuwise_table_name}
                            WHERE NOT ({total_where_sql})
                        )
                        UPDATE public.{skuwise_table_name} AS s
                        SET
                            cm2_profit = ROUND(pt.cm2_profit::numeric, 2),
                            cm2_profit_percentage = ROUND((CASE
                                WHEN {total_net_sales_expr} <> 0 THEN
                                    (pt.cm2_profit / {total_net_sales_expr}) * 100.0
                                ELSE 0.0
                            END)::numeric, 2),
                            cm2_margins = ROUND((CASE
                                WHEN {total_net_sales_expr} <> 0 THEN
                                    (pt.cm2_profit / {total_net_sales_expr}) * 100.0
                                ELSE 0.0
                            END)::numeric, 2),
                            cm2_profit_per_unit = ROUND((CASE
                                WHEN {total_units_expr} <> 0 THEN
                                    pt.cm2_profit / {total_units_expr}
                                ELSE 0.0
                            END)::numeric, 2),
                            total_ads = ROUND(({total_ads_expr})::numeric, 2),
                            total_cm2_profit = ROUND(({total_cm2_expr})::numeric, 2),
                            total_cm2_margins = ROUND((CASE
                                WHEN {total_net_sales_expr} <> 0 THEN
                                    ({total_cm2_expr} / {total_net_sales_expr}) * 100.0
                                ELSE 0.0
                            END)::numeric, 2),
                            tacos_total_advertising_cost_of_sale = ROUND((CASE
                                WHEN {total_net_sales_expr} <> 0 THEN
                                    ({total_ads_expr} / {total_net_sales_expr}) * 100.0
                                ELSE 0.0
                            END)::numeric, 2)
                        FROM product_totals AS pt
                        WHERE {total_where_sql}
                    """))

            db.session.commit()

        except Exception:
            db.session.rollback()
            raise

        return jsonify({
            "message": "Monthly ads table saved to DB successfully (SP + SD + SB) using DAILY rows aggregated for the month",
            "table_name": f"public.{table_name}",
            "country": country,
            "month": month,
            "year": year,
            "include": sorted(list(include)),
            "count": len(items),
            "items": items
        }), 200

    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401
    except PermissionError as e:
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        # ✅ show real error in terminal + return message
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@advertisement_api_routes_bp.route("/api/ads/daily_sp_sd_sb_to_db", methods=["POST"])
def daily_sp_sd_sb_to_db():
    """
    Save daily SKU-wise Sponsored Products + Sponsored Display + Sponsored Brands into:
      public.adsdaily_{user_id}_{country}_{month}_{year}

    Body:
      {
        "month": 5,
        "year": 2026,
        "country": "UK",
        "include": ["SP", "SD", "SB"]
      }

    IMPORTANT:
    - This expects raw SP/SD/SB reports to already be synced with time_unit="DAILY".
    - It stores all dates of the month in one table.
    - Example table: public.adsdaily_1_UK_5_2026
    - SKU-wise only: no Grand Total rows.
    """
    try:
        user_id = _require_jwt_user_id()
        payload = request.get_json(force=True) or {}

        month = int(payload.get("month") or 0)
        year = int(payload.get("year") or 0)
        country = str(payload.get("country") or "").upper().strip()

        country_aliases = [country]
        if country in ("UK", "GB"):
            country_aliases = ["UK", "GB"]

        include = payload.get("include") or ["SP", "SD", "SB"]
        include = {str(x).upper().strip() for x in include}

        if country in ("UK", "GB", "US"):
            include.add("SB")

        if not (1 <= month <= 12):
            return jsonify({"error": "month must be 1..12"}), 400
        if not (2000 <= year <= 2100):
            return jsonify({"error": "year looks invalid"}), 400
        if not country:
            return jsonify({"error": "country is required (e.g. UK/US/CA)"}), 400
        if not include.intersection({"SP", "SD", "SB"}):
            return jsonify({"error": "include must contain SP and/or SD and/or SB"}), 400

        # Create or refresh the requested daily ads table.
        table_name = _safe_ident(f"adsdaily_{user_id}_{country}_{month}_{year}")

        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])

        frames = []

        def _safe_div_local(a, b):
            try:
                a = float(a or 0.0)
                b = float(b or 0.0)
                return (a / b) if b else 0.0
            except Exception:
                return 0.0

        def _pick_attr_local(obj, names):
            for n in names:
                if hasattr(obj, n):
                    v = getattr(obj, n)
                    if v is not None and str(v).strip() != "":
                        return str(v).strip()
            return ""

        # =========================
        # 1) Sponsored Products
        # =========================
        if "SP" in include:
            sp_rows = amazon_sponsored_products.query.filter(
                amazon_sponsored_products.user_id == user_id,
                amazon_sponsored_products.country.in_(country_aliases),
                amazon_sponsored_products.start_date >= first_day,
                amazon_sponsored_products.start_date <= last_day,

                # ✅ Daily table should read DAILY rows only
                func.upper(func.coalesce(amazon_sponsored_products.time_unit, "")) == "DAILY",
            ).all()

            if sp_rows:
                sp_df = pd.DataFrame([{
                    "date": getattr(r, "start_date", None),
                    "source": "SP",
                    "advertised_sku": getattr(r, "advertised_sku", None) or "",
                    "advertised_asin": getattr(r, "advertised_asin", None) or "",
                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    "sales": float(getattr(r, "sales_7d", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders_7d", 0.0) or 0.0),
                    "units": float(getattr(r, "units_7d", 0.0) or 0.0),

                    "sp_ads_sales": float(getattr(r, "sales_7d", 0.0) or 0.0),
                    "sd_ads_sales": 0.0,
                    "sb_ads_sales": 0.0,

                    "advertised_unit_sale": float(getattr(r, "adv_sku_units_7d", 0.0) or 0.0),
                    "other_unit_sale": float(getattr(r, "other_sku_units_7d", 0.0) or 0.0),
                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),

                    "product_spend": float(getattr(r, "spend", 0.0) or 0.0),
                    "display_spend": 0.0,
                    "brand_spend": 0.0,
                } for r in sp_rows])
                frames.append(sp_df)

        # =========================
        # 2) Sponsored Display
        # =========================
        if "SD" in include:
            sd_rows = amazon_sponsored_display_advertised_products.query.filter(
                amazon_sponsored_display_advertised_products.user_id == user_id,
                amazon_sponsored_display_advertised_products.country.in_(country_aliases),
                amazon_sponsored_display_advertised_products.start_date >= first_day,
                amazon_sponsored_display_advertised_products.start_date <= last_day,

                # ✅ Daily table should read DAILY rows only
                func.upper(func.coalesce(amazon_sponsored_display_advertised_products.time_unit, "")) == "DAILY",
            ).all()

            if sd_rows:
                sd_df = pd.DataFrame([{
                    "date": getattr(r, "start_date", None),
                    "source": "SD",
                    "advertised_sku": getattr(r, "advertised_sku", None) or "",
                    "advertised_asin": getattr(r, "advertised_asin", None) or "",
                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    "sales": float(getattr(r, "sales_14d", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders_14d", 0.0) or 0.0),
                    "units": float(getattr(r, "units_14d", 0.0) or 0.0),

                    "sp_ads_sales": 0.0,
                    "sd_ads_sales": float(getattr(r, "sales_14d", 0.0) or 0.0),
                    "sb_ads_sales": 0.0,

                    "advertised_unit_sale": 0.0,
                    "other_unit_sale": 0.0,
                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),

                    "product_spend": 0.0,
                    "display_spend": float(getattr(r, "spend", 0.0) or 0.0),
                    "brand_spend": 0.0,
                } for r in sd_rows])
                frames.append(sd_df)

        # =========================
        # 3) Sponsored Brands
        # =========================
        if "SB" in include:
            sb_rows = amazon_sponsored_brands_keywords.query.filter(
                amazon_sponsored_brands_keywords.user_id == user_id,
                func.upper(func.trim(amazon_sponsored_brands_keywords.country)).in_(country_aliases),
                amazon_sponsored_brands_keywords.start_date >= first_day,
                amazon_sponsored_brands_keywords.start_date <= last_day,

                # ✅ Daily table should read DAILY rows only
                func.upper(func.coalesce(amazon_sponsored_brands_keywords.time_unit, "")) == "DAILY",
            ).all()

            if sb_rows:
                sb_df = pd.DataFrame([{
                    "date": getattr(r, "start_date", None),
                    "source": "SB",

                    # SB keyword data usually cannot map to product SKU/ASIN unless model has those fields.
                    "advertised_sku": _pick_attr_local(r, ["advertised_sku", "sku", "product_sku"]) or "",
                    "advertised_asin": _pick_attr_local(r, ["advertised_asin", "asin", "product_asin"]) or "",

                    "currency": getattr(r, "currency", None),

                    "impressions": int(getattr(r, "impressions", 0) or 0),
                    "clicks": int(getattr(r, "clicks", 0) or 0),
                    "spend": float(getattr(r, "spend", 0.0) or 0.0),

                    "sales": float(getattr(r, "sales", 0.0) or 0.0),
                    "orders": float(getattr(r, "orders", 0.0) or 0.0),
                    "units": float(getattr(r, "units", 0.0) or 0.0),

                    "sp_ads_sales": 0.0,
                    "sd_ads_sales": 0.0,
                    "sb_ads_sales": float(getattr(r, "sales", 0.0) or 0.0),

                    "advertised_unit_sale": 0.0,
                    "other_unit_sale": 0.0,
                    "new_to_brand_sales": float(getattr(r, "new_to_brand_sales", 0.0) or 0.0),

                    "product_spend": 0.0,
                    "display_spend": 0.0,
                    "brand_spend": float(getattr(r, "spend", 0.0) or 0.0),
                } for r in sb_rows])
                frames.append(sb_df)

        if not frames:
            return jsonify({
                "error": "No daily ads rows found. First sync SP/SD/SB with time_unit='DAILY' for this month."
            }), 404

        df = pd.concat(frames, ignore_index=True)

        # Safety
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
        df = df[df["date"].notna()]

        for col in [
            "impressions", "clicks", "spend", "sales", "orders", "units",
            "advertised_unit_sale", "other_unit_sale", "new_to_brand_sales",
            "product_spend", "display_spend", "brand_spend",
            "sp_ads_sales", "sd_ads_sales", "sb_ads_sales",
        ]:
            if col not in df.columns:
                df[col] = 0.0
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        df["advertised_sku"] = df.get("advertised_sku", "").fillna("").astype(str).str.strip()
        df["advertised_asin"] = df.get("advertised_asin", "").fillna("").astype(str).str.strip()

        # Optional but useful: remove rows that have neither SKU nor ASIN.
        # This prevents unmapped SB keyword rows from creating blank SKU rows.
        df = df[
            (df["advertised_sku"].str.strip() != "") |
            (df["advertised_asin"].str.strip() != "")
        ]

        if df.empty:
            return jsonify({
                "error": "No SKU-wise daily ads rows found after removing blank SKU/ASIN rows."
            }), 404

        # ✅ DAILY + SKU-WISE GROUPING
        g = df.groupby(["date", "advertised_sku", "advertised_asin"], as_index=False).agg({
            "impressions": "sum",
            "clicks": "sum",
            "spend": "sum",
            "sales": "sum",
            "orders": "sum",
            "units": "sum",

            "advertised_unit_sale": "sum",
            "other_unit_sale": "sum",
            "new_to_brand_sales": "sum",

            "product_spend": "sum",
            "display_spend": "sum",
            "brand_spend": "sum",

            "sp_ads_sales": "sum",
            "sd_ads_sales": "sum",
            "sb_ads_sales": "sum",

            "source": lambda s: ",".join(sorted(set([str(x).upper() for x in s if x]))),
        })

        out = pd.DataFrame()
        out["sno"] = range(1, len(g) + 1)
        out["date"] = g["date"]
        out["sku"] = g["advertised_sku"]
        out["asin"] = g["advertised_asin"]

        def _ad_type_from_source(src):
            parts = set((src or "").split(","))
            labels = []
            if "SD" in parts:
                labels.append("sponsored_display")
            if "SP" in parts:
                labels.append("sponsored_product")
            if "SB" in parts:
                labels.append("sponsored_brands")
            return ", ".join(labels) if labels else None

        out["ad_type"] = g["source"].apply(_ad_type_from_source)
        out["match_type"] = None

        out["impressions"] = g["impressions"].astype(int)
        out["clicks"] = g["clicks"].astype(int)

        out["ctr"] = [
            _safe_div_local(c, i) * 100.0
            for c, i in zip(g["clicks"].tolist(), g["impressions"].tolist())
        ]

        out["cpc"] = [
            _safe_div_local(sp, c)
            for sp, c in zip(g["spend"].tolist(), g["clicks"].tolist())
        ]

        out["spend"] = g["spend"].astype(float)

        out["product_spend"] = g["product_spend"].astype(float)
        out["display_spend"] = g["display_spend"].astype(float)
        out["brand_spend"] = g["brand_spend"].astype(float)

        out["sale_units"] = g["units"].astype(float)
        out["sale_amount"] = g["sales"].astype(float)

        out["sp_ads_sales"] = g["sp_ads_sales"].astype(float)
        out["sd_ads_sales"] = g["sd_ads_sales"].astype(float)
        out["sb_ads_sales"] = g["sb_ads_sales"].astype(float)

        # ✅ total ads sales + total ads spend
        out["ads_sales_total"] = (
            out["sp_ads_sales"] +
            out["sd_ads_sales"] +
            out["sb_ads_sales"]
        )

        out["ads_spend_total"] = (
            out["product_spend"] +
            out["display_spend"] +
            out["brand_spend"]
        )

        out["advertised_unit_sale"] = g["advertised_unit_sale"].astype(float)
        out["other_unit_sale"] = g["other_unit_sale"].astype(float)
        out["new_to_brand_sales"] = g["new_to_brand_sales"].astype(float)

        out["conversion_rate"] = [
            _safe_div_local(o, c) * 100.0
            for o, c in zip(g["orders"].tolist(), g["clicks"].tolist())
        ]

        out["roas"] = [
            _safe_div_local(sa, sp)
            for sa, sp in zip(g["sales"].tolist(), g["spend"].tolist())
        ]

        out["acos"] = [
            _safe_div_local(sp, sa) * 100.0
            for sp, sa in zip(g["spend"].tolist(), g["sales"].tolist())
        ]

        # ✅ SKU-wise only: no Grand Total rows

        out = out.sort_values(
            by=["date", "sku"],
            key=lambda s: s.astype(str)
        ).reset_index(drop=True)

        items = out.where(pd.notnull(out), None).to_dict(orient="records")

        # ✅ target daily table is created automatically before insert

        create_sql = f"""
        CREATE TABLE IF NOT EXISTS public.{table_name} (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            country VARCHAR(10) NOT NULL,
            month INT NOT NULL,
            year INT NOT NULL,
            date DATE NOT NULL,

            sno INT,
            sku TEXT,
            asin VARCHAR(32),
            ad_type TEXT,
            match_type TEXT,

            impressions BIGINT,
            clicks BIGINT,
            ctr DOUBLE PRECISION,
            cpc DOUBLE PRECISION,
            spend DOUBLE PRECISION,

            product_spend DOUBLE PRECISION,
            display_spend DOUBLE PRECISION,
            brand_spend DOUBLE PRECISION,

            sale_units DOUBLE PRECISION,
            sale_amount DOUBLE PRECISION,

            sp_ads_sales DOUBLE PRECISION DEFAULT 0,
            sd_ads_sales DOUBLE PRECISION DEFAULT 0,
            sb_ads_sales DOUBLE PRECISION DEFAULT 0,

            ads_sales_total DOUBLE PRECISION DEFAULT 0,
            ads_spend_total DOUBLE PRECISION DEFAULT 0,

            advertised_unit_sale DOUBLE PRECISION,
            other_unit_sale DOUBLE PRECISION,
            new_to_brand_sales DOUBLE PRECISION,

            conversion_rate DOUBLE PRECISION,
            roas DOUBLE PRECISION,
            acos DOUBLE PRECISION,

            created_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW()
        );
        """

        insert_sql = f"""
        INSERT INTO public.{table_name} (
            user_id, country, month, year, date,
            sno, sku, asin, ad_type, match_type,
            impressions, clicks, ctr, cpc, spend,
            product_spend, display_spend, brand_spend,
            sale_units, sale_amount,
            sp_ads_sales, sd_ads_sales, sb_ads_sales,
            ads_sales_total, ads_spend_total,
            advertised_unit_sale, other_unit_sale, new_to_brand_sales,
            conversion_rate, roas, acos
        ) VALUES (
            :user_id, :country, :month, :year, :date,
            :sno, :sku, :asin, :ad_type, :match_type,
            :impressions, :clicks, :ctr, :cpc, :spend,
            :product_spend, :display_spend, :brand_spend,
            :sale_units, :sale_amount,
            :sp_ads_sales, :sd_ads_sales, :sb_ads_sales,
            :ads_sales_total, :ads_spend_total,
            :advertised_unit_sale, :other_unit_sale, :new_to_brand_sales,
            :conversion_rate, :roas, :acos
        );
        """

        try:
            # Create the daily table automatically when it is missing.
            db.session.execute(text(create_sql))

            # ✅ if table already exists, add missing columns safely
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS date DATE;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sku TEXT;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS product_spend DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS display_spend DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS brand_spend DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sp_ads_sales DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sd_ads_sales DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS sb_ads_sales DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS ads_sales_total DOUBLE PRECISION DEFAULT 0;'))
            db.session.execute(text(f'ALTER TABLE public.{table_name} ADD COLUMN IF NOT EXISTS ads_spend_total DOUBLE PRECISION DEFAULT 0;'))

            # ✅ recreate full month every time
            db.session.execute(text(f"TRUNCATE TABLE public.{table_name};"))

            params = []
            for r in items:
                params.append({
                    "user_id": user_id,
                    "country": country,
                    "month": month,
                    "year": year,
                    "date": r.get("date"),

                    "sno": r.get("sno"),
                    "sku": r.get("sku"),
                    "asin": r.get("asin"),
                    "ad_type": r.get("ad_type"),
                    "match_type": r.get("match_type"),

                    "impressions": int(r.get("impressions") or 0),
                    "clicks": int(r.get("clicks") or 0),
                    "ctr": float(r.get("ctr") or 0.0),
                    "cpc": float(r.get("cpc") or 0.0),
                    "spend": float(r.get("spend") or 0.0),

                    "product_spend": float(r.get("product_spend") or 0.0),
                    "display_spend": float(r.get("display_spend") or 0.0),
                    "brand_spend": float(r.get("brand_spend") or 0.0),

                    "sale_units": float(r.get("sale_units") or 0.0),
                    "sale_amount": float(r.get("sale_amount") or 0.0),

                    "sp_ads_sales": float(r.get("sp_ads_sales") or 0.0),
                    "sd_ads_sales": float(r.get("sd_ads_sales") or 0.0),
                    "sb_ads_sales": float(r.get("sb_ads_sales") or 0.0),

                    "ads_sales_total": float(r.get("ads_sales_total") or 0.0),
                    "ads_spend_total": float(r.get("ads_spend_total") or 0.0),

                    "advertised_unit_sale": float(r.get("advertised_unit_sale") or 0.0),
                    "other_unit_sale": float(r.get("other_unit_sale") or 0.0),
                    "new_to_brand_sales": float(r.get("new_to_brand_sales") or 0.0),

                    "conversion_rate": float(r.get("conversion_rate") or 0.0),
                    "roas": float(r.get("roas") or 0.0),
                    "acos": float(r.get("acos") or 0.0),
                })

            if params:
                db.session.execute(text(insert_sql), params)

            db.session.commit()

        except Exception:
            db.session.rollback()
            raise

        return jsonify({
            "message": "Daily ads table saved to DB successfully (SP + SD + SB) - SKU-wise only",
            "table_name": f"public.{table_name}",
            "country": country,
            "month": month,
            "year": year,
            "include": sorted(list(include)),
            "count": len(items),
            "items": items,
        }), 200

    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401
    except PermissionError as e:
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


#------------------------------------------  "monthly" | "quarterly" | "yearly" routes ------------------------------------------#

def resolve_date_ranges(period: str, year: int, month: int | None, quarter: int | None):
    """
    Returns list of (start_date: date, end_date: date) ranges.
    - monthly: one range for given month
    - quarterly: three month ranges for the quarter
    - yearly: 12 month ranges
    """
    period = (period or "").lower().strip()
    ranges: list[tuple[date, date]] = []

    if period == "monthly":
        if month is None:
            raise ValueError("month is required for period=monthly")
        if not (1 <= int(month) <= 12):
            raise ValueError("month must be 1..12")
        m = int(month)
        start = date(year, m, 1)
        end = date(year, m, calendar.monthrange(year, m)[1])
        ranges.append((start, end))

    elif period == "quarterly":
        if quarter is None:
            raise ValueError("quarter is required for period=quarterly")
        q = int(quarter)
        if q not in {1, 2, 3, 4}:
            raise ValueError("quarter must be 1..4")
        start_month = (q - 1) * 3 + 1
        for m in range(start_month, start_month + 3):
            start = date(year, m, 1)
            end = date(year, m, calendar.monthrange(year, m)[1])
            ranges.append((start, end))

    elif period == "yearly":
        for m in range(1, 13):
            start = date(year, m, 1)
            end = date(year, m, calendar.monthrange(year, m)[1])
            ranges.append((start, end))

    else:
        raise ValueError('period must be one of: "monthly" | "quarterly" | "yearly"')

    return ranges


@advertisement_api_routes_bp.route("/api/ads/manager/sp_advertised_product_report_period", methods=["POST"])
def sp_advertised_product_report_period():
    """
    New route that supports period-based fetch:
      period = "monthly" | "quarterly" | "yearly"

    Body examples:

    Monthly:
      {"period":"monthly","year":2025,"month":12,"countries":["UK"],"time_unit":"SUMMARY","return_excel":false}

    Quarterly:
      {"period":"quarterly","year":2025,"quarter":4,"countries":["UK"],"time_unit":"SUMMARY","return_excel":false}

    Yearly:
      {"period":"yearly","year":2025,"countries":["UK"],"time_unit":"SUMMARY","return_excel":false}

    Notes:
    - This route still respects Ads retention. If you request old months Amazon may 400.
    - It merges all months in the period into ONE aggregated dataframe and saves to amazon_sponsored_products.
    """
    try:
        user_id = _require_jwt_user_id()
        u = _get_user_row(user_id)

        if not u.amazon_ads_refresh_token:
            return jsonify({"error": "Amazon Ads not connected for this user."}), 400

        data = request.get_json(force=True) or {}
        period = (data.get("period") or "monthly").lower()
        year = int(data.get("year") or 0)
        month = data.get("month")
        quarter = data.get("quarter")

        time_unit = (data.get("time_unit") or "SUMMARY").upper()
        return_excel = bool(data.get("return_excel", False))  # this route returns JSON by default

        wanted_countries = normalize_ads_country_filter(data.get("countries"))

        if year < 2000 or year > 2100:
            return jsonify({"error": "year looks invalid"}), 400
        if time_unit not in {"SUMMARY", "DAILY"}:
            return jsonify({"error": "time_unit must be SUMMARY or DAILY"}), 400

        # ✅ resolve to month ranges
        try:
            date_ranges = resolve_date_ranges(period, year, int(month) if month is not None else None,
                                              int(quarter) if quarter is not None else None)
        except Exception as ve:
            return jsonify({"error": str(ve)}), 400

        access_token = get_ads_access_token_from_refresh(u.amazon_ads_refresh_token)

        top_profiles = list_top_level_profiles_all_regions(access_token)
        manager_profile_id = find_manager_profile_id(top_profiles)

        if manager_profile_id:
            child_by_region = list_child_profiles_all_regions(access_token, manager_profile_id)
        else:
            child_by_region = top_profiles

        all_profiles = flatten_ads_profiles(child_by_region, wanted_countries)

        if not all_profiles:
            return jsonify({"error": "No advertiser profiles found (or country filter removed all)."}), 400

        merged_rows = []

        # ✅ fetch each month-range for each profile
        download_errors = []
        max_workers = int(data.get("max_workers") or 4)

        for start_dt, end_dt in date_ranges:
            start_str = start_dt.isoformat()
            end_str = end_dt.isoformat()

            rows, errors = fetch_report_rows_for_profiles(
                access_token=access_token,
                profiles=all_profiles,
                create_method_name="create_sp_advertised_product_report",
                start_date=start_str,
                end_date=end_str,
                time_unit=time_unit,
                max_wait_seconds=int(data.get("max_wait_seconds") or 1800),
                poll_every_seconds=int(data.get("poll_every_seconds") or 10),
                max_workers=max_workers,
                strict_row_dicts=True,
                row_extra={"_range_start": start_str, "_range_end": end_str},
            )
            merged_rows.extend(rows)
            download_errors.extend(errors)

        if not merged_rows:
            return jsonify({
                "error": "Reports returned no rows",
                "download_errors": download_errors[:50],
            }), 400

        df = pd.DataFrame(merged_rows)

        # numeric safety
        numeric_cols = [
            "impressions", "clicks", "cost", "costPerClick", "clickThroughRate",
            "sales7d", "purchases7d", "unitsSoldClicks7d",
            "acosClicks7d", "roasClicks7d",
            "attributedSalesSameSku7d", "salesOtherSku7d",
            "purchasesSameSku7d", "unitsSoldSameSku7d", "unitsSoldOtherSku7d",
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        # ✅ aggregate period into one "out" table (group by country/profile/campaign/adgroup/sku/asin)
        group_cols = [
            "_country", "_profileId",
            "campaignId", "campaignName",
            "adGroupId", "adGroupName",
            "portfolioId",
            "campaignBudgetCurrencyCode",
            "advertisedSku", "advertisedAsin",
        ]
        for c in group_cols:
            if c not in df.columns:
                df[c] = ""

        g = df.groupby(group_cols, as_index=False).agg({
            "impressions": "sum",
            "clicks": "sum",
            "cost": "sum",
            "sales7d": "sum",
            "purchases7d": "sum",
            "unitsSoldClicks7d": "sum",
            "acosClicks7d": "mean",   # optional - you can recompute later
            "roasClicks7d": "mean",   # optional - you can recompute later
            "attributedSalesSameSku7d": "sum",
            "salesOtherSku7d": "sum",
            "purchasesSameSku7d": "sum",
            "unitsSoldSameSku7d": "sum",
            "unitsSoldOtherSku7d": "sum",
        })

        # recompute ctr/cpc/conv/acos/roas safely based on totals
        g["ctr"] = [_safe_div(c, i) for c, i in zip(g["clicks"], g["impressions"])]
        g["cpc"] = [_safe_div(cost, c) for cost, c in zip(g["cost"], g["clicks"])]
        g["conv_rate_7d"] = [_safe_div(o, c) for o, c in zip(g["purchases7d"], g["clicks"])]
        g["roas"] = [_safe_div(sales, cost) for sales, cost in zip(g["sales7d"], g["cost"])]
        g["acos"] = [_safe_div(cost, sales) for cost, sales in zip(g["cost"], g["sales7d"])]

        # ✅ SAVE: delete + bulk insert into amazon_sponsored_products
        # We save period rows using start_date/end_date = min/max of requested ranges
        sd = min([d[0] for d in date_ranges])
        ed = max([d[1] for d in date_ranges])
        now = datetime.utcnow()

        q = amazon_sponsored_products.query.filter(
            amazon_sponsored_products.user_id == user_id,
            amazon_sponsored_products.start_date == sd,
            amazon_sponsored_products.end_date == ed,
        )
        if wanted_countries:
            q = q.filter(amazon_sponsored_products.country.in_(list(wanted_countries)))

        q.delete(synchronize_session=False)
        db.session.commit()

        rows_to_insert = []
        for rec in g.to_dict(orient="records"):
            rows_to_insert.append({
                "user_id": user_id,
                "created_at": now,
                "updated_at": now,

                "start_date": sd,
                "end_date": ed,
                "country": rec.get("_country") or None,
                "profile_id": str(rec.get("_profileId") or "") or None,

                "portfolio_id": str(rec.get("portfolioId") or "") or None,
                "currency": rec.get("campaignBudgetCurrencyCode") or None,

                "campaign_id": str(rec.get("campaignId") or "") or None,
                "campaign_name": rec.get("campaignName") or None,

                "ad_group_id": str(rec.get("adGroupId") or "") or None,
                "ad_group_name": rec.get("adGroupName") or None,

                "advertised_sku": rec.get("advertisedSku") or None,
                "advertised_asin": rec.get("advertisedAsin") or None,

                "impressions": _to_int(rec.get("impressions")),
                "clicks": _to_int(rec.get("clicks")),

                "ctr": _to_float(rec.get("ctr")),
                "cpc": _to_float(rec.get("cpc")),

                "spend": _to_float(rec.get("cost")),

                "sales_7d": _to_float(rec.get("sales7d")),
                "orders_7d": _to_float(rec.get("purchases7d")),
                "units_7d": _to_float(rec.get("unitsSoldClicks7d")),

                "acos": _to_float(rec.get("acos")),
                "roas": _to_float(rec.get("roas")),
                "conv_rate_7d": _to_float(rec.get("conv_rate_7d")),

                "adv_sku_sales_7d": _to_float(rec.get("attributedSalesSameSku7d")),
                "other_sku_sales_7d": _to_float(rec.get("salesOtherSku7d")),

                "adv_sku_orders_7d": _to_float(rec.get("purchasesSameSku7d")),
                "adv_sku_units_7d": _to_float(rec.get("unitsSoldSameSku7d")),
                "other_sku_units_7d": _to_float(rec.get("unitsSoldOtherSku7d")),
            })

        if rows_to_insert:
            db.session.bulk_insert_mappings(amazon_sponsored_products, rows_to_insert)
            db.session.commit()

        # ✅ JSON items
        items = pd.DataFrame(rows_to_insert).where(pd.notnull(pd.DataFrame(rows_to_insert)), None).to_dict(orient="records")

        return jsonify({
            "message": "Saved Sponsored Products period report rows",
            "period": period,
            "requested_year": year,
            "requested_month": int(month) if month is not None else None,
            "requested_quarter": int(quarter) if quarter is not None else None,
            "start_date": sd.isoformat(),
            "end_date": ed.isoformat(),
            "countries": list(wanted_countries) if wanted_countries else None,
            "rows_saved": len(rows_to_insert),
            "count": len(items),
            "items": items,
        }), 200

    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401
    except PermissionError as e:
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

#------------------------------------------ Sponsored Display Advertised Product Report ------------------------------------------#

@advertisement_api_routes_bp.route("/api/ads/manager/sd_advertised_product_report/sync",methods=["POST"])
def manager_sd_advertised_product_report_sync_one_hit_country_only():
    """
    ONE-HIT (blocking) + COUNTRY ONLY:
    - Requires countries[]: ["UK"] / ["US"] / ["UK","US"]
    - Creates report(s) only for those countries
    - Polls until COMPLETED/SUCCESS or timeout
    - Downloads + saves to DB
    - Never returns 202
    """

    try:
        user_id = _require_jwt_user_id()
        u = _get_user_row(user_id)

        data = request.get_json(force=True) or {}

        start_date = data.get("start_date")
        end_date = data.get("end_date")
        time_unit = (data.get("time_unit") or "SUMMARY").upper()

        # ✅ Increase default window for real "one hit"
        max_wait_seconds = int(data.get("max_wait_seconds") or 1800)
        poll_every_seconds = int(data.get("poll_every_seconds") or 10)

        if time_unit not in {"DAILY", "SUMMARY"}:
            return jsonify({"error": "time_unit must be DAILY or SUMMARY"}), 400
        if not start_date or not end_date:
            return jsonify({"error": "start_date and end_date required"}), 400
        if not u.amazon_ads_refresh_token:
            return jsonify({"error": "Amazon Ads not connected"}), 400

        # ✅ REQUIRE countries
        wanted = normalize_ads_country_filter(data.get("countries"))
        if not wanted:
            return jsonify({"error": "countries[] is required. Example: {\"countries\":[\"UK\"]}"}), 400

        # map countries -> regions (your logic)
        regions_to_use = set()
        if "UK" in wanted:
            regions_to_use.add("EU")
        if "US" in wanted:
            regions_to_use.add("NA")

        if not regions_to_use:
            return jsonify({"error": "Unsupported countries. Use UK and/or US (or GB for UK)."}), 400

        access_token = get_ads_access_token_from_refresh(u.amazon_ads_refresh_token)

        # -----------------------------
        # Find profiles (manager -> child)
        # -----------------------------
        top_profiles = list_top_level_profiles_all_regions(access_token)
        manager_profile_id = find_manager_profile_id(top_profiles)
        child_by_region = (
            list_child_profiles_all_regions(access_token, manager_profile_id)
            if manager_profile_id else top_profiles
        )

        all_profiles = flatten_ads_profiles(child_by_region, wanted, regions=regions_to_use)

        if not all_profiles:
            return jsonify({
                "error": "No advertiser profiles found for requested countries.",
                "countries": sorted(list(wanted)),
                "regions_used": sorted(list(regions_to_use)),
            }), 400

        rows_all, download_errors = fetch_report_rows_for_profiles(
            access_token=access_token,
            profiles=all_profiles,
            create_method_name="create_sd_advertised_product_report",
            start_date=start_date,
            end_date=end_date,
            time_unit=time_unit,
            max_wait_seconds=max_wait_seconds,
            poll_every_seconds=poll_every_seconds,
            max_workers=int(data.get("max_workers") or 4),
            strict_row_dicts=False,
        )

        if not rows_all:
            return jsonify({
                "error": "Report completed but returned no rows",
                "download_errors": download_errors[:50],
            }), 400

        df = pd.DataFrame(rows_all)

        # Delete existing rows safely
        sd_req = _to_date(start_date, strict=True, field_name="start_date")
        ed_req = _to_date(end_date, strict=True, field_name="end_date")

        delete_q = db.session.query(amazon_sponsored_display_advertised_products).filter(
            amazon_sponsored_display_advertised_products.user_id == user_id,
        )

        if time_unit == "DAILY":
            # ✅ DAILY reports save one row per day, so delete the whole requested date range
            delete_q = delete_q.filter(
                amazon_sponsored_display_advertised_products.start_date >= sd_req,
                amazon_sponsored_display_advertised_products.start_date <= ed_req,
            )
        else:
            # ✅ SUMMARY reports save one row for the full requested range
            delete_q = delete_q.filter(
                amazon_sponsored_display_advertised_products.start_date == sd_req,
                amazon_sponsored_display_advertised_products.end_date == ed_req,
            )

        if wanted:
            delete_q = delete_q.filter(
                amazon_sponsored_display_advertised_products.country.in_(list(wanted))
            )

        delete_q.delete(synchronize_session=False)
        db.session.commit()

        # numeric conversions
        for col in ["impressions", "clicks", "cost", "sales", "purchases", "unitsSold"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        def _safe_div(a, b):
            a = float(a or 0.0)
            b = float(b or 0.0)
            return (a / b) if b else 0.0

        now = datetime.utcnow()
        inserts = []

        for rec in df.to_dict("records"):
            cost = _to_float(rec.get("cost"))
            clicks = _to_int(rec.get("clicks"))
            impressions = _to_int(rec.get("impressions"))
            sales = _to_float(rec.get("sales"))

            inserts.append({
                "user_id": user_id,
                "created_at": now,
                "updated_at": now,
                "time_unit": time_unit,
                "start_date": _to_date(rec.get("date") or rec.get("startDate") or start_date),
                "end_date": _to_date(rec.get("date") or rec.get("endDate") or rec.get("startDate") or end_date),

                "country": rec.get("_country"),
                "profile_id": str(rec.get("_profileId") or ""),

                "campaign_id": str(rec.get("campaignId") or ""),
                "campaign_name": rec.get("campaignName"),
                "ad_group_id": str(rec.get("adGroupId") or ""),
                "ad_group_name": rec.get("adGroupName"),

                "advertised_sku": rec.get("promotedSku"),
                "advertised_asin": rec.get("promotedAsin"),

                "currency": rec.get("campaignBudgetCurrencyCode"),

                "impressions": _to_int(rec.get("impressions")),
                "clicks": clicks,
                "spend": cost,

                "cpc": _safe_div(cost, clicks),
                "ctr": _safe_div(clicks, impressions),

                "sales_14d": sales,
                "orders_14d": _to_int(rec.get("purchases")),
                "units_14d": _to_int(rec.get("unitsSold")),

                "acos": _safe_div(cost, sales),
                "roas": _safe_div(sales, cost),
            })

        if inserts:
            db.session.bulk_insert_mappings(amazon_sponsored_display_advertised_products, inserts)
            db.session.commit()

        return jsonify({
            "message": "SD advertised product report synced and saved (one-hit, country-only)",
            "start_date": start_date,
            "end_date": end_date,
            "time_unit": time_unit,
            "countries": sorted(list(wanted)),
            "profiles_used": len(all_profiles),
            "rows_saved": len(inserts),
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@advertisement_api_routes_bp.route("/api/ads/manager/sb_keyword_report", methods=["POST"])
def manager_sb_keyword_report():
    """
    Body:
      {
        "start_date": "YYYY-MM-DD",
        "end_date": "YYYY-MM-DD",
        "time_unit": "SUMMARY" | "DAILY",
        "countries": ["UK","US"],      # optional
        "return_excel": true           # optional (default true)
      }
    """
    try:
        user_id = _require_jwt_user_id()
        u = _get_user_row(user_id)

        data = request.get_json(force=True) or {}
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        time_unit = (data.get("time_unit") or "SUMMARY").upper()
        return_excel = bool(data.get("return_excel", True))

        wanted_countries = normalize_ads_country_filter(data.get("countries"))

        if time_unit not in {"DAILY", "SUMMARY"}:
            return jsonify({"error": "time_unit must be DAILY or SUMMARY"}), 400
        if not start_date or not end_date:
            return jsonify({"error": "start_date and end_date required (YYYY-MM-DD)"}), 400

        # validate dates once
        try:
            sd = _to_date(start_date, strict=True, field_name="start_date")
            ed = _to_date(end_date, strict=True, field_name="end_date")
        except ValueError as ve:
            return jsonify({"error": str(ve)}), 400

        if not u.amazon_ads_refresh_token:
            return jsonify({"error": "Amazon Ads not connected for this user."}), 400

        access_token = get_ads_access_token_from_refresh(u.amazon_ads_refresh_token)

        # manager -> child advertiser profiles
        top_profiles = list_top_level_profiles_all_regions(access_token)
        manager_profile_id = find_manager_profile_id(top_profiles)
        child_by_region = (
            list_child_profiles_all_regions(access_token, manager_profile_id)
            if manager_profile_id else top_profiles
        )

        all_profiles = flatten_ads_profiles(child_by_region, wanted_countries)

        if not all_profiles:
            return jsonify({"error": "No advertiser profiles found (or your country filter removed all)."}), 400

        merged_rows = []
        join_maps = {}
        download_errors = []
        join_warnings = []

        for p in all_profiles:
            profile_id = p.get("profileId")
            if not profile_id:
                continue
            join_maps[str(profile_id)] = {
                "campaign_to_portfolio": {},
                "portfolioid_to_name": {},
            }

        merged_rows, download_errors = fetch_report_rows_for_profiles(
            access_token=access_token,
            profiles=all_profiles,
            create_method_name="create_sb_keyword_report",
            start_date=start_date,
            end_date=end_date,
            time_unit=time_unit,
            max_wait_seconds=int(data.get("max_wait_seconds") or 1800),
            poll_every_seconds=int(data.get("poll_every_seconds") or 10),
            max_workers=int(data.get("max_workers") or 4),
            continue_on_error=True,
        )

        if not merged_rows:
            return jsonify({
                "error": "Reports returned no rows",
                "download_errors": download_errors[:50],
            }), 400

        df = pd.DataFrame(merged_rows)

        # ---------------------------
        # 2) Normalize / compute fields
        # ---------------------------
        # safe numeric conversions (ONLY columns we requested)
        for col in [
            "impressions", "clicks",
            "viewableImpressions",
            "topOfSearchImpressionShare",
            "cost",
        ]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

        # targeting text (keywordText preferred)
        for c in ["keywordText", "targetingText", "targetingExpression"]:
            if c not in df.columns:
                df[c] = ""
        df["__targeting_value"] = (
            df["keywordText"].fillna("").astype(str)
              .where(df["keywordText"].fillna("").astype(str) != "", None)
        )
        df["__targeting_value"] = df["__targeting_value"].fillna(df["targetingText"].fillna("").astype(str))
        df["__targeting_value"] = df["__targeting_value"].fillna(df["targetingExpression"].fillna("").astype(str))
        df["__targeting_value"] = df["__targeting_value"].fillna("")

        # currency
        if "campaignBudgetCurrencyCode" not in df.columns:
            df["campaignBudgetCurrencyCode"] = ""
        df["__currency_value"] = df["campaignBudgetCurrencyCode"].fillna("").astype(str)

        # portfolio join (per profileId + campaignId)
        if "campaignId" not in df.columns:
            df["campaignId"] = ""
        df["__portfolio_name"] = ""

        for pid, sub_idx in df.groupby("_profileId").groups.items():
            maps = join_maps.get(str(pid)) or {}
            campaign_to_portfolio = maps.get("campaign_to_portfolio") or {}
            portfolioid_to_name = maps.get("portfolioid_to_name") or {}

            sub = df.loc[sub_idx]
            portfolio_id_series = sub["campaignId"].astype(str).map(campaign_to_portfolio).fillna("")
            portfolio_name_series = portfolio_id_series.astype(str).map(portfolioid_to_name).fillna("")
            df.loc[sub_idx, "__portfolio_name"] = portfolio_name_series

        # ensure ID columns exist (Amazon often omits keywordId/targetingId for sbTargeting)
        for id_col in ["campaignId", "adGroupId", "keywordId", "targetingId"]:
            if id_col not in df.columns:
                df[id_col] = ""

        # ---------------------------
        # 3) Build output with all available Sponsored Brands columns
        # ---------------------------

        def _num(col_name, default=0.0):
            if col_name in df.columns:
                return pd.to_numeric(df[col_name], errors="coerce").fillna(default)
            return pd.Series([default] * len(df))

        def _txt(col_name, default=""):
            if col_name in df.columns:
                return df[col_name].fillna("").astype(str)
            return pd.Series([default] * len(df))

        def _safe_series_div(a, b):
            a = pd.to_numeric(a, errors="coerce").fillna(0.0)
            b = pd.to_numeric(b, errors="coerce").fillna(0.0).replace({0: pd.NA})
            return (a / b).fillna(0.0)

        out = pd.DataFrame()

        # ✅ IMPORTANT for DAILY reports:
        # If Amazon returns row-level startDate/endDate, keep those.
        # Otherwise fallback to request-level start_date/end_date.
        if time_unit == "DAILY":
            out["Start Date"] = _txt("date").where(_txt("date").str.strip() != "", start_date)
            out["End Date"] = out["Start Date"]
        else:
            out["Start Date"] = _txt("startDate").where(_txt("startDate").str.strip() != "", start_date)
            out["End Date"] = _txt("endDate").where(_txt("endDate").str.strip() != "", out["Start Date"])

        out["Country"] = _txt("_country")
        out["Profile ID"] = _txt("_profileId")

        out["Portfolio name"] = df.get("__portfolio_name", "")
        out["Currency"] = df.get("__currency_value", "")

        out["Campaign ID"] = _txt("campaignId")
        out["Campaign Name"] = _txt("campaignName")
        out["Ad Group ID"] = _txt("adGroupId")
        out["Ad Group Name"] = _txt("adGroupName")

        out["Keyword ID"] = _txt("keywordId")
        out["Targeting ID"] = _txt("targetingId")
        out["Targeting"] = df.get("__targeting_value", "")

        out["Match Type"] = _txt("matchType")
        out["Cost Type"] = _txt("costType")

        out["Impressions"] = _num("impressions").astype(int)
        out["Top-of-search impression share"] = _num("topOfSearchImpressionShare")
        out["Viewable impressions"] = _num("viewableImpressions").astype(int)
        out["Clicks"] = _num("clicks").astype(int)

        out["Click-Thru Rate (CTR)"] = _safe_series_div(out["Clicks"], out["Impressions"]) * 100.0

        out["Spend"] = _num("cost")
        out["Cost Per Click (CPC)"] = _safe_series_div(out["Spend"], out["Clicks"])

        # Extra SB columns if Amazon sends them
        extra_metric_map = {
            "campaignStatus": "Campaign Status",
            "campaignBudgetAmount": "Campaign Budget Amount",
            "campaignBudgetType": "Campaign Budget Type",

            "keywordBid": "Keyword Bid",
            "adKeywordStatus": "Keyword Status",
            "keywordType": "Keyword Type",
            "targetingType": "Targeting Type",

            "viewabilityRate": "Viewability Rate",
            "viewClickThroughRate": "View Click-Through Rate",

            "brandedSearches": "Brand Searches",
            "brandedSearchesClicks": "Brand Searches Clicks",

            "detailPageViews": "Detail Page Views",
            "detailPageViewsClicks": "Detail Page Views Clicks",

            "addToCart": "Add to Cart",
            "addToCartClicks": "Add to Cart Clicks",
            "addToCartRate": "Add to Cart Rate",
            "eCPAddToCart": "eCP Add to Cart",

            "purchases": "Purchases",
            "purchasesClicks": "Purchases Clicks",
            "purchasesPromoted": "Purchases Promoted",

            "sales": "Sales",
            "salesClicks": "Sales Clicks",
            "salesPromoted": "Sales Promoted",

            "unitsSold": "Units Sold",
            "unitsSoldClicks": "Units Sold Clicks",

            "newToBrandPurchases": "New-to-brand Purchases",
            "newToBrandPurchasesClicks": "New-to-brand Purchases Clicks",
            "newToBrandPurchasesPercentage": "New-to-brand Purchases %",
            "newToBrandPurchasesRate": "New-to-brand Purchases Rate",

            "newToBrandSales": "New-to-brand Sales",
            "newToBrandSalesClicks": "New-to-brand Sales Clicks",
            "newToBrandSalesPercentage": "New-to-brand Sales %",

            "newToBrandUnitsSold": "New-to-brand Units Sold",
            "newToBrandUnitsSoldClicks": "New-to-brand Units Sold Clicks",
            "newToBrandUnitsSoldPercentage": "New-to-brand Units Sold %",

            "newToBrandDetailPageViews": "New-to-brand Detail Page Views",
            "newToBrandDetailPageViewsClicks": "New-to-brand Detail Page Views Clicks",
            "newToBrandDetailPageViewRate": "New-to-brand Detail Page View Rate",
            "newToBrandECPDetailPageView": "New-to-brand eCP Detail Page View",

            "video5SecondViews": "Video 5-second views",
            "video5SecondViewRate": "Video 5-second view rate",
            "videoFirstQuartileViews": "Video first quartile views",
            "videoMidpointViews": "Video midpoint views",
            "videoThirdQuartileViews": "Video third quartile views",
            "videoCompleteViews": "Video complete views",
            "videoUnmutes": "Video unmutes",

            "qualifiedBorrows": "Qualified Borrows",
            "qualifiedBorrowsFromClicks": "Qualified Borrows from Clicks",
            "royaltyQualifiedBorrows": "Royalty Qualified Borrows",
            "royaltyQualifiedBorrowsFromClicks": "Royalty Qualified Borrows from Clicks",

            "addToList": "Add to List",
            "addToListFromClicks": "Add to List from Clicks",
        }

        for raw_col, pretty_col in extra_metric_map.items():
            if raw_col in df.columns:
                out[pretty_col] = df[raw_col]

        if "sales" in df.columns:
            sales_series = _num("sales")
            out["Total Advertising Cost of Sales (ACOS)"] = _safe_series_div(out["Spend"], sales_series) * 100.0
            out["Total Return on Advertising Spend (ROAS)"] = _safe_series_div(sales_series, out["Spend"])

        if "purchases" in df.columns:
            out["Conversion Rate"] = _safe_series_div(_num("purchases"), out["Clicks"]) * 100.0

        # Keep every raw Amazon column also, so no column is lost
        already_used_raw_cols = {
            "_profileId", "_country",

            "startDate", "endDate",

            "campaignId", "campaignName", "campaignStatus",
            "campaignBudgetAmount", "campaignBudgetCurrencyCode", "campaignBudgetType",

            "adGroupId", "adGroupName",

            "keywordId", "keywordBid", "adKeywordStatus",
            "keywordText", "keywordType",

            "targetingId", "targetingExpression", "targetingText", "targetingType",

            "matchType", "costType",

            "impressions", "topOfSearchImpressionShare",
            "viewableImpressions", "viewabilityRate",
            "viewClickThroughRate",

            "clicks", "cost",

            "brandedSearches", "brandedSearchesClicks",

            "detailPageViews", "detailPageViewsClicks",

            "addToCart", "addToCartClicks",
            "addToCartRate", "eCPAddToCart",

            "purchases", "purchasesClicks", "purchasesPromoted",

            "sales", "salesClicks", "salesPromoted",

            "unitsSold", "unitsSoldClicks",

            "newToBrandPurchases",
            "newToBrandPurchasesClicks",
            "newToBrandPurchasesPercentage",
            "newToBrandPurchasesRate",

            "newToBrandSales",
            "newToBrandSalesClicks",
            "newToBrandSalesPercentage",

            "newToBrandUnitsSold",
            "newToBrandUnitsSoldClicks",
            "newToBrandUnitsSoldPercentage",

            "newToBrandDetailPageViews",
            "newToBrandDetailPageViewsClicks",
            "newToBrandDetailPageViewRate",
            "newToBrandECPDetailPageView",

            "video5SecondViews",
            "video5SecondViewRate",
            "videoFirstQuartileViews",
            "videoMidpointViews",
            "videoThirdQuartileViews",
            "videoCompleteViews",
            "videoUnmutes",

            "qualifiedBorrows",
            "qualifiedBorrowsFromClicks",
            "royaltyQualifiedBorrows",
            "royaltyQualifiedBorrowsFromClicks",

            "addToList",
            "addToListFromClicks",
        }

        for raw_col in df.columns:
            if raw_col.startswith("__"):
                continue
            if raw_col in already_used_raw_cols:
                continue

            export_col = f"Amazon Raw - {raw_col}"
            if export_col not in out.columns:
                out[export_col] = df[raw_col]

        # cast ints where appropriate
        out["Impressions"] = out["Impressions"].astype(int)
        out["Clicks"] = out["Clicks"].astype(int)
        out["Viewable impressions"] = out["Viewable impressions"].astype(int)

        # ---------------------------
        # 4) Dedupe + stable synthetic IDs (CRITICAL FIX)
        # ---------------------------
        out["_campaignId"] = df["campaignId"].fillna("").astype(str)
        out["_adGroupId"] = df["adGroupId"].fillna("").astype(str)
        out["_keywordId"] = df["keywordId"].fillna("").astype(str)
        out["_targetingId"] = df["targetingId"].fillna("").astype(str)

        out_export = out.drop(
            columns=["_campaignId", "_adGroupId", "_keywordId", "_targetingId"],
            errors="ignore"
        )

        def _stable_targeting_hash(row: pd.Series) -> str:
            parts = [
                row.get("Start Date", ""),
                row.get("End Date", ""),
                row.get("Country", ""),
                row.get("Profile ID", ""),
                row.get("_campaignId", ""),
                row.get("_adGroupId", ""),
                row.get("Campaign Name", ""),
                row.get("Ad Group Name", ""),
                row.get("Targeting", ""),
                row.get("Match Type", ""),
                row.get("Cost Type", ""),
                row.get("Portfolio name", ""),
            ]
            s = "|".join(str(p or "").strip() for p in parts)
            return hashlib.sha1(s.encode("utf-8")).hexdigest()[:24]

        # If Amazon didn’t return keywordId/targetingId, synthesize targetingId
        missing_ids = (
            (out["_keywordId"].str.strip() == "") &
            (out["_targetingId"].str.strip() == "")
        )
        if missing_ids.any():
            out.loc[missing_ids, "_targetingId"] = out.loc[missing_ids].apply(_stable_targeting_hash, axis=1)

        key_cols = [
            "Start Date", "End Date", "Country", "Profile ID",
            "_campaignId", "_adGroupId", "_keywordId", "_targetingId",
            "Match Type", "Cost Type"
        ]

        for c in key_cols:
            out[c] = out[c].fillna("").astype(str)

        out = out.drop_duplicates(subset=key_cols, keep="last").reset_index(drop=True)

        # remove hidden id cols from final export (keep internal if you want)
        out_export = out.drop(columns=["_campaignId", "_adGroupId", "_keywordId", "_targetingId"], errors="ignore")

        # ---------------------------
        # 5) UPSERT into DB (now safe)
        # ---------------------------
        now = datetime.utcnow()

        rows_to_insert = []
        for rec in out.to_dict(orient="records"):
            rows_to_insert.append({
                "user_id": user_id,
                "created_at": now,
                "updated_at": now,
                "time_unit": time_unit,

                # ✅ IMPORTANT for DAILY reports:
                # Save each Amazon row's own date instead of the request-level range.
                "start_date": _to_date(
                    rec.get("Start Date") or start_date,
                    strict=True,
                    field_name="Start Date"
                ),
                "end_date": _to_date(
                    rec.get("End Date") or rec.get("Start Date") or end_date,
                    strict=True,
                    field_name="End Date"
                ),

                "country": (rec.get("Country") or None),
                "profile_id": str(rec.get("Profile ID") or "") or None,

                # IDs (synthetic targeting_id included if Amazon omitted)
                "campaign_id": str(rec.get("_campaignId") or "") or None,
                "ad_group_id": str(rec.get("_adGroupId") or "") or None,
                "keyword_id": str(rec.get("_keywordId") or "") or None,
                "targeting_id": str(rec.get("_targetingId") or "") or None,

                "portfolio_name": (rec.get("Portfolio name") or None),
                "currency": (rec.get("Currency") or None),

                "campaign_name": (rec.get("Campaign Name") or None),
                "ad_group_name": (rec.get("Ad Group Name") or None),

                "targeting": (rec.get("Targeting") or None),
                "match_type": (rec.get("Match Type") or None),
                "cost_type": (rec.get("Cost Type") or None),

                "impressions": _to_int(rec.get("Impressions")),
                "top_of_search_impression_share": _to_float(rec.get("Top-of-search impression share")),
                "viewable_impressions": _to_int(rec.get("Viewable impressions")),

                "clicks": _to_int(rec.get("Clicks")),
                "ctr": _to_float(rec.get("Click-Thru Rate (CTR)")),  # % value
                "spend": _to_float(rec.get("Spend")),
                "cpc": _to_float(rec.get("Cost Per Click (CPC)")),
                "campaign_status": rec.get("Campaign Status") or None,
                "campaign_budget_amount": _to_float(rec.get("Campaign Budget Amount")),
                "campaign_budget_type": rec.get("Campaign Budget Type") or None,

                "keyword_bid": _to_float(rec.get("Keyword Bid")),
                "keyword_status": rec.get("Keyword Status") or None,
                "keyword_type": rec.get("Keyword Type") or None,
                "targeting_type": rec.get("Targeting Type") or None,

                "viewability_rate": _to_float(rec.get("Viewability Rate")),
                "view_click_through_rate": _to_float(rec.get("View Click-Through Rate")),

                "sales": _to_float(rec.get("Sales")),
                "sales_clicks": _to_float(rec.get("Sales Clicks")),
                "sales_promoted": _to_float(rec.get("Sales Promoted")),

                "orders": _to_float(rec.get("Purchases")),
                "orders_clicks": _to_float(rec.get("Purchases Clicks")),
                "orders_promoted": _to_float(rec.get("Purchases Promoted")),

                "units": _to_float(rec.get("Units Sold")),
                "units_clicks": _to_float(rec.get("Units Sold Clicks")),

                "acos": _to_float(rec.get("Total Advertising Cost of Sales (ACOS)")),
                "roas": _to_float(rec.get("Total Return on Advertising Spend (ROAS)")),
                "conversion_rate": _to_float(rec.get("Conversion Rate")),

                "branded_searches": _to_float(rec.get("Brand Searches")),
                "branded_searches_clicks": _to_float(rec.get("Brand Searches Clicks")),

                "detail_page_views": _to_float(rec.get("Detail Page Views")),
                "detail_page_views_clicks": _to_float(rec.get("Detail Page Views Clicks")),

                "add_to_cart": _to_float(rec.get("Add to Cart")),
                "add_to_cart_clicks": _to_float(rec.get("Add to Cart Clicks")),
                "add_to_cart_rate": _to_float(rec.get("Add to Cart Rate")),
                "ecp_add_to_cart": _to_float(rec.get("eCP Add to Cart")),

                "new_to_brand_sales": _to_float(rec.get("New-to-brand Sales")),
                "new_to_brand_sales_clicks": _to_float(rec.get("New-to-brand Sales Clicks")),
                "new_to_brand_sales_percentage": _to_float(rec.get("New-to-brand Sales %")),

                "new_to_brand_purchases": _to_float(rec.get("New-to-brand Purchases")),
                "new_to_brand_purchases_clicks": _to_float(rec.get("New-to-brand Purchases Clicks")),
                "new_to_brand_purchases_percentage": _to_float(rec.get("New-to-brand Purchases %")),
                "new_to_brand_purchases_rate": _to_float(rec.get("New-to-brand Purchases Rate")),

                "new_to_brand_units_sold": _to_float(rec.get("New-to-brand Units Sold")),
                "new_to_brand_units_sold_clicks": _to_float(rec.get("New-to-brand Units Sold Clicks")),
                "new_to_brand_units_sold_percentage": _to_float(rec.get("New-to-brand Units Sold %")),

                "new_to_brand_detail_page_views": _to_float(rec.get("New-to-brand Detail Page Views")),
                "new_to_brand_detail_page_views_clicks": _to_float(rec.get("New-to-brand Detail Page Views Clicks")),
                "new_to_brand_detail_page_view_rate": _to_float(rec.get("New-to-brand Detail Page View Rate")),
                "new_to_brand_ecp_detail_page_view": _to_float(rec.get("New-to-brand eCP Detail Page View")),

                "video_5_second_views": _to_float(rec.get("Video 5-second views")),
                "video_5_second_view_rate": _to_float(rec.get("Video 5-second view rate")),
                "video_first_quartile_views": _to_float(rec.get("Video first quartile views")),
                "video_midpoint_views": _to_float(rec.get("Video midpoint views")),
                "video_third_quartile_views": _to_float(rec.get("Video third quartile views")),
                "video_complete_views": _to_float(rec.get("Video complete views")),
                "video_unmutes": _to_float(rec.get("Video unmutes")),

                "qualified_borrows": _to_float(rec.get("Qualified Borrows")),
                "qualified_borrows_from_clicks": _to_float(rec.get("Qualified Borrows from Clicks")),
                "royalty_qualified_borrows": _to_float(rec.get("Royalty Qualified Borrows")),
                "royalty_qualified_borrows_from_clicks": _to_float(rec.get("Royalty Qualified Borrows from Clicks")),

                "add_to_list": _to_float(rec.get("Add to List")),
                "add_to_list_from_clicks": _to_float(rec.get("Add to List from Clicks")),
            })

        if any(r["start_date"] is None or r["end_date"] is None for r in rows_to_insert):
            return jsonify({"error": "Sanity check failed: start_date/end_date became null before insert"}), 500

        if rows_to_insert:
            table = amazon_sponsored_brands_keywords.__table__
            stmt = insert(table).values(rows_to_insert)

            # conflict key must match your UNIQUE constraint in amazon_sponsored_brands_keywords
            conflict_cols = [
                "user_id", "start_date", "end_date", "country", "profile_id",
                "campaign_id", "ad_group_id", "keyword_id", "targeting_id",
            ]

            update_cols = {
                "updated_at": stmt.excluded.updated_at,
                "time_unit": stmt.excluded.time_unit,
                "currency": stmt.excluded.currency,

                "portfolio_name": stmt.excluded.portfolio_name,
                "campaign_name": stmt.excluded.campaign_name,
                "ad_group_name": stmt.excluded.ad_group_name,
                "targeting": stmt.excluded.targeting,
                "match_type": stmt.excluded.match_type,
                "cost_type": stmt.excluded.cost_type,

                "impressions": stmt.excluded.impressions,
                "top_of_search_impression_share": stmt.excluded.top_of_search_impression_share,
                "viewable_impressions": stmt.excluded.viewable_impressions,

                "clicks": stmt.excluded.clicks,
                "ctr": stmt.excluded.ctr,
                "spend": stmt.excluded.spend,
                "cpc": stmt.excluded.cpc,
                "campaign_status": stmt.excluded.campaign_status,
                "campaign_budget_amount": stmt.excluded.campaign_budget_amount,
                "campaign_budget_type": stmt.excluded.campaign_budget_type,

                "keyword_bid": stmt.excluded.keyword_bid,
                "keyword_status": stmt.excluded.keyword_status,
                "keyword_type": stmt.excluded.keyword_type,
                "targeting_type": stmt.excluded.targeting_type,

                "viewability_rate": stmt.excluded.viewability_rate,
                "view_click_through_rate": stmt.excluded.view_click_through_rate,

                "sales": stmt.excluded.sales,
                "sales_clicks": stmt.excluded.sales_clicks,
                "sales_promoted": stmt.excluded.sales_promoted,

                "orders": stmt.excluded.orders,
                "orders_clicks": stmt.excluded.orders_clicks,
                "orders_promoted": stmt.excluded.orders_promoted,

                "units": stmt.excluded.units,
                "units_clicks": stmt.excluded.units_clicks,

                "acos": stmt.excluded.acos,
                "roas": stmt.excluded.roas,
                "conversion_rate": stmt.excluded.conversion_rate,

                "branded_searches": stmt.excluded.branded_searches,
                "branded_searches_clicks": stmt.excluded.branded_searches_clicks,

                "detail_page_views": stmt.excluded.detail_page_views,
                "detail_page_views_clicks": stmt.excluded.detail_page_views_clicks,

                "add_to_cart": stmt.excluded.add_to_cart,
                "add_to_cart_clicks": stmt.excluded.add_to_cart_clicks,
                "add_to_cart_rate": stmt.excluded.add_to_cart_rate,
                "ecp_add_to_cart": stmt.excluded.ecp_add_to_cart,

                "new_to_brand_sales": stmt.excluded.new_to_brand_sales,
                "new_to_brand_sales_clicks": stmt.excluded.new_to_brand_sales_clicks,
                "new_to_brand_sales_percentage": stmt.excluded.new_to_brand_sales_percentage,

                "new_to_brand_purchases": stmt.excluded.new_to_brand_purchases,
                "new_to_brand_purchases_clicks": stmt.excluded.new_to_brand_purchases_clicks,
                "new_to_brand_purchases_percentage": stmt.excluded.new_to_brand_purchases_percentage,
                "new_to_brand_purchases_rate": stmt.excluded.new_to_brand_purchases_rate,

                "new_to_brand_units_sold": stmt.excluded.new_to_brand_units_sold,
                "new_to_brand_units_sold_clicks": stmt.excluded.new_to_brand_units_sold_clicks,
                "new_to_brand_units_sold_percentage": stmt.excluded.new_to_brand_units_sold_percentage,

                "new_to_brand_detail_page_views": stmt.excluded.new_to_brand_detail_page_views,
                "new_to_brand_detail_page_views_clicks": stmt.excluded.new_to_brand_detail_page_views_clicks,
                "new_to_brand_detail_page_view_rate": stmt.excluded.new_to_brand_detail_page_view_rate,
                "new_to_brand_ecp_detail_page_view": stmt.excluded.new_to_brand_ecp_detail_page_view,

                "video_5_second_views": stmt.excluded.video_5_second_views,
                "video_5_second_view_rate": stmt.excluded.video_5_second_view_rate,
                "video_first_quartile_views": stmt.excluded.video_first_quartile_views,
                "video_midpoint_views": stmt.excluded.video_midpoint_views,
                "video_third_quartile_views": stmt.excluded.video_third_quartile_views,
                "video_complete_views": stmt.excluded.video_complete_views,
                "video_unmutes": stmt.excluded.video_unmutes,

                "qualified_borrows": stmt.excluded.qualified_borrows,
                "qualified_borrows_from_clicks": stmt.excluded.qualified_borrows_from_clicks,
                "royalty_qualified_borrows": stmt.excluded.royalty_qualified_borrows,
                "royalty_qualified_borrows_from_clicks": stmt.excluded.royalty_qualified_borrows_from_clicks,

                "add_to_list": stmt.excluded.add_to_list,
                "add_to_list_from_clicks": stmt.excluded.add_to_list_from_clicks,
            }

            stmt = stmt.on_conflict_do_update(index_elements=conflict_cols, set_=update_cols)
            db.session.execute(stmt)
            db.session.commit()

        # ---------------------------
        # 6) JSON only (optional)
        # ---------------------------
        if not return_excel:
            return jsonify({
                "message": "Saved Sponsored Brands keyword report rows (UPSERT)",
                "rows_saved": len(rows_to_insert),
                "start_date": start_date,
                "end_date": end_date,
                "time_unit": time_unit,
                "countries": sorted(list(wanted_countries)) if wanted_countries else None,
                "excel_columns": list(out_export.columns),
                "amazon_raw_columns": list(df.columns),
                "join_warnings": join_warnings[:50],
                "download_errors": download_errors[:50],
            }), 200

        # ---------------------------
        # 7) Return Excel
        # ---------------------------
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            out_export.to_excel(writer, index=False, sheet_name="SB_Keywords")
            ws = writer.book["SB_Keywords"]

            if "Spend" in out_export.columns:
                spend_col_idx = list(out_export.columns).index("Spend") + 1
                spend_letter = get_column_letter(spend_col_idx)

                cur = ""
                if "Currency" in out_export.columns:
                    non_empty = [c for c in out_export["Currency"].tolist() if str(c).strip()]
                    cur = non_empty[0] if non_empty else ""

                money_fmt = u"£#,##0.00" if str(cur).upper() in {"GBP", "UK", "GB"} else u"#,##0.00"

                for r in range(2, ws.max_row + 1):
                    ws[f"{spend_letter}{r}"].number_format = money_fmt

        output.seek(0)
        filename = f"SB_Keyword_{start_date}_to_{end_date}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    
