from __future__ import annotations

import os, re, io, csv, time, gzip, logging, json, jwt, calendar, requests
from datetime import datetime, date, timezone
from typing import Optional
from dotenv import find_dotenv, load_dotenv
from sqlalchemy.orm import load_only
from flask import request, jsonify, Blueprint, send_file
from sqlalchemy.dialects.postgresql import insert, insert as pg_insert
from sqlalchemy.exc import SQLAlchemyError
from app import db
from sqlalchemy import delete, text
from app.models.user_models import Inventory, CountryProfile, MonthwiseInventory , InventoryAged, InventoryAWD, InventoryAgedHistory
from app.utils.token_utils import get_effective_user_id_from_token
from app.utils.amazon_utils import amazon_client, _apply_region_and_marketplace_from_request
from app.utils.live_bi_utils import generate_inventory_alerts_for_all_skus
from config import Config
from contextlib import contextmanager
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# basic config
# ---------------------------------------------------------------------------

SECRET_KEY = Config.SECRET_KEY

dotenv_path = find_dotenv(filename=".env", usecwd=True)
load_dotenv(dotenv_path, override=True)

logger = logging.getLogger("amazon_sp_api")
logging.basicConfig(level=logging.INFO)

db_url = os.getenv("DATABASE_URL")
db_url1 = os.getenv("DATABASE_ADMIN_URL") or db_url
if not db_url:
    raise RuntimeError("DATABASE_URL is not set")
if not db_url1:
    print("[WARN] DATABASE_ADMIN_URL not set; falling back to DATABASE_URL")

# ---------------------------------------------------------------------
# AMAZON DB ENGINE (amazon_db)
# ---------------------------------------------------------------------

user_engine = create_engine(db_url, pool_pre_ping=True)
admin_engine = create_engine(db_url1, pool_pre_ping=True)

DATABASE_AMAZON_URL = os.getenv("DATABASE_AMAZON_URL") 
if not DATABASE_AMAZON_URL:
    raise RuntimeError("DATABASE_AMAZON_URL is missing in .env")

amazon_engine = create_engine(DATABASE_AMAZON_URL, pool_pre_ping=True)

inventory_bp = Blueprint("inventory", __name__)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _safe_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, dict):
        return sum(int(v or 0) for v in val.values() if isinstance(v, (int, float)))
    try:
        return int(val)
    except Exception:
        return 0


# -------------------------- FBA inventory summaries -------------------------

def _fetch_fba_inventory_summaries(mp: str) -> list[dict]:
    """
    Returns a list of normalized inventory rows shaped for the Inventory table.
    Uses the FBA Inventory Summaries API with details=True and paginates via nextToken.
    """
    rows: list[dict] = []

    params = {
        "granularityType": "Marketplace",
        "granularityId": mp,
        "marketplaceIds": [mp],
        "details": "true",
        "pageSize": 100,
    }

    def _normalize(summary: dict) -> dict:
        s = summary or {}
        det = s.get("inventoryDetails") or {}

        inbound_total = (
            _safe_int(det.get("inboundWorkingQuantity")) +
            _safe_int(det.get("inboundShippedQuantity")) +
            _safe_int(det.get("inboundReceivingQuantity"))
        )

        available_qty = s.get("availableQuantity") or det.get("fulfillableQuantity")
        total_qty = s.get("totalQuantity")
        if total_qty is None:
            total_qty = (
                _safe_int(available_qty)
                + _safe_int(det.get("reservedQuantity"))
                + inbound_total
            )

        candidate_name = (
            s.get("productName")
            or s.get("title")
            or s.get("itemName")
        )

        return {
            "asin": s.get("asin"),
            "seller_sku": s.get("sellerSku"),
            "marketplace_id": mp,
            "product_name": candidate_name,
            "total_quantity": _safe_int(total_qty),
            "inbound_quantity": inbound_total,
            "available_quantity": _safe_int(available_qty),
            "reserved_quantity": _safe_int(det.get("reservedQuantity")),
            "fulfillable_quantity": _safe_int(det.get("fulfillableQuantity")),
            "synced_at": datetime.utcnow(),
            "inventory_age_days": 0,
        }

    res = amazon_client.make_api_call("/fba/inventory/v1/summaries", "GET", params)
    if not res or "error" in res:
        logger.warning("Inventory summaries fetch failed: %s", res)
        return rows

    payload = res.get("payload") or res
    for s in (payload.get("inventorySummaries") or []):
        rows.append(_normalize(s))

    nxt = payload.get("pagination", {}).get("nextToken")
    while nxt:
        page = amazon_client.make_api_call(
            "/fba/inventory/v1/summaries",
            "GET",
            {"nextToken": nxt, "marketplaceIds": [mp], "details": "true"},
        )
        if not page or "error" in page:
            logger.warning("Inventory pagination failed: %s", page)
            break
        p2 = page.get("payload") or page
        for s in (p2.get("inventorySummaries") or []):
            rows.append(_normalize(s))
        nxt = p2.get("pagination", {}).get("nextToken")

    # Enrich names via Catalog if missing
    try:
        _enrich_product_names(rows, mp)
    except Exception as e:
        logger.warning("Product-name enrichment failed: %s", e)

    return rows


# ----------------------------- Catalog enrichment ----------------------------

def _extract_catalog_title(item: dict) -> Optional[str]:
    summaries = (item or {}).get("summaries") or []
    if summaries:
        cand = (
            summaries[0].get("itemName")
            or summaries[0].get("title")
            or summaries[0].get("displayName")
        )
        if cand:
            return str(cand)

    attrs = (item or {}).get("attributes") or {}
    for k in ("item_name", "title", "item_name_en", "item_title"):
        v = attrs.get(k)
        if isinstance(v, list) and v and isinstance(v[0], dict):
            val = v[0].get("value")
            if val:
                return str(val)
        elif isinstance(v, str) and v:
            return v
    return None


def _enrich_product_names(rows: list[dict], mp: str) -> None:
    """Mutates rows in-place: fills product_name using Catalog Items v2022 by ASIN."""
    needed = {r["asin"] for r in rows if r.get("asin") and not r.get("product_name")}
    if not needed:
        return

    BATCH = 20
    asin_to_title: dict[str, str] = {}

    pending = list(needed)
    while pending:
        chunk = pending[:BATCH]
        pending = pending[BATCH:]

        params = {
            "identifiers": ",".join(chunk),
            "identifiersType": "ASIN",
            "marketplaceIds": [mp],
            "includedData": "summaries,attributes",
        }
        res = amazon_client.make_api_call("/catalog/2022-04-01/items", "GET", params)
        if not res or "error" in res:
            logger.warning("Catalog items fetch failed for %s: %s", chunk, res)
            continue

        items = res.get("items") or res.get("payload") or []
        for it in items:
            identifiers = it.get("identifiers") or {}
            asin = identifiers.get("asin") or it.get("asin")
            title = _extract_catalog_title(it)
            if asin and title:
                asin_to_title[asin] = title

    for r in rows:
        if not r.get("product_name") and r.get("asin"):
            r["product_name"] = asin_to_title.get(r["asin"])


def backfill_inventory_product_names(mp: str) -> int:
    """Fill product_name for existing rows with NULL/empty name for a marketplace."""
    missing = (
        db.session.query(Inventory.asin)
        .filter(
            Inventory.marketplace_id == mp,
            (Inventory.product_name.is_(None) | (Inventory.product_name == "")),
        )
        .distinct()
        .all()
    )
    asins = [a for (a,) in missing if a]
    if not asins:
        return 0

    temp_rows = [{"asin": a, "marketplace_id": mp, "product_name": None} for a in asins]
    _enrich_product_names(temp_rows, mp)

    updates = {
        r["asin"]: r.get("product_name") for r in temp_rows if r.get("product_name")
    }
    if not updates:
        return 0

    for asin, name in updates.items():
        db.session.query(Inventory).filter(
            Inventory.marketplace_id == mp,
            Inventory.asin == asin,
        ).update({"product_name": name}, synchronize_session=False)
    db.session.commit()
    return len(updates)


# ---------------------------- Inventory health report ------------------------

def _request_inventory_age_report(mp: str, retry_count: int = 0) -> Optional[str]:
    body = {
        "reportType": "GET_FBA_INVENTORY_PLANNING_DATA",
        "marketplaceIds": [mp],
    }

    try:
        res = amazon_client.make_api_call(
            "/reports/2021-06-30/reports",
            "POST",
            {},
            body,
        )

        if not res:
            raise RuntimeError("Empty response from SP-API")

        if "error" in res:
            error = res.get("error", {})
            msg = str(error)

            # 🚨 1. HARD STOP ON 403
            if "403" in msg or "Unauthorized" in msg or "forbidden" in msg.lower():
                logger.error(
                    "❌ SP-API AUTH ERROR (NO RETRY) | marketplace=%s | error=%s",
                    mp, msg
                )
                raise RuntimeError(
                    "SP-API 403 Unauthorized. Fix Seller Central permissions. DO NOT RETRY."
                )

            # 🔁 2. RETRY ONLY TRANSIENT ERRORS
            transient = [
                "429", "QuotaExceeded", "Too Many Requests",
                "500", "502", "503", "504"
            ]

            if any(t in msg for t in transient):
                if retry_count < 2:
                    wait = 2 ** retry_count
                    logger.warning(
                        "Retrying SP-API report (attempt=%s, wait=%ss)",
                        retry_count + 1, wait
                    )
                    time.sleep(wait)
                    return _request_inventory_age_report(mp, retry_count + 1)

            # ❌ OTHER ERRORS → NO RETRY
            logger.error("Non-retryable SP-API error: %s", msg)
            raise RuntimeError(msg)

        payload = res.get("payload") or res
        report_id = payload.get("reportId")

        if not report_id:
            raise RuntimeError(f"No reportId returned: {res}")

        logger.info("✅ Created inventory health report: %s", report_id)
        return report_id

    except Exception as e:
        logger.exception("Inventory report creation failed")
        raise

def _wait_for_report(report_id: str, timeout_sec: int = 600, poll_interval: int = 20) -> Optional[str]:
    """
    Poll report until DONE, return reportDocumentId.
    Enhanced with better error handling and logging.
    """
    deadline = time.time() + timeout_sec
    attempts = 0

    while time.time() < deadline:
        attempts += 1
        res = amazon_client.make_api_call(
            f"/reports/2021-06-30/reports/{report_id}",
            "GET",
            {},
        )

        if not res or "error" in res:
            logger.warning(
                "Error polling report %s (attempt %d): %s",
                report_id,
                attempts,
                res,
            )
            time.sleep(poll_interval)
            continue

        payload = res.get("payload") or res
        status = payload.get("processingStatus")

        logger.info(
            "Report %s status: %s (attempt %d)",
            report_id,
            status,
            attempts,
        )

        if status == "DONE":
            doc_id = payload.get("reportDocumentId")
            logger.info("Report %s completed successfully, document: %s", report_id, doc_id)
            return doc_id

        if status in ("FATAL", "CANCELLED"):
            logger.error(
                "Report %s ended with status %s. Full response: %s",
                report_id,
                status,
                payload,
            )

            # Log any error details if available
            if "processingEndTime" in payload:
                logger.error("Processing ended at: %s", payload["processingEndTime"])

            return None

        time.sleep(poll_interval)

    logger.warning("Timed out waiting for report %s after %d attempts", report_id, attempts)
    return None


def _download_report_document(report_document_id: str) -> Optional[bytes]:
    """
    Download the report document and return raw bytes (after de-compression if needed).
    Enhanced with better error handling.
    """
    try:
        meta = amazon_client.make_api_call(
            f"/reports/2021-06-30/documents/{report_document_id}",
            "GET",
            {},
        )

        if not meta or "error" in meta:
            logger.warning("Error getting report document meta: %s", meta)
            return None

        payload = meta.get("payload") or meta
        url = payload.get("url")
        compression = payload.get("compressionAlgorithm")

        if not url:
            logger.warning("No URL found in report document metadata")
            return None

        logger.info("Downloading report document from: %s", url[:100] + "...")

        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        content = resp.content

        if compression == "GZIP":
            logger.info("Decompressing GZIP content")
            content = gzip.decompress(content)

        logger.info("Successfully downloaded %d bytes", len(content))
        return content

    except requests.RequestException as e:
        logger.error("Failed to download report document: %s", e)
        return None
    except Exception as e:
        logger.error("Unexpected error downloading report: %s", e)
        return None


def _int(val) -> int:
    try:
        return int(val or 0)
    except Exception:
        return 0


def _compute_inventory_age_from_row(row: dict) -> int:
    """
    Compute a single inventory_age_days bucket for one CSV row from the
    FBA Manage Inventory Health Report (GET_FBA_INVENTORY_PLANNING_DATA).

    Primary buckets (coarse) we expect in this report:

      - inv-age-0-to-90-days
      - inv-age-91-to-180-days
      - inv-age-181-to-270-days
      - inv-age-271-to-365-days
      - inv-age-365-plus-days

    We return the LOWER bound of the oldest non-empty bucket:
      365, 271, 181, 91, or 0.

    If those coarse buckets are missing, we fall back to the older,
    fine-grained columns (0–30, 31–60, 61–90, etc.) and return a
    similar "oldest bucket lower bound" value.
    """

    # --- 1) Prefer the new coarse buckets ---
    c_0_90     = _int(row.get("inv-age-0-to-90-days"))
    c_91_180   = _int(row.get("inv-age-91-to-180-days"))
    c_181_270  = _int(row.get("inv-age-181-to-270-days"))
    c_271_365  = _int(row.get("inv-age-271-to-365-days"))
    c_365_plus = _int(row.get("inv-age-365-plus-days"))

    if any([c_0_90, c_91_180, c_181_270, c_271_365, c_365_plus]):
        # Oldest first
        if c_365_plus > 0:
            return 365
        if c_271_365 > 0:
            return 271
        if c_181_270 > 0:
            return 181
        if c_91_180 > 0:
            return 91
        if c_0_90 > 0:
            return 0
        return 0

    # --- 2) Fallback to the old fine-grained buckets if present ---
    a_0_30   = _int(row.get("inv-age-0-to-30-days"))
    a_31_60  = _int(row.get("inv-age-31-to-60-days"))
    a_61_90  = _int(row.get("inv-age-61-to-90-days"))
    a_91_180 = _int(row.get("inv-age-91-to-180-days"))

    a_181_270 = _int(row.get("inv-age-181-to-270-days"))
    a_271_365 = _int(row.get("inv-age-271-to-365-days"))
    a_181_330 = _int(row.get("inv-age-181-to-330-days"))
    a_331_365 = _int(row.get("inv-age-331-to-365-days"))
    a_365_plus = _int(row.get("inv-age-365-plus-days"))

    # Combine into broader ranges (similar to your previous logic)
    b_0_60    = a_0_30 + a_31_60
    b_61_90   = a_61_90
    b_91_180  = a_91_180
    b_181_330 = a_181_330 + a_181_270 + max(0, a_271_365 - a_331_365)
    b_331_365 = a_331_365
    b_365_plus = a_365_plus

    # Check from oldest -> youngest
    if b_365_plus > 0:
        return 365
    if b_331_365 > 0:
        return 331
    if b_181_330 > 0:
        return 181
    if b_91_180 > 0:
        return 91
    if b_61_90 > 0:
        return 61
    if b_0_60 > 0:
        return 0

    # If all zero, treat as 0 days
    return 0


def _fetch_inventory_age_by_sku(mp: str) -> dict[str, int]:
    """
    Returns {seller_sku: inventory_age_days} using GET_FBA_INVENTORY_PLANNING_DATA
    (FBA Manage Inventory Health Report).

    If the report fails or the structure is unexpected, returns {} and
    we leave inventory_age_days = 0.
    """
    logger.info("Starting inventory health (age) report request for marketplace: %s", mp)

    try:
        try:
            report_id = _request_inventory_age_report(mp)
        except RuntimeError as e:
            return jsonify({
                "success": False,
                "error": str(e),
                "action": "Reauthorize Amazon permissions OR disable aged inventory sync"
            }), 403

        doc_id = _wait_for_report(report_id)
        if not doc_id:
            logger.warning("No reportDocumentId for GET_FBA_INVENTORY_PLANNING_DATA")
            # Optionally: try some fallback like summaries
            return _estimate_age_from_summaries(mp)

        content = _download_report_document(doc_id)
        if not content:
            logger.warning(
                "No content downloaded for inventory health report document %s",
                doc_id,
            )
            return {}

        text = content.decode("utf-8-sig", errors="replace")
        f = io.StringIO(text)
        reader = csv.DictReader(f, delimiter="\t")

        if not reader.fieldnames:
            logger.warning("Inventory health report has no header row")
            return {}

        age_by_sku: dict[str, int] = {}
        rows_count = 0

        for row in reader:
            rows_count += 1
            sku = (
                row.get("sku")
                or row.get("seller-sku")
                or row.get("seller_sku")
            )
            if not sku:
                continue
            age_by_sku[sku] = _compute_inventory_age_from_row(row)

        logger.info(
            "Processed %d rows from inventory health report, mapped %d SKUs",
            rows_count,
            len(age_by_sku),
        )

        if not age_by_sku:
            logger.warning(
                "Parsed %d rows from inventory health report but did not map any sku -> age",
                rows_count,
            )

        return age_by_sku

    except Exception as e:
        logger.error("Exception while fetching inventory health/age: %s", e, exc_info=True)
        return {}


def _estimate_age_from_summaries(mp: str) -> dict[str, int]:
    """
    Fallback method: estimate age based on inventory summary data.
    This is less accurate but better than nothing. Currently a stub.
    """
    logger.info("Using fallback method to estimate inventory age from summaries")
    return {}


# ------------------------------- DB upsert logic -----------------------------

def _upsert_inventory_rows(rows: list[dict], user_id: int | None) -> int:
    if not rows:
        return 0
    for r in rows:
        r["user_id"] = user_id

    stmt = insert(Inventory).values(rows)
    stmt = stmt.on_conflict_do_update(
        constraint="uq_inventory_sku_mkt",
        set_={
            "user_id": stmt.excluded.user_id,
            "asin": stmt.excluded.asin,
            "product_name": stmt.excluded.product_name,
            "total_quantity": stmt.excluded.total_quantity,
            "inbound_quantity": stmt.excluded.inbound_quantity,
            "available_quantity": stmt.excluded.available_quantity,
            "reserved_quantity": stmt.excluded.reserved_quantity,
            "fulfillable_quantity": stmt.excluded.fulfillable_quantity,
            "inventory_age_days": stmt.excluded.inventory_age_days,
            "synced_at": stmt.excluded.synced_at,
        },
    )
    db.session.execute(stmt)
    db.session.commit()
    return len(rows)


# ----------------------------------- Route -----------------------------------

@inventory_bp.route("/amazon_api/inventory", methods=["GET"])
def inventory_all():
    """
    Fetch FBA inventory summaries (with details) for the given marketplace,
    enrich with Catalog titles AND inventory age (via GET_FBA_INVENTORY_PLANNING_DATA
    / FBA Manage Inventory Health Report), and upsert into the `inventory` table.
    """

    # --- auth ---
    auth_header = request.headers.get("Authorization")
    user_id = None
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        try:
            payload, user_id, member_id = get_effective_user_id_from_token(token)
            user_id = payload.get("user_id")
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401

    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    store_in_db = request.args.get("store_in_db", "true").lower() != "false"
    refresh_names = request.args.get("refresh_names", "false").lower() == "true"
    skip_age_report = request.args.get("skip_age_report", "false").lower() == "true"

    # --- 1) fetch inventory summaries ---
    rows = _fetch_fba_inventory_summaries(mp)

    # --- 2) fetch inventory health/age data and attach inventory_age_days ---
    inventory_age_notice = None

    if not skip_age_report:
        try:
            age_by_sku = _fetch_inventory_age_by_sku(mp)
            if not age_by_sku:
                inventory_age_notice = (
                    "Inventory health/age report failed or returned no data. "
                    "This may be due to Amazon API issues. "
                    "Inventory age data will be set to 0. "
                    "You can retry this request later or use skip_age_report=true "
                    "to skip this step."
                )
            else:
                for r in rows:
                    sku = r.get("seller_sku")
                    r["inventory_age_days"] = age_by_sku.get(sku, 0)

        except Exception as e:
            logger.error("Failed to fetch inventory health/age report: %s", e, exc_info=True)
            inventory_age_notice = (
                f"Error fetching inventory age data: {str(e)}. "
                "Inventory age will be set to 0."
            )
    else:
        logger.info("Skipping inventory health/age report as requested")
        inventory_age_notice = "Inventory health/age report skipped per request parameter."

    out = {
        "success": True,
        "marketplace_id": mp,
        "count": len(rows),
        "items": rows,
        "db": {"saved_inventory": 0},
    }
    if inventory_age_notice:
        out["inventory_age_notice"] = inventory_age_notice

    # --- 3) persist ---
    if store_in_db and rows:
        try:
            saved = _upsert_inventory_rows(rows, user_id)
            out["db"]["saved_inventory"] = saved
        except Exception as e:
            logger.exception("Failed to upsert inventory rows")
            out["db"]["error"] = str(e)

    if refresh_names:
        try:
            updated = backfill_inventory_product_names(mp)
            out["db"]["backfilled_names"] = updated
        except Exception as e:
            out["db"]["backfill_error"] = str(e)

    if not rows:
        out["empty_message"] = "No inventory found for this seller account."

    return jsonify(out), 200


# --------------------------------------------- helpers ------------------------------------------

def _safe_float(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except Exception:
        return None


def _parse_snapshot_date(val) -> Optional[date]:
    """
    snapshot-date is typically 'YYYY-MM-DD' (sometimes with time component).
    """
    if not val:
        return None
    try:
        # strip off any time part if present
        val = str(val).split("T", 1)[0]
        return datetime.strptime(val, "%Y-%m-%d").date()
    except Exception:
        return None


def _row_to_inventory_aged(row: dict) -> "InventoryAged":
    """
    Map one TSV row from GET_FBA_INVENTORY_PLANNING_DATA into an InventoryAged instance.
    Assumes InventoryAged model is defined in this module or imported.
    """
    snapshot_date = _parse_snapshot_date(row.get("snapshot-date"))

    sku = (
        row.get("sku")
        or row.get("seller-sku")
        or row.get("seller_sku")
    )

    inv = InventoryAged(
        # basic identifiers
        snapshot_date=snapshot_date,
        sku=sku,
        fnsku=row.get("fnsku"),
        asin=row.get("asin"),
        product_name=row.get("product-name"),  # will be overwritten by our SKU table
        condition=row.get("condition"),

        # quantities & age buckets (main)
        available=_int(row.get("available")),
        pending_removal_quantity=_int(row.get("pending-removal-quantity")),
        inv_age_0_90=_int(row.get("inv-age-0-to-90-days")),
        inv_age_91_180=_int(row.get("inv-age-91-to-180-days")),
        inv_age_181_270=_int(row.get("inv-age-181-to-270-days")),
        inv_age_271_365=_int(row.get("inv-age-271-to-365-days")),
        inv_age_365_plus=(
            _int(row.get("inv-age-365-plus-days"))
            + _int(row.get("inv-age-366-to-455-days"))
            + _int(row.get("inv-age-456-plus-days"))
        ),
        currency=row.get("currency"),

        # shipped units
        units_shipped_t7=_int(row.get("units-shipped-t7")),
        units_shipped_t30=_int(row.get("units-shipped-t30")),
        units_shipped_t60=_int(row.get("units-shipped-t60")),
        units_shipped_t90=_int(row.get("units-shipped-t90")),

        # pricing & alerts
        alert=row.get("alert"),
        your_price=_safe_float(row.get("your-price")),
        sales_price=_safe_float(row.get("sales-price")),
        lowest_price_new_plus_shipping=_safe_float(
            row.get("lowest-price-new-plus-shipping")
        ),
        lowest_price_used=_safe_float(row.get("lowest-price-used")),
        recommended_action=row.get("recommended-action"),
        healthy_inventory_level=_safe_float(row.get("healthy-inventory-level")),
        recommended_sales_price=_safe_float(row.get("recommended-sales-price")),
        recommended_sale_duration_days=_int(row.get("recommended-sale-duration-days")),
        recommended_removal_quantity=_int(row.get("recommended-removal-quantity")),
        estimated_cost_savings_recommended_actions=_safe_float(
            row.get("estimated-cost-savings-of-recommended-actions")
        ),

        sell_through=_safe_float(row.get("sell-through")),

        # volume & storage
        item_volume=_safe_float(row.get("item-volume")),
        volume_unit_measurement=row.get("volume-unit-measurement"),
        storage_type=row.get("storage-type"),
        storage_volume=_safe_float(row.get("storage-volume")),

        # catalog / marketplace
        marketplace=row.get("marketplace"),
        product_group=row.get("product-group"),
        sales_rank=_int(row.get("sales-rank")),

        # supply / excess / cover
        days_of_supply=_safe_float(row.get("days-of-supply")),
        estimated_excess_quantity=_int(row.get("estimated-excess-quantity")),
        weeks_of_cover_t30=_safe_float(row.get("weeks-of-cover-t30")),
        weeks_of_cover_t90=_safe_float(row.get("weeks-of-cover-t90")),

        featuredoffer_price=_safe_float(row.get("featuredoffer-price")),

        sales_shipped_last_7_days=_int(row.get("sales-shipped-last-7-days")),
        sales_shipped_last_30_days=_int(row.get("sales-shipped-last-30-days")),
        sales_shipped_last_60_days=_int(row.get("sales-shipped-last-60-days")),
        sales_shipped_last_90_days=_int(row.get("sales-shipped-last-90-days")),

        # more detailed age buckets
        inv_age_0_30=_int(row.get("inv-age-0-to-30-days")),
        inv_age_31_60=_int(row.get("inv-age-31-to-60-days")),
        inv_age_61_90=_int(row.get("inv-age-61-to-90-days")),
        inv_age_181_330=_int(row.get("inv-age-181-to-330-days")),
        inv_age_331_365=_int(row.get("inv-age-331-to-365-days")),

        estimated_storage_cost_next_month=_safe_float(
            row.get("estimated-storage-cost-next-month")
        ),

        # inbound / reserved / unfulfillable
        inbound_quantity=_int(row.get("inbound-quantity")),
        inbound_working=_int(row.get("inbound-working")),
        inbound_shipped=_int(row.get("inbound-shipped")),
        inbound_received=_int(row.get("inbound-received")),

        total_reserved_quantity=_int(row.get("Total Reserved Quantity")),
        unfulfillable_quantity=_int(row.get("unfulfillable-quantity")),

        qty_charged_ais_241_270=_int(
            row.get("quantity-to-be-charged-ais-241-270-days")
        ),
        est_ais_241_270=_safe_float(row.get("estimated-ais-241-270-days")),

        qty_charged_ais_271_300=_int(
            row.get("quantity-to-be-charged-ais-271-300-days")
        ),
        est_ais_271_300=_safe_float(row.get("estimated-ais-271-300-days")),

        qty_charged_ais_301_330=_int(
            row.get("quantity-to-be-charged-ais-301-330-days")
        ),
        est_ais_301_330=_safe_float(row.get("estimated-ais-301-330-days")),

        qty_charged_ais_331_365=_int(
            row.get("quantity-to-be-charged-ais-331-365-days")
        ),
        est_ais_331_365=_safe_float(row.get("estimated-ais-331-365-days")),

        qty_charged_ais_365_plus=_int(
            row.get("quantity-to-be-charged-ais-365-plus-days")
        ),
        est_ais_365_plus=_safe_float(row.get("estimated-ais-365-plus-days")),

        # historical supply / recommendations
        historical_days_of_supply=_safe_float(row.get("historical-days-of-supply")),
        recommended_ship_in_quantity=_int(
            row.get("Recommended ship-in quantity")
        ),
        recommended_ship_in_date=_parse_snapshot_date(
            row.get("Recommended ship-in date")
        ),
        last_updated_historical_dos=_parse_snapshot_date(
            row.get("Last updated date for Historical Days of Supply")
        ),
        short_term_historical_dos=_safe_float(
            row.get("Short term historical days of supply")
        ),
        long_term_historical_dos=_safe_float(
            row.get("Long term historical days of supply")
        ),
        inventory_age_snapshot_date=_parse_snapshot_date(
            row.get("Inventory age snapshot date")
        ),

        # inventory / reserved at FBA
        inventory_supply_at_fba=_int(row.get("Inventory Supply at FBA")),
        reserved_fc_transfer=_int(row.get("Reserved FC Transfer")),
        reserved_fc_processing=_int(row.get("Reserved FC Processing")),
        reserved_customer_order=_int(row.get("Reserved Customer Order")),
        total_days_of_supply_incl_open_shipments=_safe_float(
            row.get("Total Days of Supply (including units from open shipments)")
        ),
        fc_transfer=_int(row.get("fc-transfer")),

        inv_age_366_455=_int(row.get("inv-age-366-to-455-days")),
        inv_age_456_plus=_int(row.get("inv-age-456-plus-days")),

        deprecated_healthy_inventory_level=_safe_float(
            row.get("DEPRECATED healthy-inventory-level")
        ),

        no_sale_last_6_months=row.get("no-sale-last-6-months"),

        qty_charged_ais_181_210=_int(
            row.get("quantity-to-be-charged-ais-181-210-days")
        ),
        est_ais_181_210=_safe_float(row.get("estimated-ais-181-210-days")),

        qty_charged_ais_211_240=_int(
            row.get("quantity-to-be-charged-ais-211-240-days")
        ),
        est_ais_211_240=_safe_float(row.get("estimated-ais-211-240-days")),

        qty_charged_ais_366_455=_int(
            row.get("quantity-to-be-charged-ais-366-455-days")
        ),
        est_ais_366_455=_safe_float(row.get("estimated-ais-366-455-days")),

        qty_charged_ais_456_plus=_int(
            row.get("quantity-to-be-charged-ais-456-plus-days")
        ),
        est_ais_456_plus=_safe_float(row.get("estimated-ais-456-plus-days")),

        fba_minimum_inventory_level=_int(row.get("fba-minimum-inventory-level")),
        fba_inventory_level_health_status=row.get(
            "fba-inventory-level-health-status"
        ),

        exempted_low_inventory_fee=row.get(
            "Exempted from Low-Inventory-Level fee?"
        ),
        low_inventory_fee_current_week=row.get(
            "Low-Inventory-Level fee applied in current week?"
        ),

        reserved_staging=_int(row.get("Reserved Staging")),

        supplier=row.get("supplier"),
        is_seasonal_next_3_months=row.get("is-seasonal-in-next-3-months"),
        season_name=row.get("season-name"),
        season_start_date=row.get("season-start-date"),
        season_end_date=row.get("season-end-date"),
            )

    return inv


# -------- helper to map SKU -> product_name from public.sku_{user_id}_data_table --------

def _attach_sku_product_names(objs: list["InventoryAged"], user_id: int | None) -> None:
    """
    Overwrite InventoryAged.product_name from public.sku_{user_id}_data_table
    where inventory_aged.sku == sku_uk.
    """
    if not objs or not user_id:
        return

    skus = sorted({obj.sku for obj in objs if obj.sku})
    if not skus:
        return

    sku_table = f"public.sku_{user_id}_data_table"

    placeholders = ", ".join(f":sku{i}" for i in range(len(skus)))
    sql = text(f"""
        SELECT sku_uk, product_name
        FROM {sku_table}
        WHERE sku_uk IN ({placeholders})
    """)

    params = {f"sku{i}": sku for i, sku in enumerate(skus)}
    result = db.session.execute(sql, params)

    mapping = {row.sku_uk: row.product_name for row in result}

    for obj in objs:
        if obj.sku in mapping:
            obj.product_name = mapping[obj.sku]


# ------------------------------- route: sync InventoryAged --------------------------

MARKETPLACE_TO_COUNTRY = {
    "A1F83G8C2ARO7P": "uk",
    "ATVPDKIKX0DER": "us",
    "A2EUQ1WTGCTBG2": "ca",
    "A1PA6795UKMFR9": "de",
    "A13V1IB3VIYZZH": "fr",
    "A1RKKUPIHCS9HS": "es",
    "APJ6JRA9NG5V4": "it",
}

# Ledger Summary View can return multiple North America locations even when
# the requested marketplace is US. Keep each marketplace isolated by the
# report's Location column.
MARKETPLACE_TO_LEDGER_LOCATION = {
    "ATVPDKIKX0DER": "US",
    "A2EUQ1WTGCTBG2": "CA",
    "A1F83G8C2ARO7P": "GB",
}

def _ledger_location_for_marketplace(mp: str | None) -> str | None:
    return MARKETPLACE_TO_LEDGER_LOCATION.get((mp or "").strip())

@inventory_bp.route("/amazon_api/inventory/aged", methods=["GET"])
def sync_inventory_aged():
    """
    Fetch GET_FBA_INVENTORY_PLANNING_DATA and store ONLY the latest snapshot_date
    into inventory_aged.

    Behavior:
    - Parses the report TSV
    - Detects the latest snapshot_date present in the report (treated as "current")
    - Deletes ALL previous snapshot_date data for this user + marketplace
    - Deletes any existing rows for the current snapshot_date (to prevent duplicates)
    - Inserts only rows from the current snapshot_date
    """

    # ---------------- AUTH ----------------
    auth_header = request.headers.get("Authorization")
    user_id = None

    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        try:
            payload, user_id, member_id = get_effective_user_id_from_token(token)
            user_id = payload.get("user_id")
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    # ---------------- MARKETPLACE ----------------
    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)

    logger.info("Starting InventoryAged sync for marketplace %s (user_id=%s)", mp, user_id)

    # ---------------- REQUEST REPORT ----------------
    try:
        report_id = _request_inventory_age_report(mp)
    except RuntimeError as e:
        return jsonify({
            "success": False,
            "error": str(e),
            "action": "Reauthorize Amazon permissions OR disable aged inventory sync"
        }), 403

    doc_id = _wait_for_report(report_id)
    if not doc_id:
        return jsonify({
            "success": False,
            "error": "Inventory health report did not complete",
        }), 502

    content = _download_report_document(doc_id)
    if not content:
        return jsonify({
            "success": False,
            "error": "Failed to download inventory health report document",
        }), 502

    # ---------------- PARSE TSV ----------------
    text_content = content.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text_content)
    reader = csv.DictReader(f, delimiter="\t")

    if not reader.fieldnames:
        return jsonify({
            "success": False,
            "error": "Inventory health report has no header row",
        }), 500

    parsed_objs: list[InventoryAged] = []
    rows_count = 0
    snapshot_dates: set[date] = set()

    for row in reader:
        rows_count += 1

        inv = _row_to_inventory_aged(row)
        if not inv or not inv.sku:
            continue

        inv.user_id = user_id
        inv.marketplace = mp

        parsed_objs.append(inv)

        if inv.snapshot_date:
            snapshot_dates.add(inv.snapshot_date)

    # ---------------- DETERMINE "CURRENT" SNAPSHOT DATE ----------------
    # Prefer the report's latest snapshot_date (best definition of "current")
    if snapshot_dates:
        current_snapshot = max(snapshot_dates)
    else:
        # fallback if Amazon doesn't provide snapshot-date in the file
        current_snapshot = date.today()

    # Keep only current snapshot rows
    objs = [o for o in parsed_objs if o.snapshot_date == current_snapshot]

    logger.info(
        "InventoryAged parsed rows=%d, snapshot_dates=%s, current_snapshot=%s, keeping_rows=%d",
        rows_count,
        sorted([d.isoformat() for d in snapshot_dates]),
        current_snapshot.isoformat(),
        len(objs),
    )

    # ---------------- DELETE OLD DATA (KEEP ONLY CURRENT SNAPSHOT) ----------------
    try:
        # 1) delete all other snapshot dates for this user+marketplace
        db.session.execute(
            delete(InventoryAged).where(
                InventoryAged.user_id == user_id,
                InventoryAged.marketplace == mp,
                InventoryAged.snapshot_date != current_snapshot,
            )
        )

        # 2) delete current snapshot too (so re-sync doesn't duplicate)
        db.session.execute(
            delete(InventoryAged).where(
                InventoryAged.user_id == user_id,
                InventoryAged.marketplace == mp,
                InventoryAged.snapshot_date == current_snapshot,
            )
        )

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.exception("Failed to delete old InventoryAged data")
        return jsonify({
            "success": False,
            "error": f"Failed to cleanup old data: {str(e)}",
        }), 500

    # ---------------- ATTACH PRODUCT NAMES ----------------
    try:
        _attach_sku_product_names(objs, user_id)
    except Exception as e:
        logger.exception("Failed to attach SKU product names: %s", e)

    # ---------------- INSERT NEW DATA ----------------
    try:
        if objs:
            db.session.bulk_save_objects(objs)
            db.session.commit()
        saved = len(objs)
    except Exception as e:
        db.session.rollback()
        logger.exception("Failed to save InventoryAged rows")
        return jsonify({
            "success": False,
            "error": f"Failed to save rows: {str(e)}",
        }), 500

    # ---------------- INVENTORY ALERTS ----------------
    try:
        country = MARKETPLACE_TO_COUNTRY.get(mp, "uk")
        inventory_alerts = generate_inventory_alerts_for_all_skus(
            user_id=user_id,
            country=country,
        )
    except Exception:
        logger.exception("Failed to generate inventory alerts")
        inventory_alerts = {}

    # ---------------- RESPONSE ----------------
    return jsonify({
        "success": True,
        "marketplace_id": mp,
        "report_id": report_id,
        "document_id": doc_id,
        "rows_in_report": rows_count,
        "rows_saved": saved,
        "current_snapshot_date": current_snapshot.isoformat(),
        "inventory_alerts": inventory_alerts,
    }), 200



@inventory_bp.route("/amazon_api/inventory/aged/columns", methods=["GET"])
def get_inventory_aged_selected_columns():
    # --- auth ---
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({"error": "Missing Authorization header"}), 401

    token = auth_header.split(" ")[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = payload.get("user_id")
    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Invalid token payload"}), 401

    # --- filters ---
    marketplace_id = request.args.get("marketplace_id")
    snapshot_date = request.args.get("snapshot_date")  # YYYY-MM-DD
    latest = request.args.get("latest", "0") == "1"

    q = InventoryAged.query.filter(InventoryAged.user_id == user_id)

    if marketplace_id:
        q = q.filter(InventoryAged.marketplace == marketplace_id)

    if latest:
        latest_date = (
            db.session.query(InventoryAged.snapshot_date)
            .filter(InventoryAged.user_id == user_id)
            .order_by(InventoryAged.snapshot_date.desc())
            .limit(1)
            .scalar()
        )
        if latest_date:
            q = q.filter(InventoryAged.snapshot_date == latest_date)

    elif snapshot_date:
        q = q.filter(InventoryAged.snapshot_date == snapshot_date)

    rows = q.order_by(InventoryAged.id.asc()).all()

    # --- ONLY required fields ---
    data = []
    for r in rows:
        data.append({
            "fnsku": getattr(r, "fnsku", None),
            "asin": getattr(r, "asin", None),
            "product-name": getattr(r, "product_name", None),
            "condition": getattr(r, "condition", None),
            "available": getattr(r, "available", 0),
            "pending-removal-quantity": getattr(r, "pending_removal_quantity", 0),
            "inv-age-0-to-90-days": getattr(r, "inv_age_0_90", 0),
            "inv-age-91-to-180-days": getattr(r, "inv_age_91_180", 0),
            "inv-age-181-to-270-days": getattr(r, "inv_age_181_270", 0),
            "inv-age-271-to-365-days": getattr(r, "inv_age_271_365", 0),
            "inv-age-365-plus-days": getattr(r, "inv_age_365_plus", 0),
            "currency": getattr(r, "currency", None),
            "estimated-storage-cost-next-month": getattr(r, "estimated_storage_cost_next_month", 0.0),
            "fc-transfer": getattr(r, "fc_transfer", 0),

            "inv-age-366-to-455-days": getattr(r, "inv_age_366_455", 0),
            "inv-age-456-plus-days": getattr(r, "inv_age_456_plus", 0),

            "DEPRECATED healthy-inventory-level": getattr(
                r, "deprecated_healthy_inventory_level", None
            ),
            "no-sale-last-6-months": getattr(r, "no_sale_last_6_months", None),

            "quantity-to-be-charged-ais-181-210-days": getattr(
                r, "qty_charged_ais_181_210", 0
            ),
            "estimated-ais-181-210-days": getattr(r, "est_ais_181_210", None),

            "quantity-to-be-charged-ais-211-240-days": getattr(
                r, "qty_charged_ais_211_240", 0
            ),
            "estimated-ais-211-240-days": getattr(r, "est_ais_211_240", None),

            "quantity-to-be-charged-ais-366-455-days": getattr(
                r, "qty_charged_ais_366_455", 0
            ),
            "estimated-ais-366-455-days": getattr(r, "est_ais_366_455", None),

            "quantity-to-be-charged-ais-456-plus-days": getattr(
                r, "qty_charged_ais_456_plus", 0
            ),
            "estimated-ais-456-plus-days": getattr(r, "est_ais_456_plus", None),

            "fba-minimum-inventory-level": getattr(
                r, "fba_minimum_inventory_level", 0
            ),
            "fba-inventory-level-health-status": getattr(
                r, "fba_inventory_level_health_status", None
            ),

            "Exempted from Low-Inventory-Level fee?": getattr(
                r, "exempted_low_inventory_fee", None
            ),
            "Low-Inventory-Level fee applied in current week?": getattr(
                r, "low_inventory_fee_current_week", None
            ),

            "Reserved Staging": getattr(r, "reserved_staging", 0),

            "supplier": getattr(r, "supplier", None),
            "is-seasonal-in-next-3-months": getattr(r, "is_seasonal_next_3_months", None),
            "season-name": getattr(r, "season_name", None),
            "season-start-date": getattr(r, "season_start_date", None),
            "season-end-date": getattr(r, "season_end_date", None),
        })

    return jsonify({
        "success": True,
        "count": len(data),
        "data": data
    }), 200


@inventory_bp.route("/country-profile", methods=["POST"])
def upsert_country_profile():
    auth_header = request.headers.get("Authorization")

    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({
            "error": "Authorization token is missing or invalid"
        }), 401

    token = auth_header.split(" ")[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = payload.get("user_id")
    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Invalid user"}), 401

    data = request.get_json(silent=True) or {}

    country = str(data.get("country") or "").strip().lower()
    marketplace = str(data.get("marketplace") or "").strip()

    ship_time_weeks = data.get("ship_time_weeks")
    air_time_weeks = data.get("air_time_weeks")
    stock_unit_weeks = data.get("stock_unit_weeks")

    errors = {}

    if not country:
        errors["country"] = "country is required"

    if not marketplace:
        errors["marketplace"] = "marketplace is required"

    def validate_positive_integer(value, field_name):
        try:
            parsed_value = int(value)

            if parsed_value <= 0:
                raise ValueError

            return parsed_value
        except (TypeError, ValueError):
            errors[field_name] = (
                f"{field_name} must be a positive integer"
            )
            return None

    ship_time_weeks = validate_positive_integer(
        ship_time_weeks,
        "ship_time_weeks",
    )
    air_time_weeks = validate_positive_integer(
        air_time_weeks,
        "air_time_weeks",
    )
    stock_unit_weeks = validate_positive_integer(
        stock_unit_weeks,
        "stock_unit_weeks",
    )

    if errors:
        return jsonify({
            "success": False,
            "errors": errors,
        }), 400

    try:
        profile = CountryProfile.query.filter_by(
            user_id=user_id,
            country=country,
            marketplace=marketplace,
        ).first()

        created = profile is None

        if created:
            profile = CountryProfile(
                user_id=user_id,
                country=country,
                marketplace=marketplace,
                ship_time_weeks=ship_time_weeks,
                air_time_weeks=air_time_weeks,
                stock_unit_weeks=stock_unit_weeks,
            )
            db.session.add(profile)
        else:
            profile.ship_time_weeks = ship_time_weeks
            profile.air_time_weeks = air_time_weeks
            profile.stock_unit_weeks = stock_unit_weeks

        db.session.commit()

        return jsonify({
            "success": True,
            "created": created,
            "profile": {
                "id": profile.id,
                "user_id": profile.user_id,
                "country": profile.country,
                "marketplace": profile.marketplace,
                "ship_time_weeks": profile.ship_time_weeks,
                "air_time_weeks": profile.air_time_weeks,
                "stock_unit_weeks": profile.stock_unit_weeks,
                "ship_alert_threshold_weeks": (
                    int(profile.ship_time_weeks or 0)
                    + int(profile.stock_unit_weeks or 0)
                ),
                "air_alert_threshold_weeks": (
                    int(profile.air_time_weeks or 0)
                    + int(profile.stock_unit_weeks or 0)
                ),
            },
        }), 201 if created else 200

    except SQLAlchemyError as exc:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": "Database error",
            "detail": str(exc),
        }), 500


@inventory_bp.route("/country-profile", methods=["GET"])
def get_country_profile():
    auth_header = request.headers.get("Authorization")

    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({
            "error": "Authorization token is missing or invalid"
        }), 401

    token = auth_header.split(" ")[1]

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = payload.get("user_id")
    except jwt.ExpiredSignatureError:
        return jsonify({"error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Invalid token payload"}), 401

    country = (request.args.get("country") or "").strip().lower()
    marketplace = (request.args.get("marketplace") or "").strip()

    if not country or not marketplace:
        return jsonify({
            "error": "country and marketplace are required"
        }), 400

    profile = CountryProfile.query.filter_by(
        user_id=user_id,
        country=country,
        marketplace=marketplace,
    ).first()

    if not profile:
        return jsonify({
            "exists": False,
            "profile": None,
        }), 200

    return jsonify({
        "exists": True,
        "profile": {
            "id": profile.id,
            "user_id": profile.user_id,
            "country": profile.country,
            "marketplace": profile.marketplace,
            "ship_time_weeks": profile.ship_time_weeks,
            "air_time_weeks": profile.air_time_weeks,
            "stock_unit_weeks": profile.stock_unit_weeks,
            "ship_alert_threshold_weeks": (
                int(profile.ship_time_weeks or 0)
                + int(profile.stock_unit_weeks or 0)
            ),
            "air_alert_threshold_weeks": (
                int(profile.air_time_weeks or 0)
                + int(profile.stock_unit_weeks or 0)
            ),
        },
    }), 200


#------------------------------------------------------------------------------ MonthwiseInventory upsert logic --------------------------------------
def _safe_int(v):
    try:
        if v is None:
            return 0
        s = str(v).strip()
        if s == "":
            return 0
        s = s.replace(",", "")
        return int(float(s))
    except Exception:
        return 0


def _parse_date_str(value: str) -> date:
    """
    Accepts:
      - '2025-10-31'
      - '10/31/2025'
      - '30/11/2025'
      - ISO like '2025-12-01T00:00:00Z'
    """
    value = (value or "").strip()
    if not value:
        raise ValueError("empty date")

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    # ISO timestamps
    try:
        s2 = value.replace("Z", "+00:00")
        return datetime.fromisoformat(s2).date()
    except Exception:
        pass

    raise ValueError(
        f"Invalid date format: {value}. "
        "Use YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY, or ISO timestamp."
    )


def _filter_first_last_dates(rows: list[dict]) -> list[dict]:
    """
    Keep ONLY the first date and last date present in rows.
    """
    if not rows:
        return rows

    dates = sorted({r["date"] for r in rows if isinstance(r.get("date"), date)})
    if not dates:
        return rows

    first_date, last_date = dates[0], dates[-1]
    keep = {first_date, last_date}
    return [r for r in rows if r.get("date") in keep]


def _month_range(year: int, month: int) -> tuple[date, date]:
    if not (1 <= month <= 12):
        raise ValueError("month must be between 1 and 12")
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")

    last_dom = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_dom)


def _quarter_range(year: int, quarter: int) -> tuple[date, date]:
    if quarter not in (1, 2, 3, 4):
        raise ValueError("quarter must be 1, 2, 3, or 4")
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")

    start_month = {1: 1, 2: 4, 3: 7, 4: 10}[quarter]
    end_month = start_month + 2

    start_date = date(year, start_month, 1)
    end_dom = calendar.monthrange(year, end_month)[1]
    end_date = date(year, end_month, end_dom)
    return start_date, end_date


def _year_range(year: int) -> tuple[date, date]:
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")
    return date(year, 1, 1), date(year, 12, 31)


# =============================================================================
# ENRICH PRODUCT NAME
# =============================================================================

def _attach_product_names_to_rows(
    rows: list[dict],
    user_id: int | None,
    country: str | None = None,
    marketplace_id: str | None = None,
) -> None:
    """
    Fill product_name from public.sku_{user_id}_data_table.

    Country mapping:
      US -> sku_us
      UK -> sku_uk
      CA -> sku_canada / sku_ca

    Existing non-empty product names are preserved when the master table
    does not contain a matching value. Matching is case-insensitive and
    ignores surrounding spaces.
    """
    if not rows or not user_id:
        return

    country_key = (country or MARKETPLACE_TO_COUNTRY.get(marketplace_id or "", "")).strip().lower()
    candidates_by_country = {
        "us": ["sku_us", "sku_usa", "sku"],
        "usa": ["sku_us", "sku_usa", "sku"],
        "uk": ["sku_uk", "sku_gb", "sku"],
        "gb": ["sku_uk", "sku_gb", "sku"],
        "ca": ["sku_canada", "sku_ca", "sku"],
        "canada": ["sku_canada", "sku_ca", "sku"],
    }

    table_name = f"sku_{int(user_id)}_data_table"
    if not re.fullmatch(r"[A-Za-z0-9_]+", table_name):
        return

    try:
        column_rows = db.session.execute(text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
        """), {"table_name": table_name}).all()
        available_columns = {row[0] for row in column_rows}
    except Exception:
        logger.exception("Could not inspect SKU master table %s", table_name)
        return

    sku_column = next(
        (c for c in candidates_by_country.get(country_key, [f"sku_{country_key}", "sku"]) if c in available_columns),
        None,
    )
    if not sku_column or "product_name" not in available_columns:
        logger.warning("No usable SKU/product columns found in public.%s for country=%s", table_name, country_key)
        return

    normalized_mskus = sorted({
        str(r.get("msku") or "").strip().upper()
        for r in rows
        if str(r.get("msku") or "").strip()
    })
    if not normalized_mskus:
        return

    placeholders = ", ".join(f":sku{i}" for i in range(len(normalized_mskus)))
    sql = text(f"""
        SELECT
            UPPER(TRIM(CAST("{sku_column}" AS TEXT))) AS normalized_sku,
            product_name
        FROM public."{table_name}"
        WHERE UPPER(TRIM(CAST("{sku_column}" AS TEXT))) IN ({placeholders})
          AND product_name IS NOT NULL
          AND TRIM(CAST(product_name AS TEXT)) <> ''
    """)
    params = {f"sku{i}": sku for i, sku in enumerate(normalized_mskus)}

    try:
        result = db.session.execute(sql, params).mappings().all()
    except Exception:
        logger.exception("Could not read product names from public.%s", table_name)
        return

    mapping = {
        str(row.get("normalized_sku") or "").strip().upper(): str(row.get("product_name") or "").strip()
        for row in result
        if str(row.get("normalized_sku") or "").strip()
        and str(row.get("product_name") or "").strip()
    }

    for row in rows:
        normalized_sku = str(row.get("msku") or "").strip().upper()
        mapped_name = mapping.get(normalized_sku)
        current_name = str(row.get("product_name") or "").strip()
        if mapped_name:
            row["product_name"] = mapped_name
        elif current_name.lower() in {"nan", "none", "null", "[null]", "<na>"}:
            row["product_name"] = None


def _get_sku_product_name_lookup(
    user_id: int | None,
    country: str | None = None,
    marketplace_id: str | None = None,
) -> dict[str, str]:
    if not user_id:
        return {}

    country_key = (country or MARKETPLACE_TO_COUNTRY.get(marketplace_id or "", "")).strip().lower()
    candidates_by_country = {
        "us": ["sku_us", "sku_usa", "sku"],
        "usa": ["sku_us", "sku_usa", "sku"],
        "uk": ["sku_uk", "sku_gb", "sku"],
        "gb": ["sku_uk", "sku_gb", "sku"],
        "ca": ["sku_canada", "sku_ca", "sku"],
        "canada": ["sku_canada", "sku_ca", "sku"],
    }

    table_name = f"sku_{int(user_id)}_data_table"
    if not re.fullmatch(r"[A-Za-z0-9_]+", table_name):
        return {}

    try:
        column_rows = db.session.execute(text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
        """), {"table_name": table_name}).all()
        available_columns = {row[0] for row in column_rows}
    except Exception:
        logger.exception("Could not inspect SKU master table %s", table_name)
        return {}

    sku_column = next(
        (c for c in candidates_by_country.get(country_key, [f"sku_{country_key}", "sku"]) if c in available_columns),
        None,
    )

    if not sku_column or "product_name" not in available_columns:
        return {}

    try:
        result = db.session.execute(text(f"""
            SELECT
                "{sku_column}" AS sku,
                product_name
            FROM public."{table_name}"
            WHERE "{sku_column}" IS NOT NULL
              AND TRIM(CAST("{sku_column}" AS TEXT)) <> ''
              AND product_name IS NOT NULL
              AND TRIM(CAST(product_name AS TEXT)) <> ''
        """)).mappings().all()
    except Exception:
        logger.exception("Could not read SKU product names from public.%s", table_name)
        return {}

    lookup: dict[str, str] = {}
    invalid_product_names = {"", "-", "0", "nan", "none", "null", "undefined", "total", "others", "other skus"}

    for row in result:
        product_name = str(row.get("product_name") or "").strip()
        if product_name.lower() in invalid_product_names:
            continue

        for sku in re.split(r"[,;\n]+", str(row.get("sku") or "")):
            normalized_sku = sku.strip().upper()
            if normalized_sku and normalized_sku not in lookup:
                lookup[normalized_sku] = product_name

    return lookup


# =============================================================================
# UPSERT
# =============================================================================

def _upsert_monthwise_inventory_rows(rows: list[dict], user_id: int | None) -> int:
    """
    Upsert into MonthwiseInventory.
    IMPORTANT: constraint uq_monthwise_inv_key must match model UniqueConstraint.
    """
    if not rows:
        return 0

    # SQLAlchemy multi-row INSERT requires every row to contain the same keys.
    # Normalize product_name for all rows before enrichment. Use the Amazon title
    # only as a fallback; the SKU master mapping below takes priority.
    for r in rows:
        current_name = str(r.get("product_name") or "").strip()
        if current_name.lower() in {"", "nan", "none", "null", "[null]", "<na>"}:
            fallback_title = str(r.get("title") or "").strip()
            r["product_name"] = fallback_title or None
        else:
            r["product_name"] = current_name

    _attach_product_names_to_rows(
        rows,
        user_id,
        marketplace_id=(rows[0].get("marketplace_id") if rows else None),
    )

    for r in rows:
        r["user_id"] = user_id
        # Keep the key present even when no master/title match exists.
        r.setdefault("product_name", None)

    stmt = pg_insert(MonthwiseInventory).values(rows)

    stmt = stmt.on_conflict_do_update(
        constraint="uq_monthwise_inv_key",
        set_={
            "fnsku": stmt.excluded.fnsku,
            "title": stmt.excluded.title,
            "product_name": stmt.excluded.product_name,

            "starting_warehouse_balance": stmt.excluded.starting_warehouse_balance,
            "in_transit_between_warehouses": stmt.excluded.in_transit_between_warehouses,
            "receipts": stmt.excluded.receipts,
            "customer_shipments": stmt.excluded.customer_shipments,
            "customer_returns": stmt.excluded.customer_returns,
            "vendor_returns": stmt.excluded.vendor_returns,
            "warehouse_transfer_in_out": stmt.excluded.warehouse_transfer_in_out,
            "found": stmt.excluded.found,
            "lost": stmt.excluded.lost,
            "damaged": stmt.excluded.damaged,
            "disposed": stmt.excluded.disposed,
            "other_events": stmt.excluded.other_events,
            "ending_warehouse_balance": stmt.excluded.ending_warehouse_balance,
            "unknown_events": stmt.excluded.unknown_events,
            "synced_at": stmt.excluded.synced_at,
        },
    )

    db.session.execute(stmt)
    db.session.commit()
    return len(rows)


# =============================================================================
# FETCH REPORT
# =============================================================================

def _fetch_ledger_summary_rows(
    mp: str,
    start_date: date,
    end_date: date,
    time_period: str = "DAILY",   # keep DAILY so we can do first+last filtering ourselves
) -> list[dict]:
    body = {
        "reportType": "GET_LEDGER_SUMMARY_VIEW_DATA",
        "marketplaceIds": [mp],
        "dataStartTime": start_date.isoformat() + "T00:00:00Z",
        "dataEndTime": end_date.isoformat() + "T23:59:59Z",
        "reportOptions": {
            "aggregateByLocation": "COUNTRY",
            "aggregatedByTimePeriod": time_period,
        },
    }

    create = amazon_client.make_api_call(
        "/reports/2021-06-30/reports",
        method="POST",
        params=None,
        data=body,
    )

    if not create or create.get("error"):
        msg = str(create)

        if "QuotaExceeded" in msg or "429" in msg:
            raise RuntimeError("AMAZON_RATE_LIMIT_EXCEEDED")

        if (
            "status_code': 403" in msg
            or '"status_code": 403' in msg
            or "Unauthorized" in msg
            or "Access to requested resource is denied" in msg
        ):
            raise RuntimeError("CONNECTED_AMAZON_ACCOUNT_ACCESS_DENIED")

        raise RuntimeError(f"Failed to create ledger summary report: {create}")

    create_payload = create.get("payload") or create
    report_id = create_payload.get("reportId")
    if not report_id:
        raise RuntimeError(f"Missing reportId in ledger summary response: {create}")

    # poll
    status = None
    meta_payload = None
    report_meta = None

    for _ in range(60):
        report_meta = amazon_client.make_api_call(
            f"/reports/2021-06-30/reports/{report_id}", "GET", None
        )

        if not report_meta or report_meta.get("error"):
            raise RuntimeError(f"Error polling ledger summary report: {report_meta}")

        meta_payload = report_meta.get("payload") or report_meta
        status = meta_payload.get("processingStatus")

        if status in ("DONE", "FATAL", "CANCELLED"):
            break

        time.sleep(10)

    if status != "DONE":
        raise RuntimeError(f"Report not completed. Status={status}, meta={report_meta}")

    doc_id = meta_payload.get("reportDocumentId")
    if not doc_id:
        raise RuntimeError(f"Missing reportDocumentId in meta: {report_meta}")

    doc = amazon_client.make_api_call(
        f"/reports/2021-06-30/documents/{doc_id}", "GET", None
    )
    if not doc or doc.get("error"):
        raise RuntimeError(f"Error getting report document: {doc}")

    doc_payload = doc.get("payload") or doc
    url = doc_payload.get("url")
    if not url:
        raise RuntimeError(f"Missing URL in report document: {doc_payload}")

    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    raw_bytes = resp.content

    if doc_payload.get("compressionAlgorithm") == "GZIP":
        raw_bytes = gzip.decompress(raw_bytes)

    text_data = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text_data), delimiter="\t")

    rows: list[dict] = []
    for r in reader:
        raw_date = (r.get("Date") or "").strip()
        if not raw_date:
            continue

        try:
            row_date = _parse_date_str(raw_date)
        except ValueError:
            continue

        rows.append(
            {
                "date": row_date,
                "fnsku": (r.get("FNSKU") or r.get("FnSku") or "").strip(),
                "asin": (r.get("ASIN") or "").strip(),
                "msku": (r.get("MSKU") or "").strip(),
                "title": r.get("Title") or "",
                "disposition": (r.get("Disposition") or "").strip(),
                "location": (r.get("Location") or "").strip(),
                "marketplace_id": mp,
                "synced_at": datetime.utcnow(),

                "starting_warehouse_balance": _safe_int(r.get("Starting Warehouse Balance")),
                "in_transit_between_warehouses": _safe_int(r.get("In Transit Between Warehouses")),
                "receipts": _safe_int(r.get("Receipts")),
                "customer_shipments": _safe_int(r.get("Customer Shipments")),
                "customer_returns": _safe_int(r.get("Customer Returns")),
                "vendor_returns": _safe_int(r.get("Vendor Returns")),
                "warehouse_transfer_in_out": _safe_int(r.get("Warehouse Transfer In/Out")),
                "found": _safe_int(r.get("Found")),
                "lost": _safe_int(r.get("Lost")),
                "damaged": _safe_int(r.get("Damaged")),
                "disposed": _safe_int(r.get("Disposed")),
                "other_events": _safe_int(r.get("Other Events")),
                "ending_warehouse_balance": _safe_int(r.get("Ending Warehouse Balance")),
                "unknown_events": _safe_int(r.get("Unknown Events")),
            }
        )
    return rows



# =============================================================================
# ROUTE
# =============================================================================

@inventory_bp.route("/amazon_api/inventory/ledger-summary", methods=["GET"])
def inventory_ledger_summary():
    # ---- Auth ----
    auth_header = request.headers.get("Authorization")
    user_id = None
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        try:
            payload, user_id, member_id = get_effective_user_id_from_token(token)
            user_id = payload.get("user_id")
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401

    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    store_in_db = request.args.get("store_in_db", "true").lower() != "false"

    # If true, filter first+last even for custom start/end range
    keep_first_last = request.args.get("keep_first_last", "false").lower() == "true"

    # ---- Params ----
    date_str = request.args.get("date")
    start_str = request.args.get("start_date")
    end_str = request.args.get("end_date")

    month_param = request.args.get("month")      # "12" or "2025-12"
    quarter_param = request.args.get("quarter")  # "1" or "2025-Q1"
    year_param = request.args.get("year")        # "2025"

    mode = None  # "month" | "quarter" | "year" | None

    try:
        # 1) Month mode
        if month_param and not (date_str or start_str or end_str or quarter_param):
            if "-" in month_param:
                y_str, m_str = month_param.split("-", 1)
                year = int(year_param or y_str)
                month = int(m_str)
            else:
                month = int(month_param)
                year = int(year_param) if year_param else datetime.utcnow().year

            start_date, end_date = _month_range(year, month)
            mode = "month"

        # 2) Quarter mode
        elif quarter_param and not (date_str or start_str or end_str or month_param):
            # Allow "2025-Q1" or "Q1" with ?year=2025 or "1" with ?year=2025
            qp = quarter_param.strip().upper()

            if "-Q" in qp:  # "2025-Q1"
                y_str, q_str = qp.split("-Q", 1)
                year = int(year_param or y_str)
                quarter = int(q_str)
            elif qp.startswith("Q"):  # "Q1"
                quarter = int(qp[1:])
                year = int(year_param) if year_param else datetime.utcnow().year
            else:  # "1"
                quarter = int(qp)
                year = int(year_param) if year_param else datetime.utcnow().year

            start_date, end_date = _quarter_range(year, quarter)
            mode = "quarter"

        # 3) Year mode (full year) -> only when year is provided AND no other date params
        elif year_param and not (date_str or start_str or end_str or month_param or quarter_param):
            year = int(year_param)
            start_date, end_date = _year_range(year)
            mode = "year"

        # 4) Single date
        elif date_str:
            start_date = end_date = _parse_date_str(date_str)

        # 5) Explicit range
        else:
            if not (start_str and end_str):
                return jsonify({
                    "error": (
                        "Provide either ?date=..., both ?start_date= and ?end_date=, "
                        "or ?month=..., or ?quarter=...&year=..., or ?year=YYYY."
                    )
                }), 400

            start_date = _parse_date_str(start_str)
            end_date = _parse_date_str(end_str)
            if end_date < start_date:
                raise ValueError("end_date cannot be before start_date")

    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    # ---- Fetch DAILY from Amazon ----
    try:
        rows = _fetch_ledger_summary_rows(mp, start_date, end_date, time_period="DAILY")

        # Amazon NA ledger reports may include US, CA and MX rows together.
        # For a US request, retain only Location=US; for CA retain only CA.
        expected_location = _ledger_location_for_marketplace(mp)
        if expected_location:
            rows = [
                r for r in rows
                if str(r.get("location") or "").strip().upper() == expected_location
            ]
    except Exception as e:
        logger.exception("Failed to fetch ledger summary")
        msg = str(e)

        if "AMAZON_RATE_LIMIT_EXCEEDED" in msg:
            return jsonify({
                "success": False,
                "error": "Amazon rate limit exceeded. Please try again later."
            }), 429

        if "CONNECTED_AMAZON_ACCOUNT_ACCESS_DENIED" in msg:
            return jsonify({
                "success": False,
                "error": "Amazon access denied. Please reconnect your account."
            }), 403

        return jsonify({
            "success": False,
            "error": msg
        }), 500

    # ---- Optional: Filter ONLY first+last when keep_first_last=true ----
    if rows and keep_first_last:
        rows = _filter_first_last_dates(rows)

        # update response start/end to match filtered data
        dates = sorted({r["date"] for r in rows if isinstance(r.get("date"), date)})
        if dates:
            start_date, end_date = dates[0], dates[-1]
    else:
        # keep full range; still normalize start/end based on returned data if you want
        dates = sorted({r["date"] for r in rows if isinstance(r.get("date"), date)})
        if dates:
            start_date, end_date = dates[0], dates[-1]

    # ---- Format output ----
    display_rows = []
    for r in rows:
        r2 = dict(r)
        if isinstance(r2.get("date"), date):
            r2["date"] = r2["date"].strftime("%m/%d/%Y")
        display_rows.append(r2)

    out = {
        "success": True,
        "marketplace_id": mp,
        "mode": mode or "range",
        "start_date": start_date.strftime("%m/%d/%Y"),
        "end_date": end_date.strftime("%m/%d/%Y"),
        "count": len(display_rows),
        "db": {"saved_rows": 0},
        "items": display_rows,
    }

    # ---- Store to DB ----
    if store_in_db and rows:
        try:
            # 1. First save raw Amazon ledger rows into public.monthwise_inventory
            saved_rows = _upsert_monthwise_inventory_rows(rows, user_id)
            out["db"]["saved_rows"] = saved_rows

            # 2. Then create/update monthly summary table
            country = (
                request.args.get("country")
                or MARKETPLACE_TO_COUNTRY.get(mp, "us")
            ).strip().lower()

            out["db"]["inventorymonthly_tables"] = _create_inventorymonthly_after_fetch(
                user_id=user_id,
                country=country,
                mp=mp,
                rows=rows,
            )

        except Exception as e:
            logger.exception("Failed to save monthwise inventory or create inventorymonthly table")
            out["db"]["error"] = str(e)

    if not display_rows:
        out["empty_message"] = "No ledger summary rows for this date range."

    return jsonify(out), 200

# ------------------------------------------------------------
# Helper: SKU-wise MONTHLY (function only)
# ------------------------------------------------------------

def _unique_names(names: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for name in names:
        safe_name = (name or "").strip().lower()
        if safe_name and safe_name not in seen:
            seen.add(safe_name)
            out.append(safe_name)
    return out


def _skuwise_db_sources():
    return [
        ("user", user_engine),
        ("admin", admin_engine),
        ("amazon", amazon_engine),
    ]


def _read_skuwise_units_rows(
    table_candidates: list[str],
    *,
    select_sql: str,
    where_sql: str,
    order_sql: str,
    params: dict,
) -> dict:
    last_errors: list[str] = []

    for table_name in _unique_names(table_candidates):
        if not re.fullmatch(r"[a-z0-9_]+", table_name):
            continue

        sql = text(f"""
            SELECT
                {select_sql}
            FROM public."{table_name}"
            {where_sql}
            {order_sql}
        """)

        for source_name, engine in _skuwise_db_sources():
            try:
                with engine.connect() as conn:
                    rows = conn.execute(sql, params).mappings().all()

                return {
                    "success": True,
                    "table": table_name,
                    "source": source_name,
                    "rows": rows,
                }
            except Exception as exc:
                if len(last_errors) < 4:
                    last_errors.append(f"{source_name}:{table_name}: {exc}")

    return {
        "success": False,
        "error": "Could not read any SKU-wise units table",
        "candidates": _unique_names(table_candidates),
        "details": " | ".join(last_errors),
    }


def get_skuwise_monthly_from_db(
    user_id: int,
    country: str,
    month_param: str,
    year_param: str,
    sku_filter: str = None,
    product_filter: str = None,
):
    """
    Reads from: public.skuwisemonthly_{user_id}_{country}_{monthName}{year}

    Supports:
      - month=december&year=2025
      - month=12&year=2025
      - month=2025-12 (year inferred unless year= overrides)

    Returns dict:
      { success, country, month, year, table, count, items }
    """
    country = (country or "").strip().lower()
    month_param = (month_param or "").strip().lower()
    year_param = (year_param or "").strip()

    if not country or not month_param:
        return {"success": False, "error": "country and month are required for skuwise_monthly"}

    # ---- parse month/year ----
    try:
        if "-" in month_param:  # "2025-12"
            y_str, m_str = month_param.split("-", 1)
            year = int(year_param or y_str)
            m_int = int(m_str)
            month_name = datetime(year, m_int, 1).strftime("%B").lower()  # "december"
        else:
            if month_param.isdigit():  # "12"
                if not year_param:
                    return {"success": False, "error": "year is required when month is numeric (e.g. month=12&year=2025)"}
                year = int(year_param)
                m_int = int(month_param)
                month_name = datetime(year, m_int, 1).strftime("%B").lower()
            else:  # "december"
                if not year_param:
                    return {"success": False, "error": "year is required when month is a name (e.g. month=december&year=2025)"}
                year = int(year_param)
                month_name = month_param
    except Exception:
        return {"success": False, "error": "Invalid month/year format for skuwise_monthly"}

    # ---- validate (prevent SQL injection) ----
    if not re.fullmatch(r"[a-z0-9_]+", country):
        return {"success": False, "error": "Invalid country value"}
    if not re.fullmatch(r"[a-z0-9_]+", month_name):
        return {"success": False, "error": "Invalid month value"}
    if year < 2000 or year > 2100:
        return {"success": False, "error": "Invalid year"}

    base_table_name = f"skuwisemonthly_{user_id}_{country}_{month_name}{year}"
    table_candidates = [
        base_table_name,
        f"{base_table_name}_table",
        f"skuwisemonthly_{user_id}_{country}_{month_name}_{year}",
        f"skuwisemonthly_{user_id}_{country}_{month_name}_{year}_table",
    ]

    # ---- optional filters ----
    where_clauses = []
    params = {}

    if sku_filter:
        where_clauses.append("sku = :sku")
        params["sku"] = sku_filter

    if product_filter:
        where_clauses.append("product_name = :product_name")
        params["product_name"] = product_filter

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    read_result = _read_skuwise_units_rows(
        table_candidates,
        select_sql="""
            sku,
            product_name,
            quantity,
            return_quantity,
            total_quantity
        """,
        where_sql=where_sql,
        order_sql="ORDER BY product_name ASC, sku ASC",
        params=params,
    )

    if not read_result.get("success"):
        return {
            "success": False,
            "error": f"Could not read table {base_table_name}",
            "details": read_result.get("details"),
            "candidates": read_result.get("candidates"),
        }

    rows = read_result["rows"]

    items = [{
        "sku": r["sku"],
        "product_name": r["product_name"],
        "quantity": int(r["quantity"] or 0),
        "return_quantity": int(r["return_quantity"] or 0),
        "total_quantity": int(r["total_quantity"] or 0),
    } for r in rows]

    return {
        "success": True,
        "country": country,
        "month": month_name,
        "year": year,
        "table": read_result["table"],
        "source": read_result["source"],
        "count": len(items),
        "items": items,
    }


# ------------------------------------------------------------
# Helper: SKU-wise QUARTERLY (function only)
# ------------------------------------------------------------
def get_skuwise_quarterly_from_db(
    user_id: int,
    country: str,
    quarter_param: str,
    year_param: str,
    sku_filter: str = None,
    product_filter: str = None,
):
    """
    Reads from: public.quarter{Q}_{user_id}_{country}_{year}_table
      example: quarter4_1_uk_2025_table

    Supports quarter formats:
      - quarter=4&year=2025
      - quarter=Q4&year=2025
      - quarter=2025-Q4  (year inferred unless year= overrides)

    Returns dict:
      { success, country, quarter, year, table, count, items }
    """
    country = (country or "").strip().lower()
    quarter_param = (quarter_param or "").strip().upper()
    year_param = (year_param or "").strip()

    if not country or not quarter_param:
        return {"success": False, "error": "country and quarter are required for skuwise_quarterly"}

    # ---- parse quarter/year ----
    try:
        if "-Q" in quarter_param:  # "2025-Q4"
            y_str, q_str = quarter_param.split("-Q", 1)
            year = int(year_param or y_str)
            quarter = int(q_str)
        elif quarter_param.startswith("Q"):  # "Q4"
            quarter = int(quarter_param[1:])
            if not year_param:
                return {"success": False, "error": "year is required when quarter is like Q4 (e.g. quarter=Q4&year=2025)"}
            year = int(year_param)
        else:  # "4"
            quarter = int(quarter_param)
            if not year_param:
                return {"success": False, "error": "year is required when quarter is numeric (e.g. quarter=4&year=2025)"}
            year = int(year_param)
    except Exception:
        return {"success": False, "error": "Invalid quarter/year format for skuwise_quarterly"}

    if quarter not in (1, 2, 3, 4):
        return {"success": False, "error": "quarter must be 1-4"}

    # ---- validate (prevent SQL injection) ----
    if not re.fullmatch(r"[a-z0-9_]+", country):
        return {"success": False, "error": "Invalid country value"}
    if year < 2000 or year > 2100:
        return {"success": False, "error": "Invalid year"}

    base_table_name = f"quarter{quarter}_{user_id}_{country}_{year}"
    table_candidates = [
        f"{base_table_name}_table",
        base_table_name,
    ]

    # ---- optional filters ----
    where_clauses = []
    params = {}

    if sku_filter:
        where_clauses.append("sku = :sku")
        params["sku"] = sku_filter

    if product_filter:
        where_clauses.append("product_name = :product_name")
        params["product_name"] = product_filter

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    read_result = _read_skuwise_units_rows(
        table_candidates,
        select_sql="""
            product_name,
            sku,
            quantity,
            return_quantity,
            total_quantity
        """,
        where_sql=where_sql,
        order_sql="ORDER BY product_name ASC, sku ASC",
        params=params,
    )

    if not read_result.get("success"):
        return {
            "success": False,
            "error": f"Could not read table {base_table_name}_table",
            "details": read_result.get("details"),
            "candidates": read_result.get("candidates"),
        }

    rows = read_result["rows"]

    items = []
    for r in rows:
        items.append({
            "product_name": r.get("product_name"),
            "sku": r.get("sku"),
            "quantity": int(r.get("quantity") or 0),
            "return_quantity": int(r.get("return_quantity") or 0),
            "total_quantity": int(r.get("total_quantity") or 0),
        })

    return {
        "success": True,
        "country": country,
        "quarter": quarter,
        "year": year,
        "table": read_result["table"],
        "source": read_result["source"],
        "count": len(items),
        "items": items,
    }


# ------------------------------------------------------------
# Helper: SKU-wise YEARLY (function only)
# ------------------------------------------------------------
def get_skuwise_yearly_from_db(
    user_id: int,
    country: str,
    year_param: str,
    sku_filter: str = None,
    product_filter: str = None,
):
    """
    Reads from: public.skuwiseyearly_{user_id}_{country}_{year}_table
      example: skuwiseyearly_1_uk_2025_table

    Returns dict:
      { success, country, year, table, count, items }
    """
    country = (country or "").strip().lower()
    year_param = (year_param or "").strip()

    if not country or not year_param:
        return {"success": False, "error": "country and year are required for skuwise_yearly"}

    try:
        year = int(year_param)
    except Exception:
        return {"success": False, "error": "Invalid year format for skuwise_yearly"}

    # ---- validate (prevent SQL injection) ----
    if not re.fullmatch(r"[a-z0-9_]+", country):
        return {"success": False, "error": "Invalid country value"}
    if year < 2000 or year > 2100:
        return {"success": False, "error": "Invalid year"}

    base_table_name = f"skuwiseyearly_{user_id}_{country}_{year}"
    table_candidates = [
        f"{base_table_name}_table",
        base_table_name,
    ]

    # ---- optional filters ----
    where_clauses = []
    params = {}

    if sku_filter:
        where_clauses.append("sku = :sku")
        params["sku"] = sku_filter

    if product_filter:
        where_clauses.append("product_name = :product_name")
        params["product_name"] = product_filter

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    read_result = _read_skuwise_units_rows(
        table_candidates,
        select_sql="""
            product_name,
            sku,
            quantity,
            return_quantity,
            total_quantity
        """,
        where_sql=where_sql,
        order_sql="ORDER BY product_name ASC, sku ASC",
        params=params,
    )

    if not read_result.get("success"):
        return {
            "success": False,
            "error": f"Could not read table {base_table_name}_table",
            "details": read_result.get("details"),
            "candidates": read_result.get("candidates"),
        }

    rows = read_result["rows"]

    items = []
    for r in rows:
        items.append({
            "product_name": r.get("product_name"),
            "sku": r.get("sku"),
            "quantity": int(r.get("quantity") or 0),
            "return_quantity": int(r.get("return_quantity") or 0),
            "total_quantity": int(r.get("total_quantity") or 0),
        })

    return {
        "success": True,
        "country": country,
        "year": year,
        "table": read_result["table"],
        "source": read_result["source"],
        "count": len(items),
        "items": items,
    }


@contextmanager
def amazon_conn():
    conn = amazon_engine.connect()
    trans = conn.begin()
    try:
        yield conn
        trans.commit()
    except Exception:
        trans.rollback()
        raise
    finally:
        conn.close()


# ---------------------------------------------------------------------
# DATE HELPERS
# ---------------------------------------------------------------------

def _parse_date_str(value: str) -> date:
    value = (value or "").strip()
    if not value:
        raise ValueError("empty date")

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    try:
        s2 = value.replace("Z", "+00:00")
        return datetime.fromisoformat(s2).date()
    except Exception:
        pass

    raise ValueError(
        f"Invalid date format: {value}. "
        "Use YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY, or ISO timestamp."
    )


def _month_range(year: int, month: int) -> tuple[date, date]:
    if not (1 <= month <= 12):
        raise ValueError("month must be between 1 and 12")
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")
    last_dom = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_dom)


def _quarter_range(year: int, quarter: int) -> tuple[date, date]:
    if quarter not in (1, 2, 3, 4):
        raise ValueError("quarter must be 1, 2, 3, or 4")
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")
    start_month = {1: 1, 2: 4, 3: 7, 4: 10}[quarter]
    end_month = start_month + 2
    start_date = date(year, start_month, 1)
    end_dom = calendar.monthrange(year, end_month)[1]
    end_date = date(year, end_month, end_dom)
    return start_date, end_date


def _year_range(year: int) -> tuple[date, date]:
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")
    return date(year, 1, 1), date(year, 12, 31)


# ---------------------------------------------------------------------
# AUTH (same pattern you already use)
# ---------------------------------------------------------------------

def _get_user_id_from_bearer() -> int | None:
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    token = auth_header.split(" ")[1]
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        return payload.get("user_id")
    except Exception:
        return None


# ---------------------------------------------------------------------
# SAFE IDENTIFIERS (for dynamic table names)
# ---------------------------------------------------------------------

def _safe_ident(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        raise ValueError("invalid identifier")
    return s


def _monthly_table_name(user_id: int, country: str, month: int, year: int) -> str:
    return f"inventorymonthly_{user_id}_{_safe_ident(country)}_{month:02d}_{year}"


def _quarterly_table_name(user_id: int, country: str, quarter: int, year: int) -> str:
    return f"inventoryquarterly_{user_id}_{_safe_ident(country)}_q{quarter}_{year}"


def _yearly_table_name(user_id: int, country: str, year: int) -> str:
    return f"inventoryyearly_{user_id}_{_safe_ident(country)}_{year}"


# ---------------------------------------------------------------------
# SOURCE TABLE CHECK (your real table is public.monthwise_inventory)
# ---------------------------------------------------------------------

def _table_exists(conn, schema: str, table: str) -> bool:
    q = text("""
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = :schema AND table_name = :table
        LIMIT 1
    """)
    return conn.execute(q, {"schema": schema, "table": table}).first() is not None


def _get_source_table(conn) -> str:
    # IMPORTANT: your table (seen in pgAdmin) is monthwise_inventory
    if _table_exists(conn, "public", "monthwise_inventory"):
        return "public.monthwise_inventory"
    # fallback if someone created a different name
    if _table_exists(conn, "public", "monthwiseinventory"):
        return "public.monthwiseinventory"
    raise RuntimeError("Source table not found in amazon_db (expected public.monthwise_inventory)")


# ---------------------------------------------------------------------
# AGGREGATION (GROUP BY MSKU) + GRAND TOTAL
# ---------------------------------------------------------------------

def _aggregate_from_monthwise_inventory(conn, user_id: int, mp: str, start_date: date, end_date: date) -> list[dict]:
    src = _get_source_table(conn)
    location_code = _ledger_location_for_marketplace(mp)
    location_filter_sql = (
        "AND UPPER(TRIM(COALESCE(mi.location, ''))) = :location_code"
        if location_code else ""
    )

    sql = text(f"""
        SELECT
            mi.msku AS msku,
            COALESCE(
                MAX(NULLIF(TRIM(mi.product_name), '')),
                MAX(NULLIF(TRIM(mi.title), '')),
                mi.msku
            ) AS product_name,

            SUM(COALESCE(mi.receipts, 0))           AS sum_receipts,
            SUM(COALESCE(mi.customer_shipments, 0)) AS sum_customer_shipments,
            SUM(COALESCE(mi.customer_returns, 0))   AS sum_customer_returns,
            SUM(COALESCE(mi.vendor_returns, 0))     AS sum_vendor_returns,
            SUM(COALESCE(mi.found, 0))              AS sum_found,
            SUM(COALESCE(mi.lost, 0))               AS sum_lost,
            SUM(COALESCE(mi.damaged, 0))            AS sum_damaged,
            SUM(COALESCE(mi.disposed, 0))           AS sum_disposed,

            -- ✅ NEW columns
            SUM(COALESCE(mi.in_transit_between_warehouses, 0)) AS sum_in_transit_between_warehouses,
            SUM(COALESCE(mi.warehouse_transfer_in_out, 0))     AS sum_warehouse_transfer_in_out,
            SUM(COALESCE(mi.other_events, 0))                  AS sum_other_events,
            SUM(COALESCE(mi.unknown_events, 0))                AS sum_unknown_events,

            -- ✅ FIRST DAY snapshots
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'DEFECTIVE')
                AS defective_sum_first,
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'SELLABLE')
                AS sellable_sum_first,
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'WAREHOUSE_DAMAGED')
                AS warehouse_damaged_sum_first,
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'EXPIRED')
                AS expired_sum_first,
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'CUSTOMER_DAMAGED')
                AS customer_damaged_sum_first,
            SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'DISTRIBUTOR_DAMAGED')
                AS distributor_damaged_sum_first,

            -- ✅ LAST DAY snapshots
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'DEFECTIVE')
                AS defective_sum_last,
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'SELLABLE')
                AS sellable_sum_last,
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'WAREHOUSE_DAMAGED')
                AS warehouse_damaged_sum_last,
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'EXPIRED')
                AS expired_sum_last,
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'CUSTOMER_DAMAGED')
                AS customer_damaged_sum_last,
            SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'DISTRIBUTOR_DAMAGED')
                AS distributor_damaged_sum_last,

            -- ✅ Derived totals
            (
                COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'SELLABLE'), 0)
              + COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'DEFECTIVE'), 0)
              + COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'WAREHOUSE_DAMAGED'), 0)
              + COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'EXPIRED'), 0)
              + COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'CUSTOMER_DAMAGED'), 0)
              + COALESCE(SUM(COALESCE(mi.starting_warehouse_balance, 0)) FILTER (WHERE mi.date = :start_date AND mi.disposition = 'DISTRIBUTOR_DAMAGED'), 0)
            ) AS beginning_total,

            COALESCE(SUM(COALESCE(mi.receipts, 0)), 0) AS transit_total,

            -- Displayed Other Items formula (same as the Excel columns):
            -- Disposed + Damaged + Unknown + Other Events + Vendor Return
            -- + Lost - Found. Warehouse Transfer is NOT an Other Items column.
            (
                ABS(COALESCE(SUM(COALESCE(mi.disposed, 0)), 0))
              + ABS(COALESCE(SUM(COALESCE(mi.damaged, 0)), 0))
              + ABS(COALESCE(SUM(COALESCE(mi.unknown_events, 0)), 0))
              + ABS(COALESCE(SUM(COALESCE(mi.other_events, 0)), 0))
              + ABS(COALESCE(SUM(COALESCE(mi.vendor_returns, 0)), 0))
              + ABS(COALESCE(SUM(COALESCE(mi.lost, 0)), 0))
              - ABS(COALESCE(SUM(COALESCE(mi.found, 0)), 0))
            ) AS other_total,

            (
            COALESCE(SUM(COALESCE(mi.customer_shipments, 0)), 0)
            + COALESCE(SUM(COALESCE(mi.customer_returns, 0)), 0)
            ) AS sold_total,

            (
                COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'SELLABLE'), 0)
              + COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'DEFECTIVE'), 0)
              + COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'WAREHOUSE_DAMAGED'), 0)
              + COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'EXPIRED'), 0)
              + COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'CUSTOMER_DAMAGED'), 0)
              + COALESCE(SUM(COALESCE(mi.ending_warehouse_balance, 0)) FILTER (WHERE mi.date = :end_date AND mi.disposition = 'DISTRIBUTOR_DAMAGED'), 0)
            ) AS ending_total,

            (
                COALESCE(
                    SUM(COALESCE(mi.starting_warehouse_balance, 0))
                    FILTER (
                        WHERE mi.date = :start_date
                        AND mi.disposition IN (
                            'SELLABLE','DEFECTIVE','WAREHOUSE_DAMAGED',
                            'EXPIRED','CUSTOMER_DAMAGED','DISTRIBUTOR_DAMAGED'
                        )
                    ),
                0)
                + COALESCE(SUM(COALESCE(mi.receipts, 0)), 0)

                -- Reconciliation must use Amazon's raw signed movements.
                -- Include Warehouse Transfer here even though it is not part of
                -- the displayed other_total column.
                + (
                    COALESCE(SUM(COALESCE(mi.disposed, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.damaged, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.unknown_events, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.other_events, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.vendor_returns, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.lost, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.found, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.warehouse_transfer_in_out, 0)), 0)
                )

                -- sold_total is stored with Amazon's original sign.
                -- Shipments are negative and returns are positive, so ADD the
                -- signed sold_total to match Excel: Beginning + Transit
                -- - Other Items - Net Units Sold - Ending Inventory.
                + (
                    COALESCE(SUM(COALESCE(mi.customer_shipments, 0)), 0)
                  + COALESCE(SUM(COALESCE(mi.customer_returns, 0)), 0)
                )

                - COALESCE(
                    SUM(COALESCE(mi.ending_warehouse_balance, 0))
                    FILTER (
                        WHERE mi.date = :end_date
                        AND mi.disposition IN (
                            'SELLABLE','DEFECTIVE','WAREHOUSE_DAMAGED',
                            'EXPIRED','CUSTOMER_DAMAGED','DISTRIBUTOR_DAMAGED'
                        )
                    ),
                0)
            ) AS difference_total

        FROM {src} mi
        WHERE mi.user_id = :user_id
          AND mi.marketplace_id = :mp
          {location_filter_sql}
          AND mi.date >= :start_date
          AND mi.date <= :end_date
          AND mi.msku IS NOT NULL
          AND NULLIF(TRIM(mi.msku), '') IS NOT NULL
        GROUP BY mi.msku
        ORDER BY mi.msku
    """)

    rows = conn.execute(sql, {
        "user_id": user_id,
        "mp": mp,
        "location_code": location_code,
        "start_date": start_date,
        "end_date": end_date,
    }).mappings().all()

    return [dict(r) for r in rows]


def _attach_awd_quantities_to_rows(
    conn,
    items: list[dict],
    user_id: int,
    marketplace_id: str,
    period_start: date,
    period_end: date,
    country: str | None = None,
) -> None:
    """Attach open AWD inbound shipment quantities to monthly SKU summary rows.

    Values come from public.inventory_awd_inbound_shipments and are matched by:
      user_id + marketplace_id + normalized SKU + ship_by date in the selected month.
    Missing AWD SKUs receive zero values.
    """
    if not items:
        return

    _ensure_awd_inbound_shipments_table(conn)

    awd_rows = conn.execute(text("""
        SELECT
            UPPER(TRIM(COALESCE(sku, ''))) AS normalized_sku,
            SUM(COALESCE(expected_unit_quantity, 0)) AS total_inbound_quantity
        FROM public.inventory_awd_inbound_shipments
        WHERE user_id = :user_id
          AND marketplace_id = :marketplace_id
          AND TRIM(COALESCE(sku, '')) <> ''
          AND COALESCE(expected_unit_quantity, 0) > 0
          AND UPPER(REPLACE(COALESCE(shipment_status, ''), '-', '_')) NOT IN ('CANCELLED', 'CLOSED', 'DELIVERED')
          AND ship_by IS NOT NULL
          AND CAST(ship_by AS DATE) >= :period_start
          AND CAST(ship_by AS DATE) <= :period_end
        GROUP BY UPPER(TRIM(COALESCE(sku, '')))
    """), {
        "user_id": int(user_id),
        "marketplace_id": marketplace_id,
        "period_start": period_start,
        "period_end": period_end,
    }).mappings().all()

    awd_by_sku = {
        str(row["normalized_sku"]): row
        for row in awd_rows
        if row.get("normalized_sku")
    }

    # AWD is only valid for explicitly supported marketplaces (currently US).
    # Never read AWD/on-hand values from a UK currentinventory table because those
    # columns may contain stale US AWD data copied during an earlier inventory sync.
    if marketplace_id in AWD_SUPPORTED_MARKETPLACES:
        current_inventory_awd_by_sku = _fetch_current_inventory_awd_by_sku(
            conn=conn,
            user_id=user_id,
            country=country,
            snapshot_date=period_end,
            marketplace_id=marketplace_id,
        )
        awd_inventory_by_sku = _fetch_awd_inventory_by_sku(conn, user_id, marketplace_id)
    else:
        # Non-AWD marketplace (for example UK): remove any stale AWD values
        # that may have been copied into the currentinventory table, including
        # their contribution to total_stock / total_transit.
        try:
            _sanitize_non_awd_current_inventory_table(
                conn=conn,
                user_id=user_id,
                country=country,
                snapshot_date=period_end,
            )
        except Exception:
            logger.exception(
                "Failed sanitizing non-AWD currentinventory table for country=%s marketplace=%s",
                country,
                marketplace_id,
            )

        current_inventory_awd_by_sku = {}
        awd_inventory_by_sku = {}
        awd_by_sku = {}

    for item in items:
        sku = str(item.get("msku") or "").strip().upper()
        awd = awd_by_sku.get(sku, {})
        awd_inventory = current_inventory_awd_by_sku.get(sku) or awd_inventory_by_sku.get(sku, {})
        item["total_onhand_quantity"] = int(awd_inventory.get("total_onhand_quantity") or 0)
        item["total_inbound_quantity"] = int(awd.get("total_inbound_quantity") or 0)
        item["available_distributable_quantity"] = int(awd_inventory.get("available_distributable_quantity") or 0)
        item["reserved_distributable_quantity"] = int(awd_inventory.get("reserved_distributable_quantity") or 0)
        item["replenishment_quantity"] = int(awd_inventory.get("replenishment_quantity") or 0)


def _current_inventory_table_name(user_id: int, country: str, snapshot_date: date) -> str:
    country_key = _safe_ident((country or "").strip().lower().replace(" ", "_"))
    month_name = snapshot_date.strftime("%B").lower()
    return f"currentinventory_{int(user_id)}_{country_key}_{month_name}{snapshot_date.year}_table"



def _sanitize_non_awd_current_inventory_table(
    conn,
    user_id: int,
    country: str | None,
    snapshot_date: date,
) -> int:
    """Remove AWD-only quantities from a non-AWD currentinventory table.

    For marketplaces such as UK, total_onhand_quantity / total_inbound_quantity
    must not contain US AWD values. Recalculate total_stock without AWD on-hand:
        total_stock = available + fc-transfer
    and total_transit without AWD inbound:
        total_transit = inbound-shipped
    """
    if not country:
        return 0

    try:
        table_name = _current_inventory_table_name(user_id, country, snapshot_date)
    except Exception:
        return 0

    if not _table_exists(conn, "public", table_name):
        return 0

    columns = {
        row["column_name"]
        for row in conn.execute(text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :table_name
        """), {"table_name": table_name}).mappings().all()
    }

    def pick(*candidates):
        return next((c for c in candidates if c in columns), None)

    available_col = pick("available", "available_quantity")
    fc_transfer_col = pick("fc-transfer", "fc_transfer", "fc_transfer_quantity")
    inbound_shipped_col = pick("inbound-shipped", "inbound_shipped")

    set_parts = []
    for col in (
        "total_onhand_quantity",
        "total_inbound_quantity",
        "available_distributable_quantity",
        "reserved_distributable_quantity",
        "replenishment_quantity",
    ):
        if col in columns:
            set_parts.append(f'"{col}" = 0')

    if "total_stock" in columns:
        stock_terms = []
        if available_col:
            stock_terms.append(f'COALESCE("{available_col}"::numeric, 0)')
        if fc_transfer_col:
            stock_terms.append(f'COALESCE("{fc_transfer_col}"::numeric, 0)')
        set_parts.append(
            f'"total_stock" = {" + ".join(stock_terms) if stock_terms else "0"}'
        )

    if "total_transit" in columns:
        transit_expr = (
            f'COALESCE("{inbound_shipped_col}"::numeric, 0)'
            if inbound_shipped_col else "0"
        )
        set_parts.append(f'"total_transit" = {transit_expr}')

    if not set_parts:
        return 0

    result = conn.execute(text(
        f'UPDATE public."{table_name}" SET ' + ", ".join(set_parts)
    ))
    return int(result.rowcount or 0)

def _fetch_current_inventory_awd_by_sku(
    conn,
    user_id: int,
    country: str | None,
    snapshot_date: date,
    marketplace_id: str | None = None,
) -> dict[str, dict]:
    if not country:
        return {}

    # Safety guard: currentinventory AWD columns must only be consumed for
    # marketplaces that genuinely support AWD. This prevents US AWD quantities
    # from leaking into UK dispatch/inventory views.
    if marketplace_id and marketplace_id not in AWD_SUPPORTED_MARKETPLACES:
        return {}

    try:
        table_name = _current_inventory_table_name(user_id, country, snapshot_date)
    except Exception:
        return {}

    def read_from(read_conn) -> dict[str, dict]:
        if not _table_exists(read_conn, "public", table_name):
            return {}

        columns = {
            row["column_name"]
            for row in read_conn.execute(text("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = :table_name
            """), {"table_name": table_name}).mappings().all()
        }

        sku_col = next((c for c in ["SKU", "sku", "msku", "seller_sku", "merchant_sku"] if c in columns), None)
        onhand_col = next((
            c for c in [
                "total_onhand_quantity",
                "total_on_hand_quantity",
                "onhand_quantity",
                "on_hand_quantity",
            ] if c in columns
        ), None)

        if not sku_col or not onhand_col:
            return {}

        rows = read_conn.execute(text(f"""
            SELECT
                UPPER(TRIM(COALESCE("{sku_col}"::text, ''))) AS normalized_sku,
                SUM(COALESCE("{onhand_col}"::numeric, 0)) AS total_onhand_quantity
            FROM public."{table_name}"
            WHERE TRIM(COALESCE("{sku_col}"::text, '')) <> ''
            GROUP BY UPPER(TRIM(COALESCE("{sku_col}"::text, '')))
        """)).mappings().all()

        return {
            str(row["normalized_sku"]): {
                "total_onhand_quantity": int(row.get("total_onhand_quantity") or 0),
            }
            for row in rows
            if row.get("normalized_sku")
        }

    from_current_conn = read_from(conn)
    if from_current_conn:
        return from_current_conn

    try:
        with user_engine.connect() as main_conn:
            return read_from(main_conn)
    except Exception:
        logger.exception("Failed reading current inventory AWD snapshot for %s", table_name)
        return {}


def _fetch_awd_inventory_by_sku(conn, user_id: int, marketplace_id: str) -> dict[str, dict]:
    if not _table_exists(conn, "public", "inventory_awd"):
        return {}

    rows = conn.execute(text("""
        SELECT
            UPPER(TRIM(COALESCE(sku, ''))) AS normalized_sku,
            SUM(COALESCE(total_onhand_quantity, 0)) AS total_onhand_quantity,
            SUM(COALESCE(available_distributable_quantity, 0)) AS available_distributable_quantity,
            SUM(COALESCE(reserved_distributable_quantity, 0)) AS reserved_distributable_quantity,
            SUM(COALESCE(replenishment_quantity, 0)) AS replenishment_quantity
        FROM public.inventory_awd
        WHERE user_id = :user_id
          AND marketplace_id = :marketplace_id
          AND TRIM(COALESCE(sku, '')) <> ''
        GROUP BY UPPER(TRIM(COALESCE(sku, '')))
    """), {
        "user_id": int(user_id),
        "marketplace_id": marketplace_id,
    }).mappings().all()

    return {
        str(row["normalized_sku"]): dict(row)
        for row in rows
        if row.get("normalized_sku")
    }


def _ledger_units_fallback(row: dict) -> dict:
    return {
        "quantity": abs(int(row.get("sum_customer_shipments") or 0)),
        "return_quantity": abs(int(row.get("sum_customer_returns") or 0)),
        "total_quantity": abs(int(row.get("sold_total") or 0)),
    }


def _normalize_sku_key(value) -> str:
    return str(value or "").strip().upper()


def _is_skuwise_total_row(row: dict) -> bool:
    sku = _normalize_sku_key(row.get("sku"))
    product_name = _normalize_sku_key(row.get("product_name"))
    return sku in {"TOTAL", "GRAND TOTAL"} or product_name in {"TOTAL", "GRAND TOTAL"}


def _skuwise_units(row: dict) -> dict:
    return {
        "quantity": int(row.get("quantity") or 0),
        "return_quantity": int(row.get("return_quantity") or 0),
        "total_quantity": int(row.get("total_quantity") or 0),
    }


def _attach_skuwise_units_to_rows(
    items: list[dict],
    user_id: int,
    country: str,
    *,
    mode: str,
    year: int,
    month: int | None = None,
    quarter: int | None = None,
) -> dict | None:
    """Attach displayed Units Sold values from SKU-wise period tables.

    These values drive the Units Sold group and the displayed Difference column.
    """
    if not items:
        return None

    for item in items:
        item.update(_ledger_units_fallback(item))

    try:
        if mode == "month":
            if not month:
                return None
            result = get_skuwise_monthly_from_db(user_id, country, str(month), str(year))
        elif mode == "quarter":
            if not quarter:
                return None
            result = get_skuwise_quarterly_from_db(user_id, country, str(quarter), str(year))
        elif mode == "year":
            result = get_skuwise_yearly_from_db(user_id, country, str(year))
        else:
            return None
    except Exception:
        logger.exception("Could not read SKU-wise units for inventory reconciliation")
        return None

    if not result or not result.get("success"):
        logger.warning(
            "SKU-wise units unavailable for inventory reconciliation: %s",
            (result or {}).get("details") or (result or {}).get("error"),
        )
        return None

    skuwise_rows = result.get("items") or []
    if not skuwise_rows:
        return None

    units_by_sku: dict[str, dict] = {}
    summed_total = {"quantity": 0, "return_quantity": 0, "total_quantity": 0}
    source_total = None

    for skuwise_row in skuwise_rows:
        units = _skuwise_units(skuwise_row)

        if _is_skuwise_total_row(skuwise_row):
            source_total = units
            continue

        sku = _normalize_sku_key(skuwise_row.get("sku"))
        if not sku:
            continue

        acc = units_by_sku.setdefault(
            sku,
            {"quantity": 0, "return_quantity": 0, "total_quantity": 0},
        )
        for key in ("quantity", "return_quantity", "total_quantity"):
            acc[key] += units[key]
            summed_total[key] += units[key]

    if not units_by_sku and source_total is None:
        return None

    for item in items:
        sku = _normalize_sku_key(item.get("msku"))
        units = units_by_sku.get(sku, {"quantity": 0, "return_quantity": 0, "total_quantity": 0})
        item.update(units)

    return source_total or summed_total


def _apply_skuwise_totals(row: dict, totals: dict | None) -> None:
    if not totals:
        return

    for key in ("quantity", "return_quantity", "total_quantity"):
        row[key] = int(totals.get(key) or 0)


def _display_abs_int(row: dict, key: str) -> int:
    return abs(int(row.get(key) or 0))


def _display_net_units_sold(row: dict) -> int:
    if "total_quantity" in row:
        return abs(int(row.get("total_quantity") or 0))
    return abs(int(row.get("sold_total") or 0))


def _display_units_inwarded_total(row: dict) -> int:
    return _display_abs_int(row, "transit_total") + _display_abs_int(row, "total_inbound_quantity")


def _compute_display_difference(row: dict) -> int:
    return (
        _display_abs_int(row, "beginning_total")
        + _display_abs_int(row, "total_onhand_quantity")
        + _display_units_inwarded_total(row)
        - _display_net_units_sold(row)
        - _display_abs_int(row, "other_total")
        - _display_abs_int(row, "ending_total")
        - _display_abs_int(row, "total_onhand_quantity")
    )


def _apply_display_difference(row: dict) -> None:
    row["difference_total"] = _compute_display_difference(row)


def _apply_display_differences(items: list[dict]) -> None:
    for item in items:
        _apply_display_difference(item)


def _compute_grand_total(items: list[dict]) -> dict:
    gt = {
        "msku": "Grand Total",
        "product_name": "Grand Total",

        "sum_receipts": 0,
        "sum_customer_shipments": 0,
        "sum_customer_returns": 0,
        "sum_vendor_returns": 0,
        "sum_found": 0,
        "sum_lost": 0,
        "sum_damaged": 0,
        "sum_disposed": 0,
        # ✅ NEW
        "sum_in_transit_between_warehouses": 0,
        "sum_warehouse_transfer_in_out": 0,
        "sum_other_events": 0,
        "sum_unknown_events": 0,
        "defective_sum_first": 0,
        "defective_sum_last": 0,
        "sellable_sum_first": 0,
        "sellable_sum_last": 0,
        "warehouse_damaged_sum_first": 0,
        "warehouse_damaged_sum_last": 0,
        "expired_sum_first": 0,
        "expired_sum_last": 0,
        "customer_damaged_sum_first": 0,
        "customer_damaged_sum_last": 0,
        "distributor_damaged_sum_first": 0,
        "distributor_damaged_sum_last": 0,
        # ✅ add these
        "beginning_total": 0,
        "transit_total": 0,
        "other_total": 0,
        "sold_total": 0,
        "quantity": 0,
        "return_quantity": 0,
        "total_quantity": 0,
        "ending_total": 0,
        "difference_total": 0,
        "inventory_coverage_ratio": 0.0,

        # AWD inventory snapshot values
        "total_onhand_quantity": 0,
        "total_inbound_quantity": 0,
        "available_distributable_quantity": 0,
        "reserved_distributable_quantity": 0,
        "replenishment_quantity": 0,


    }
    for r in items:
        gt["sum_receipts"] += int(r.get("sum_receipts") or 0)
        gt["sum_customer_shipments"] += int(r.get("sum_customer_shipments") or 0)
        gt["sum_customer_returns"] += int(r.get("sum_customer_returns") or 0)
        gt["sum_vendor_returns"] += int(r.get("sum_vendor_returns") or 0)
        gt["sum_found"] += int(r.get("sum_found") or 0)
        gt["sum_lost"] += int(r.get("sum_lost") or 0)
        gt["sum_damaged"] += int(r.get("sum_damaged") or 0)
        gt["sum_disposed"] += int(r.get("sum_disposed") or 0)
        # ✅ NEW
        gt["sum_in_transit_between_warehouses"] += int(r.get("sum_in_transit_between_warehouses") or 0)
        gt["sum_warehouse_transfer_in_out"] += int(r.get("sum_warehouse_transfer_in_out") or 0)
        gt["sum_other_events"] += int(r.get("sum_other_events") or 0)
        gt["sum_unknown_events"] += int(r.get("sum_unknown_events") or 0)
        gt["defective_sum_first"] += int(r.get("defective_sum_first") or 0)
        gt["defective_sum_last"] += int(r.get("defective_sum_last") or 0)
        gt["sellable_sum_first"] += int(r.get("sellable_sum_first") or 0)
        gt["sellable_sum_last"] += int(r.get("sellable_sum_last") or 0)
        gt["warehouse_damaged_sum_first"] += int(r.get("warehouse_damaged_sum_first") or 0)
        gt["warehouse_damaged_sum_last"] += int(r.get("warehouse_damaged_sum_last") or 0)
        gt["expired_sum_first"] += int(r.get("expired_sum_first") or 0)
        gt["expired_sum_last"] += int(r.get("expired_sum_last") or 0)
        gt["customer_damaged_sum_first"] += int(r.get("customer_damaged_sum_first") or 0)
        gt["customer_damaged_sum_last"] += int(r.get("customer_damaged_sum_last") or 0)
        gt["distributor_damaged_sum_first"] += int(r.get("distributor_damaged_sum_first") or 0)
        gt["distributor_damaged_sum_last"] += int(r.get("distributor_damaged_sum_last") or 0)
        gt["beginning_total"] += int(r.get("beginning_total") or 0)
        gt["transit_total"] += int(r.get("transit_total") or 0)
        gt["other_total"] += int(r.get("other_total") or 0)
        gt["sold_total"] += int(r.get("sold_total") or 0)
        gt["quantity"] += int(r.get("quantity") or 0)
        gt["return_quantity"] += int(r.get("return_quantity") or 0)
        gt["total_quantity"] += int(r.get("total_quantity") or 0)
        gt["ending_total"] += int(r.get("ending_total") or 0)
        gt["difference_total"] += int(r.get("difference_total") or 0)
        gt["total_onhand_quantity"] += int(r.get("total_onhand_quantity") or 0)
        gt["total_inbound_quantity"] += int(r.get("total_inbound_quantity") or 0)
        gt["available_distributable_quantity"] += int(r.get("available_distributable_quantity") or 0)
        gt["reserved_distributable_quantity"] += int(r.get("reserved_distributable_quantity") or 0)
        gt["replenishment_quantity"] += int(r.get("replenishment_quantity") or 0)
        gt["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(gt["ending_total"], gt["sold_total"])


    return gt


# ---------------------------------------------------------------------
# CREATE + UPSERT SUMMARY TABLE (and save GRAND TOTAL row also)
# ---------------------------------------------------------------------

def _ensure_inventory_summary_table_exists(conn, table_name: str) -> None:
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS public.{table_name} (
            id BIGSERIAL PRIMARY KEY,

            msku TEXT NOT NULL UNIQUE,
            product_name TEXT,

            sum_receipts BIGINT NOT NULL DEFAULT 0,
            sum_customer_shipments BIGINT NOT NULL DEFAULT 0,
            sum_customer_returns BIGINT NOT NULL DEFAULT 0,
            sum_vendor_returns BIGINT NOT NULL DEFAULT 0,
            sum_found BIGINT NOT NULL DEFAULT 0,
            sum_lost BIGINT NOT NULL DEFAULT 0,
            sum_damaged BIGINT NOT NULL DEFAULT 0,
            sum_disposed BIGINT NOT NULL DEFAULT 0,

            sum_in_transit_between_warehouses BIGINT DEFAULT 0,
            sum_warehouse_transfer_in_out BIGINT DEFAULT 0,
            sum_other_events BIGINT DEFAULT 0,
            sum_unknown_events BIGINT DEFAULT 0,

            -- ✅ snapshots
            defective_sum_first BIGINT DEFAULT 0,
            defective_sum_last BIGINT DEFAULT 0,
            sellable_sum_first BIGINT DEFAULT 0,
            sellable_sum_last BIGINT DEFAULT 0,
            warehouse_damaged_sum_first BIGINT DEFAULT 0,
            warehouse_damaged_sum_last BIGINT DEFAULT 0,
            expired_sum_first BIGINT DEFAULT 0,
            expired_sum_last BIGINT DEFAULT 0,
            customer_damaged_sum_first BIGINT DEFAULT 0,
            customer_damaged_sum_last BIGINT DEFAULT 0,
            distributor_damaged_sum_first BIGINT DEFAULT 0,
            distributor_damaged_sum_last BIGINT DEFAULT 0,

            -- ✅ derived totals
            beginning_total BIGINT DEFAULT 0,
            transit_total BIGINT DEFAULT 0,
            other_total BIGINT DEFAULT 0,
            sold_total BIGINT DEFAULT 0,
            quantity BIGINT DEFAULT 0,
            return_quantity BIGINT DEFAULT 0,
            total_quantity BIGINT DEFAULT 0,
            ending_total BIGINT DEFAULT 0,
            difference_total BIGINT DEFAULT 0,
            inventory_coverage_ratio DOUBLE PRECISION,

            -- AWD inventory snapshot values
            total_onhand_quantity BIGINT DEFAULT 0,
            total_inbound_quantity BIGINT DEFAULT 0,
            available_distributable_quantity BIGINT DEFAULT 0,
            reserved_distributable_quantity BIGINT DEFAULT 0,
            replenishment_quantity BIGINT DEFAULT 0,

            computed_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
        );
    """))

    # ✅ one ALTER statement, commas between clauses, one semicolon at end
    conn.execute(text(f"""
        ALTER TABLE public.{table_name}
            ADD COLUMN IF NOT EXISTS product_name TEXT,
            ADD COLUMN IF NOT EXISTS sum_in_transit_between_warehouses BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sum_warehouse_transfer_in_out BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sum_other_events BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sum_unknown_events BIGINT DEFAULT 0,

            ADD COLUMN IF NOT EXISTS defective_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS defective_sum_last BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sellable_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sellable_sum_last BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS warehouse_damaged_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS warehouse_damaged_sum_last BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS expired_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS expired_sum_last BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS customer_damaged_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS customer_damaged_sum_last BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS distributor_damaged_sum_first BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS distributor_damaged_sum_last BIGINT DEFAULT 0,

            -- ✅ derived totals
            ADD COLUMN IF NOT EXISTS beginning_total BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS transit_total BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS other_total BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS sold_total BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS return_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS total_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS ending_total BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS inventory_coverage_ratio DOUBLE PRECISION,
            ADD COLUMN IF NOT EXISTS difference_total BIGINT DEFAULT 0,

            ADD COLUMN IF NOT EXISTS total_onhand_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS total_inbound_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS available_distributable_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS reserved_distributable_quantity BIGINT DEFAULT 0,
            ADD COLUMN IF NOT EXISTS replenishment_quantity BIGINT DEFAULT 0;
    """))

def _compute_inventory_coverage_ratio(ending_total, sold_total):
    try:
        ending = float(ending_total or 0)
        sold = float(sold_total or 0)
        if sold == 0:
            return None
        return ending / abs(sold)
    except Exception:
        return None

def _upsert_inventory_summary_rows(conn, table_name: str, rows: list[dict]) -> int:
    if not rows:
        return 0

    upsert_sql = text(f"""
        INSERT INTO public.{table_name} (
            msku,
            product_name,

            sum_receipts,
            sum_customer_shipments,
            sum_customer_returns,
            sum_vendor_returns,
            sum_found,
            sum_lost,
            sum_damaged,
            sum_disposed,

            sum_in_transit_between_warehouses,
            sum_warehouse_transfer_in_out,
            sum_other_events,
            sum_unknown_events,

            defective_sum_first,
            defective_sum_last,
            sellable_sum_first,
            sellable_sum_last,
            warehouse_damaged_sum_first,
            warehouse_damaged_sum_last,
            expired_sum_first,
            expired_sum_last,
            customer_damaged_sum_first,
            customer_damaged_sum_last,
            distributor_damaged_sum_first,
            distributor_damaged_sum_last,
            beginning_total,
            transit_total,
            other_total,
            sold_total,
            quantity,
            return_quantity,
            total_quantity,
            ending_total,
            difference_total,
            inventory_coverage_ratio,
            total_onhand_quantity,
            total_inbound_quantity,
            available_distributable_quantity,
            reserved_distributable_quantity,
            replenishment_quantity,
            computed_at
        )
        VALUES (
            :msku,
            :product_name,

            :sum_receipts,
            :sum_customer_shipments,
            :sum_customer_returns,
            :sum_vendor_returns,
            :sum_found,
            :sum_lost,
            :sum_damaged,
            :sum_disposed,

            :sum_in_transit_between_warehouses,
            :sum_warehouse_transfer_in_out,
            :sum_other_events,
            :sum_unknown_events,

            :defective_sum_first,
            :defective_sum_last,
            :sellable_sum_first,
            :sellable_sum_last,
            :warehouse_damaged_sum_first,
            :warehouse_damaged_sum_last,
            :expired_sum_first,
            :expired_sum_last,
            :customer_damaged_sum_first,
            :customer_damaged_sum_last,
            :distributor_damaged_sum_first,
            :distributor_damaged_sum_last,
            :beginning_total,
            :transit_total,
            :other_total,
            :sold_total,
            :quantity,
            :return_quantity,
            :total_quantity,
            :ending_total,
            :difference_total,
            :inventory_coverage_ratio,
            :total_onhand_quantity,
            :total_inbound_quantity,
            :available_distributable_quantity,
            :reserved_distributable_quantity,
            :replenishment_quantity,
            NOW()
        )
        ON CONFLICT (msku) DO UPDATE SET
            product_name = EXCLUDED.product_name,

            sum_receipts = EXCLUDED.sum_receipts,
            sum_customer_shipments = EXCLUDED.sum_customer_shipments,
            sum_customer_returns = EXCLUDED.sum_customer_returns,
            sum_vendor_returns = EXCLUDED.sum_vendor_returns,
            sum_found = EXCLUDED.sum_found,
            sum_lost = EXCLUDED.sum_lost,
            sum_damaged = EXCLUDED.sum_damaged,
            sum_disposed = EXCLUDED.sum_disposed,

            sum_in_transit_between_warehouses = EXCLUDED.sum_in_transit_between_warehouses,
            sum_warehouse_transfer_in_out = EXCLUDED.sum_warehouse_transfer_in_out,
            sum_other_events = EXCLUDED.sum_other_events,
            sum_unknown_events = EXCLUDED.sum_unknown_events,

            defective_sum_first = EXCLUDED.defective_sum_first,
            defective_sum_last  = EXCLUDED.defective_sum_last,
            sellable_sum_first  = EXCLUDED.sellable_sum_first,
            sellable_sum_last   = EXCLUDED.sellable_sum_last,
            warehouse_damaged_sum_first = EXCLUDED.warehouse_damaged_sum_first,
            warehouse_damaged_sum_last  = EXCLUDED.warehouse_damaged_sum_last,
            expired_sum_first   = EXCLUDED.expired_sum_first,
            expired_sum_last    = EXCLUDED.expired_sum_last,
            customer_damaged_sum_first = EXCLUDED.customer_damaged_sum_first,
            customer_damaged_sum_last  = EXCLUDED.customer_damaged_sum_last,
            distributor_damaged_sum_first = EXCLUDED.distributor_damaged_sum_first,
            distributor_damaged_sum_last  = EXCLUDED.distributor_damaged_sum_last,
            beginning_total = EXCLUDED.beginning_total,
            transit_total = EXCLUDED.transit_total, 
            other_total = EXCLUDED.other_total,
            sold_total = EXCLUDED.sold_total,   
            quantity = EXCLUDED.quantity,
            return_quantity = EXCLUDED.return_quantity,
            total_quantity = EXCLUDED.total_quantity,
            ending_total = EXCLUDED.ending_total,
            difference_total = EXCLUDED.difference_total,
            inventory_coverage_ratio = EXCLUDED.inventory_coverage_ratio,
            total_onhand_quantity = EXCLUDED.total_onhand_quantity,
            total_inbound_quantity = EXCLUDED.total_inbound_quantity,
            available_distributable_quantity = EXCLUDED.available_distributable_quantity,
            reserved_distributable_quantity = EXCLUDED.reserved_distributable_quantity,
            replenishment_quantity = EXCLUDED.replenishment_quantity,
            computed_at = NOW();
    """)

    for r in rows:
        conn.execute(upsert_sql, {
            "msku": (r.get("msku") or "").strip(),
            "product_name": (r.get("product_name") or None),

            "sum_receipts": int(r.get("sum_receipts") or 0),
            "sum_customer_shipments": int(r.get("sum_customer_shipments") or 0),
            "sum_customer_returns": int(r.get("sum_customer_returns") or 0),
            "sum_vendor_returns": int(r.get("sum_vendor_returns") or 0),
            "sum_found": int(r.get("sum_found") or 0),
            "sum_lost": int(r.get("sum_lost") or 0),
            "sum_damaged": int(r.get("sum_damaged") or 0),
            "sum_disposed": int(r.get("sum_disposed") or 0),

            "sum_in_transit_between_warehouses": int(r.get("sum_in_transit_between_warehouses") or 0),
            "sum_warehouse_transfer_in_out": int(r.get("sum_warehouse_transfer_in_out") or 0),
            "sum_other_events": int(r.get("sum_other_events") or 0),
            "sum_unknown_events": int(r.get("sum_unknown_events") or 0),

            "defective_sum_first": int(r.get("defective_sum_first") or 0),
            "defective_sum_last": int(r.get("defective_sum_last") or 0),
            "sellable_sum_first": int(r.get("sellable_sum_first") or 0),
            "sellable_sum_last": int(r.get("sellable_sum_last") or 0),
            "warehouse_damaged_sum_first": int(r.get("warehouse_damaged_sum_first") or 0),
            "warehouse_damaged_sum_last": int(r.get("warehouse_damaged_sum_last") or 0),
            "expired_sum_first": int(r.get("expired_sum_first") or 0),
            "expired_sum_last": int(r.get("expired_sum_last") or 0),
            "customer_damaged_sum_first": int(r.get("customer_damaged_sum_first") or 0),
            "customer_damaged_sum_last": int(r.get("customer_damaged_sum_last") or 0),
            "distributor_damaged_sum_first": int(r.get("distributor_damaged_sum_first") or 0),
            "distributor_damaged_sum_last": int(r.get("distributor_damaged_sum_last") or 0),
            "beginning_total": int(r.get("beginning_total") or 0),
            "transit_total": int(r.get("transit_total") or 0),
            "other_total": int(r.get("other_total") or 0),
            "sold_total": int(r.get("sold_total") or 0),
            "quantity": int(r.get("quantity") or 0),
            "return_quantity": int(r.get("return_quantity") or 0),
            "total_quantity": int(r.get("total_quantity") or 0),
            "ending_total": int(r.get("ending_total") or 0),
            "difference_total": int(r.get("difference_total") or 0),
            "inventory_coverage_ratio": float(r.get("inventory_coverage_ratio") or 0.0),
            "total_onhand_quantity": int(r.get("total_onhand_quantity") or 0),
            "total_inbound_quantity": int(r.get("total_inbound_quantity") or 0),
            "available_distributable_quantity": int(r.get("available_distributable_quantity") or 0),
            "reserved_distributable_quantity": int(r.get("reserved_distributable_quantity") or 0),
            "replenishment_quantity": int(r.get("replenishment_quantity") or 0),

        })

    return len(rows)

def _create_inventorymonthly_after_fetch(
    user_id: int,
    country: str,
    mp: str,
    rows: list[dict],
) -> list[dict]:
    """
    After saving raw rows in public.monthwise_inventory,
    create/update inventorymonthly_{user_id}_{country}_{MM}_{YYYY}.
    """
    if not rows:
        return []

    dates = sorted({
        r.get("date")
        for r in rows
        if isinstance(r.get("date"), date)
    })

    if not dates:
        return []

    # group rows by month
    month_groups: dict[tuple[int, int], list[date]] = {}

    for d in dates:
        month_groups.setdefault((d.year, d.month), []).append(d)

    created_tables = []

    with amazon_conn() as conn:
        for (year, month), month_dates in sorted(month_groups.items()):
            start_date = min(month_dates)
            end_date = max(month_dates)
            month_start, month_end = _month_range(year, month)

            table_name = _monthly_table_name(
                user_id=user_id,
                country=country,
                month=month,
                year=year,
            )

            items = _aggregate_from_monthwise_inventory(
                conn=conn,
                user_id=user_id,
                mp=mp,
                start_date=start_date,
                end_date=end_date,
            )

            # Prefer the SKU master product name for this country. The
            # aggregation already falls back to Amazon's ledger Title/MSKU.
            _attach_product_names_to_rows(
                items,
                user_id,
                country=country,
                marketplace_id=mp,
            )

            # Add open AWD inbound shipment quantities by matching SKU to MSKU.
            _attach_awd_quantities_to_rows(
                conn=conn,
                items=items,
                user_id=user_id,
                marketplace_id=mp,
                period_start=month_start,
                period_end=month_end,
                country=country,
            )

            skuwise_totals = _attach_skuwise_units_to_rows(
                items,
                user_id,
                country,
                mode="month",
                month=month,
                year=year,
            )
            _apply_display_differences(items)

            for r in items:
                r["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                    r.get("ending_total"),
                    r.get("sold_total"),
                )

            grand_total = _compute_grand_total(items)
            _apply_skuwise_totals(grand_total, skuwise_totals)
            _apply_display_difference(grand_total)
            grand_total["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                grand_total.get("ending_total"),
                grand_total.get("sold_total"),
            )

            to_save = items + [grand_total]

            _ensure_inventory_summary_table_exists(conn, table_name)
            saved = _upsert_inventory_summary_rows(conn, table_name, to_save)

            created_tables.append({
                "table": f"public.{table_name}",
                "month": month,
                "year": year,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "saved_rows": saved,
            })

    return created_tables

def _safe_ident(name: str) -> str:
    # only allow identifiers you generate internally
    if not re.fullmatch(r"[A-Za-z0-9_]+", name or ""):
        raise ValueError("Invalid table name")
    return name

def _read_inventory_summary_table(conn, table_name: str, sort_order="desc"):
    t = _safe_ident(table_name)

    order_clause = "DESC" if sort_order.lower() == "desc" else "ASC"

    sql = text(f'''
        SELECT *
        FROM public."{t}"
        ORDER BY
            CASE WHEN UPPER(COALESCE(msku, '')) = 'GRAND TOTAL' THEN 1 ELSE 0 END,
            ABS(COALESCE(total_quantity, sold_total)) {order_clause}
    ''')

    result = conn.execute(sql)
    return [dict(r) for r in result.mappings().all()]


# # ============================================================
# # ROUTE 1: READ amazon_db -> GROUP BY MSKU + return GRAND TOTAL
# # GET /amazon_api/inventory/ledger-summary/db?month=11&year=2025
# # ============================================================

@inventory_bp.route("/amazon_api/inventory/ledger-summary/db/store-month", methods=["GET"])
def inventory_ledger_summary_store_month():
    user_id = _get_user_id_from_bearer()
    if not user_id:
        return jsonify({"error": "Invalid or missing token"}), 401

    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    country = (request.args.get("country") or "us").strip().lower()
    sort_order = request.args.get("sort", "desc")

    try:
        month = int(request.args.get("month", "0"))
        year = int(request.args.get("year", "0"))
        requested_start, requested_end = _month_range(year, month)
    except Exception:
        return jsonify({"error": "Provide valid ?country=xx&month=MM&year=YYYY"}), 400

    table_name = _monthly_table_name(user_id, country, month, year)

    try:
        with amazon_conn() as conn:
            src = _get_source_table(conn)
            location_code = _ledger_location_for_marketplace(mp)
            location_filter_sql = (
                "AND UPPER(TRIM(COALESCE(mi.location, ''))) = :location_code"
                if location_code else ""
            )

            # Rebuild the monthly table on every store-month request so old rows
            # created with an earlier difference_total formula are corrected.
            available_range = conn.execute(text(f"""
                SELECT MIN(mi.date) AS first_date, MAX(mi.date) AS last_date
                FROM {src} mi
                WHERE mi.user_id = :user_id
                  AND mi.marketplace_id = :mp
                  {location_filter_sql}
                  AND mi.date >= :requested_start
                  AND mi.date <= :requested_end
            """), {
                "user_id": user_id,
                "mp": mp,
                "location_code": location_code,
                "requested_start": requested_start,
                "requested_end": requested_end,
            }).mappings().first()

            if not available_range or not available_range.get("first_date") or not available_range.get("last_date"):
                return jsonify({
                    "success": False,
                    "error": "No monthwise inventory data found for the requested month.",
                    "table": f"public.{table_name}",
                    "marketplace_id": mp,
                    "mode": "month",
                    "country": country,
                    "start_date": requested_start.isoformat(),
                    "end_date": requested_end.isoformat(),
                    "count": 0,
                    "items": [],
                }), 404

            start_date = available_range["first_date"]
            end_date = available_range["last_date"]

            items = _aggregate_from_monthwise_inventory(
                conn=conn,
                user_id=user_id,
                mp=mp,
                start_date=start_date,
                end_date=end_date,
            )

            # Re-apply the US/UK/CA SKU master mapping while rebuilding the
            # summary table, so existing NULL product names are backfilled.
            _attach_product_names_to_rows(
                items,
                user_id,
                country=country,
                marketplace_id=mp,
            )

            # Attach open AWD inbound shipment quantities before calculating
            # the Grand Total and upserting the rebuilt inventorymonthly table.
            _attach_awd_quantities_to_rows(
                conn=conn,
                items=items,
                user_id=user_id,
                marketplace_id=mp,
                period_start=requested_start,
                period_end=requested_end,
                country=country,
            )

            skuwise_totals = _attach_skuwise_units_to_rows(
                items,
                user_id,
                country,
                mode="month",
                month=month,
                year=year,
            )
            _apply_display_differences(items)

            for row in items:
                row["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                    row.get("ending_total"), row.get("sold_total")
                )

            grand_total = _compute_grand_total(items)
            _apply_skuwise_totals(grand_total, skuwise_totals)
            _apply_display_difference(grand_total)
            grand_total["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                grand_total.get("ending_total"), grand_total.get("sold_total")
            )

            _ensure_inventory_summary_table_exists(conn, table_name)
            saved = _upsert_inventory_summary_rows(
                conn,
                table_name,
                items + [grand_total],
            )

            read_items = _read_inventory_summary_table(
                conn,
                table_name,
                sort_order=sort_order,
            )

        return jsonify({
            "success": True,
            "marketplace_id": mp,
            "mode": "month",
            "country": country,
            "table": f"public.{table_name}",
            "created_or_updated": True,
            "saved_rows": saved,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "count": max(len(read_items) - 1, 0) if read_items else 0,
            "items": read_items,
        }), 200

    except Exception as e:
        logger.exception("Failed to rebuild monthly inventory summary")
        return jsonify({"success": False, "error": str(e)}), 500


def _quarter_range_upto_latest_completed_month_end(
    conn,
    user_id: int,
    mp: str,
    year: int,
    quarter: int
) -> tuple[date, date]:
    """
    Quarter summary rule (CORRECT):

    - start = first day of quarter
    - if full quarter exists → use quarter end (e.g. 30 Jun)
    - else → use latest COMPLETED month-end inside that quarter
    - ignore partial current month

    Examples:
      Q2:
        only April → end = 30 Apr
        April+May → end = 31 May
        full → end = 30 Jun
    """

    if quarter not in (1, 2, 3, 4):
        raise ValueError("quarter must be 1-4")

    start_date, quarter_end = _quarter_range(year, quarter)
    src = _get_source_table(conn)
    location_code = _ledger_location_for_marketplace(mp)
    location_filter_sql = (
        "AND UPPER(TRIM(COALESCE(mi.location, ''))) = :location_code"
        if location_code else ""
    )

    # 1️⃣ Check if full quarter end exists (e.g. 30 Jun)
    full_q_sql = text(f"""
        SELECT 1
        FROM {src} mi
        WHERE mi.user_id = :user_id
          AND mi.marketplace_id = :mp
          {location_filter_sql}
          AND mi.date = :quarter_end
        LIMIT 1
    """)

    full_exists = conn.execute(full_q_sql, {
        "user_id": user_id,
        "mp": mp,
        "location_code": location_code,
        "quarter_end": quarter_end,
    }).first()

    if full_exists:
        return start_date, quarter_end

    # 2️⃣ Get latest completed month-end inside quarter
    month_end_sql = text(f"""
        SELECT MAX(x.month_end_date) AS last_completed_month_end
        FROM (
            SELECT DISTINCT mi.date AS month_end_date
            FROM {src} mi
            WHERE mi.user_id = :user_id
              AND mi.marketplace_id = :mp
              {location_filter_sql}
              AND mi.date >= :start_date
              AND mi.date <= :quarter_end
              AND mi.date = (
                  date_trunc('month', mi.date)::date
                  + INTERVAL '1 month'
                  - INTERVAL '1 day'
              )::date
        ) x
    """)

    last_completed = conn.execute(month_end_sql, {
        "user_id": user_id,
        "mp": mp,
        "location_code": location_code,
        "start_date": start_date,
        "quarter_end": quarter_end,
    }).scalar()

    if last_completed:
        return start_date, last_completed

    # 3️⃣ fallback (no month-end found)
    fallback_sql = text(f"""
        SELECT MAX(mi.date)
        FROM {src} mi
        WHERE mi.user_id = :user_id
          AND mi.marketplace_id = :mp
          {location_filter_sql}
          AND mi.date >= :start_date
          AND mi.date <= :quarter_end
    """)

    fallback = conn.execute(fallback_sql, {
        "user_id": user_id,
        "mp": mp,
        "location_code": location_code,
        "start_date": start_date,
        "quarter_end": quarter_end,
    }).scalar()

    if fallback:
        return start_date, fallback

    return start_date, quarter_end


@inventory_bp.route("/amazon_api/inventory/ledger-summary/db/store-quarter", methods=["GET"])
def inventory_ledger_summary_store_quarter():
    user_id = _get_user_id_from_bearer()
    if not user_id:
        return jsonify({"error": "Invalid or missing token"}), 401

    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    country = (request.args.get("country") or "us").strip().lower()

    try:
        quarter = int(request.args.get("quarter", "0"))
        year = int(request.args.get("year", "0"))

        if quarter not in (1, 2, 3, 4):
            raise ValueError("Invalid quarter")
        if year < 2000 or year > 2100:
            raise ValueError("Invalid year")
    except Exception:
        return jsonify({"error": "Provide valid ?country=xx&quarter=Q&year=YYYY"}), 400

    table_name = _quarterly_table_name(user_id, country, quarter, year)

    try:
        with amazon_conn() as conn:
            # ✅ dynamic quarter end date
            start_date, end_date = _quarter_range_upto_latest_completed_month_end(
                conn, user_id, mp, year, quarter
            )

            items = _aggregate_from_monthwise_inventory(conn, user_id, mp, start_date, end_date)

            # Use the same short product_name mapping as the monthly table.
            # This overwrites long Amazon ledger titles with product_name from
            # public.sku_{user_id}_data_table (sku_us / sku_uk / sku_canada).
            _attach_product_names_to_rows(
                items,
                user_id,
                country=country,
                marketplace_id=mp,
            )

            _attach_awd_quantities_to_rows(
                conn=conn,
                items=items,
                user_id=user_id,
                marketplace_id=mp,
                period_start=start_date,
                period_end=end_date,
                country=country,
            )

            skuwise_totals = _attach_skuwise_units_to_rows(
                items,
                user_id,
                country,
                mode="quarter",
                quarter=quarter,
                year=year,
            )
            _apply_display_differences(items)

            for r in items:
                r["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                    r.get("ending_total"), r.get("sold_total")
                )

            grand_total = _compute_grand_total(items)
            _apply_skuwise_totals(grand_total, skuwise_totals)
            _apply_display_difference(grand_total)
            grand_total["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                grand_total.get("ending_total"), grand_total.get("sold_total")
            )

            to_save = items + [grand_total]

            _ensure_inventory_summary_table_exists(conn, table_name)
            saved = _upsert_inventory_summary_rows(conn, table_name, to_save)
            sort_order = request.args.get("sort", "desc")

            read_items = _read_inventory_summary_table(conn, table_name, sort_order=sort_order)

        return jsonify({
            "success": True,
            "marketplace_id": mp,
            "mode": "quarter",
            "country": country,
            "table": f"public.{table_name}",
            "saved_rows": saved,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "count": max(len(read_items) - 1, 0) if read_items else 0,
            "items": read_items,
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
def _year_range_upto_latest_completed_month_end(
    conn,
    user_id: int,
    mp: str,
    year: int
) -> tuple[date, date]:
    """
    Yearly summary rule:
    - start = 1 Jan of requested year
    - if 31 Dec exists, use it
    - otherwise use the latest completed month-end date inside that year
    - ignore partial current-month dates

    Example:
      if today is 2026-04-08 and data exists up to 2026-04-08,
      yearly end_date should be 2026-03-31, not 2026-04-08.
    """
    if year < 2000 or year > 2100:
        raise ValueError("year must be between 2000 and 2100")

    start_date = date(year, 1, 1)
    dec_31 = date(year, 12, 31)

    src = _get_source_table(conn)
    location_code = _ledger_location_for_marketplace(mp)
    location_filter_sql = (
        "AND UPPER(TRIM(COALESCE(mi.location, ''))) = :location_code"
        if location_code else ""
    )

    # 1) if full year-end exists, use 31-Dec
    dec31_sql = text(f"""
        SELECT 1
        FROM {src} mi
        WHERE mi.user_id = :user_id
          AND mi.marketplace_id = :mp
          {location_filter_sql}
          AND mi.date = :dec_31
        LIMIT 1
    """)
    dec31_exists = conn.execute(
        dec31_sql,
        {
            "user_id": user_id,
            "mp": mp,
            "location_code": location_code,
            "dec_31": dec_31,
        },
    ).first()

    if dec31_exists:
        return start_date, dec_31

    # 2) otherwise pick latest month-end date that actually exists in this year
    month_end_sql = text(f"""
        SELECT MAX(x.month_end_date) AS last_completed_month_end
        FROM (
            SELECT DISTINCT mi.date AS month_end_date
            FROM {src} mi
            WHERE mi.user_id = :user_id
              AND mi.marketplace_id = :mp
              {location_filter_sql}
              AND mi.date >= :start_date
              AND mi.date < :dec_31
              AND mi.date = (
                  date_trunc('month', mi.date)::date
                  + INTERVAL '1 month'
                  - INTERVAL '1 day'
              )::date
        ) x
    """)

    last_completed_month_end = conn.execute(
        month_end_sql,
        {
            "user_id": user_id,
            "mp": mp,
            "location_code": location_code,
            "start_date": start_date,
            "dec_31": dec_31,
        },
    ).scalar()

    if last_completed_month_end:
        return start_date, last_completed_month_end

    # 3) fallback: no month-end found, use latest available date
    fallback_sql = text(f"""
        SELECT MAX(mi.date) AS last_available_date
        FROM {src} mi
        WHERE mi.user_id = :user_id
          AND mi.marketplace_id = :mp
          {location_filter_sql}
          AND mi.date >= :start_date
          AND mi.date <= :dec_31
    """)

    last_available_date = conn.execute(
        fallback_sql,
        {
            "user_id": user_id,
            "mp": mp,
            "location_code": location_code,
            "start_date": start_date,
            "dec_31": dec_31,
        },
    ).scalar()

    if not last_available_date:
        return start_date, dec_31

    return start_date, last_available_date



@inventory_bp.route("/amazon_api/inventory/ledger-summary/db/store-year", methods=["GET"])
def inventory_ledger_summary_store_year():
    user_id = _get_user_id_from_bearer()
    if not user_id:
        return jsonify({"error": "Invalid or missing token"}), 401

    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    country = (request.args.get("country") or "us").strip().lower()

    try:
        year = int(request.args.get("year", "0"))
        if year < 2000 or year > 2100:
            raise ValueError("Invalid year")
    except Exception:
        return jsonify({"error": "Provide valid ?country=xx&year=YYYY"}), 400

    table_name = _yearly_table_name(user_id, country, year)

    try:
        with amazon_conn() as conn:
            # ✅ use latest completed month-end, not partial current month
            start_date, end_date = _year_range_upto_latest_completed_month_end(
                conn, user_id, mp, year
            )

            items = _aggregate_from_monthwise_inventory(conn, user_id, mp, start_date, end_date)

            # Use the same short product_name mapping as the monthly table.
            # This overwrites long Amazon ledger titles with product_name from
            # public.sku_{user_id}_data_table (sku_us / sku_uk / sku_canada).
            _attach_product_names_to_rows(
                items,
                user_id,
                country=country,
                marketplace_id=mp,
            )

            _attach_awd_quantities_to_rows(
                conn=conn,
                items=items,
                user_id=user_id,
                marketplace_id=mp,
                period_start=start_date,
                period_end=end_date,
                country=country,
            )

            skuwise_totals = _attach_skuwise_units_to_rows(
                items,
                user_id,
                country,
                mode="year",
                year=year,
            )
            _apply_display_differences(items)

            for r in items:
                r["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                    r.get("ending_total"), r.get("sold_total")
                )

            grand_total = _compute_grand_total(items)
            _apply_skuwise_totals(grand_total, skuwise_totals)
            _apply_display_difference(grand_total)
            grand_total["inventory_coverage_ratio"] = _compute_inventory_coverage_ratio(
                grand_total.get("ending_total"), grand_total.get("sold_total")
            )

            to_save = items + [grand_total]

            _ensure_inventory_summary_table_exists(conn, table_name)
            saved = _upsert_inventory_summary_rows(conn, table_name, to_save)
            sort_order = request.args.get("sort", "desc")

            read_items = _read_inventory_summary_table(conn, table_name, sort_order=sort_order)

        return jsonify({
            "success": True,
            "marketplace_id": mp,
            "mode": "year",
            "country": country,
            "table": f"public.{table_name}",
            "saved_rows": saved,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "count": max(len(read_items) - 1, 0) if read_items else 0,
            "items": read_items,
        }), 200

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



# =============================================================================
# AWD INVENTORY - FETCH FROM AMAZON SP-API AND STORE IN inventory_awd
# =============================================================================

def _safe_int_awd(value) -> int:
    try:
        if value is None:
            return 0

        if isinstance(value, dict):
            return sum(
                int(v or 0)
                for v in value.values()
                if isinstance(v, (int, float))
            )

        s = str(value).strip().replace(",", "")
        if s == "":
            return 0

        return int(float(s))
    except Exception:
        return 0


def _extract_awd_inventory_list(payload: dict) -> list[dict]:
    """
    Amazon AWD API response can be wrapped by your amazon_client in either:
      - {"payload": {...}}
      - direct payload {...}

    This helper safely finds the inventory array.
    """
    payload = payload or {}

    # Most expected shape
    if isinstance(payload.get("inventory"), list):
        return payload.get("inventory") or []

    # Defensive fallback names
    if isinstance(payload.get("items"), list):
        return payload.get("items") or []

    if isinstance(payload.get("inventoryItems"), list):
        return payload.get("inventoryItems") or []

    return []


def _get_awd_next_token(payload: dict):
    payload = payload or {}

    # Most expected shape
    if payload.get("nextToken"):
        return payload.get("nextToken")

    # Defensive fallback if Amazon wraps pagination differently
    pagination = payload.get("pagination") or {}
    if pagination.get("nextToken"):
        return pagination.get("nextToken")

    return None


def _normalize_awd_inventory_item(item: dict) -> dict | None:
    item = item or {}

    sku = (
        item.get("sku")
        or item.get("sellerSku")
        or item.get("merchantSku")
        or item.get("msku")
    )

    if not sku:
        return None

    inventory_details = item.get("inventoryDetails") or {}

    return {
        "sku": str(sku).strip(),

        "total_onhand_quantity": _safe_int_awd(
            item.get("totalOnhandQuantity")
        ),
        "total_inbound_quantity": _safe_int_awd(
            item.get("totalInboundQuantity")
        ),

        "available_distributable_quantity": _safe_int_awd(
            inventory_details.get("availableDistributableQuantity")
        ),
        "reserved_distributable_quantity": _safe_int_awd(
            inventory_details.get("reservedDistributableQuantity")
        ),
        "replenishment_quantity": _safe_int_awd(
            inventory_details.get("replenishmentQuantity")
        ),

        "expiration_details": item.get("expirationDetails") or [],
    }


def _fetch_awd_inventory_rows_from_amazon(mp: str, sku: str | None = None) -> list[dict]:
    """
    Fetch AWD inventory from:
      GET /awd/2024-05-09/inventory

    Available Amazon-side fields:
      sku
      totalOnhandQuantity
      totalInboundQuantity
      inventoryDetails.availableDistributableQuantity
      inventoryDetails.reservedDistributableQuantity
      inventoryDetails.replenishmentQuantity
      expirationDetails
    """
    rows: list[dict] = []

    params = {
        "details": "SHOW",
        "maxResults": 200,
        "sortOrder": "ASCENDING",
    }

    if sku:
        params["sku"] = sku

    while True:
        res = amazon_client.make_api_call(
            "/awd/2024-05-09/inventory",
            "GET",
            params,
        )

        if not res:
            raise RuntimeError("Empty response from Amazon AWD inventory API")

        if isinstance(res, dict) and res.get("error"):
            msg = str(res)

            if "403" in msg or "Unauthorized" in msg or "forbidden" in msg.lower():
                raise RuntimeError(
                    "SP-API 403 Unauthorized for AWD API. "
                    "Check app authorization and AWD API permissions."
                )

            if "429" in msg or "QuotaExceeded" in msg or "Too Many Requests" in msg:
                raise RuntimeError("Amazon AWD API rate limit exceeded. Try again later.")

            raise RuntimeError(f"Failed to fetch AWD inventory: {res}")

        payload = res.get("payload") if isinstance(res, dict) else None
        if payload is None:
            payload = res

        inventory_items = _extract_awd_inventory_list(payload)

        for item in inventory_items:
            normalized = _normalize_awd_inventory_item(item)
            if normalized:
                rows.append(normalized)

        next_token = _get_awd_next_token(payload)
        if not next_token:
            break

        params = {
            "nextToken": next_token,
            "details": "SHOW",
            "maxResults": 200,
        }

    return rows


def _upsert_inventory_awd_rows(
    rows: list[dict],
    user_id: int,
    marketplace_id: str,
) -> int:
    if not rows:
        return 0

    now = datetime.utcnow()
    db_rows = []

    for r in rows:
        sku = (r.get("sku") or "").strip()
        if not sku:
            continue

        db_rows.append({
            "user_id": user_id,
            "marketplace_id": marketplace_id,
            "sku": sku,

            "total_onhand_quantity": _safe_int_awd(
                r.get("total_onhand_quantity")
            ),
            "total_inbound_quantity": _safe_int_awd(
                r.get("total_inbound_quantity")
            ),
            "available_distributable_quantity": _safe_int_awd(
                r.get("available_distributable_quantity")
            ),
            "reserved_distributable_quantity": _safe_int_awd(
                r.get("reserved_distributable_quantity")
            ),
            "replenishment_quantity": _safe_int_awd(
                r.get("replenishment_quantity")
            ),
            "expiration_details": r.get("expiration_details") or [],

            "synced_at": now,
            "updated_at": now,
        })

    if not db_rows:
        return 0

    stmt = pg_insert(InventoryAWD).values(db_rows)

    stmt = stmt.on_conflict_do_update(
        constraint="uq_inventory_awd_user_marketplace_sku",
        set_={
            "total_onhand_quantity": stmt.excluded.total_onhand_quantity,
            "total_inbound_quantity": stmt.excluded.total_inbound_quantity,
            "available_distributable_quantity": stmt.excluded.available_distributable_quantity,
            "reserved_distributable_quantity": stmt.excluded.reserved_distributable_quantity,
            "replenishment_quantity": stmt.excluded.replenishment_quantity,
            "expiration_details": stmt.excluded.expiration_details,
            "synced_at": stmt.excluded.synced_at,
            "updated_at": stmt.excluded.updated_at,
        },
    )

    db.session.execute(stmt)
    db.session.commit()

    return len(db_rows)


# AWD is currently enabled only for marketplaces where the seller account
# has AWD access. UK does not expose AWD inventory for this integration, so
# never call Amazon's /awd API for the UK marketplace.
AWD_SUPPORTED_MARKETPLACES = {
    "ATVPDKIKX0DER",  # US
}


@inventory_bp.route("/amazon_api/inventory/awd", methods=["GET"])
def sync_inventory_awd_from_amazon():
    """
    Fetch AWD current inventory from Amazon SP-API and save it into inventory_awd.

    Query params:
      marketplace_id optional
      sku optional
      store_in_db=true/false optional, default true

    Example:
      GET /amazon_api/inventory/awd?marketplace_id=ATVPDKIKX0DER
      GET /amazon_api/inventory/awd?marketplace_id=ATVPDKIKX0DER&sku=ABC123
    """

    # ---------------- AUTH ----------------
    auth_header = request.headers.get("Authorization")
    user_id = None

    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ", 1)[1]
        try:
            payload, user_id, member_id = get_effective_user_id_from_token(token)
            user_id = payload.get("user_id") or user_id
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    # ---------------- MARKETPLACE ----------------
    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({
            "success": False,
            "error": "Unsupported marketplace",
        }), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    sku = request.args.get("sku")
    store_in_db = request.args.get("store_in_db", "true").lower() != "false"

    # IMPORTANT: Do not call the Amazon AWD API for UK or any marketplace
    # that is not explicitly supported. Returning HTTP 200 keeps the live
    # dashboard refresh flow running without producing a false 500 error.
    if mp not in AWD_SUPPORTED_MARKETPLACES:
        logger.info(
            "Skipping AWD inventory sync for unsupported marketplace=%s user_id=%s",
            mp,
            user_id,
        )
        return jsonify({
            "success": True,
            "skipped": True,
            "message": "AWD inventory sync is not available for this marketplace",
            "marketplace_id": mp,
            "sku_filter": sku,
            "count": 0,
            "db": {
                "saved_rows": 0,
            },
            "items": [],
        }), 200

    logger.info(
        "Starting AWD inventory sync for marketplace=%s user_id=%s sku=%s",
        mp,
        user_id,
        sku,
    )

    try:
        rows = _fetch_awd_inventory_rows_from_amazon(
            mp=mp,
            sku=sku,
        )

        saved = 0
        if store_in_db:
            saved = _upsert_inventory_awd_rows(
                rows=rows,
                user_id=user_id,
                marketplace_id=mp,
            )

        return jsonify({
            "success": True,
            "message": "AWD inventory fetched from Amazon successfully",
            "marketplace_id": mp,
            "sku_filter": sku,
            "count": len(rows),
            "db": {
                "saved_rows": saved,
            },
            "items": rows,
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.exception("Failed to fetch AWD inventory from Amazon")

        return jsonify({
            "success": False,
            "error": str(e),
        }), 500
    
# ---------------------------------------------------------------------
# Aged Inventory Surcharge Report
# Seller Central page:
# Amazon Fulfillment Reports -> Aged Inventory Surcharge report
# Stores into: inventory_aged_history
# ---------------------------------------------------------------------

AGED_INVENTORY_SURCHARGE_REPORT_TYPE = "GET_FBA_FULFILLMENT_LONGTERM_STORAGE_FEE_CHARGES_DATA"



def _safe_str(val):
    if val is None:
        return None

    val = str(val).strip()
    return val if val != "" else None


def _safe_float_0(val):
    if val is None:
        return 0.0

    try:
        return float(str(val).replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def _safe_int_0(val):
    if val is None:
        return 0

    try:
        return int(float(str(val).replace(",", "").strip() or 0))
    except Exception:
        return 0


def _parse_snapshot_datetime(value):
    """
    Amazon report examples:
      2026-02-15T00:04:00+00:00
      2026-02-15T00:04:00Z
      2026-02-15

    DB value:
      naive UTC datetime
    """

    if not value:
        return None

    try:
        s = str(value).strip()

        if not s:
            return None

        if s.endswith("Z"):
            s = s[:-1] + "+00:00"

        if "T" in s:
            dt = datetime.fromisoformat(s)

            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)

            return dt

        return datetime.strptime(s[:10], "%Y-%m-%d")

    except Exception as e:
        logger.warning(
            "Could not parse aged surcharge snapshot-date=%s | error=%s",
            value,
            e,
        )
        return None


def _resolve_aged_surcharge_date_range():
    """
    Supports:
      ?month=march&year=2026
      ?month=3&year=2026
      ?month=03&year=2026
      ?month=2026-03
      ?start_date=2026-03-01&end_date=2026-03-31

    Returns:
      start_date, end_date, mode
    """

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    month_param = request.args.get("month")
    year_param = request.args.get("year")

    # ---------------- MONTH MODE ----------------
    if month_param and not (start_date or end_date):
        m_raw = month_param.strip().lower()

        if "-" in m_raw:
            # month=2026-03
            y_str, m_str = m_raw.split("-", 1)
            year = int(year_param or y_str)
            month_num = int(m_str)

        elif m_raw.isdigit():
            # month=3&year=2026
            if not year_param:
                raise ValueError(
                    "year is required when month is numeric, example: ?month=3&year=2026"
                )

            year = int(year_param)
            month_num = int(m_raw)

        else:
            # month=march&year=2026
            if not year_param:
                raise ValueError(
                    "year is required when month is a name, example: ?month=march&year=2026"
                )

            year = int(year_param)

            try:
                month_num = datetime.strptime(m_raw.capitalize(), "%B").month
            except ValueError:
                month_num = datetime.strptime(m_raw.capitalize(), "%b").month

        month_start, month_end = _month_range(year, month_num)

        return month_start.isoformat(), month_end.isoformat(), "month"

    # ---------------- RANGE MODE ----------------
    if start_date or end_date:
        if not start_date or not end_date:
            raise ValueError("Provide both start_date and end_date, or use month/year")

        parsed_start = _parse_date_str(start_date)
        parsed_end = _parse_date_str(end_date)

        if parsed_end < parsed_start:
            raise ValueError("end_date cannot be before start_date")

        return parsed_start.isoformat(), parsed_end.isoformat(), "range"

    # ---------------- ALL MODE ----------------
    return None, None, "all"


def _request_aged_inventory_surcharge_report(
    mp: str,
    data_start_time: str | None = None,
    data_end_time: str | None = None,
) -> str:
    """
    Creates Aged Inventory Surcharge report request.
    """

    body = {
        "reportType": AGED_INVENTORY_SURCHARGE_REPORT_TYPE,
        "marketplaceIds": [mp],
    }

    if data_start_time:
        body["dataStartTime"] = data_start_time

    if data_end_time:
        body["dataEndTime"] = data_end_time

    res = amazon_client.make_api_call(
        "/reports/2021-06-30/reports",
        "POST",
        {},
        body,
    )

    if not res:
        raise RuntimeError("Empty response from SP-API while creating aged surcharge report")

    if "error" in res:
        msg = str(res)

        if (
            "403" in msg
            or "Unauthorized" in msg
            or "forbidden" in msg.lower()
            or "Access to requested resource is denied" in msg
        ):
            raise RuntimeError(
                "SP-API 403 Unauthorized for Aged Inventory Surcharge report. "
                "Please check Seller Central report permissions."
            )

        if "QuotaExceeded" in msg or "429" in msg or "Too Many Requests" in msg:
            raise RuntimeError(
                "Amazon rate limit exceeded for Aged Inventory Surcharge report. "
                "Please retry later."
            )

        raise RuntimeError(f"Failed to create aged surcharge report: {msg}")

    payload = res.get("payload") or res
    report_id = payload.get("reportId")

    if not report_id:
        raise RuntimeError(f"No reportId returned for aged surcharge report: {res}")

    logger.info("Created aged inventory surcharge report: %s", report_id)
    return report_id


def _parse_aged_inventory_surcharge_rows(content: bytes) -> list[dict]:
    """
    Parses Amazon Aged Inventory Surcharge report.

    Expected columns:
      snapshot-date
      sku
      fnsku
      asin
      product-name
      condition
      per-unit-volume
      currency
      volume-unit
      country
      qty-charged
      amount-charged
      surcharge-age-tier
      rate-surcharge
    """

    text_content = content.decode("utf-8-sig", errors="replace")

    sample = text_content[:2048]
    delimiter = "\t" if "\t" in sample else ","

    reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)

    if not reader.fieldnames:
        raise RuntimeError("Aged Inventory Surcharge report has no header row")

    rows = []

    for row in reader:
        item = {
            "snapshot-date": _safe_str(row.get("snapshot-date")),
            "sku": _safe_str(row.get("sku")),
            "fnsku": _safe_str(row.get("fnsku")),
            "asin": _safe_str(row.get("asin")),
            "product-name": _safe_str(row.get("product-name")),
            "condition": _safe_str(row.get("condition")),
            "per-unit-volume": _safe_float_0(row.get("per-unit-volume")),
            "currency": _safe_str(row.get("currency")),
            "volume-unit": _safe_str(row.get("volume-unit")),
            "country": _safe_str(row.get("country")),
            "qty-charged": _safe_int_0(row.get("qty-charged")),
            "amount-charged": _safe_float_0(row.get("amount-charged")),
            "surcharge-age-tier": _safe_str(row.get("surcharge-age-tier")),
            "rate-surcharge": _safe_float_0(row.get("rate-surcharge")),
        }

        if not item["sku"] and not item["asin"] and not item["fnsku"]:
            continue

        rows.append(item)

    return rows


def _filter_aged_surcharge_rows_by_date(
    rows: list[dict],
    start_date: str | None,
    end_date: str | None,
) -> list[dict]:
    """
    Local safety filter using snapshot-date.
    start_date/end_date should be YYYY-MM-DD.
    """

    if not start_date and not end_date:
        return rows

    filtered = []

    for r in rows:
        snap = r.get("snapshot-date")
        if not snap:
            continue

        snap_date = str(snap)[:10]

        if start_date and snap_date < start_date:
            continue

        if end_date and snap_date > end_date:
            continue

        filtered.append(r)

    return filtered


def _get_sku_product_column_for_marketplace(marketplace_id: str) -> str | None:
    """
    Decides which SKU column to use from public.sku_{user_id}_data_table.
    """

    country = MARKETPLACE_TO_COUNTRY.get(marketplace_id, "").lower()

    if country == "uk":
        return "sku_uk"

    if country == "us":
        return "sku_us"

    return None


def _attach_product_names_to_aged_surcharge_rows(
    rows: list[dict],
    user_id: int,
    marketplace_id: str,
) -> int:
    """
    Overwrite row['product-name'] from public.sku_{user_id}_data_table.

    UK:
      row['sku'] == sku_uk

    US:
      row['sku'] == sku_us

    Returns count of mapped product names.
    """

    if not rows or not user_id:
        return 0

    sku_column = _get_sku_product_column_for_marketplace(marketplace_id)

    if not sku_column:
        logger.info(
            "Skipping aged surcharge product-name mapping for marketplace_id=%s",
            marketplace_id,
        )
        return 0

    skus = sorted({
        (r.get("sku") or "").strip()
        for r in rows
        if (r.get("sku") or "").strip()
    })

    if not skus:
        return 0

    sku_table = f"public.sku_{user_id}_data_table"

    placeholders = ", ".join(f":sku{i}" for i in range(len(skus)))

    sql = text(f"""
        SELECT {sku_column} AS sku, product_name
        FROM {sku_table}
        WHERE {sku_column} IN ({placeholders})
    """)

    params = {f"sku{i}": sku for i, sku in enumerate(skus)}

    try:
        result = db.session.execute(sql, params).mappings().all()
    except Exception as e:
        logger.exception(
            "Failed to read SKU product mapping table=%s column=%s",
            sku_table,
            sku_column,
        )
        return 0

    mapping = {
        (r.get("sku") or "").strip(): r.get("product_name")
        for r in result
        if (r.get("sku") or "").strip() and r.get("product_name")
    }

    mapped_count = 0

    for r in rows:
        sku = (r.get("sku") or "").strip()
        mapped_name = mapping.get(sku)

        if mapped_name:
            r["product-name"] = mapped_name
            mapped_count += 1

    return mapped_count


def _row_to_inventory_aged_history(
    row: dict,
    user_id: int,
    marketplace_id: str,
    report_id: str | None = None,
    document_id: str | None = None,
) -> dict:
    snapshot_dt = _parse_snapshot_datetime(row.get("snapshot-date"))

    return {
        "user_id": user_id,
        "marketplace_id": marketplace_id,
        "report_id": report_id,
        "document_id": document_id,

        "snapshot_date": snapshot_dt,

        "sku": row.get("sku"),
        "fnsku": row.get("fnsku"),
        "asin": row.get("asin"),
        "product_name": row.get("product-name"),
        "condition": row.get("condition"),

        "per_unit_volume": _safe_float_0(row.get("per-unit-volume")),
        "currency": row.get("currency"),
        "volume_unit": row.get("volume-unit"),
        "country": row.get("country"),

        "qty_charged": _safe_int_0(row.get("qty-charged")),
        "amount_charged": _safe_float_0(row.get("amount-charged")),
        "surcharge_age_tier": row.get("surcharge-age-tier"),
        "rate_surcharge": _safe_float_0(row.get("rate-surcharge")),

        "synced_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
    }


def _upsert_inventory_aged_history_rows(
    rows: list[dict],
    user_id: int,
    marketplace_id: str,
    report_id: str | None = None,
    document_id: str | None = None,
) -> tuple[int, int, int]:
    """
    Upsert parsed surcharge report rows into inventory_aged_history.

    Returns:
      saved_rows, skipped_missing_snapshot, mapped_product_names
    """

    if not rows:
        return 0, 0, 0

    mapped_product_names = _attach_product_names_to_aged_surcharge_rows(
        rows=rows,
        user_id=user_id,
        marketplace_id=marketplace_id,
    )

    db_rows = []
    skipped_missing_snapshot = 0

    for r in rows:
        sku = (r.get("sku") or "").strip()
        asin = (r.get("asin") or "").strip()
        fnsku = (r.get("fnsku") or "").strip()

        if not sku and not asin and not fnsku:
            continue

        mapped = _row_to_inventory_aged_history(
            row=r,
            user_id=user_id,
            marketplace_id=marketplace_id,
            report_id=report_id,
            document_id=document_id,
        )

        if not mapped.get("snapshot_date"):
            skipped_missing_snapshot += 1

            logger.warning(
                "Skipping aged surcharge row because snapshot_date is missing/unparseable. "
                "sku=%s raw_snapshot=%s",
                sku,
                r.get("snapshot-date"),
            )
            continue

        db_rows.append(mapped)

    if not db_rows:
        return 0, skipped_missing_snapshot, mapped_product_names

    stmt = pg_insert(InventoryAgedHistory).values(db_rows)

    stmt = stmt.on_conflict_do_update(
        constraint="uq_inventory_aged_history_key",
        set_={
            "report_id": stmt.excluded.report_id,
            "document_id": stmt.excluded.document_id,

            "product_name": stmt.excluded.product_name,
            "condition": stmt.excluded.condition,
            "per_unit_volume": stmt.excluded.per_unit_volume,
            "currency": stmt.excluded.currency,
            "volume_unit": stmt.excluded.volume_unit,
            "country": stmt.excluded.country,

            "qty_charged": stmt.excluded.qty_charged,
            "amount_charged": stmt.excluded.amount_charged,
            "rate_surcharge": stmt.excluded.rate_surcharge,

            "synced_at": stmt.excluded.synced_at,
            "updated_at": stmt.excluded.updated_at,
        },
    )

    db.session.execute(stmt)
    db.session.commit()

    return len(db_rows), skipped_missing_snapshot, mapped_product_names


@inventory_bp.route("/amazon_api/inventory/aged-surcharge", methods=["GET"])
def fetch_aged_inventory_surcharge():
    # ---------------- AUTH ----------------
    auth_header = request.headers.get("Authorization")
    user_id = None

    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]

        try:
            payload, user_id, member_id = get_effective_user_id_from_token(token)
            user_id = payload.get("user_id")

        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401

        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    # ---------------- MARKETPLACE ----------------
    _apply_region_and_marketplace_from_request()

    if amazon_client.marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({
            "success": False,
            "error": "Unsupported marketplace",
        }), 400

    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
    store_in_db = request.args.get("store_in_db", "true").lower() != "false"

    # ---------------- DATE / MONTH PARAMS ----------------
    try:
        start_date, end_date, mode = _resolve_aged_surcharge_date_range()

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Invalid date/month parameters: {str(e)}",
        }), 400

    data_start_time = f"{start_date}T00:00:00Z" if start_date else None
    data_end_time = f"{end_date}T23:59:59Z" if end_date else None

    # ---------------- REQUEST REPORT ----------------
    try:
        report_id = _request_aged_inventory_surcharge_report(
            mp=mp,
            data_start_time=data_start_time,
            data_end_time=data_end_time,
        )

    except RuntimeError as e:
        msg = str(e)
        msg_lower = msg.lower()

        if (
            "403" in msg
            or "unauthorized" in msg_lower
            or "access to requested resource is denied" in msg_lower
            or "forbidden" in msg_lower
        ):
            status = 403

        elif "rate limit" in msg_lower or "quota" in msg_lower or "429" in msg:
            status = 429

        else:
            status = 500

        return jsonify({
            "success": False,
            "error": msg,
        }), status

    # ---------------- POLL REPORT ----------------
    doc_id = _wait_for_report(report_id, timeout_sec=600, poll_interval=20)

    if not doc_id:
        return jsonify({
            "success": False,
            "error": "Aged Inventory Surcharge report did not complete",
            "report_id": report_id,
            "start_date": start_date,
            "end_date": end_date,
        }), 502

    # ---------------- DOWNLOAD REPORT ----------------
    content = _download_report_document(doc_id)

    if not content:
        return jsonify({
            "success": False,
            "error": "Failed to download Aged Inventory Surcharge report document",
            "report_id": report_id,
            "document_id": doc_id,
            "start_date": start_date,
            "end_date": end_date,
        }), 502

    # ---------------- PARSE REPORT ----------------
    try:
        rows = _parse_aged_inventory_surcharge_rows(content)

    except Exception as e:
        logger.exception("Failed to parse aged surcharge report")

        return jsonify({
            "success": False,
            "error": f"Failed to parse report: {str(e)}",
            "report_id": report_id,
            "document_id": doc_id,
        }), 500

    rows_in_report = len(rows)

    # ---------------- LOCAL DATE FILTER ----------------
    rows = _filter_aged_surcharge_rows_by_date(
        rows=rows,
        start_date=start_date,
        end_date=end_date,
    )

    # ---------------- SAVE DB ----------------
    saved_rows = 0
    skipped_missing_snapshot = 0
    mapped_product_names = 0

    if store_in_db:
        try:
            (
                saved_rows,
                skipped_missing_snapshot,
                mapped_product_names,
            ) = _upsert_inventory_aged_history_rows(
                rows=rows,
                user_id=user_id,
                marketplace_id=mp,
                report_id=report_id,
                document_id=doc_id,
            )

        except Exception as e:
            db.session.rollback()
            logger.exception("Failed to save aged surcharge rows")

            return jsonify({
                "success": False,
                "error": f"Failed to save inventory_aged_history rows: {str(e)}",
                "report_id": report_id,
                "document_id": doc_id,
                "start_date": start_date,
                "end_date": end_date,
            }), 500

    # ---------------- RESPONSE ----------------
    return jsonify({
        "success": True,
        "marketplace_id": mp,
        "country": MARKETPLACE_TO_COUNTRY.get(mp),
        "report_type": AGED_INVENTORY_SURCHARGE_REPORT_TYPE,
        "report_id": report_id,
        "document_id": doc_id,

        "mode": mode,
        "start_date": start_date,
        "end_date": end_date,

        "rows_in_report": rows_in_report,
        "count": len(rows),

        "db": {
            "table": "inventory_aged_history",
            "store_in_db": store_in_db,
            "saved_rows": saved_rows,
            "skipped_missing_snapshot": skipped_missing_snapshot,
            "mapped_product_names": mapped_product_names,
        },

        "data": rows,
    }), 200




# # =============================================================================
# # AWD EXCEL EXPORT HELPERS
# # =============================================================================

# _AWD_XLSX_MIMETYPE = (
#     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )
# _AWD_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
# _AWD_HEADER_FONT = Font(color="FFFFFF", bold=True)
# _AWD_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
# _AWD_ROW_BORDER = Border(bottom=Side(style="thin", color="D9E1F2"))


# def _awd_quantity_value(value) -> float:
#     if isinstance(value, dict):
#         value = value.get("quantity")
#     try:
#         return float(value or 0)
#     except (TypeError, ValueError):
#         return 0.0


# def _awd_quantity_unit(value) -> str:
#     if not isinstance(value, dict):
#         return ""
#     return str(value.get("unitOfMeasurement") or "")


# def _awd_product_attribute(product: dict, name: str):
#     for attribute in product.get("attributes") or []:
#         if str(attribute.get("name") or "").strip().lower() == name.lower():
#             return attribute.get("value")
#     return None


# def _awd_style_sheet(ws, freeze_panes: str = "A2") -> None:
#     ws.sheet_view.showGridLines = False
#     ws.freeze_panes = freeze_panes

#     if ws.max_row >= 1 and ws.max_column >= 1:
#         ws.auto_filter.ref = ws.dimensions

#     for cell in ws[1]:
#         cell.fill = _AWD_HEADER_FILL
#         cell.font = _AWD_HEADER_FONT
#         cell.alignment = Alignment(
#             horizontal="center",
#             vertical="center",
#             wrap_text=True,
#         )

#     ws.row_dimensions[1].height = 30

#     for row in ws.iter_rows(min_row=2):
#         for cell in row:
#             cell.alignment = Alignment(vertical="top", wrap_text=True)
#             cell.border = _AWD_ROW_BORDER

#     for column_cells in ws.columns:
#         letter = get_column_letter(column_cells[0].column)
#         max_length = 0
#         for cell in column_cells:
#             if cell.value is not None:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)


# def _awd_workbook_stream(workbook: Workbook) -> io.BytesIO:
#     output = io.BytesIO()
#     workbook.save(output)
#     output.seek(0)
#     return output


# def _build_awd_shipment_list_excel(shipments: list[dict], marketplace_id: str) -> io.BytesIO:
#     workbook = Workbook()
#     ws = workbook.active
#     ws.title = "Inbound Shipments"
#     ws.append([
#         "Marketplace ID",
#         "Shipment ID",
#         "Order ID",
#         "External Reference ID",
#         "Shipment Status",
#         "Created At",
#         "Updated At",
#     ])

#     for shipment in shipments:
#         ws.append([
#             marketplace_id,
#             shipment.get("shipmentId"),
#             shipment.get("orderId"),
#             shipment.get("externalReferenceId"),
#             shipment.get("shipmentStatus"),
#             shipment.get("createdAt"),
#             shipment.get("updatedAt"),
#         ])

#     _awd_style_sheet(ws)

#     summary_ws = workbook.create_sheet("Summary")
#     summary_ws.append(["Field", "Value"])
#     summary_ws.append(["Marketplace ID", marketplace_id])
#     summary_ws.append(["Total Shipments", len(shipments)])

#     status_counts: dict[str, int] = {}
#     for shipment in shipments:
#         status = str(shipment.get("shipmentStatus") or "UNKNOWN")
#         status_counts[status] = status_counts.get(status, 0) + 1

#     for status, count in sorted(status_counts.items()):
#         summary_ws.append([f"{status} Shipments", count])

#     _awd_style_sheet(summary_ws)
#     return _awd_workbook_stream(workbook)


# def _build_awd_shipment_detail_excel(shipment: dict, marketplace_id: str) -> io.BytesIO:
#     workbook = Workbook()

#     summary_ws = workbook.active
#     summary_ws.title = "Shipment Summary"
#     summary_ws.append(["Field", "Value"])

#     carrier = shipment.get("carrierCode") or {}
#     received_quantities = shipment.get("receivedQuantity") or []
#     total_received = sum(_awd_quantity_value(item) for item in received_quantities)
#     received_unit = _awd_quantity_unit(received_quantities[0]) if received_quantities else ""

#     summary_rows = [
#         ("Marketplace ID", marketplace_id),
#         ("Shipment ID", shipment.get("shipmentId")),
#         ("Order ID", shipment.get("orderId")),
#         ("External Reference ID", shipment.get("externalReferenceId")),
#         ("Warehouse Reference ID", shipment.get("warehouseReferenceId")),
#         ("Shipment Status", shipment.get("shipmentStatus")),
#         ("Created At", shipment.get("createdAt")),
#         ("Updated At", shipment.get("updatedAt")),
#         ("Ship By", shipment.get("shipBy")),
#         ("Tracking ID", shipment.get("trackingId")),
#         ("Destination Region", shipment.get("destinationRegion")),
#         ("Carrier Code Type", carrier.get("carrierCodeType")),
#         ("Carrier Code Value", carrier.get("carrierCodeValue")),
#         ("Total Received Quantity", total_received),
#         ("Received Unit", received_unit),
#     ]
#     for label, value in summary_rows:
#         summary_ws.append([label, "" if value is None else value])
#     _awd_style_sheet(summary_ws)

#     sku_ws = workbook.create_sheet("SKU Quantities")
#     sku_ws.append([
#         "Marketplace ID",
#         "Shipment ID",
#         "SKU",
#         "Expected Quantity",
#         "Expected Unit",
#         "Received Quantity",
#         "Received Unit",
#         "Difference",
#         "Receipt Percentage",
#     ])

#     sku_rows = shipment.get("shipmentSkuQuantities") or []
#     for item in sku_rows:
#         expected_obj = item.get("expectedQuantity") or {}
#         received_obj = item.get("receivedQuantity") or {}
#         expected = _awd_quantity_value(expected_obj)
#         received = _awd_quantity_value(received_obj)
#         receipt_percentage = (received / expected) if expected else None
#         sku_ws.append([
#             marketplace_id,
#             shipment.get("shipmentId"),
#             item.get("sku"),
#             expected,
#             _awd_quantity_unit(expected_obj),
#             received,
#             _awd_quantity_unit(received_obj),
#             received - expected,
#             receipt_percentage,
#         ])

#     if sku_rows:
#         total_row = sku_ws.max_row + 2
#         sku_ws.cell(total_row, 3, "Total")
#         sku_ws.cell(total_row, 4, f"=SUM(D2:D{sku_ws.max_row})")
#         sku_ws.cell(total_row, 6, f"=SUM(F2:F{sku_ws.max_row})")
#         sku_ws.cell(total_row, 8, f"=SUM(H2:H{sku_ws.max_row})")
#         for cell in sku_ws[total_row]:
#             cell.font = Font(bold=True)
#             cell.fill = _AWD_TOTAL_FILL

#     for cell in sku_ws["I"][1:]:
#         cell.number_format = "0.0%"
#     _awd_style_sheet(sku_ws)

#     container_ws = workbook.create_sheet("Container Details")
#     container_ws.append([
#         "Shipment ID",
#         "SKU",
#         "ASIN",
#         "Product Group",
#         "Container Type",
#         "Case Count",
#         "Units Per Case",
#         "Expected Product Units",
#         "Expiration",
#         "Prep Category",
#         "Prep Owner",
#         "Label Owner",
#         "Prep Instructions",
#         "Length",
#         "Width",
#         "Height",
#         "Dimension Unit",
#         "Weight",
#         "Weight Unit",
#     ])

#     for container in shipment.get("shipmentContainerQuantities") or []:
#         distribution_package = container.get("distributionPackage") or {}
#         contents = distribution_package.get("contents") or {}
#         products = contents.get("products") or []
#         measurements = distribution_package.get("measurements") or {}
#         dimensions = measurements.get("dimensions") or {}
#         weight = measurements.get("weight") or {}
#         case_count = _awd_quantity_value(container.get("count"))

#         if not products:
#             products = [{}]

#         for product in products:
#             units_per_case = _awd_quantity_value(product.get("quantity"))
#             prep = product.get("prepDetails") or {}
#             prep_text = ", ".join(
#                 str(instruction.get("prepType"))
#                 for instruction in prep.get("prepInstructions") or []
#                 if instruction.get("prepType")
#             )
#             container_ws.append([
#                 shipment.get("shipmentId"),
#                 product.get("sku"),
#                 _awd_product_attribute(product, "asin"),
#                 _awd_product_attribute(product, "glProductGroupType"),
#                 distribution_package.get("type"),
#                 case_count,
#                 units_per_case,
#                 case_count * units_per_case,
#                 product.get("expiration"),
#                 prep.get("prepCategory"),
#                 prep.get("prepOwner"),
#                 prep.get("labelOwner"),
#                 prep_text,
#                 dimensions.get("length"),
#                 dimensions.get("width"),
#                 dimensions.get("height"),
#                 dimensions.get("unitOfMeasurement"),
#                 weight.get("weight"),
#                 weight.get("unitOfMeasurement"),
#             ])
#     _awd_style_sheet(container_ws)

#     address_ws = workbook.create_sheet("Addresses")
#     address_ws.append([
#         "Address Type",
#         "Name",
#         "Address Line 1",
#         "Address Line 2",
#         "Address Line 3",
#         "City",
#         "District",
#         "County",
#         "State / Region",
#         "Postal Code",
#         "Country Code",
#         "Phone Number",
#     ])
#     for address_type, address in (
#         ("Origin", shipment.get("originAddress") or {}),
#         ("Destination", shipment.get("destinationAddress") or {}),
#     ):
#         address_ws.append([
#             address_type,
#             address.get("name"),
#             address.get("addressLine1"),
#             address.get("addressLine2"),
#             address.get("addressLine3"),
#             address.get("city"),
#             address.get("district"),
#             address.get("county"),
#             address.get("stateOrRegion"),
#             address.get("postalCode"),
#             address.get("countryCode"),
#             address.get("phoneNumber"),
#         ])
#     _awd_style_sheet(address_ws)

#     return _awd_workbook_stream(workbook)


# def _awd_extract_shipments(payload: dict) -> list[dict]:
#     payload = payload or {}
#     shipments = payload.get("shipments") or payload.get("inboundShipments") or payload.get("items") or []
#     return shipments if isinstance(shipments, list) else []


# def _awd_extract_next_token(payload: dict):
#     payload = payload or {}
#     pagination = payload.get("pagination") or {}
#     return pagination.get("nextToken") or payload.get("nextToken")


# def _awd_iso_year(value) -> int | None:
#     if not value:
#         return None
#     try:
#         return int(str(value)[:4])
#     except (TypeError, ValueError):
#         return None


# def _fetch_all_awd_inbound_shipments(params: dict) -> tuple[list[dict], int]:
#     """Fetch every AWD inbound-shipment page returned by Amazon."""
#     endpoint = "/awd/2024-05-09/inboundShipments"
#     request_params = dict(params or {})
#     all_shipments: list[dict] = []
#     pages_fetched = 0
#     seen_tokens: set[str] = set()

#     while True:
#         response = amazon_client.make_api_call(endpoint, "GET", request_params)
#         pages_fetched += 1

#         if not response:
#             raise RuntimeError("Amazon returned an empty response")
#         if response.get("error"):
#             error = RuntimeError(str(response.get("error") or "Amazon AWD request failed"))
#             error.amazon_response = response
#             raise error

#         payload = response.get("payload") or response
#         all_shipments.extend(_awd_extract_shipments(payload))

#         next_token = _awd_extract_next_token(payload)
#         if not next_token:
#             break
#         if str(next_token) in seen_tokens:
#             logger.warning("Stopping AWD shipment pagination because nextToken repeated")
#             break

#         seen_tokens.add(str(next_token))
#         request_params = dict(params or {})
#         request_params["nextToken"] = next_token

#     # Deduplicate defensively by shipmentId.
#     deduped: dict[str, dict] = {}
#     no_id_rows: list[dict] = []
#     for shipment in all_shipments:
#         shipment_id = str(shipment.get("shipmentId") or "").strip()
#         if shipment_id:
#             deduped[shipment_id] = shipment
#         else:
#             no_id_rows.append(shipment)

#     rows = list(deduped.values()) + no_id_rows
#     rows.sort(key=lambda item: str(item.get("createdAt") or ""), reverse=True)
#     return rows, pages_fetched


# def _fetch_awd_shipment_detail_for_export(shipment_id: str) -> dict:
#     response = amazon_client.make_api_call(
#         f"/awd/2024-05-09/inboundShipments/{shipment_id}",
#         "GET",
#         {"skuQuantities": "SHOW"},
#     )
#     if not response:
#         raise RuntimeError("Empty detail response")
#     if response.get("error"):
#         raise RuntimeError(str(response.get("error") or "Shipment detail request failed"))
#     return response.get("payload") or response


# def _build_awd_complete_year_excel(
#     shipments: list[dict],
#     shipment_details: list[dict],
#     marketplace_id: str,
#     requested_year: int,
#     detail_errors: list[dict],
# ) -> io.BytesIO:
#     """Build one workbook containing list rows and complete details for all shipments."""
#     workbook = Workbook()

#     list_ws = workbook.active
#     list_ws.title = "Inbound Shipments"
#     list_ws.append([
#         "Marketplace ID", "Shipment ID", "Order ID", "External Reference ID",
#         "Shipment Status", "Created At", "Updated At",
#     ])
#     for shipment in shipments:
#         list_ws.append([
#             marketplace_id, shipment.get("shipmentId"), shipment.get("orderId"),
#             shipment.get("externalReferenceId"), shipment.get("shipmentStatus"),
#             shipment.get("createdAt"), shipment.get("updatedAt"),
#         ])
#     _awd_style_sheet(list_ws)

#     summary_ws = workbook.create_sheet("Summary")
#     summary_ws.append(["Field", "Value"])
#     summary_ws.append(["Marketplace ID", marketplace_id])
#     summary_ws.append(["Requested Year", requested_year])
#     summary_ws.append(["Shipment Count", len(shipments)])
#     summary_ws.append(["Detail Records Loaded", len(shipment_details)])
#     summary_ws.append(["Detail Errors", len(detail_errors)])
#     for status in sorted({str(x.get("shipmentStatus") or "UNKNOWN") for x in shipments}):
#         summary_ws.append([
#             f"{status} Shipments",
#             sum(1 for x in shipments if str(x.get("shipmentStatus") or "UNKNOWN") == status),
#         ])
#     _awd_style_sheet(summary_ws)

#     sku_ws = workbook.create_sheet("SKU Quantities")
#     sku_ws.append([
#         "Marketplace ID", "Shipment ID", "Order ID", "Shipment Status", "Created At",
#         "SKU", "Expected Quantity", "Expected Unit", "Received Quantity",
#         "Received Unit", "Difference", "Receipt Percentage",
#     ])

#     container_ws = workbook.create_sheet("Container Details")
#     container_ws.append([
#         "Marketplace ID", "Shipment ID", "Order ID", "Shipment Status", "SKU", "ASIN",
#         "Product Group", "Container Type", "Case Count", "Units Per Case",
#         "Expected Product Units", "Expiration", "Prep Category", "Prep Owner",
#         "Label Owner", "Prep Instructions", "Length", "Width", "Height",
#         "Dimension Unit", "Weight", "Weight Unit",
#     ])

#     address_ws = workbook.create_sheet("Addresses")
#     address_ws.append([
#         "Marketplace ID", "Shipment ID", "Address Type", "Name", "Address Line 1",
#         "Address Line 2", "Address Line 3", "City", "District", "County",
#         "State / Region", "Postal Code", "Country Code", "Phone Number",
#     ])

#     detail_ws = workbook.create_sheet("Shipment Details")
#     detail_ws.append([
#         "Marketplace ID", "Shipment ID", "Order ID", "External Reference ID",
#         "Warehouse Reference ID", "Shipment Status", "Created At", "Updated At",
#         "Ship By", "Tracking ID", "Destination Region", "Carrier Code Type",
#         "Carrier Code Value", "Total Received Quantity", "Received Unit",
#     ])

#     for shipment in shipment_details:
#         carrier = shipment.get("carrierCode") or {}
#         received_quantities = shipment.get("receivedQuantity") or []
#         total_received = sum(_awd_quantity_value(x) for x in received_quantities)
#         received_unit = _awd_quantity_unit(received_quantities[0]) if received_quantities else ""
#         detail_ws.append([
#             marketplace_id, shipment.get("shipmentId"), shipment.get("orderId"),
#             shipment.get("externalReferenceId"), shipment.get("warehouseReferenceId"),
#             shipment.get("shipmentStatus"), shipment.get("createdAt"), shipment.get("updatedAt"),
#             shipment.get("shipBy"), shipment.get("trackingId"), shipment.get("destinationRegion"),
#             carrier.get("carrierCodeType"), carrier.get("carrierCodeValue"), total_received, received_unit,
#         ])

#         for item in shipment.get("shipmentSkuQuantities") or []:
#             expected_obj = item.get("expectedQuantity") or {}
#             received_obj = item.get("receivedQuantity") or {}
#             expected = _awd_quantity_value(expected_obj)
#             received = _awd_quantity_value(received_obj)
#             sku_ws.append([
#                 marketplace_id, shipment.get("shipmentId"), shipment.get("orderId"),
#                 shipment.get("shipmentStatus"), shipment.get("createdAt"), item.get("sku"),
#                 expected, _awd_quantity_unit(expected_obj), received,
#                 _awd_quantity_unit(received_obj), received - expected,
#                 (received / expected) if expected else None,
#             ])

#         for container in shipment.get("shipmentContainerQuantities") or []:
#             package = container.get("distributionPackage") or {}
#             contents = package.get("contents") or {}
#             products = contents.get("products") or [{}]
#             measurements = package.get("measurements") or {}
#             dimensions = measurements.get("dimensions") or {}
#             weight = measurements.get("weight") or {}
#             case_count = _awd_quantity_value(container.get("count"))
#             for product in products:
#                 units_per_case = _awd_quantity_value(product.get("quantity"))
#                 prep = product.get("prepDetails") or {}
#                 prep_text = ", ".join(
#                     str(i.get("prepType")) for i in prep.get("prepInstructions") or [] if i.get("prepType")
#                 )
#                 container_ws.append([
#                     marketplace_id, shipment.get("shipmentId"), shipment.get("orderId"),
#                     shipment.get("shipmentStatus"), product.get("sku"),
#                     _awd_product_attribute(product, "asin"),
#                     _awd_product_attribute(product, "glProductGroupType"), package.get("type"),
#                     case_count, units_per_case, case_count * units_per_case,
#                     product.get("expiration"), prep.get("prepCategory"), prep.get("prepOwner"),
#                     prep.get("labelOwner"), prep_text, dimensions.get("length"),
#                     dimensions.get("width"), dimensions.get("height"),
#                     dimensions.get("unitOfMeasurement"), weight.get("weight"),
#                     weight.get("unitOfMeasurement"),
#                 ])

#         for address_type, address in (
#             ("Origin", shipment.get("originAddress") or {}),
#             ("Destination", shipment.get("destinationAddress") or {}),
#         ):
#             address_ws.append([
#                 marketplace_id, shipment.get("shipmentId"), address_type, address.get("name"),
#                 address.get("addressLine1"), address.get("addressLine2"), address.get("addressLine3"),
#                 address.get("city"), address.get("district"), address.get("county"),
#                 address.get("stateOrRegion"), address.get("postalCode"), address.get("countryCode"),
#                 address.get("phoneNumber"),
#             ])

#     if detail_errors:
#         error_ws = workbook.create_sheet("Detail Errors")
#         error_ws.append(["Shipment ID", "Error"])
#         for item in detail_errors:
#             error_ws.append([item.get("shipment_id"), item.get("error")])
#         _awd_style_sheet(error_ws)

#     for ws in (sku_ws, container_ws, address_ws, detail_ws):
#         _awd_style_sheet(ws)
#     for cell in sku_ws["L"][1:]:
#         cell.number_format = "0.0%"

#     return _awd_workbook_stream(workbook)


# @inventory_bp.route("/amazon_api/awd/inbound-shipments", methods=["GET"])
# def list_awd_inbound_shipments():
#     """
#     List all AWD inbound shipments, with automatic pagination.

#     Excel example for the complete 2026 data available from Amazon:
#       /amazon_api/awd/inbound-shipments
#         ?marketplace_id=ATVPDKIKX0DER
#         &year=2026
#         &format=excel
#         &include_details=true
#     """
#     auth_header = request.headers.get("Authorization")
#     if not auth_header or not auth_header.startswith("Bearer "):
#         return jsonify({"success": False, "error": "Missing Authorization header"}), 401

#     token = auth_header.split(" ", 1)[1].strip()
#     try:
#         payload, user_id, member_id = get_effective_user_id_from_token(token)
#         user_id = payload.get("user_id")
#     except jwt.ExpiredSignatureError:
#         return jsonify({"success": False, "error": "Token has expired"}), 401
#     except jwt.InvalidTokenError:
#         return jsonify({"success": False, "error": "Invalid token"}), 401

#     if not user_id:
#         return jsonify({"success": False, "error": "Invalid token payload"}), 401

#     _apply_region_and_marketplace_from_request()
#     mp = request.args.get("marketplace_id", amazon_client.marketplace_id)
#     if mp not in amazon_client.ALLOWED_MARKETPLACES:
#         return jsonify({
#             "success": False,
#             "error": "Unsupported marketplace",
#             "marketplace_id": mp,
#         }), 400

#     shipment_status = (
#         request.args.get("shipment_status") or request.args.get("status") or ""
#     ).strip().upper()
#     allowed_statuses = {
#         "CREATED", "SHIPPED", "IN_TRANSIT", "RECEIVING",
#         "DELIVERED", "CLOSED", "CANCELLED",
#     }
#     if shipment_status and shipment_status not in allowed_statuses:
#         return jsonify({
#             "success": False,
#             "error": "Invalid shipment_status",
#             "allowed_statuses": sorted(allowed_statuses),
#         }), 400

#     try:
#         requested_year = int(request.args.get("year") or datetime.utcnow().year)
#         if requested_year < 2000 or requested_year > 2100:
#             raise ValueError
#     except ValueError:
#         return jsonify({"success": False, "error": "year must be between 2000 and 2100"}), 400

#     try:
#         page_size = int(request.args.get("max_results") or 100)
#         if page_size < 1:
#             raise ValueError
#         page_size = min(page_size, 100)
#     except ValueError:
#         return jsonify({"success": False, "error": "max_results must be a positive integer"}), 400

#     params = {"maxResults": page_size}
#     if shipment_status:
#         params["shipmentStatus"] = shipment_status

#     try:
#         shipments, pages_fetched = _fetch_all_awd_inbound_shipments(params)
#     except Exception as exc:
#         logger.exception("AWD listInboundShipments request failed")
#         amazon_response = getattr(exc, "amazon_response", None) or {}
#         status_code = int(amazon_response.get("status_code") or 502)
#         return jsonify({
#             "success": False,
#             "error": "Failed to list AWD inbound shipments",
#             "detail": str(exc),
#             "amazon_errors": (amazon_response.get("response_json") or {}).get("errors", []),
#         }), status_code

#     # Amazon's list operation does not guarantee a year filter, so filter locally.
#     shipments = [
#         shipment for shipment in shipments
#         if _awd_iso_year(shipment.get("createdAt")) == requested_year
#     ]

#     db_result = {"saved_rows": 0, "table": "public.inventory_awd_inbound_shipments"}
    if store_in_db and complete_items:
        try:
            db_result["saved_rows"] = _save_awd_inbound_shipments(
                complete_items=complete_items,
                user_id=user_id,
                marketplace_id=mp,
            )
        except Exception as exc:
            logger.exception("Failed to save AWD inbound shipments")
            db_result["error"] = str(exc)

    export_format = (request.args.get("format") or "json").strip().lower()
#     include_details = (request.args.get("include_details") or "true").strip().lower() != "false"

#     if export_format == "excel":
#         details: list[dict] = []
#         detail_errors: list[dict] = []

#         if include_details:
#             for shipment in shipments:
#                 shipment_id = str(shipment.get("shipmentId") or "").strip()
#                 if not shipment_id:
#                     continue
#                 try:
#                     details.append(_fetch_awd_shipment_detail_for_export(shipment_id))
#                 except Exception as exc:
#                     logger.exception("Could not fetch AWD detail for %s", shipment_id)
#                     detail_errors.append({"shipment_id": shipment_id, "error": str(exc)})
#         else:
#             details = shipments

#         excel_stream = _build_awd_complete_year_excel(
#             shipments=shipments,
#             shipment_details=details,
#             marketplace_id=mp,
#             requested_year=requested_year,
#             detail_errors=detail_errors,
#         )
#         return send_file(
#             excel_stream,
#             as_attachment=True,
#             download_name=f"AWD_Inbound_Shipments_{mp}_{requested_year}_Complete.xlsx",
#             mimetype=_AWD_XLSX_MIMETYPE,
#             max_age=0,
#         )

#     if export_format not in {"json", ""}:
#         return jsonify({"success": False, "error": "format must be json or excel"}), 400

#     return jsonify({
#         "success": True,
#         "marketplace_id": mp,
#         "year": requested_year,
#         "pages_fetched": pages_fetched,
#         "count": len(shipments),
#         "next_token": None,
#         "items": shipments,
#     }), 200

# @inventory_bp.route(
#     "/amazon_api/awd/inbound-shipment/<string:shipment_id>",
#     methods=["GET"],
# )
# def get_awd_inbound_shipment(shipment_id):
#     """
#     Fetch one AWD inbound shipment using a real shipment ID.

#     Example:
#     GET /amazon_api/awd/inbound-shipment/<REAL_ID>
#         ?marketplace_id=ATVPDKIKX0DER
#         &sku_quantities=SHOW
#     """

#     # ---------------- AUTH ----------------
#     auth_header = request.headers.get("Authorization")

#     if not auth_header or not auth_header.startswith("Bearer "):
#         return jsonify({
#             "success": False,
#             "error": "Missing Authorization header",
#         }), 401

#     token = auth_header.split(" ", 1)[1].strip()

#     try:
#         payload, user_id, member_id = get_effective_user_id_from_token(token)
#         user_id = payload.get("user_id")
#     except jwt.ExpiredSignatureError:
#         return jsonify({
#             "success": False,
#             "error": "Token has expired",
#         }), 401
#     except jwt.InvalidTokenError:
#         return jsonify({
#             "success": False,
#             "error": "Invalid token",
#         }), 401

#     if not user_id:
#         return jsonify({
#             "success": False,
#             "error": "Invalid token payload",
#         }), 401

#     shipment_id = (shipment_id or "").strip()

#     if not shipment_id:
#         return jsonify({
#             "success": False,
#             "error": "shipment_id is required",
#         }), 400

#     if shipment_id.upper() in {
#         "SHIPMENT_ID",
#         "{SHIPMENTID}",
#         "{SHIPMENT_ID}",
#     }:
#         return jsonify({
#             "success": False,
#             "error": (
#                 "Replace SHIPMENT_ID with a real AWD shipment ID returned "
#                 "by /amazon_api/awd/inbound-shipments."
#             ),
#         }), 400

#     # ---------------- MARKETPLACE / REGION ----------------
#     _apply_region_and_marketplace_from_request()

#     mp = request.args.get(
#         "marketplace_id",
#         amazon_client.marketplace_id,
#     )

#     if mp not in amazon_client.ALLOWED_MARKETPLACES:
#         return jsonify({
#             "success": False,
#             "error": "Unsupported marketplace",
#             "marketplace_id": mp,
#         }), 400

#     sku_quantities = (
#         request.args.get("sku_quantities", "SHOW")
#         .strip()
#         .upper()
#     )

#     if sku_quantities not in {"SHOW", "HIDE"}:
#         return jsonify({
#             "success": False,
#             "error": "sku_quantities must be SHOW or HIDE",
#         }), 400

#     endpoint = (
#         f"/awd/2024-05-09/inboundShipments/{shipment_id}"
#     )

#     params = {
#         "skuQuantities": sku_quantities,
#     }

#     try:
#         amazon_response = amazon_client.make_api_call(
#             endpoint,
#             "GET",
#             params,
#         )
#     except Exception as exc:
#         logger.exception(
#             "Failed to fetch AWD inbound shipment %s",
#             shipment_id,
#         )

#         return jsonify({
#             "success": False,
#             "shipment_id": shipment_id,
#             "error": "AWD inbound shipment request failed",
#             "detail": str(exc),
#         }), 500

#     if not amazon_response:
#         return jsonify({
#             "success": False,
#             "shipment_id": shipment_id,
#             "error": "Amazon returned an empty response",
#         }), 502

#     if amazon_response.get("error"):
#         upstream_status = int(
#             amazon_response.get("status_code") or 502
#         )

#         amazon_errors = (
#             amazon_response
#             .get("response_json", {})
#             .get("errors", [])
#         )

#         message = "Amazon could not fetch the AWD shipment."

#         if upstream_status == 404:
#             message = (
#                 "The AWD shipment was not found, or the connected "
#                 "seller account does not have access to it."
#             )
#         elif upstream_status == 403:
#             message = (
#                 "Amazon denied access. Confirm the connected seller "
#                 "account has AWD API permissions."
#             )
#         elif upstream_status == 429:
#             message = (
#                 "Amazon rate limit exceeded. Try again later."
#             )

#         return jsonify({
#             "success": False,
#             "shipment_id": shipment_id,
#             "marketplace_id": mp,
#             "error": message,
#             "amazon_status_code": upstream_status,
#             "amazon_request_id": amazon_response.get("amzn_request_id"),
#             "amazon_errors": amazon_errors,
#         }), upstream_status

#     shipment = amazon_response.get("payload") or amazon_response
#     export_format = (request.args.get("format") or "json").strip().lower()

#     if export_format == "excel":
#         excel_stream = _build_awd_shipment_detail_excel(
#             shipment=shipment,
#             marketplace_id=mp,
#         )
#         safe_shipment_id = re.sub(r"[^A-Za-z0-9_-]+", "_", shipment_id)
#         return send_file(
#             excel_stream,
#             as_attachment=True,
#             download_name=f"AWD_Inbound_Shipment_{safe_shipment_id}.xlsx",
#             mimetype=_AWD_XLSX_MIMETYPE,
#             max_age=0,
#         )

#     if export_format not in {"json", ""}:
#         return jsonify({
#             "success": False,
#             "error": "format must be json or excel",
#         }), 400

#     return jsonify({
#         "success": True,
#         "marketplace_id": mp,
#         "shipment_id": shipment_id,
#         "shipment": shipment,
#     }), 200


# =============================================================================
# AWD INBOUND SHIPMENTS - COMBINED LIST + DETAIL ROUTE
# =============================================================================

def _awd_combined_extract_shipments(payload: dict) -> list[dict]:
    payload = payload or {}
    rows = (
        payload.get("shipments")
        or payload.get("inboundShipments")
        or payload.get("items")
        or []
    )
    return rows if isinstance(rows, list) else []


def _awd_combined_next_token(payload: dict):
    payload = payload or {}
    pagination = payload.get("pagination") or {}
    return pagination.get("nextToken") or payload.get("nextToken")


def _awd_combined_error_response(response: dict, default_message: str):
    status_code = int((response or {}).get("status_code") or 502)
    errors = ((response or {}).get("response_json") or {}).get("errors", [])
    return jsonify({
        "success": False,
        "error": default_message,
        "amazon_status_code": status_code,
        "amazon_request_id": (response or {}).get("amzn_request_id"),
        "amazon_errors": errors,
    }), status_code




def _build_awd_complete_shipments_excel(
    complete_items: list[dict],
    marketplace_id: str,
    detail_errors: list[dict] | None = None,
) -> io.BytesIO:
    """Create an Excel workbook for complete AWD inbound shipment data."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Shipments"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(c.value or "")) for c in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

    shipment_headers = [
        "Marketplace ID", "Shipment ID", "Order ID", "External Reference ID",
        "Shipment Status", "Warehouse Reference ID", "Carrier Code Type",
        "Carrier Code Value", "Created At", "Updated At", "Ship By",
        "Expected Cases", "Received Cases", "Case Difference",
        "Origin Name", "Origin Address", "Destination Name", "Destination Address",
    ]
    ws_summary.append(shipment_headers)

    ws_skus = wb.create_sheet("SKU Quantities")
    ws_skus.append([
        "Marketplace ID", "Shipment ID", "Order ID", "Shipment Status",
        "SKU", "ASIN", "Expected Cases", "Received Cases", "Case Difference",
        "Units Per Case", "Expected Units", "Received Units", "Expiration",
        "Prep Category", "Prep Owner", "Label Owner",
    ])

    ws_containers = wb.create_sheet("Containers")
    ws_containers.append([
        "Marketplace ID", "Shipment ID", "Order ID", "Shipment Status",
        "SKU", "ASIN", "Container Count", "Package Type", "Units Per Case",
        "Total Units", "Expiration", "Length", "Width", "Height",
        "Dimension Unit", "Weight", "Weight Unit", "Prep Category",
    ])

    def address_text(addr):
        if not isinstance(addr, dict):
            return ""
        parts = [
            addr.get("addressLine1"), addr.get("addressLine2"), addr.get("addressLine3"),
            addr.get("city"), addr.get("district"), addr.get("stateOrRegion"),
            addr.get("postalCode"), addr.get("countryCode"),
        ]
        return ", ".join(str(x).strip() for x in parts if str(x or "").strip())

    for item in complete_items or []:
        summary = item.get("summary") or {}
        shipment = item.get("shipment") or {}
        shipment_id = item.get("shipment_id") or shipment.get("shipmentId") or summary.get("shipmentId")
        order_id = shipment.get("orderId") or summary.get("orderId")
        status = shipment.get("shipmentStatus") or summary.get("shipmentStatus")
        carrier = shipment.get("carrierCode") or {}
        received_total = sum(
            _safe_int(x.get("quantity"))
            for x in (shipment.get("receivedQuantity") or [])
            if isinstance(x, dict)
        )
        sku_quantities = shipment.get("shipmentSkuQuantities") or []
        expected_total = sum(
            _safe_int((x.get("expectedQuantity") or {}).get("quantity"))
            for x in sku_quantities if isinstance(x, dict)
        )
        origin = shipment.get("originAddress") or {}
        destination = shipment.get("destinationAddress") or {}

        ws_summary.append([
            marketplace_id, shipment_id, order_id,
            shipment.get("externalReferenceId") or summary.get("externalReferenceId"),
            status, shipment.get("warehouseReferenceId"),
            carrier.get("carrierCodeType"), carrier.get("carrierCodeValue"),
            shipment.get("createdAt") or summary.get("createdAt"),
            shipment.get("updatedAt") or summary.get("updatedAt"),
            shipment.get("shipBy"), expected_total, received_total,
            expected_total - received_total, origin.get("name"), address_text(origin),
            destination.get("name"), address_text(destination),
        ])

        container_by_sku = {}
        for container in shipment.get("shipmentContainerQuantities") or []:
            if not isinstance(container, dict):
                continue
            count = _safe_int(container.get("count"))
            package = container.get("distributionPackage") or {}
            measurements = package.get("measurements") or {}
            dimensions = measurements.get("dimensions") or {}
            weight = measurements.get("weight") or {}
            products = ((package.get("contents") or {}).get("products") or [])
            for product in products:
                if not isinstance(product, dict):
                    continue
                sku = product.get("sku")
                attributes = product.get("attributes") or []
                asin = next((a.get("value") for a in attributes if isinstance(a, dict) and a.get("name") == "asin"), None)
                units_per_case = _safe_int(product.get("quantity"))
                prep = product.get("prepDetails") or {}
                container_by_sku[str(sku or "").strip()] = {
                    "asin": asin, "units_per_case": units_per_case,
                    "expiration": product.get("expiration"),
                    "prep_category": prep.get("prepCategory"),
                    "prep_owner": prep.get("prepOwner"),
                    "label_owner": prep.get("labelOwner"),
                }
                ws_containers.append([
                    marketplace_id, shipment_id, order_id, status, sku, asin, count,
                    package.get("type"), units_per_case, count * units_per_case,
                    product.get("expiration"), dimensions.get("length"), dimensions.get("width"),
                    dimensions.get("height"), dimensions.get("unitOfMeasurement"),
                    weight.get("weight"), weight.get("unitOfMeasurement"),
                    prep.get("prepCategory"),
                ])

        for sku_row in sku_quantities:
            if not isinstance(sku_row, dict):
                continue
            sku = sku_row.get("sku")
            expected = _safe_int((sku_row.get("expectedQuantity") or {}).get("quantity"))
            received = _safe_int((sku_row.get("receivedQuantity") or {}).get("quantity"))
            extra = container_by_sku.get(str(sku or "").strip(), {})
            units_per_case = _safe_int(extra.get("units_per_case"))
            ws_skus.append([
                marketplace_id, shipment_id, order_id, status, sku, extra.get("asin"),
                expected, received, expected - received, units_per_case,
                expected * units_per_case, received * units_per_case,
                extra.get("expiration"), extra.get("prep_category"),
                extra.get("prep_owner"), extra.get("label_owner"),
            ])

    if detail_errors:
        ws_errors = wb.create_sheet("Detail Errors")
        ws_errors.append(["Shipment ID", "Error", "Amazon Status Code", "Amazon Request ID", "Amazon Errors"])
        for error in detail_errors:
            ws_errors.append([
                error.get("shipment_id"), error.get("error"),
                error.get("amazon_status_code"), error.get("amazon_request_id"),
                str(error.get("amazon_errors") or ""),
            ])
        style_sheet(ws_errors)

    for ws in (ws_summary, ws_skus, ws_containers):
        style_sheet(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream




def _ensure_awd_inbound_shipments_table(conn) -> None:
    """Create/upgrade SKU-level AWD inbound shipment table in DATABASE_AMAZON_URL."""
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS public.inventory_awd_inbound_shipments (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            marketplace_id TEXT NOT NULL,
            shipment_id TEXT NOT NULL,
            order_id TEXT,
            external_reference_id TEXT,
            shipment_status TEXT,
            warehouse_reference_id TEXT,
            carrier_code_type TEXT,
            carrier_code_value TEXT,
            sku TEXT,
            asin TEXT,
            expected_case_quantity BIGINT NOT NULL DEFAULT 0,
            received_case_quantity BIGINT NOT NULL DEFAULT 0,
            case_difference BIGINT NOT NULL DEFAULT 0,
            units_per_case BIGINT NOT NULL DEFAULT 0,
            expected_unit_quantity BIGINT NOT NULL DEFAULT 0,
            received_unit_quantity BIGINT NOT NULL DEFAULT 0,
            expiration_date TEXT,
            created_at TIMESTAMPTZ,
            updated_at TIMESTAMPTZ,
            ship_by TIMESTAMPTZ,
            dispatch_date DATE,
            shipment_type TEXT,
            expected_reach_date DATE,
            synced_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
    """))

    conn.execute(text("""
        ALTER TABLE public.inventory_awd_inbound_shipments
            ADD COLUMN IF NOT EXISTS sku TEXT,
            ADD COLUMN IF NOT EXISTS asin TEXT,
            ADD COLUMN IF NOT EXISTS expected_case_quantity BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS received_case_quantity BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS case_difference BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS units_per_case BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS expected_unit_quantity BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS received_unit_quantity BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS expiration_date TEXT,
            ADD COLUMN IF NOT EXISTS dispatch_date DATE,
            ADD COLUMN IF NOT EXISTS shipment_type TEXT,
            ADD COLUMN IF NOT EXISTS expected_reach_date DATE,
            ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    """))

    conn.execute(text("""
        ALTER TABLE public.inventory_awd_inbound_shipments
            DROP COLUMN IF EXISTS expected_quantity,
            DROP COLUMN IF EXISTS received_quantity,
            DROP COLUMN IF EXISTS remaining_quantity,
            DROP COLUMN IF EXISTS origin_address,
            DROP COLUMN IF EXISTS destination_address,
            DROP COLUMN IF EXISTS received_quantities,
            DROP COLUMN IF EXISTS shipment_sku_quantities,
            DROP COLUMN IF EXISTS shipment_container_quantities,
            DROP COLUMN IF EXISTS summary_json,
            DROP COLUMN IF EXISTS shipment_json,
            DROP COLUMN IF EXISTS lot_code,
            DROP COLUMN IF EXISTS prep_category,
            DROP COLUMN IF EXISTS prep_owner,
            DROP COLUMN IF EXISTS label_owner,
            DROP COLUMN IF EXISTS sku_quantity_json,
            DROP COLUMN IF EXISTS container_product_json
    """))

    # Remove the old one-row-per-shipment uniqueness rule, then create SKU-level uniqueness.
    conn.execute(text("""
        ALTER TABLE public.inventory_awd_inbound_shipments
        DROP CONSTRAINT IF EXISTS uq_inventory_awd_inbound_shipment
    """))
    conn.execute(text("""
        DROP INDEX IF EXISTS public.uq_inventory_awd_inbound_shipment_idx
    """))
    conn.execute(text("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_inventory_awd_inbound_shipment_sku_idx
        ON public.inventory_awd_inbound_shipments
            (user_id, marketplace_id, shipment_id, sku)
        WHERE sku IS NOT NULL AND TRIM(sku) <> ''
    """))


def _get_auth_user_id_from_request() -> int | None:
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None

    token = auth_header.split(" ", 1)[1].strip()
    payload, user_id, member_id = get_effective_user_id_from_token(token)
    return int(payload.get("user_id") or user_id)


@inventory_bp.route("/amazon_api/sku-product-names", methods=["GET"])
def get_sku_product_names():
    try:
        user_id = _get_auth_user_id_from_request()
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"success": False, "error": "Authorization token is missing or invalid"}), 401

    marketplace_id = (request.args.get("marketplace_id") or amazon_client.marketplace_id or "").strip()
    country = request.args.get("country")
    lookup = _get_sku_product_name_lookup(user_id, country=country, marketplace_id=marketplace_id)

    return jsonify({
        "success": True,
        "marketplace_id": marketplace_id or None,
        "country": country or MARKETPLACE_TO_COUNTRY.get(marketplace_id),
        "count": len(lookup),
        "items": [
            {"sku": sku, "product_name": product_name}
            for sku, product_name in sorted(lookup.items())
        ],
    }), 200


@inventory_bp.route("/amazon_api/awd/inbound-shipments/dispatch-inputs", methods=["GET"])
def get_awd_inbound_dispatch_inputs():
    try:
        user_id = _get_auth_user_id_from_request()
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"success": False, "error": "Authorization token is missing or invalid"}), 401

    marketplace_id = (request.args.get("marketplace_id") or amazon_client.marketplace_id or "").strip()
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace", "marketplace_id": marketplace_id}), 400

    try:
        with amazon_conn() as conn:
            _ensure_awd_inbound_shipments_table(conn)
            rows = conn.execute(text("""
                WITH sku_rows AS (
                    SELECT
                        shipment_id,
                        sku,
                        STRING_AGG(DISTINCT COALESCE(asin, ''), ', ')
                            FILTER (WHERE COALESCE(asin, '') <> '') AS asin,
                        SUM(COALESCE(expected_unit_quantity, 0)) AS expected_unit_quantity,
                        MAX(shipment_status) AS shipment_status,
                        MIN(created_at) AS created_at,
                        MAX(updated_at) AS updated_at,
                        MAX(ship_by) AS ship_by,
                        MAX(dispatch_date) AS dispatch_date,
                        MAX(shipment_type) AS shipment_type,
                        MAX(expected_reach_date) AS expected_reach_date
                    FROM public.inventory_awd_inbound_shipments
                    WHERE user_id = :user_id
                      AND marketplace_id = :marketplace_id
                      AND COALESCE(expected_unit_quantity, 0) > 0
                      AND UPPER(REPLACE(COALESCE(shipment_status, ''), '-', '_')) NOT IN ('CANCELLED', 'CLOSED', 'DELIVERED')
                    GROUP BY shipment_id, sku
                )
                SELECT
                    shipment_id,
                    MAX(shipment_status) AS shipment_status,
                    STRING_AGG(DISTINCT sku, ', ') AS sku,
                    STRING_AGG(DISTINCT asin, ', ') FILTER (WHERE COALESCE(asin, '') <> '') AS asin,
                    SUM(COALESCE(expected_unit_quantity, 0)) AS expected_unit_quantity,
                    MIN(created_at) AS created_at,
                    MAX(updated_at) AS updated_at,
                    MAX(ship_by) AS ship_by,
                    MAX(dispatch_date) AS dispatch_date,
                    MAX(shipment_type) AS shipment_type,
                    MAX(expected_reach_date) AS expected_reach_date,
                    JSONB_AGG(
                        JSONB_BUILD_OBJECT(
                            'sku', sku,
                            'expected_unit_quantity', expected_unit_quantity
                        )
                        ORDER BY sku
                    ) FILTER (WHERE COALESCE(sku, '') <> '') AS sku_quantities
                FROM sku_rows
                GROUP BY shipment_id
                ORDER BY COALESCE(MAX(ship_by), MAX(updated_at), MIN(created_at)) ASC NULLS LAST, shipment_id ASC
            """), {
                "user_id": int(user_id),
                "marketplace_id": marketplace_id,
            }).mappings().all()
    except Exception as exc:
        logger.exception("Failed to read AWD dispatch inputs")
        return jsonify({"success": False, "error": str(exc)}), 500

    items = []
    for row in rows:
        item = dict(row)
        for key in ("created_at", "updated_at", "ship_by"):
            value = item.get(key)
            item[key] = value.isoformat() if value else None
        dispatch_date = item.get("dispatch_date")
        item["dispatch_date"] = dispatch_date.isoformat() if dispatch_date else None
        reach_date = item.get("expected_reach_date")
        item["expected_reach_date"] = reach_date.isoformat() if reach_date else None
        item["expected_unit_quantity"] = int(item.get("expected_unit_quantity") or 0)
        items.append(item)

    return jsonify({
        "success": True,
        "marketplace_id": marketplace_id,
        "items": items,
    }), 200


@inventory_bp.route("/amazon_api/fba/inbound-shipments/dispatch-inputs", methods=["GET"])
def get_fba_inbound_dispatch_inputs():
    try:
        user_id = _get_auth_user_id_from_request()
    except jwt.ExpiredSignatureError:
        return jsonify({
            "success": False,
            "error": "Token has expired"
        }), 401
    except jwt.InvalidTokenError:
        return jsonify({
            "success": False,
            "error": "Invalid token"
        }), 401

    if not user_id:
        return jsonify({
            "success": False,
            "error": "Authorization token is missing or invalid"
        }), 401

    # ---------------------------------------------------------
    # Marketplace
    # ---------------------------------------------------------
    marketplace_id = (
        request.args.get("marketplace_id")
        or amazon_client.marketplace_id
        or ""
    ).strip()

    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({
            "success": False,
            "error": "Unsupported marketplace",
            "marketplace_id": marketplace_id
        }), 400

    # ---------------------------------------------------------
    # Shipment statuses
    #
    # Default Dispatch statuses:
    # WORKING
    # READY_TO_SHIP
    # ACTIVE
    # IN_TRANSIT
    #
    # You can still override them using:
    # ?shipment_statuses=WORKING,IN_TRANSIT
    # ---------------------------------------------------------
    status_values = [
        value.strip().upper().replace("-", "_")
        for value in (
            request.args.get("shipment_statuses")
            or request.args.get("statuses")
            or "WORKING,READY_TO_SHIP,ACTIVE,IN_TRANSIT"
        ).split(",")
        if value.strip()
    ]

    allowed_statuses = {
        "WORKING",
        "READY_TO_SHIP",
        "ACTIVE",
        "SHIPPED",
        "RECEIVING",
        "CANCELLED",
        "DELETED",
        "CLOSED",
        "ERROR",
        "IN_TRANSIT",
        "DELIVERED",
        "CHECKED_IN",
    }

    invalid_statuses = [
        value
        for value in status_values
        if value not in allowed_statuses
    ]

    if invalid_statuses:
        return jsonify({
            "success": False,
            "error": "Invalid shipment status",
            "invalid_statuses": invalid_statuses,
            "allowed_statuses": sorted(allowed_statuses),
        }), 400

    if not status_values:
        return jsonify({
            "success": False,
            "error": "At least one shipment status is required"
        }), 400

    # Because values have already been validated against a fixed whitelist,
    # this is safe to insert into the SQL IN clause.
    status_sql = ", ".join(
        f"'{value}'"
        for value in status_values
    )

    # ---------------------------------------------------------
    # Get shipment data
    # ---------------------------------------------------------
    try:
        with amazon_conn() as conn:

            _ensure_inventory_fba_inbound_shipments_table(conn)

            rows = conn.execute(
                text(f"""
                    WITH normalized AS (
                        SELECT

                            COALESCE(
                                NULLIF(TRIM("shipmentId"), ''),
                                shipment_id
                            ) AS shipment_id,

                            UPPER(
                                REPLACE(
                                    COALESCE(
                                        NULLIF(TRIM(status), ''),
                                        shipment_status,
                                        ''
                                    ),
                                    '-',
                                    '_'
                                )
                            ) AS shipment_status,

                            COALESCE(
                                NULLIF(TRIM(msku), ''),
                                seller_sku
                            ) AS sku,

                            fnsku,
                            asin,

                            COALESCE(
                                quantity,
                                quantity_shipped,
                                0
                            ) AS quantity,

                            "createdAt" AS created_at,

                            COALESCE(
                                "lastUpdatedAt",
                                updated_at::TEXT
                            ) AS updated_at,

                            dispatch_date,
                            shipment_type,
                            expected_reach_date,
                            name

                        FROM public.inventory_fba_inbound_shipments

                        WHERE user_id = :user_id

                          AND marketplace_id = :marketplace_id

                          AND COALESCE(
                                quantity,
                                quantity_shipped,
                                0
                              ) > 0

                          AND UPPER(
                                REPLACE(
                                    COALESCE(
                                        NULLIF(TRIM(status), ''),
                                        shipment_status,
                                        ''
                                    ),
                                    '-',
                                    '_'
                                )
                              ) IN ({status_sql})
                    )

                    SELECT

                        shipment_id,

                        MAX(shipment_status)
                            AS shipment_status,

                        STRING_AGG(
                            DISTINCT sku,
                            ', '
                        ) AS sku,

                        STRING_AGG(
                            DISTINCT COALESCE(fnsku, ''),
                            ', '
                        )
                        FILTER (
                            WHERE COALESCE(fnsku, '') <> ''
                        ) AS fnsku,

                        STRING_AGG(
                            DISTINCT COALESCE(asin, ''),
                            ', '
                        )
                        FILTER (
                            WHERE COALESCE(asin, '') <> ''
                        ) AS asin,

                        SUM(
                            COALESCE(quantity, 0)
                        ) AS quantity,

                        MIN(created_at)
                            AS created_at,

                        MAX(updated_at)
                            AS updated_at,

                        MAX(dispatch_date)
                            AS dispatch_date,

                        MAX(shipment_type)
                            AS shipment_type,

                        MAX(expected_reach_date)
                            AS expected_reach_date,

                        MAX(name)
                            AS name

                    FROM normalized

                    WHERE COALESCE(shipment_id, '') <> ''

                      AND COALESCE(sku, '') <> ''

                    GROUP BY shipment_id

                    ORDER BY
                        MIN(created_at) ASC NULLS LAST,
                        shipment_id ASC
                """),
                {
                    "user_id": int(user_id),
                    "marketplace_id": marketplace_id,
                }
            ).mappings().all()

    except Exception as exc:
        logger.exception(
            "Failed to read FBA dispatch inputs"
        )

        return jsonify({
            "success": False,
            "error": str(exc)
        }), 500

    # ---------------------------------------------------------
    # Prepare response
    # ---------------------------------------------------------
    items = []

    for row in rows:
        item = dict(row)

        item["quantity"] = int(
            item.get("quantity") or 0
        )

        # created_at
        created_at = item.get("created_at")

        if hasattr(created_at, "isoformat"):
            item["created_at"] = created_at.isoformat()
        elif created_at:
            item["created_at"] = str(created_at)
        else:
            item["created_at"] = None

        # updated_at
        updated_at = item.get("updated_at")

        if hasattr(updated_at, "isoformat"):
            item["updated_at"] = updated_at.isoformat()
        elif updated_at:
            item["updated_at"] = str(updated_at)
        else:
            item["updated_at"] = None

        # dispatch_date
        dispatch_date = item.get("dispatch_date")

        if hasattr(dispatch_date, "isoformat"):
            item["dispatch_date"] = dispatch_date.isoformat()
        elif dispatch_date:
            item["dispatch_date"] = str(dispatch_date)
        else:
            item["dispatch_date"] = None

        # expected reach date
        expected_reach_date = item.get(
            "expected_reach_date"
        )

        if hasattr(expected_reach_date, "isoformat"):
            item["expected_reach_date"] = (
                expected_reach_date.isoformat()
            )
        elif expected_reach_date:
            item["expected_reach_date"] = str(
                expected_reach_date
            )
        else:
            item["expected_reach_date"] = None

        items.append(item)

    # ---------------------------------------------------------
    # Response
    # ---------------------------------------------------------
    return jsonify({
        "success": True,
        "marketplace_id": marketplace_id,
        "shipment_statuses": status_values,
        "count": len(items),
        "items": items,
    }), 200



@inventory_bp.route("/amazon_api/awd/inbound-shipments/dispatch-inputs", methods=["POST"])
def save_awd_inbound_dispatch_inputs():
    try:
        user_id = _get_auth_user_id_from_request()
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"success": False, "error": "Authorization token is missing or invalid"}), 401

    payload = request.get_json(silent=True) or {}
    marketplace_id = (payload.get("marketplace_id") or request.args.get("marketplace_id") or amazon_client.marketplace_id or "").strip()
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace", "marketplace_id": marketplace_id}), 400

    shipments = payload.get("shipments") or []
    if not isinstance(shipments, list):
        return jsonify({"success": False, "error": "shipments must be a list"}), 400

    normalized_rows = []
    for item in shipments:
        if not isinstance(item, dict):
            continue
        shipment_id = str(item.get("shipment_id") or "").strip()
        dispatch_date = str(item.get("dispatch_date") or "").strip()
        shipment_type = str(item.get("shipment_type") or "").strip().upper()
        expected_reach_date = str(item.get("expected_reach_date") or "").strip()
        if not shipment_id:
            continue
        try:
            datetime.strptime(dispatch_date, "%Y-%m-%d")
        except Exception:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have dispatch_date YYYY-MM-DD"}), 400
        if shipment_type not in {"SEA", "AIR"}:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have shipment_type SEA or AIR"}), 400
        try:
            datetime.strptime(expected_reach_date, "%Y-%m-%d")
        except Exception:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have expected_reach_date YYYY-MM-DD"}), 400
        normalized_rows.append({
            "shipment_id": shipment_id,
            "dispatch_date": dispatch_date,
            "shipment_type": shipment_type,
            "expected_reach_date": expected_reach_date,
        })

    try:
        saved = 0
        with amazon_conn() as conn:
            _ensure_awd_inbound_shipments_table(conn)
            for row in normalized_rows:
                result = conn.execute(text("""
                    UPDATE public.inventory_awd_inbound_shipments
                    SET dispatch_date = CAST(:dispatch_date AS DATE),
                        shipment_type = :shipment_type,
                        expected_reach_date = CAST(:expected_reach_date AS DATE)
                    WHERE user_id = :user_id
                      AND marketplace_id = :marketplace_id
                      AND shipment_id = :shipment_id
                """), {
                    "user_id": int(user_id),
                    "marketplace_id": marketplace_id,
                    **row,
                })
                saved += int(result.rowcount or 0)
    except Exception as exc:
        logger.exception("Failed to save AWD dispatch inputs")
        return jsonify({"success": False, "error": str(exc)}), 500

    return jsonify({
        "success": True,
        "marketplace_id": marketplace_id,
        "updated_rows": saved,
    }), 200


@inventory_bp.route("/amazon_api/fba/inbound-shipments/dispatch-inputs", methods=["POST"])
def save_fba_inbound_dispatch_inputs():
    try:
        user_id = _get_auth_user_id_from_request()
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401

    if not user_id:
        return jsonify({"success": False, "error": "Authorization token is missing or invalid"}), 401

    payload = request.get_json(silent=True) or {}
    marketplace_id = (payload.get("marketplace_id") or request.args.get("marketplace_id") or amazon_client.marketplace_id or "").strip()
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace", "marketplace_id": marketplace_id}), 400

    shipments = payload.get("shipments") or []
    if not isinstance(shipments, list):
        return jsonify({"success": False, "error": "shipments must be a list"}), 400

    normalized_rows = []
    for item in shipments:
        if not isinstance(item, dict):
            continue
        shipment_id = str(item.get("shipment_id") or "").strip()
        dispatch_date = str(item.get("dispatch_date") or "").strip()
        shipment_type = str(item.get("shipment_type") or "").strip().upper()
        expected_reach_date = str(item.get("expected_reach_date") or "").strip()
        if not shipment_id:
            continue
        try:
            datetime.strptime(dispatch_date, "%Y-%m-%d")
        except Exception:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have dispatch_date YYYY-MM-DD"}), 400
        if shipment_type not in {"SEA", "AIR"}:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have shipment_type SEA or AIR"}), 400
        try:
            datetime.strptime(expected_reach_date, "%Y-%m-%d")
        except Exception:
            return jsonify({"success": False, "error": f"Shipment {shipment_id} must have expected_reach_date YYYY-MM-DD"}), 400
        normalized_rows.append({
            "shipment_id": shipment_id,
            "dispatch_date": dispatch_date,
            "shipment_type": shipment_type,
            "expected_reach_date": expected_reach_date,
        })

    try:
        saved = 0
        with amazon_conn() as conn:
            _ensure_inventory_fba_inbound_shipments_table(conn)
            for row in normalized_rows:
                result = conn.execute(text("""
                    UPDATE public.inventory_fba_inbound_shipments
                    SET dispatch_date = CAST(:dispatch_date AS DATE),
                        shipment_type = :shipment_type,
                        expected_reach_date = CAST(:expected_reach_date AS DATE)
                    WHERE user_id = :user_id
                      AND marketplace_id = :marketplace_id
                      AND COALESCE(NULLIF(TRIM("shipmentId"), ''), shipment_id) = :shipment_id
                """), {
                    "user_id": int(user_id),
                    "marketplace_id": marketplace_id,
                    **row,
                })
                saved += int(result.rowcount or 0)
    except Exception as exc:
        logger.exception("Failed to save FBA dispatch inputs")
        return jsonify({"success": False, "error": str(exc)}), 500

    return jsonify({
        "success": True,
        "marketplace_id": marketplace_id,
        "updated_rows": saved,
    }), 200


def _save_awd_inbound_shipments(
    complete_items: list[dict],
    user_id: int,
    marketplace_id: str,
) -> int:
    """Replace current marketplace data with one DB row per shipment SKU."""
    if not complete_items:
        return 0

    insert_sql = text("""
        INSERT INTO public.inventory_awd_inbound_shipments (
            user_id, marketplace_id, shipment_id, order_id,
            external_reference_id, shipment_status, warehouse_reference_id,
            carrier_code_type, carrier_code_value,
            sku, asin,
            expected_case_quantity, received_case_quantity, case_difference,
            units_per_case, expected_unit_quantity, received_unit_quantity,
            expiration_date,
            created_at, updated_at, ship_by,
            synced_at
        ) VALUES (
            :user_id, :marketplace_id, :shipment_id, :order_id,
            :external_reference_id, :shipment_status, :warehouse_reference_id,
            :carrier_code_type, :carrier_code_value,
            :sku, :asin,
            :expected_case_quantity, :received_case_quantity, :case_difference,
            :units_per_case, :expected_unit_quantity, :received_unit_quantity,
            :expiration_date,
            CAST(:created_at AS TIMESTAMPTZ), CAST(:updated_at AS TIMESTAMPTZ),
            CAST(:ship_by AS TIMESTAMPTZ),
            NOW()
        )
        ON CONFLICT (user_id, marketplace_id, shipment_id, sku)
        WHERE sku IS NOT NULL AND TRIM(sku) <> ''
        DO UPDATE SET
            order_id = EXCLUDED.order_id,
            external_reference_id = EXCLUDED.external_reference_id,
            shipment_status = EXCLUDED.shipment_status,
            warehouse_reference_id = EXCLUDED.warehouse_reference_id,
            carrier_code_type = EXCLUDED.carrier_code_type,
            carrier_code_value = EXCLUDED.carrier_code_value,
            asin = EXCLUDED.asin,
            expected_case_quantity = EXCLUDED.expected_case_quantity,
            received_case_quantity = EXCLUDED.received_case_quantity,
            case_difference = EXCLUDED.case_difference,
            units_per_case = EXCLUDED.units_per_case,
            expected_unit_quantity = EXCLUDED.expected_unit_quantity,
            received_unit_quantity = EXCLUDED.received_unit_quantity,
            expiration_date = EXCLUDED.expiration_date,
            created_at = EXCLUDED.created_at,
            updated_at = EXCLUDED.updated_at,
            ship_by = EXCLUDED.ship_by,
            synced_at = NOW()
    """)

    saved = 0
    with amazon_conn() as conn:
        _ensure_awd_inbound_shipments_table(conn)

        # Clear old shipment-level rows and stale SKU rows for this user/marketplace.
        conn.execute(text("""
            DELETE FROM public.inventory_awd_inbound_shipments
            WHERE user_id = :user_id AND marketplace_id = :marketplace_id
        """), {"user_id": int(user_id), "marketplace_id": marketplace_id})

        for item in complete_items:
            summary = item.get("summary") or {}
            shipment = item.get("shipment") or {}
            shipment_id = str(
                item.get("shipment_id")
                or shipment.get("shipmentId")
                or summary.get("shipmentId")
                or ""
            ).strip()
            if not shipment_id:
                continue

            carrier = shipment.get("carrierCode") or summary.get("carrierCode") or {}
            sku_rows = shipment.get("shipmentSkuQuantities") or []

            # Container details provide ASIN, units/case, and expiration.
            container_by_sku: dict[str, dict] = {}
            for container in shipment.get("shipmentContainerQuantities") or []:
                if not isinstance(container, dict):
                    continue
                package = container.get("distributionPackage") or {}
                products = ((package.get("contents") or {}).get("products") or [])
                for product in products:
                    if not isinstance(product, dict):
                        continue
                    sku_key = str(product.get("sku") or "").strip()
                    if not sku_key:
                        continue
                    attributes = product.get("attributes") or []
                    asin = next((a.get("value") for a in attributes if isinstance(a, dict) and a.get("name") == "asin"), None)
                    container_by_sku[sku_key] = {
                        "asin": asin,
                        "units_per_case": _safe_int(product.get("quantity")),
                        "expiration": product.get("expiration"),
                    }

            for sku_row in sku_rows:
                if not isinstance(sku_row, dict):
                    continue
                sku = str(sku_row.get("sku") or "").strip()
                if not sku:
                    continue

                expected_cases = _safe_int((sku_row.get("expectedQuantity") or {}).get("quantity"))
                received_cases = _safe_int((sku_row.get("receivedQuantity") or {}).get("quantity"))
                extra = container_by_sku.get(sku, {})
                units_per_case = _safe_int(extra.get("units_per_case"))

                conn.execute(insert_sql, {
                    "user_id": int(user_id),
                    "marketplace_id": marketplace_id,
                    "shipment_id": shipment_id,
                    "order_id": shipment.get("orderId") or summary.get("orderId"),
                    "external_reference_id": shipment.get("externalReferenceId") or summary.get("externalReferenceId"),
                    "shipment_status": shipment.get("shipmentStatus") or summary.get("shipmentStatus") or shipment.get("status") or summary.get("status"),
                    "warehouse_reference_id": shipment.get("warehouseReferenceId") or summary.get("warehouseReferenceId"),
                    "carrier_code_type": carrier.get("carrierCodeType") if isinstance(carrier, dict) else None,
                    "carrier_code_value": carrier.get("carrierCodeValue") if isinstance(carrier, dict) else None,
                    "sku": sku,
                    "asin": extra.get("asin"),
                    "expected_case_quantity": expected_cases,
                    "received_case_quantity": received_cases,
                    "case_difference": expected_cases - received_cases,
                    "units_per_case": units_per_case,
                    "expected_unit_quantity": expected_cases * units_per_case,
                    "received_unit_quantity": received_cases * units_per_case,
                    "expiration_date": extra.get("expiration"),
                    "created_at": shipment.get("createdAt") or summary.get("createdAt"),
                    "updated_at": shipment.get("updatedAt") or summary.get("updatedAt"),
                    "ship_by": shipment.get("shipBy") or summary.get("shipBy"),
                })
                saved += 1

    return saved


@inventory_bp.route("/amazon_api/awd/inbound-shipments-complete", methods=["GET"])
def get_all_awd_inbound_shipments_complete():
    """
    One route that uses both Amazon operations:
      1. listInboundShipments
      2. getInboundShipment for every returned shipment ID

    Example:
      GET /amazon_api/awd/inbound-shipments-complete
          ?marketplace_id=ATVPDKIKX0DER
          &shipment_status=CLOSED
          &sku_quantities=SHOW
          &max_results=100

    Optional:
      shipment_status=CREATED|SHIPPED|IN_TRANSIT|RECEIVING|DELIVERED|CLOSED|CANCELLED
      sku_quantities=SHOW|HIDE
      max_results=1..100
      next_token=<Amazon token>   # fetch from a specific page only
    """

    # ---------------- AUTH ----------------
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({
            "success": False,
            "error": "Missing Authorization header",
        }), 401

    token = auth_header.split(" ", 1)[1].strip()

    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = payload.get("user_id")
    except jwt.ExpiredSignatureError:
        return jsonify({
            "success": False,
            "error": "Token has expired",
        }), 401
    except jwt.InvalidTokenError:
        return jsonify({
            "success": False,
            "error": "Invalid token",
        }), 401

    if not user_id:
        return jsonify({
            "success": False,
            "error": "Invalid token payload",
        }), 401

    # ---------------- MARKETPLACE ----------------
    _apply_region_and_marketplace_from_request()
    mp = request.args.get("marketplace_id", amazon_client.marketplace_id)

    if mp not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({
            "success": False,
            "error": "Unsupported marketplace",
            "marketplace_id": mp,
        }), 400

    # ---------------- PARAMS ----------------
    shipment_status = (
        request.args.get("shipment_status")
        or request.args.get("status")
        or ""
    ).strip().upper()

    allowed_statuses = {
        "CREATED",
        "SHIPPED",
        "IN_TRANSIT",
        "RECEIVING",
        "DELIVERED",
        "CLOSED",
        "CANCELLED",
    }

    if shipment_status and shipment_status not in allowed_statuses:
        return jsonify({
            "success": False,
            "error": "Invalid shipment_status",
            "allowed_statuses": sorted(allowed_statuses),
        }), 400

    sku_quantities = (
        request.args.get("sku_quantities", "SHOW")
        .strip()
        .upper()
    )

    if sku_quantities not in {"SHOW", "HIDE"}:
        return jsonify({
            "success": False,
            "error": "sku_quantities must be SHOW or HIDE",
        }), 400

    try:
        max_results = int(request.args.get("max_results") or 100)
        if max_results < 1:
            raise ValueError
        max_results = min(max_results, 100)
    except ValueError:
        return jsonify({
            "success": False,
            "error": "max_results must be a positive integer",
        }), 400

    supplied_next_token = (request.args.get("next_token") or "").strip()
    store_in_db = request.args.get("store_in_db", "true").strip().lower() != "false"

    base_params = {"maxResults": max_results}
    if shipment_status:
        base_params["shipmentStatus"] = shipment_status

    # ---------------- LIST ALL PAGES ----------------
    list_endpoint = "/awd/2024-05-09/inboundShipments"
    shipments: list[dict] = []
    pages_fetched = 0
    seen_tokens: set[str] = set()
    next_token = supplied_next_token or None

    while True:
        page_params = dict(base_params)
        if next_token:
            page_params["nextToken"] = next_token

        try:
            list_response = amazon_client.make_api_call(
                list_endpoint,
                "GET",
                page_params,
            )
        except Exception as exc:
            logger.exception("AWD listInboundShipments request failed")
            return jsonify({
                "success": False,
                "error": "Failed to list AWD inbound shipments",
                "detail": str(exc),
            }), 500

        if not list_response:
            return jsonify({
                "success": False,
                "error": "Amazon returned an empty shipment-list response",
            }), 502

        if list_response.get("error"):
            return _awd_combined_error_response(
                list_response,
                "Amazon could not list AWD inbound shipments",
            )

        pages_fetched += 1
        list_payload = list_response.get("payload") or list_response
        shipments.extend(_awd_combined_extract_shipments(list_payload))

        returned_next_token = _awd_combined_next_token(list_payload)

        # When caller supplied next_token, fetch only that page.
        if supplied_next_token:
            next_token = returned_next_token
            break

        if not returned_next_token:
            next_token = None
            break

        token_key = str(returned_next_token)
        if token_key in seen_tokens:
            logger.warning("Stopping AWD pagination because nextToken repeated")
            next_token = returned_next_token
            break

        seen_tokens.add(token_key)
        next_token = returned_next_token

    # Deduplicate by shipmentId.
    deduped_shipments: list[dict] = []
    seen_shipment_ids: set[str] = set()

    for shipment in shipments:
        shipment_id = str(shipment.get("shipmentId") or "").strip()
        if shipment_id and shipment_id in seen_shipment_ids:
            continue
        if shipment_id:
            seen_shipment_ids.add(shipment_id)
        deduped_shipments.append(shipment)

    # ---------------- FETCH EACH SHIPMENT DETAIL ----------------
    complete_items: list[dict] = []
    detail_errors: list[dict] = []

    for shipment_summary in deduped_shipments:
        shipment_id = str(shipment_summary.get("shipmentId") or "").strip()

        if not shipment_id:
            detail_errors.append({
                "shipment_id": None,
                "error": "Shipment list row does not contain shipmentId",
                "summary": shipment_summary,
            })
            continue

        detail_endpoint = f"/awd/2024-05-09/inboundShipments/{shipment_id}"

        try:
            detail_response = amazon_client.make_api_call(
                detail_endpoint,
                "GET",
                {"skuQuantities": sku_quantities},
            )
        except Exception as exc:
            logger.exception("Failed to fetch AWD shipment detail %s", shipment_id)
            detail_errors.append({
                "shipment_id": shipment_id,
                "error": str(exc),
            })
            continue

        if not detail_response:
            detail_errors.append({
                "shipment_id": shipment_id,
                "error": "Amazon returned an empty shipment-detail response",
            })
            continue

        if detail_response.get("error"):
            detail_errors.append({
                "shipment_id": shipment_id,
                "amazon_status_code": detail_response.get("status_code"),
                "amazon_request_id": detail_response.get("amzn_request_id"),
                "amazon_errors": (
                    (detail_response.get("response_json") or {}).get("errors", [])
                ),
            })
            continue

        shipment_detail = detail_response.get("payload") or detail_response

        complete_items.append({
            "shipment_id": shipment_id,
            "summary": shipment_summary,
            "shipment": shipment_detail,
        })

    # ---------------- STORE COMPLETE SHIPMENT DATA ----------------
    db_result = {
        "table": "public.inventory_awd_inbound_shipments",
        "saved_rows": 0,
    }

    if store_in_db and complete_items:
        try:
            db_result["saved_rows"] = _save_awd_inbound_shipments(
                complete_items=complete_items,
                user_id=int(user_id),
                marketplace_id=mp,
            )
        except Exception as exc:
            logger.exception("Failed to save AWD inbound shipments")
            db_result["error"] = str(exc)

    export_format = (request.args.get("format") or "json").strip().lower()

    if export_format == "excel":
        excel_stream = _build_awd_complete_shipments_excel(
            complete_items=complete_items,
            marketplace_id=mp,
            detail_errors=detail_errors,
        )
        status_suffix = f"_{shipment_status}" if shipment_status else ""
        filename = f"AWD_Inbound_Shipments_Complete_{mp}{status_suffix}.xlsx"
        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )

    if export_format not in {"json", ""}:
        return jsonify({
            "success": False,
            "error": "format must be json or excel",
        }), 400

    return jsonify({
        "success": True,
        "marketplace_id": mp,
        "shipment_status_filter": shipment_status or None,
        "sku_quantities": sku_quantities,
        "pages_fetched": pages_fetched,
        "shipment_count": len(deduped_shipments),
        "detail_count": len(complete_items),
        "detail_error_count": len(detail_errors),
        "store_in_db": store_in_db,
        "db": db_result,
        "next_token": next_token,
        "items": complete_items,
        "detail_errors": detail_errors,
    }), 200

# =============================================================================
# FBA INBOUND SHIPMENTS + ITEM QUANTITIES (Fulfillment Inbound v0)
# Stores one row per shipment + SKU in public.inventory_fba_inbound_shipments.
# Default shipment statuses: WORKING and CLOSED.
# =============================================================================

def _ensure_inventory_fba_inbound_shipments_table(conn) -> None:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS public.inventory_fba_inbound_shipments (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            marketplace_id TEXT NOT NULL,
            shipment_id TEXT NOT NULL,
            shipment_name TEXT,
            shipment_status TEXT,

            seller_sku TEXT NOT NULL,
            fulfillment_network_sku TEXT,
            quantity_shipped BIGINT NOT NULL DEFAULT 0,
            quantity_received BIGINT NOT NULL DEFAULT 0,
            quantity_in_case BIGINT NOT NULL DEFAULT 0,

            synced_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
            created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
            updated_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
        );
    """))

    # Migrate an older shipment-header-only table without deleting its data.
    conn.execute(text("""
        ALTER TABLE public.inventory_fba_inbound_shipments
            ADD COLUMN IF NOT EXISTS seller_sku TEXT,
            ADD COLUMN IF NOT EXISTS fulfillment_network_sku TEXT,
            ADD COLUMN IF NOT EXISTS quantity_shipped BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS quantity_received BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS quantity_in_case BIGINT NOT NULL DEFAULT 0;
    """))

    conn.execute(text("""
        ALTER TABLE public.inventory_fba_inbound_shipments
            ADD COLUMN IF NOT EXISTS "inboundPlanId" TEXT,
            ADD COLUMN IF NOT EXISTS status TEXT,
            ADD COLUMN IF NOT EXISTS "createdAt" TEXT,
            ADD COLUMN IF NOT EXISTS "lastUpdatedAt" TEXT,
            ADD COLUMN IF NOT EXISTS "marketplaceIds" JSONB,
            ADD COLUMN IF NOT EXISTS shipment_count BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS plan_box_count BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS msku TEXT,
            ADD COLUMN IF NOT EXISTS fnsku TEXT,
            ADD COLUMN IF NOT EXISTS asin TEXT,
            ADD COLUMN IF NOT EXISTS quantity BIGINT NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS "shipmentId" TEXT,
            ADD COLUMN IF NOT EXISTS "boxId" TEXT,
            ADD COLUMN IF NOT EXISTS "packageId" TEXT,
            ADD COLUMN IF NOT EXISTS name TEXT,
            ADD COLUMN IF NOT EXISTS dispatch_date DATE,
            ADD COLUMN IF NOT EXISTS shipment_type TEXT,
            ADD COLUMN IF NOT EXISTS expected_reach_date DATE;
    """))

    conn.execute(text("""
        ALTER TABLE public.inventory_fba_inbound_shipments
            DROP COLUMN IF EXISTS destination_fulfillment_center_id,
            DROP COLUMN IF EXISTS label_prep_type,
            DROP COLUMN IF EXISTS are_cases_required,
            DROP COLUMN IF EXISTS confirmed_need_by_date,
            DROP COLUMN IF EXISTS box_contents_source,
            DROP COLUMN IF EXISTS ship_from_name,
            DROP COLUMN IF EXISTS ship_from_address_line1,
            DROP COLUMN IF EXISTS ship_from_address_line2,
            DROP COLUMN IF EXISTS ship_from_district_or_county,
            DROP COLUMN IF EXISTS ship_from_city,
            DROP COLUMN IF EXISTS ship_from_state_or_province_code,
            DROP COLUMN IF EXISTS ship_from_country_code,
            DROP COLUMN IF EXISTS ship_from_postal_code,
            DROP COLUMN IF EXISTS estimated_box_contents_total_units,
            DROP COLUMN IF EXISTS fee_per_unit_currency_code,
            DROP COLUMN IF EXISTS fee_per_unit_value,
            DROP COLUMN IF EXISTS total_fee_currency_code,
            DROP COLUMN IF EXISTS total_fee_value,
            DROP COLUMN IF EXISTS shipment_json,
            DROP COLUMN IF EXISTS shipment_item_json;
    """))

    # The old table used one row per shipment. Replace that unique key with
    # shipment + SKU so every SKU quantity can be stored separately.
    conn.execute(text("""
        ALTER TABLE public.inventory_fba_inbound_shipments
            DROP CONSTRAINT IF EXISTS uq_inventory_fba_inbound_shipments;
    """))
    conn.execute(text("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_inventory_fba_inbound_shipment_sku
        ON public.inventory_fba_inbound_shipments
            (user_id, marketplace_id, shipment_id, seller_sku);
    """))


def _parse_fba_inbound_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _to_fba_decimal(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def _extract_fba_shipment_items(payload_data) -> list[dict]:
    if not isinstance(payload_data, dict):
        return []
    items = (
        payload_data.get("ItemData")
        or payload_data.get("ShipmentItemData")
        or payload_data.get("items")
        or []
    )
    return [x for x in items if isinstance(x, dict)] if isinstance(items, list) else []


def _fetch_fba_shipment_items(*, shipment_id: str, marketplace_id: str) -> tuple[list[dict], int]:
    endpoint = f"/fba/inbound/v0/shipments/{shipment_id}/items"
    all_items: list[dict] = []
    next_token = None
    pages = 0
    seen_tokens: set[str] = set()

    while True:
        params = {"MarketplaceId": marketplace_id}
        if next_token:
            params["NextToken"] = next_token

        response = amazon_client.make_api_call(endpoint, "GET", params)
        if not response:
            raise RuntimeError(f"Amazon returned an empty item response for shipment {shipment_id}")
        if response.get("error"):
            raise RuntimeError(json.dumps(response, default=str))

        payload_data = response.get("payload") or response
        all_items.extend(_extract_fba_shipment_items(payload_data))
        pages += 1

        returned_next_token = payload_data.get("NextToken") if isinstance(payload_data, dict) else None
        if not returned_next_token or str(returned_next_token) in seen_tokens:
            break
        seen_tokens.add(str(returned_next_token))
        next_token = returned_next_token

    return all_items, pages


def _save_fba_inbound_shipment_items(
    *,
    rows: list[dict],
    user_id: int,
    marketplace_id: str,
) -> int:
    if not rows:
        return 0

    upsert_sql = text("""
        INSERT INTO public.inventory_fba_inbound_shipments (
            user_id, marketplace_id,
            shipment_id, shipment_name,
            shipment_status,
            seller_sku, fulfillment_network_sku,
            quantity_shipped, quantity_received, quantity_in_case,
            synced_at, updated_at
        ) VALUES (
            :user_id, :marketplace_id,
            :shipment_id, :shipment_name,
            :shipment_status,
            :seller_sku, :fulfillment_network_sku,
            :quantity_shipped, :quantity_received, :quantity_in_case,
            NOW(), NOW()
        )
        ON CONFLICT (user_id, marketplace_id, shipment_id, seller_sku)
        DO UPDATE SET
            shipment_name = EXCLUDED.shipment_name,
            shipment_status = EXCLUDED.shipment_status,
            fulfillment_network_sku = EXCLUDED.fulfillment_network_sku,
            quantity_shipped = EXCLUDED.quantity_shipped,
            quantity_received = EXCLUDED.quantity_received,
            quantity_in_case = EXCLUDED.quantity_in_case,
            synced_at = NOW(),
            updated_at = NOW();
    """)

    saved = 0
    with amazon_conn() as conn:
        _ensure_inventory_fba_inbound_shipments_table(conn)

        for row in rows:
            shipment = row.get("shipment") or {}
            item = row.get("item") or {}
            shipment_id = str(shipment.get("ShipmentId") or "").strip()
            seller_sku = str(item.get("SellerSKU") or "").strip()
            if not shipment_id or not seller_sku:
                continue

            conn.execute(upsert_sql, {
                "user_id": int(user_id),
                "marketplace_id": marketplace_id,
                "shipment_id": shipment_id,
                "shipment_name": shipment.get("ShipmentName"),
                "shipment_status": shipment.get("ShipmentStatus"),
                "seller_sku": seller_sku,
                "fulfillment_network_sku": item.get("FulfillmentNetworkSKU"),
                "quantity_shipped": _safe_int(item.get("QuantityShipped")),
                "quantity_received": _safe_int(item.get("QuantityReceived")),
                "quantity_in_case": _safe_int(item.get("QuantityInCase")),
            })
            saved += 1

    return saved


@inventory_bp.route("/amazon_api/fba/inbound-shipments", methods=["GET"])
def get_fba_inbound_shipments():
    try:
        return _get_fba_inbound_shipments_impl()
    except Exception as exc:
        logger.exception("Unhandled FBA inbound shipments route failure")
        return jsonify({
            "success": False,
            "error": "FBA inbound shipments request failed",
            "detail": str(exc),
        }), 500


def _get_fba_inbound_shipments_impl():
    """Fetch WORKING/CLOSED FBA shipments and their SKU quantities.

    Default behavior:
      ShipmentStatusList = WORKING,CLOSED
      One DB row per shipment + SellerSKU
    """
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({"success": False, "error": "Missing Authorization header"}), 401

    token = auth_header.split(" ", 1)[1].strip()
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
        user_id = int(user_id or payload.get("user_id"))
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401
    except Exception:
        return jsonify({"success": False, "error": "Invalid token payload"}), 401

    _apply_region_and_marketplace_from_request()
    marketplace_id = request.args.get("marketplace_id") or amazon_client.marketplace_id
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    store_in_db = request.args.get("store_in_db", "true").strip().lower() != "false"
    include_items_arg = request.args.get("include_items")
    include_items = (
        include_items_arg.strip().lower() in {"1", "true", "yes", "y"}
        if include_items_arg is not None
        else not store_in_db
    )
    status_values = [
        x.strip().upper()
        for x in (request.args.get("shipment_statuses") or "WORKING,CLOSED").split(",")
        if x.strip()
    ]
    allowed_statuses = {
        "WORKING", "READY_TO_SHIP", "SHIPPED", "RECEIVING", "CANCELLED",
        "DELETED", "CLOSED", "ERROR", "IN_TRANSIT", "DELIVERED", "CHECKED_IN",
    }
    invalid_statuses = [s for s in status_values if s not in allowed_statuses]
    if invalid_statuses:
        return jsonify({
            "success": False,
            "error": "Invalid shipment status",
            "invalid_statuses": invalid_statuses,
        }), 400

    endpoint = "/fba/inbound/v0/shipments"
    all_shipments: list[dict] = []
    next_token = None
    shipment_pages = 0
    seen_tokens: set[str] = set()

    while True:
        params = {
            "QueryType": "NEXT_TOKEN" if next_token else "SHIPMENT",
            "MarketplaceId": marketplace_id,
        }
        if next_token:
            params["NextToken"] = next_token
        else:
            params["ShipmentStatusList"] = ",".join(status_values)

        response = amazon_client.make_api_call(endpoint, "GET", params)
        if not response:
            return jsonify({"success": False, "error": "Amazon returned an empty response"}), 502
        if response.get("error"):
            return jsonify({
                "success": False,
                "error": "Amazon could not fetch FBA inbound shipments",
                "amazon_error": response,
            }), int(response.get("status_code") or 502)

        payload_data = response.get("payload") or response
        page_shipments = payload_data.get("ShipmentData") or []
        if isinstance(page_shipments, list):
            all_shipments.extend(x for x in page_shipments if isinstance(x, dict))
        shipment_pages += 1

        returned_next_token = payload_data.get("NextToken")
        if not returned_next_token or str(returned_next_token) in seen_tokens:
            break
        seen_tokens.add(str(returned_next_token))
        next_token = returned_next_token

    deduped_shipments = []
    seen_ids = set()
    for shipment in all_shipments:
        shipment_id = str(shipment.get("ShipmentId") or "").strip()
        if shipment_id and shipment_id not in seen_ids:
            seen_ids.add(shipment_id)
            deduped_shipments.append(shipment)

    complete_rows: list[dict] = []
    shipment_results: list[dict] = []
    item_errors: list[dict] = []
    item_pages_total = 0

    for shipment in deduped_shipments:
        shipment_id = str(shipment.get("ShipmentId") or "").strip()
        try:
            shipment_items, item_pages = _fetch_fba_shipment_items(
                shipment_id=shipment_id,
                marketplace_id=marketplace_id,
            )
            item_pages_total += item_pages
        except Exception as exc:
            logger.exception("Failed to fetch items for FBA shipment %s", shipment_id)
            shipment_items = []
            item_errors.append({"shipment_id": shipment_id, "error": str(exc)})

        quantity_shipped_total = sum(_safe_int(x.get("QuantityShipped")) for x in shipment_items)
        quantity_received_total = sum(_safe_int(x.get("QuantityReceived")) for x in shipment_items)

        shipment_results.append({
            **shipment,
            "ShipmentItems": shipment_items,
            "SkuCount": len(shipment_items),
            "QuantityShippedTotal": quantity_shipped_total,
            "QuantityReceivedTotal": quantity_received_total,
        })

        for item in shipment_items:
            if str(item.get("SellerSKU") or "").strip():
                complete_rows.append({"shipment": shipment, "item": item})

    db_result = {
        "table": "public.inventory_fba_inbound_shipments",
        "saved_rows": 0,
    }
    if store_in_db and complete_rows:
        try:
            db_result["saved_rows"] = _save_fba_inbound_shipment_items(
                rows=complete_rows,
                user_id=user_id,
                marketplace_id=marketplace_id,
            )
        except Exception as exc:
            logger.exception("Failed to save FBA inbound shipment items")
            return jsonify({
                "success": False,
                "error": "Amazon data was fetched but database save failed",
                "detail": str(exc),
                "shipment_count": len(shipment_results),
                "sku_row_count": len(complete_rows),
                "item_errors": item_errors,
                "db": db_result,
            }), 500

    response_payload = {
        "success": True,
        "source": "fulfillment-inbound-v0-getShipments+getShipmentItemsByShipmentId",
        "marketplace_id": marketplace_id,
        "shipment_statuses": status_values,
        "store_in_db": store_in_db,
        "include_items": include_items,
        "shipment_pages_fetched": shipment_pages,
        "item_pages_fetched": item_pages_total,
        "shipment_count": len(shipment_results),
        "sku_row_count": len(complete_rows),
        "quantity_shipped_total": sum(x.get("QuantityShippedTotal", 0) for x in shipment_results),
        "quantity_received_total": sum(x.get("QuantityReceivedTotal", 0) for x in shipment_results),
        "item_errors": item_errors,
        "db": db_result,
    }
    if include_items:
        response_payload["items"] = shipment_results
    else:
        response_payload["items_preview"] = shipment_results[:5]

    return jsonify(response_payload), 200


# =============================================================================
# FBA INBOUND PLAN SHIPMENT BOXES (Fulfillment Inbound v2024-03-20)
# Docs:
# GET /inbound/fba/2024-03-20/inboundPlans/{inboundPlanId}/shipments/{shipmentId}/boxes
# =============================================================================

def _extract_fba_2024_next_token(payload_data) -> str | None:
    if not isinstance(payload_data, dict):
        return None
    pagination = payload_data.get("pagination") or {}
    if isinstance(pagination, dict):
        return pagination.get("nextToken")
    return None


def _fetch_fba_2024_shipment_boxes(
    *,
    inbound_plan_id: str,
    shipment_id: str,
    page_size: int,
    pagination_token: str | None = None,
    fetch_all: bool = True,
) -> tuple[list[dict], int, str | None]:
    endpoint = (
        f"/inbound/fba/2024-03-20/inboundPlans/{inbound_plan_id}"
        f"/shipments/{shipment_id}/boxes"
    )
    boxes: list[dict] = []
    pages_fetched = 0
    next_token = pagination_token
    seen_tokens: set[str] = set()

    while True:
        params = {"pageSize": page_size}
        if next_token:
            params["paginationToken"] = next_token

        response = amazon_client.make_api_call(endpoint, "GET", params)
        if not response:
            raise RuntimeError("Amazon returned an empty shipment-boxes response")
        if response.get("error"):
            raise RuntimeError(json.dumps(response, default=str))

        payload_data = response.get("payload") or response
        page_boxes = payload_data.get("boxes") if isinstance(payload_data, dict) else []
        if isinstance(page_boxes, list):
            boxes.extend(x for x in page_boxes if isinstance(x, dict))

        pages_fetched += 1
        returned_next_token = _extract_fba_2024_next_token(payload_data)

        if not fetch_all or not returned_next_token:
            next_token = returned_next_token
            break

        token_key = str(returned_next_token)
        if token_key in seen_tokens:
            logger.warning("Stopping FBA 2024 shipment-box pagination because nextToken repeated")
            next_token = returned_next_token
            break

        seen_tokens.add(token_key)
        next_token = returned_next_token

    return boxes, pages_fetched, next_token


@inventory_bp.route("/amazon_api/fba/shipment-boxes", methods=["GET"])
def get_fba_2024_shipment_boxes():
    """List boxes for one Fulfillment Inbound v2024 shipment.

    Required query params:
      inbound_plan_id or inboundPlanId
      shipment_id or shipmentId

    Optional:
      marketplace_id=<id>  # chooses SP-API region
      page_size=1..1000
      pagination_token=<token>
      fetch_all=true/false
    """
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({"success": False, "error": "Missing Authorization header"}), 401

    token = auth_header.split(" ", 1)[1].strip()
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401
    except Exception:
        return jsonify({"success": False, "error": "Invalid token payload"}), 401

    _apply_region_and_marketplace_from_request()
    marketplace_id = request.args.get("marketplace_id") or amazon_client.marketplace_id
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    inbound_plan_id = (
        request.args.get("inbound_plan_id")
        or request.args.get("inboundPlanId")
        or ""
    ).strip()
    shipment_id = (
        request.args.get("shipment_id")
        or request.args.get("shipmentId")
        or ""
    ).strip()

    if not inbound_plan_id or not shipment_id:
        return jsonify({
            "success": False,
            "error": "inbound_plan_id and shipment_id are required",
        }), 400

    try:
        page_size = int(request.args.get("page_size") or request.args.get("pageSize") or 1000)
        if page_size < 1 or page_size > 1000:
            raise ValueError
    except ValueError:
        return jsonify({"success": False, "error": "page_size must be between 1 and 1000"}), 400

    pagination_token = (
        request.args.get("pagination_token")
        or request.args.get("paginationToken")
        or None
    )
    fetch_all = request.args.get("fetch_all", "true").strip().lower() != "false"

    try:
        boxes, pages_fetched, next_token = _fetch_fba_2024_shipment_boxes(
            inbound_plan_id=inbound_plan_id,
            shipment_id=shipment_id,
            page_size=page_size,
            pagination_token=pagination_token,
            fetch_all=fetch_all,
        )
    except Exception as exc:
        logger.exception("Failed to fetch FBA 2024 shipment boxes")
        return jsonify({
            "success": False,
            "error": "Amazon could not fetch FBA shipment boxes",
            "detail": str(exc),
        }), 502

    return jsonify({
        "success": True,
        "source": "fulfillment-inbound-v2024-03-20-listShipmentBoxes",
        "marketplace_id": marketplace_id,
        "inbound_plan_id": inbound_plan_id,
        "shipment_id": shipment_id,
        "page_size": page_size,
        "fetch_all": fetch_all,
        "pages_fetched": pages_fetched,
        "box_count": len(boxes),
        "next_token": next_token,
        "boxes": boxes,
    }), 200


def _fetch_fba_2024_collection(
    *,
    endpoint: str,
    collection_keys: list[str],
    page_size: int,
    pagination_token: str | None = None,
    fetch_all: bool = True,
    extra_params: dict | None = None,
) -> tuple[list[dict], int, str | None]:
    items: list[dict] = []
    pages_fetched = 0
    next_token = pagination_token
    seen_tokens: set[str] = set()

    while True:
        params = dict(extra_params or {})
        params["pageSize"] = page_size
        if next_token:
            params["paginationToken"] = next_token

        response = amazon_client.make_api_call(endpoint, "GET", params)
        if not response:
            raise RuntimeError(f"Amazon returned an empty response for {endpoint}")
        if response.get("error"):
            raise RuntimeError(json.dumps(response, default=str))

        payload_data = response.get("payload") or response
        page_items = []
        if isinstance(payload_data, dict):
            for collection_key in collection_keys:
                value = payload_data.get(collection_key)
                if isinstance(value, list):
                    page_items = value
                    break

        items.extend(x for x in page_items if isinstance(x, dict))
        pages_fetched += 1

        returned_next_token = _extract_fba_2024_next_token(payload_data)
        if not fetch_all or not returned_next_token:
            next_token = returned_next_token
            break

        token_key = str(returned_next_token)
        if token_key in seen_tokens:
            logger.warning("Stopping FBA 2024 pagination for %s because nextToken repeated", endpoint)
            next_token = returned_next_token
            break

        seen_tokens.add(token_key)
        next_token = returned_next_token

    return items, pages_fetched, next_token


def _fetch_fba_2024_object(endpoint: str) -> dict:
    response = amazon_client.make_api_call(endpoint, "GET", {})
    if not response:
        raise RuntimeError(f"Amazon returned an empty response for {endpoint}")
    if response.get("error"):
        raise RuntimeError(json.dumps(response, default=str))

    payload_data = response.get("payload") or response
    return payload_data if isinstance(payload_data, dict) else {"data": payload_data}


def _find_fba_2024_ids(obj, key_names: set[str]) -> list[str]:
    found: list[str] = []

    def walk(value):
        if isinstance(value, dict):
            for key, nested in value.items():
                if key in key_names and nested:
                    found.append(str(nested).strip())
                walk(nested)
        elif isinstance(value, list):
            for nested in value:
                walk(nested)

    walk(obj)
    return list(dict.fromkeys(x for x in found if x))


def _fba_2024_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, default=str)
    return value


def _fba_2024_append_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title[:31])
    if not rows:
        ws.append(["message"])
        ws.append(["No data"])
        return

    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    ws.append(headers)
    for row in rows:
        ws.append([_fba_2024_excel_value(row.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor="5EA88B")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9E2E7"),
        right=Side(style="thin", color="D9E2E7"),
        top=Side(style="thin", color="D9E2E7"),
        bottom=Side(style="thin", color="D9E2E7"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells[:100]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 45)


def _fba_2024_flatten_plan_summary(plan: dict) -> dict:
    summary = plan.get("summary") or {}
    detail = plan.get("detail") or {}
    return {
        "inboundPlanId": plan.get("inboundPlanId"),
        "status": detail.get("status") or summary.get("status"),
        "name": detail.get("name") or summary.get("name"),
        "createdAt": detail.get("createdAt") or summary.get("createdAt"),
        "lastUpdatedAt": detail.get("lastUpdatedAt") or summary.get("lastUpdatedAt"),
        "marketplaceIds": detail.get("marketplaceIds") or summary.get("marketplaceIds"),
        "shipment_count": len(plan.get("shipments") or []),
        "plan_item_count": len(plan.get("plan_items") or []),
        "plan_box_count": len(plan.get("plan_boxes") or []),
        "error_count": len(plan.get("errors") or []),
        "sourceAddress": detail.get("sourceAddress"),
        "summary_json": summary,
        "detail_json": detail,
    }


def _fba_2024_flatten_box(inbound_plan_id: str, box: dict, shipment_id: str = "") -> dict:
    dimensions = box.get("dimensions") or {}
    weight = box.get("weight") or {}
    return {
        "inboundPlanId": inbound_plan_id,
        "shipmentId": shipment_id,
        "boxId": box.get("boxId"),
        "packageId": box.get("packageId"),
        "externalContainerIdentifier": box.get("externalContainerIdentifier"),
        "externalContainerIdentifierType": box.get("externalContainerIdentifierType"),
        "contentInformationSource": box.get("contentInformationSource"),
        "templateName": box.get("templateName"),
        "box_quantity": box.get("quantity"),
        "length": dimensions.get("length"),
        "width": dimensions.get("width"),
        "height": dimensions.get("height"),
        "dimension_unit": dimensions.get("unitOfMeasurement"),
        "weight_value": weight.get("value"),
        "weight_unit": weight.get("unit"),
        "item_count": len(box.get("items") or []),
        "box_json": box,
    }


def _fba_2024_flatten_box_item(
    inbound_plan_id: str,
    box: dict,
    item: dict,
    shipment_id: str = "",
) -> dict:
    return {
        "inboundPlanId": inbound_plan_id,
        "shipmentId": shipment_id,
        "boxId": box.get("boxId"),
        "packageId": box.get("packageId"),
        "msku": item.get("msku"),
        "asin": item.get("asin"),
        "fnsku": item.get("fnsku"),
        "quantity": item.get("quantity"),
        "expiration": item.get("expiration"),
        "labelOwner": item.get("labelOwner"),
        "manufacturingLotCode": item.get("manufacturingLotCode"),
        "prepInstructions": item.get("prepInstructions"),
        "item_json": item,
    }


def _first_non_empty(*values):
    for value in values:
        if value is not None and str(value).strip() != "":
            return value
    return None


def _fba_2024_item_identity(item: dict) -> tuple[str, str, str]:
    msku = str(item.get("msku") or item.get("sellerSku") or item.get("SellerSKU") or "").strip()
    fnsku = str(item.get("fnsku") or item.get("fulfillmentNetworkSku") or item.get("FulfillmentNetworkSKU") or "").strip()
    asin = str(item.get("asin") or item.get("ASIN") or "").strip()
    return msku, fnsku, asin


def _collect_fba_2024_inbound_plan_db_rows(plans: list[dict]) -> list[dict]:
    rows: list[dict] = []

    for plan in plans:
        inbound_plan_id = str(plan.get("inboundPlanId") or "").strip()
        summary = plan.get("summary") or {}
        detail = plan.get("detail") or {}
        status = _first_non_empty(detail.get("status"), summary.get("status"))
        created_at = _first_non_empty(detail.get("createdAt"), summary.get("createdAt"))
        last_updated_at = _first_non_empty(detail.get("lastUpdatedAt"), summary.get("lastUpdatedAt"))
        marketplace_ids = _first_non_empty(detail.get("marketplaceIds"), summary.get("marketplaceIds"), [])
        name = _first_non_empty(detail.get("name"), summary.get("name"))
        shipment_count = len(plan.get("shipments") or [])
        plan_box_count = len(plan.get("plan_boxes") or [])

        def append_item_row(
            item: dict,
            *,
            shipment_id: str = "",
            box: dict | None = None,
            row_name=None,
            row_status=None,
            row_created_at=None,
            row_last_updated_at=None,
        ):
            if not isinstance(item, dict):
                return

            msku, fnsku, asin = _fba_2024_item_identity(item)
            if not msku:
                return

            box = box or {}
            rows.append({
                "inboundPlanId": inbound_plan_id,
                "status": _first_non_empty(row_status, status),
                "createdAt": _first_non_empty(row_created_at, created_at),
                "lastUpdatedAt": _first_non_empty(row_last_updated_at, last_updated_at),
                "marketplaceIds": marketplace_ids,
                "shipment_count": shipment_count,
                "plan_box_count": plan_box_count,
                "msku": msku,
                "fnsku": fnsku,
                "asin": asin,
                "quantity": _safe_int(item.get("quantity")),
                "shipmentId": str(shipment_id or "").strip(),
                "boxId": str(box.get("boxId") or "").strip(),
                "packageId": str(box.get("packageId") or "").strip(),
                "name": _first_non_empty(row_name, name),
            })

        for shipment in plan.get("shipments") or []:
            shipment_id = str(shipment.get("shipmentId") or "").strip()
            shipment_detail = shipment.get("detail") or {}
            shipment_status = shipment_detail.get("status")
            shipment_name = shipment_detail.get("name")
            shipment_created_at = shipment_detail.get("createdAt")
            shipment_last_updated_at = shipment_detail.get("lastUpdatedAt")
            before_shipment_rows = len(rows)

            for box in shipment.get("boxes") or []:
                for item in box.get("items") or []:
                    append_item_row(
                        item,
                        shipment_id=shipment_id,
                        box=box,
                        row_name=shipment_name,
                        row_status=shipment_status,
                        row_created_at=shipment_created_at,
                        row_last_updated_at=shipment_last_updated_at,
                    )

            if len(rows) == before_shipment_rows:
                for item in shipment.get("items") or []:
                    append_item_row(
                        item,
                        shipment_id=shipment_id,
                        row_name=shipment_name,
                        row_status=shipment_status,
                        row_created_at=shipment_created_at,
                        row_last_updated_at=shipment_last_updated_at,
                    )

        if not any(str(row.get("inboundPlanId") or "") == inbound_plan_id for row in rows):
            for box in plan.get("plan_boxes") or []:
                for item in box.get("items") or []:
                    append_item_row(item, box=box)

        if not any(str(row.get("inboundPlanId") or "") == inbound_plan_id for row in rows):
            for item in plan.get("plan_items") or []:
                append_item_row(item)

    aggregated: dict[tuple[str, str, str], dict] = {}
    for row in rows:
        shipment_id = row.get("shipmentId") or row.get("inboundPlanId") or ""
        key = (str(row.get("inboundPlanId") or ""), str(shipment_id), str(row.get("msku") or ""))
        current = aggregated.get(key)
        if not current:
            current = {**row, "shipmentId": str(shipment_id)}
            aggregated[key] = current
            continue

        current["quantity"] = _safe_int(current.get("quantity")) + _safe_int(row.get("quantity"))
        for col in ("boxId", "packageId"):
            existing = [x.strip() for x in str(current.get(col) or "").split(",") if x.strip()]
            candidate = str(row.get(col) or "").strip()
            if candidate and candidate not in existing:
                existing.append(candidate)
            current[col] = ", ".join(existing)

    return list(aggregated.values())


def _save_fba_2024_inbound_plan_rows(
    *,
    rows: list[dict],
    user_id: int,
    marketplace_id: str,
) -> int:
    if not rows:
        return 0

    upsert_sql = text("""
        INSERT INTO public.inventory_fba_inbound_shipments (
            user_id, marketplace_id,
            shipment_id, shipment_status, shipment_name,
            seller_sku, fulfillment_network_sku, quantity_shipped,
            "inboundPlanId", status, "createdAt", "lastUpdatedAt", "marketplaceIds",
            shipment_count, plan_box_count,
            msku, fnsku, asin, quantity,
            "shipmentId", "boxId", "packageId", name,
            synced_at, created_at, updated_at
        ) VALUES (
            :user_id, :marketplace_id,
            :shipment_id, :status, :name,
            :msku, :fnsku, :quantity,
            :inboundPlanId, :status, :createdAt, :lastUpdatedAt, CAST(:marketplaceIds AS JSONB),
            :shipment_count, :plan_box_count,
            :msku, :fnsku, :asin, :quantity,
            :shipmentId, :boxId, :packageId, :name,
            NOW(),
            COALESCE(CAST(NULLIF(:createdAt, '') AS TIMESTAMPTZ), NOW()),
            COALESCE(
                CAST(NULLIF(:lastUpdatedAt, '') AS TIMESTAMPTZ),
                CAST(NULLIF(:createdAt, '') AS TIMESTAMPTZ),
                NOW()
            )
        )
        ON CONFLICT (user_id, marketplace_id, shipment_id, seller_sku)
        DO UPDATE SET
            shipment_status = EXCLUDED.shipment_status,
            shipment_name = EXCLUDED.shipment_name,
            fulfillment_network_sku = EXCLUDED.fulfillment_network_sku,
            quantity_shipped = EXCLUDED.quantity_shipped,
            "inboundPlanId" = EXCLUDED."inboundPlanId",
            status = EXCLUDED.status,
            "createdAt" = EXCLUDED."createdAt",
            "lastUpdatedAt" = EXCLUDED."lastUpdatedAt",
            "marketplaceIds" = EXCLUDED."marketplaceIds",
            shipment_count = EXCLUDED.shipment_count,
            plan_box_count = EXCLUDED.plan_box_count,
            msku = EXCLUDED.msku,
            fnsku = EXCLUDED.fnsku,
            asin = EXCLUDED.asin,
            quantity = EXCLUDED.quantity,
            "shipmentId" = EXCLUDED."shipmentId",
            "boxId" = EXCLUDED."boxId",
            "packageId" = EXCLUDED."packageId",
            name = EXCLUDED.name,
            synced_at = NOW(),
            created_at = EXCLUDED.created_at,
            updated_at = EXCLUDED.updated_at;
    """)

    saved = 0
    with amazon_conn() as conn:
        _ensure_inventory_fba_inbound_shipments_table(conn)
        for row in rows:
            shipment_id = str(row.get("shipmentId") or row.get("inboundPlanId") or "").strip()
            msku = str(row.get("msku") or "").strip()
            if not shipment_id or not msku:
                continue

            conn.execute(upsert_sql, {
                "user_id": int(user_id),
                "marketplace_id": marketplace_id,
                "shipment_id": shipment_id,
                "inboundPlanId": row.get("inboundPlanId"),
                "status": row.get("status"),
                "createdAt": row.get("createdAt"),
                "lastUpdatedAt": row.get("lastUpdatedAt"),
                "marketplaceIds": json.dumps(row.get("marketplaceIds") or [], default=str),
                "shipment_count": _safe_int(row.get("shipment_count")),
                "plan_box_count": _safe_int(row.get("plan_box_count")),
                "msku": msku,
                "fnsku": row.get("fnsku"),
                "asin": row.get("asin"),
                "quantity": _safe_int(row.get("quantity")),
                "shipmentId": row.get("shipmentId"),
                "boxId": row.get("boxId"),
                "packageId": row.get("packageId"),
                "name": row.get("name"),
            })
            saved += 1

    return saved


def _build_fba_2024_inbound_plans_excel(
    *,
    marketplace_id: str,
    statuses: list[str],
    plans: list[dict],
    list_errors: list[dict],
    detail_errors: list[dict],
    totals: dict,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["Marketplace ID", marketplace_id],
        ["Statuses", ", ".join(statuses)],
        ["Plan Count", len(plans)],
        ["Plan Items", totals.get("plan_items", 0)],
        ["Plan Boxes", totals.get("plan_boxes", 0)],
        ["Shipments", totals.get("shipments", 0)],
        ["Shipment Items", totals.get("shipment_items", 0)],
        ["Shipment Boxes", totals.get("shipment_boxes", 0)],
        ["List Error Count", len(list_errors)],
        ["Detail Error Count", len(detail_errors)],
        ["Generated At", datetime.utcnow().isoformat()],
    ]
    for row in summary_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40

    plan_rows = []
    plan_item_rows = []
    plan_box_rows = []
    plan_box_item_rows = []
    shipment_rows = []
    shipment_item_rows = []
    shipment_box_rows = []
    shipment_box_item_rows = []
    error_rows = []

    for err in list_errors:
        error_rows.append({
            "scope": "listInboundPlans",
            "inboundPlanId": "",
            "shipmentId": "",
            "operation": "listInboundPlans",
            "error": err.get("error"),
            "error_json": err,
        })

    for detail_error in detail_errors:
        for err in detail_error.get("errors") or []:
            error_rows.append({
                "scope": "inboundPlan",
                "inboundPlanId": detail_error.get("inboundPlanId"),
                "shipmentId": "",
                "operation": err.get("operation"),
                "error": err.get("error"),
                "error_json": err,
            })

    for plan in plans:
        inbound_plan_id = str(plan.get("inboundPlanId") or "")
        plan_rows.append(_fba_2024_flatten_plan_summary(plan))

        for item in plan.get("plan_items") or []:
            row = {"inboundPlanId": inbound_plan_id}
            row.update(item)
            row["item_json"] = item
            plan_item_rows.append(row)

        for box in plan.get("plan_boxes") or []:
            plan_box_rows.append(_fba_2024_flatten_box(inbound_plan_id, box))
            for item in box.get("items") or []:
                plan_box_item_rows.append(_fba_2024_flatten_box_item(inbound_plan_id, box, item))

        for shipment in plan.get("shipments") or []:
            shipment_id = str(shipment.get("shipmentId") or "")
            detail = shipment.get("detail") or {}
            shipment_rows.append({
                "inboundPlanId": inbound_plan_id,
                "shipmentId": shipment_id,
                "status": detail.get("status"),
                "name": detail.get("name"),
                "createdAt": detail.get("createdAt"),
                "lastUpdatedAt": detail.get("lastUpdatedAt"),
                "item_count": len(shipment.get("items") or []),
                "box_count": len(shipment.get("boxes") or []),
                "error_count": len(shipment.get("errors") or []),
                "detail_json": detail,
            })

            for err in shipment.get("errors") or []:
                error_rows.append({
                    "scope": "shipment",
                    "inboundPlanId": inbound_plan_id,
                    "shipmentId": shipment_id,
                    "operation": err.get("operation"),
                    "error": err.get("error"),
                    "error_json": err,
                })

            for item in shipment.get("items") or []:
                row = {"inboundPlanId": inbound_plan_id, "shipmentId": shipment_id}
                row.update(item)
                row["item_json"] = item
                shipment_item_rows.append(row)

            for box in shipment.get("boxes") or []:
                shipment_box_rows.append(_fba_2024_flatten_box(inbound_plan_id, box, shipment_id))
                for item in box.get("items") or []:
                    shipment_box_item_rows.append(
                        _fba_2024_flatten_box_item(inbound_plan_id, box, item, shipment_id)
                    )

    _fba_2024_append_sheet(wb, "Plans", plan_rows)
    _fba_2024_append_sheet(wb, "Plan Items", plan_item_rows)
    _fba_2024_append_sheet(wb, "Plan Boxes", plan_box_rows)
    _fba_2024_append_sheet(wb, "Plan Box Items", plan_box_item_rows)
    _fba_2024_append_sheet(wb, "Shipments", shipment_rows)
    _fba_2024_append_sheet(wb, "Shipment Items", shipment_item_rows)
    _fba_2024_append_sheet(wb, "Shipment Boxes", shipment_box_rows)
    _fba_2024_append_sheet(wb, "Shipment Box Items", shipment_box_item_rows)
    _fba_2024_append_sheet(wb, "Errors", error_rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@inventory_bp.route("/amazon_api/fba/inbound-plans-all", methods=["GET"])
def get_fba_2024_inbound_plans_all():
    """Fetch v2024 inbound plans and related Amazon-side data.

    This is a crawler over Amazon GET operations, not a single Amazon API:
      listInboundPlans -> getInboundPlan -> list plan items/boxes
      -> get shipment/list shipment items/list shipment boxes when shipment IDs exist.
    """
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return jsonify({"success": False, "error": "Missing Authorization header"}), 401

    token = auth_header.split(" ", 1)[1].strip()
    try:
        payload, user_id, member_id = get_effective_user_id_from_token(token)
    except jwt.ExpiredSignatureError:
        return jsonify({"success": False, "error": "Token has expired"}), 401
    except jwt.InvalidTokenError:
        return jsonify({"success": False, "error": "Invalid token"}), 401
    except Exception:
        return jsonify({"success": False, "error": "Invalid token payload"}), 401

    _apply_region_and_marketplace_from_request()
    marketplace_id = request.args.get("marketplace_id") or amazon_client.marketplace_id
    if marketplace_id not in amazon_client.ALLOWED_MARKETPLACES:
        return jsonify({"success": False, "error": "Unsupported marketplace"}), 400

    status_values = [
        x.strip().upper()
        for x in (request.args.get("statuses") or request.args.get("status") or "ACTIVE,SHIPPED").split(",")
        if x.strip()
    ]
    allowed_statuses = {"ACTIVE", "VOIDED", "SHIPPED"}
    invalid_statuses = [x for x in status_values if x not in allowed_statuses]
    if invalid_statuses:
        return jsonify({
            "success": False,
            "error": "Invalid inbound plan status",
            "invalid_statuses": invalid_statuses,
            "allowed_statuses": sorted(allowed_statuses),
        }), 400

    try:
        plan_page_size = int(request.args.get("plan_page_size") or request.args.get("page_size") or 30)
        page_size = int(request.args.get("detail_page_size") or 1000)
        max_plans = int(request.args.get("max_plans") or 10)
        delay_seconds = float(request.args.get("delay_seconds") or 0.55)
        if plan_page_size < 1 or plan_page_size > 30:
            raise ValueError("plan_page_size")
        if page_size < 1 or page_size > 1000:
            raise ValueError("detail_page_size")
        if max_plans < 1 or max_plans > 200:
            raise ValueError("max_plans")
        if delay_seconds < 0:
            raise ValueError("delay_seconds")
    except ValueError as exc:
        return jsonify({
            "success": False,
            "error": "Invalid paging/limit parameter",
            "detail": str(exc),
        }), 400

    fetch_all_pages = request.args.get("fetch_all_pages", "true").strip().lower() != "false"
    include_plan_items = request.args.get("include_plan_items", "true").strip().lower() != "false"
    include_plan_boxes = request.args.get("include_plan_boxes", "true").strip().lower() != "false"
    include_shipments = request.args.get("include_shipments", "true").strip().lower() != "false"
    include_shipment_items = request.args.get("include_shipment_items", "true").strip().lower() != "false"
    include_shipment_boxes = request.args.get("include_shipment_boxes", "true").strip().lower() != "false"
    store_in_db = request.args.get("store_in_db", "true").strip().lower() != "false"

    sort_by = (request.args.get("sort_by") or request.args.get("sortBy") or "LAST_UPDATED_TIME").strip().upper()
    sort_order = (request.args.get("sort_order") or request.args.get("sortOrder") or "DESC").strip().upper()
    if sort_by not in {"LAST_UPDATED_TIME", "CREATION_TIME"}:
        return jsonify({"success": False, "error": "sort_by must be LAST_UPDATED_TIME or CREATION_TIME"}), 400
    if sort_order not in {"ASC", "DESC"}:
        return jsonify({"success": False, "error": "sort_order must be ASC or DESC"}), 400

    all_plan_summaries: list[dict] = []
    list_errors: list[dict] = []
    list_pages_fetched = 0

    for status in status_values:
        try:
            summaries, pages, next_token = _fetch_fba_2024_collection(
                endpoint="/inbound/fba/2024-03-20/inboundPlans",
                collection_keys=["inboundPlans", "plans"],
                page_size=plan_page_size,
                fetch_all=fetch_all_pages,
                extra_params={
                    "status": status,
                    "sortBy": sort_by,
                    "sortOrder": sort_order,
                },
            )
            list_pages_fetched += pages
            for summary in summaries:
                summary["_requested_status"] = status
                all_plan_summaries.append(summary)
        except Exception as exc:
            logger.exception("Failed to list FBA 2024 inbound plans for status %s", status)
            list_errors.append({"status": status, "error": str(exc)})

    deduped_summaries: list[dict] = []
    seen_plan_ids: set[str] = set()
    for summary in all_plan_summaries:
        plan_ids = _find_fba_2024_ids(summary, {"inboundPlanId"})
        plan_id = plan_ids[0] if plan_ids else ""
        if not plan_id or plan_id in seen_plan_ids:
            continue
        seen_plan_ids.add(plan_id)
        deduped_summaries.append(summary)
        if len(deduped_summaries) >= max_plans:
            break

    plans: list[dict] = []
    detail_errors: list[dict] = []
    totals = {
        "plan_items": 0,
        "plan_boxes": 0,
        "shipments": 0,
        "shipment_items": 0,
        "shipment_boxes": 0,
    }

    for summary in deduped_summaries:
        plan_id = (_find_fba_2024_ids(summary, {"inboundPlanId"}) or [""])[0]
        plan_record = {
            "inboundPlanId": plan_id,
            "summary": summary,
            "detail": None,
            "plan_items": [],
            "plan_boxes": [],
            "shipments": [],
            "errors": [],
        }

        try:
            plan_record["detail"] = _fetch_fba_2024_object(
                f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}"
            )
        except Exception as exc:
            plan_record["errors"].append({"operation": "getInboundPlan", "error": str(exc)})

        if include_plan_items:
            try:
                plan_items, _, _ = _fetch_fba_2024_collection(
                    endpoint=f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}/items",
                    collection_keys=["items"],
                    page_size=page_size,
                    fetch_all=fetch_all_pages,
                )
                plan_record["plan_items"] = plan_items
                totals["plan_items"] += len(plan_items)
                time.sleep(delay_seconds)
            except Exception as exc:
                plan_record["errors"].append({"operation": "listInboundPlanItems", "error": str(exc)})

        if include_plan_boxes:
            try:
                plan_boxes, _, _ = _fetch_fba_2024_collection(
                    endpoint=f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}/boxes",
                    collection_keys=["boxes"],
                    page_size=page_size,
                    fetch_all=fetch_all_pages,
                )
                plan_record["plan_boxes"] = plan_boxes
                totals["plan_boxes"] += len(plan_boxes)
                time.sleep(delay_seconds)
            except Exception as exc:
                plan_record["errors"].append({"operation": "listInboundPlanBoxes", "error": str(exc)})

        shipment_ids = _find_fba_2024_ids(
            {
                "summary": summary,
                "detail": plan_record.get("detail"),
                "plan_boxes": plan_record.get("plan_boxes"),
                "plan_items": plan_record.get("plan_items"),
            },
            {"shipmentId"},
        )

        if include_shipments:
            for shipment_id in shipment_ids:
                shipment_record = {
                    "shipmentId": shipment_id,
                    "detail": None,
                    "items": [],
                    "boxes": [],
                    "errors": [],
                }
                try:
                    shipment_record["detail"] = _fetch_fba_2024_object(
                        f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}/shipments/{shipment_id}"
                    )
                    time.sleep(delay_seconds)
                except Exception as exc:
                    shipment_record["errors"].append({"operation": "getShipment", "error": str(exc)})

                if include_shipment_items:
                    try:
                        shipment_items, _, _ = _fetch_fba_2024_collection(
                            endpoint=f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}/shipments/{shipment_id}/items",
                            collection_keys=["items"],
                            page_size=page_size,
                            fetch_all=fetch_all_pages,
                        )
                        shipment_record["items"] = shipment_items
                        totals["shipment_items"] += len(shipment_items)
                        time.sleep(delay_seconds)
                    except Exception as exc:
                        shipment_record["errors"].append({"operation": "listShipmentItems", "error": str(exc)})

                if include_shipment_boxes:
                    try:
                        shipment_boxes, _, _ = _fetch_fba_2024_collection(
                            endpoint=f"/inbound/fba/2024-03-20/inboundPlans/{plan_id}/shipments/{shipment_id}/boxes",
                            collection_keys=["boxes"],
                            page_size=page_size,
                            fetch_all=fetch_all_pages,
                        )
                        shipment_record["boxes"] = shipment_boxes
                        totals["shipment_boxes"] += len(shipment_boxes)
                        time.sleep(delay_seconds)
                    except Exception as exc:
                        shipment_record["errors"].append({"operation": "listShipmentBoxes", "error": str(exc)})

                plan_record["shipments"].append(shipment_record)

        totals["shipments"] += len(plan_record["shipments"])
        if plan_record["errors"]:
            detail_errors.append({"inboundPlanId": plan_id, "errors": plan_record["errors"]})
        plans.append(plan_record)
        time.sleep(delay_seconds)

    db_result = {
        "table": "public.inventory_fba_inbound_shipments",
        "saved_rows": 0,
    }
    if store_in_db:
        try:
            db_rows = _collect_fba_2024_inbound_plan_db_rows(plans)
            db_result["saved_rows"] = _save_fba_2024_inbound_plan_rows(
                rows=db_rows,
                user_id=int(user_id or payload.get("user_id")),
                marketplace_id=marketplace_id,
            )
        except Exception as exc:
            logger.exception("Failed to save FBA 2024 inbound plan rows")
            return jsonify({
                "success": False,
                "error": "Amazon data was fetched but database save failed",
                "detail": str(exc),
                "marketplace_id": marketplace_id,
                "statuses": status_values,
                "plan_count": len(plans),
                "totals": totals,
                "db": db_result,
            }), 500

    export_format = (request.args.get("format") or "").strip().lower()
    download_excel = (
        export_format in {"excel", "xlsx"}
        or request.args.get("download", "").strip().lower() in {"1", "true", "yes", "y"}
    )

    if download_excel:
        excel_stream = _build_fba_2024_inbound_plans_excel(
            marketplace_id=marketplace_id,
            statuses=status_values,
            plans=plans,
            list_errors=list_errors,
            detail_errors=detail_errors,
            totals=totals,
        )
        filename = (
            f"FBA_Inbound_Plans_All_{marketplace_id}_"
            f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )

    return jsonify({
        "success": len(list_errors) == 0,
        "source": "fulfillment-inbound-v2024-03-20-all-get-data",
        "marketplace_id": marketplace_id,
        "statuses": status_values,
        "sort_by": sort_by,
        "sort_order": sort_order,
        "fetch_all_pages": fetch_all_pages,
        "limits": {
            "max_plans": max_plans,
            "plan_page_size": plan_page_size,
            "detail_page_size": page_size,
            "delay_seconds": delay_seconds,
        },
        "list_pages_fetched": list_pages_fetched,
        "plan_count": len(plans),
        "totals": totals,
        "store_in_db": store_in_db,
        "db": db_result,
        "list_errors": list_errors,
        "detail_errors": detail_errors,
        "plans": plans,
    }), 200
